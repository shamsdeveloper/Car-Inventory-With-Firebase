import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QLabel, QHBoxLayout, QListWidget, QListWidgetItem, QStackedWidget, QLineEdit,
    QTableWidget, QTableWidgetItem, QTabWidget, QWidget, QVBoxLayout, QFormLayout, QLineEdit, QComboBox, QPushButton, QLabel, 
    QTextEdit, QFileDialog, QHBoxLayout,QScrollArea,QGridLayout,QSpacerItem,QSizePolicy, QDateEdit,QTimeEdit,QRadioButton,QAbstractScrollArea,
    QDialog, QDialogButtonBox,QHeaderView,QMessageBox,QCheckBox
)
from PyQt5.QtGui import QStandardItemModel, QStandardItem,QIcon, QColor
from PyQt5.QtWidgets import QWidget, QFileDialog, QLabel, QTreeView, QVBoxLayout, QAbstractItemView
from tempfile import NamedTemporaryFile
from reportlab.lib.pagesizes import letter, landscape
from fpdf import FPDF
from PIL import Image
from PyQt5.QtGui import QPixmap, QPainter, QBrush, QFont,QIcon,QImage
from PyQt5.QtChart import QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis
from PyQt5.QtCore import Qt,QBuffer
from PyQt5.QtCore import QEvent
from PyQt5.QtCore import Qt, QDate, QTime, QTimer
import random
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtGui import QPainter
import json
##########################################################3
from google.cloud import firestore
import os
import re
import io
import base64
import datetime
from qt_material import apply_stylesheet
##################################################################
#import inflect
from reportlab.lib.pagesizes import A4,landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import openpyxl
from openpyxl import Workbook
from collections import defaultdict
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import matplotlib.pyplot as plt
from reportlab.lib.utils import ImageReader
###################################################################################################################################
# Set environment variable for Firestore credentials
# Determine the path to the serviceAccountKey.json file
if getattr(sys, 'frozen', False):
    # If running in a bundle (exe), get the application path
    application_path = sys._MEIPASS
    service_account_path = os.path.join(application_path, 'serviceAccountKey.json')
else:
    # If running in a normal Python environment, set the path directly
    service_account_path = "serviceAccountKey.json"  # Ensure this path is correct

# Set environment variable for Firestore credentials
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = service_account_path

# Initialize Firestore
db = firestore.Client()
splash = None
##########################################
#################################################################################################################################################################################################################################
class DashboardWindow(QMainWindow):
    def __init__(self,email):
        super().__init__()

        self.setWindowTitle("Online Car Inventory Management System")
        self.setGeometry(100, 100, 1200, 800)
        
        # Apply modern material style from qt-material
        apply_stylesheet(self, theme='light_blue.xml')
        #####################################################################################
        self.Login_email=email
        #####################################
        ######(for customer)#############
        self.current_id=1  # Start ID
        ##########(for supplier)###############
        self.current_id1=1
        ###########(for sales)###############
        self.current_id2=1
        ###########(for purchase)############
        self.current_id3=1
        #####################################################################################################
        #Set the custom window icon
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))  
        icon_path = os.path.join(base_path, "dashboard_images/car_inventory.png")  # Adjust the path to your icon
        self.setWindowIcon(QIcon(QPixmap(icon_path)))  # Use the constructed path for the icon
        # Main layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        ##############################################
        ##############################################################################3
        self.My_total_purchase_field22=QLineEdit()
        ##################
        self.total_purchase_price = 0  # Initialize the instance variable
        self.totat_sale_price=0
        #################################################################################3
        # Header
        self.create_header()

        # Content
        self.create_content()

        # Footer
        self.create_footer()

        # Show the window maximized (with window bar)
        self.showMaximized()
######################################3333#############################################################################################################################
########################################################3#################################################################################################################
    def create_header(self):
        """Create a header with a circular admin logo, app name, centered search field, and user info on the right."""
        header = QWidget()
        header.setStyleSheet("background-color:white;padding: 15px; border-bottom:1px solid #ddd;")
        header.setFixedHeight(65)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(0)
        # Set the base path
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))
#################################################################################################################
        def create_circular_pixmap(file_path, size):
            pixmap = QPixmap(file_path)
            pixmap = pixmap.scaled(size, size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            mask = QPixmap(size, size)
            mask.fill(Qt.transparent)
            painter = QPainter(mask)
            painter.setRenderHint(QPainter.Antialiasing)
            painter.setBrush(QBrush(pixmap))
            painter.setPen(Qt.NoPen)
            painter.drawEllipse(0, 0, size, size)
            painter.end()
            return mask
        #################################################################
        admin_logo_size = 50
        admin_logo_path = os.path.join(base_path, "dashboard_images/admin_user.png")  # Use the constructed path
        admin_logo_pixmap = create_circular_pixmap(admin_logo_path, admin_logo_size)
        admin_logo_label = QLabel()
        admin_logo_label.setPixmap(admin_logo_pixmap)
        admin_logo_label.setFixedSize(90, 80)
        admin_logo_label.setStyleSheet("""
            QLabel {
                border-radius: 20px;
                background-color: white;
            }
        """)
        header_layout.addWidget(admin_logo_label, alignment=Qt.AlignLeft)

        self.app_name = QLabel(f"{self.Login_email}")
        self.app_name.setFont(QFont("Arial", 24, QFont.Bold))
        self.app_name.setStyleSheet("color: #4B4D5A; padding-left: 10px;")
        self.app_name.setFixedHeight(50)
        header_layout.addWidget(self.app_name, alignment=Qt.AlignLeft)

        header_layout.addStretch(1)
        search_widget = QWidget()
        search_layout = QHBoxLayout(search_widget)

        search_icon = QLabel()
        search_icon_size = 40
        search_icon_path = os.path.join(base_path, "dashboard_images/search.png")  # Use the constructed path
        search_icon_pixmap = QPixmap(search_icon_path)
        search_icon_pixmap = search_icon_pixmap.scaled(search_icon_size, search_icon_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        search_icon.setPixmap(search_icon_pixmap)
        search_icon.setFixedSize(70, 50)

        search_field = QLineEdit()
        search_field.setPlaceholderText("Search...")
        search_field.setFixedWidth(500)
        search_field.setFixedHeight(35)
        search_field.setStyleSheet("""
            QLineEdit {
                border: 1px solid #ddd;
                padding: 5px;
                border-radius: 15px;
                font-size: 14px;
            }
        """)
        search_layout.addWidget(search_icon, alignment=Qt.AlignCenter)
        search_layout.addWidget(search_field, alignment=Qt.AlignCenter)
        search_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.addWidget(search_widget, alignment=Qt.AlignCenter)
        header_layout.addStretch(2)

        team_logo_size = 50
        circular_logo_path = os.path.join(base_path, "dashboard_images/teamwork.png")  # Use the constructed path
        circular_logo_pixmap = create_circular_pixmap(circular_logo_path, team_logo_size)
        circular_logo_label = QLabel()
        circular_logo_label.setPixmap(circular_logo_pixmap)
        circular_logo_label.setFixedSize(90, 80)
        circular_logo_label.setStyleSheet("""
            QLabel {
                border-radius:20px;
                background-color:white;
            }
        """)
        header_layout.addWidget(circular_logo_label, alignment=Qt.AlignRight)

        user_info = QLabel("Login User")
        user_info.setFont(QFont("Arial", 16))
        user_info.setFixedHeight(40)
        user_info.setStyleSheet("color: #4B4D5A; padding-left: 10px;")
        header_layout.addWidget(user_info, alignment=Qt.AlignRight)
        self.main_layout.addWidget(header)
#############################################################################################################################################
#############################################################################################################################################
    def create_footer(self):
        """Create a footer like in the image."""
        footer = QLabel("© 2024 Inventory Car Management Dashboard")
        footer.setAlignment(Qt.AlignCenter)
        footer.setStyleSheet("background-color:#42a5f5;color:white;padding:10px; color: #4B4D5A; font-size: 12px; border-top: 1px solid #ddd;")
        self.main_layout.addWidget(footer)
#########################################################################################################################################################################
    def create_content(self):
        """Create the main content, including the sidebar and QStackedWidget."""
        content_layout = QHBoxLayout()
        # Sidebar
        self.sidebar = self.create_sidebar()
        content_layout.addWidget(self.sidebar)
        # QStackedWidget for switching between views
        self.stacked_widget = QStackedWidget()
        self.stacked_widget.addWidget(self.create_home_view())
        self.stacked_widget.addWidget(self.create_sales_management_view())
        self.stacked_widget.addWidget(self.create_purchase_management_view())
        self.stacked_widget.addWidget(self.create_employee_management_view())
        self.stacked_widget.addWidget(self.create_inventory_store_management_view())  # Add the new view here
        self.stacked_widget.addWidget(self.create_financial_management_view())
        self.stacked_widget.addWidget(self.create_reports_management_view())
        self.stacked_widget.addWidget(self.create_admin_user_management_view())
        self.stacked_widget.addWidget(self.create_logout_view())
        # Wrapping QStackedWidget in a QScrollArea
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)  # Make the scroll area resize as needed
        scroll_area.setWidget(self.stacked_widget)
        # Add the scroll area to the layout
        content_layout.addWidget(scroll_area)
        self.main_layout.addLayout(content_layout)
        # Set the default view to Home
        self.set_initial_view()
##########################################################################33##################################################################################
#######################################################################################################################################################
    def create_sidebar(self):
        """Create a modern, vibrant, and user-friendly sidebar with enhanced visibility."""
        # Apply the modern material style from qt-material
        apply_stylesheet(self, theme='light_blue.xml')

        # Set the base path
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))

        sidebar = QListWidget()
        sidebar.setStyleSheet("""
            /* Overall sidebar styling */
            QListWidget {
                background-color: #1E1E2F;  /* Darker background for a modern look */
                color: #FFFFFF;  /* Set default text color to white */
                border: none;
                padding: 12px;
                font-family: 'Segoe UI', sans-serif;
                font-size: 14px;  /* Slightly larger font size for better readability */
                border-radius:14px;
                font-weight:bold;
            }

            /* Menu items with clear separation and professional look */
            QListWidget::item {
                margin: 0px;  /* Removed margin for tighter spacing */
                padding: 18px 20px;  /* Larger padding for more clickable area */
                border-radius: 12px;  /* Rounded corners for smooth appearance */
                background-color: #2A2A40;  /* Subtle contrast for menu items */
                color: #FFFFFF;  /* Default text color for items */
                font-weight: bold;  /* Make the text bold */
                min-width: 230px;  /* Set a minimum width for the items */
                min-height:30px;  /* Set a minimum height for the items */
                max-width: 230px;  /* Set a maximum width for the items */
                max-height:30px;  /* Set a maximum height for the items */
            }

            /* Hover effect for a more interactive experience */
            QListWidget::item:hover {
                background-color: #4C82FB;  /* Brighter blue for hover effect */
                color: #FFFFFF;  /* Keep text color white */
                border: 1px solid #4C82FB;
                border-radius: 10px;
            }

            /* Selected item effect for strong feedback */
            QListWidget::item:selected {
                background-color: #28A745;  /* Bright green to indicate selection */
                color: #FFFFFF;  /* Ensure the text remains white */
                border: 1px solid #28A745;
                border-radius: 10px;
            }
        """)

        sidebar.setFixedWidth(300)  # Increased width for better spacing and readability

        # Define the menu items with appropriate icons and colors
        items = [
            ("Home", self.show_home_view, "Side_bar_icon/home.png", "white"),
            ("Car Delivery Management", self.show_sales_management_view, "Side_bar_icon/sale.png", "#2A2A40"),
            ("Car Purchase Management", self.show_purchase_management_view, "Side_bar_icon/purchase.png", "#2A2A40"),
            ("Employee Management", self.show_employee_management_view, "Side_bar_icon/employee.png", "#2A2A40"),
            ("Inventory Store Management", self.show_inventory_store_management_view, "Side_bar_icon/inventory.png", "#2A2A40"),
            ("Financial Management", self.show_financial_management_view, "Side_bar_icon/financial.png", "#2A2A40"),
            ("Reports Management", self.show_reports_management_view, "Side_bar_icon/report.png", "#2A2A40"),
            ("Admin User Management", self.show_admin_user_management_view, "Side_bar_icon/admin.png", "#2A2A40"),
            ("Logout", self.show_logout_view, "Side_bar_icon/logut.png", "#2A2A40"),
        ]

        # Iterate over the items and add them to the sidebar with icons
        for idx, (btn_name, action, icon_relative_path, color) in enumerate(items):
            item = QListWidgetItem(f"  {btn_name}")
            item.setText(btn_name)
            item.setData(Qt.UserRole, action)
            item.setData(Qt.BackgroundRole, color)
            # Construct the full icon path
            icon_path = os.path.join(base_path, icon_relative_path)
            # Add professional-looking icons to the items
            icon = QIcon(icon_path)
            item.setIcon(icon)
            # Set the text color of the "Home" menu item to white
            if idx == 0:  # Assuming "Home" is the first item
                item.setForeground(QColor('white'))  # Set Home item text to white
            sidebar.addItem(item)
        sidebar.itemClicked.connect(self.change_view)
        # Change text color to white when an item is clicked
        def change_item_color(item):
            # Get the clicked item
            clicked_item = sidebar.currentItem()
            # Set all items to the default color
            for i in range(sidebar.count()):
                sidebar.item(i).setForeground(QColor('#FFFFFF'))  # Set to white
            # Set the clicked item's color if it is "Home"
            if clicked_item and clicked_item.text().strip() == "Home":
                clicked_item.setForeground(QColor('white'))  # Ensure Home stays white
        
        sidebar.itemClicked.connect(change_item_color)
        return sidebar
##############################################################################################################################################################################
#####################################################################################################################################################################
    def set_initial_view(self):
        """Set the initial view and highlight the 'Home' menu item."""
        self.sidebar.setCurrentRow(0)
        self.stacked_widget.setCurrentIndex(0)
#################################################################################################################
    def change_view(self, item):
        """Change the view based on the sidebar item clicked."""
        action = item.data(Qt.UserRole)
        if action:
            action()
#################################################################################################################
    def show_home_view(self):
        """Show the home view."""
        self.stacked_widget.setCurrentIndex(0)
#################################################################################################################
    def show_sales_management_view(self):
        """Show the sales management view."""
        self.stacked_widget.setCurrentIndex(1)
#################################################################################################################
    def show_purchase_management_view(self):
        """Show the purchase management view."""
        self.stacked_widget.setCurrentIndex(2)
#################################################################################################################
    def show_employee_management_view(self):
        """Show the employee management view."""
        self.stacked_widget.setCurrentIndex(3)
#################################################################################################################
    def show_inventory_store_management_view(self):
        """Show the inventory store management view."""
        self.stacked_widget.setCurrentIndex(4)  # Update index to match the new position
#################################################################################################################
    def show_financial_management_view(self):
        """Show the financial management view."""
        self.stacked_widget.setCurrentIndex(5)
################################################################################################################3
    def show_reports_management_view(self):
        """Show the reports management view."""
        self.stacked_widget.setCurrentIndex(6)
#############################################################################################################3
    def show_admin_user_management_view(self):
        """Show the admin user management view."""
        self.stacked_widget.setCurrentIndex(7)
###################################################################################################3
    def show_logout_view(self):
        """Show the logout view."""
        self.stacked_widget.setCurrentIndex(8)
####################################################################################################################################
#######################################################################################################################################################

#########################################################################################################################################################
###########################################################################################################################################################
    def create_sales_management_view(self):
        #Create the main widget
        sales_management_widget = QWidget()
        ###############################################################
        sales_management_widget.setStyleSheet("background-color:white;")
        #####################################################################
        layout = QVBoxLayout(sales_management_widget)
        sales_management_widget.setStyleSheet("background-color:white;")
        # Step (1): Add Tab bar with name "Add Sale Car Details"
        tab_widget = QTabWidget()

        # Style for the tab bar
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #C0C0C0;
                border-radius:10px;
            }
            QTabBar::tab {
                background-color: #F1F1F1;
                border: 1px solid #C0C0C0;
                border-radius: 5px 5px 0 0;
                min-width: 120px;
                font-weight: bold;
                color: #333;
            }
            QTabBar::tab:selected {
                background-color: #007BFF;
                color: white;
            }
        """)

        #Create the form widget inside the tab
        form_widget = QWidget()
        grid_layout = QGridLayout(form_widget)
        ############################################################
        grid_layout.setSpacing(5)  # Set spacing between grid cells to 5
        grid_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins around the grid
        ############################################################################################
        # Define fixed size for fields
        field_width = 270
        field_height = 35
        button_width = 150
        button_height = 40
        # General stylesheet for the form
        form_widget.setStyleSheet("""
            QWidget {
                font-family: Arial, sans-serif;
                font-size:12px;
            }
            QLabel {
                font-weight: bold;
                color: #333;
                font-size:15px;
                font-weight:bold;
            }
            QLineEdit, QDateEdit, QTimeEdit {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #F9F9F9;
                font-weight:bold;
                font-size:14px;
            }
            QPushButton {
                background-color: #007BFF;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #003d7a;
            }
            QLabel[style*="border"] {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #F9F9F9;
            }
            QVBoxLayout,QHBoxLayout {
                spacing:0px;
            }
        """)
        # Step (2): Add the Title of the form INSIDE the Tab bar
        title_label=QLabel("Delivery Note")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 24px; font-weight: bold; margin-bottom: 20px; color: #333;")
        grid_layout.addWidget(title_label,0,0,1,4)  # Span the title across 4 columns
        # Step (3): Add date, time, day, and invoice number
        self.date_field=QDateEdit(QDate.currentDate())
        # self.date_field.setEnabled(False)  # This makes the field non-editable and grayed out.
        self.date_field.setAlignment(Qt.AlignLeft)
        self.date_field.setContentsMargins(0,0,0,30)
        ###########################################################33
        self.date_field.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QDateEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.date_field.setFixedWidth(field_width)
        self.date_field.setFixedHeight(field_height)
        #########################################################
        self.time_field=QTimeEdit(QTime.currentTime())
        self.time_field.setAlignment(Qt.AlignLeft)
        self.time_field.setReadOnly(True)  # Set the field to read-only
        #####################################################
        self.time_field.setFixedWidth(field_width)
        self.time_field.setFixedHeight(field_height)
        self.time_field.setContentsMargins(0,0,0,30)
        self.time_field.setStyleSheet("""
            QTimeEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QTimeEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        #################################################################33
        self.day_field=QLineEdit(QDate.currentDate().toString("dddd"))  # Automatically fill day
        self.day_field.setReadOnly(True)  # Set the field to read-only
        self.day_field.setAlignment(Qt.AlignLeft)
        self.day_field.setFixedWidth(field_width)
        self.day_field.setFixedHeight(field_height)
        self.day_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        #########################################################################################################3
        self.invoice_number_field=QLineEdit()  # Random invoice number
        self.current_id2=self.get_latest_sales_id_from_db()
        self.invoice_number_field.setText(str(self.current_id2).zfill(6))
        ##############################################################
        self.invoice_number_field.setAlignment(Qt.AlignLeft)
        self.invoice_number_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.invoice_number_field.setFixedWidth(field_width)
        self.invoice_number_field.setFixedHeight(field_height)
        #################################################################
        self.invoice_number_field.mousePressEvent=lambda event: self.generate_dynamic_sales_id()
        ######################################################################################
        # Add fields to grid layout
        grid_layout.addWidget(QLabel("Date:"),1,0)
        grid_layout.addWidget(self.date_field,1,1)
        grid_layout.addWidget(QLabel("Time:"),1,2)
        grid_layout.addWidget(self.time_field,1,3)
        grid_layout.addWidget(QLabel("Day:"),2,0)
        grid_layout.addWidget(self.day_field,2,1)
        grid_layout.addWidget(QLabel("Invoice Number:"), 2, 2)
        grid_layout.addWidget(self.invoice_number_field, 2, 3)
        ######################################################################################################################
        ##############################################################################################################################
        ###################################################################################################################################
        # Step (4): Add customer details
        self.customer_name_field=QComboBox()
        # Populate the combo box with example customer names
        #self.customer_name_field.addItems(["Select Customer Name"]+customer_names)
        #customer_name_field = QLineEdit()
        self.customer_name_field.setFixedHeight(field_height)
        self.customer_name_field.setFixedWidth(field_width)
        self.customer_name_field.currentIndexChanged.connect(self.on_customer_selection_changed)
        self.populate_customer_combo_box()
        #############################################################
        self.customer_name_field.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ####################################################
        self.father_name_field=QLineEdit(self)
        self.father_name_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.father_name_field.setFixedWidth(field_width)
        self.father_name_field.setFixedHeight(field_height)
        self.father_name_field.setAlignment(Qt.AlignLeft)
        ######################################################
        self.address_field=QLineEdit(self)
        self.address_field.setFixedWidth(field_width)
        self.address_field.setFixedHeight(field_height)
        self.address_field.setAlignment(Qt.AlignLeft)
        self.address_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        #######################################################################################
        self.customer_phone_number1=QLineEdit(self)
        self.customer_phone_number1.setFixedWidth(field_width)
        self.customer_phone_number1.setFixedHeight(field_height)
        self.customer_phone_number1.setAlignment(Qt.AlignLeft)
        self.customer_phone_number1.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ##########################################################
        self.cnic_field = QLineEdit(self)
        self.cnic_field.setFixedWidth(field_width)
        self.cnic_field.setFixedHeight(field_height)
        self.cnic_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.cnic_field.setAlignment(Qt.AlignLeft)
        ##########################################################
        # Add fields to grid layout
        grid_layout.addWidget(QLabel("Purchaser Name:"), 3, 0)
        grid_layout.addWidget(self.customer_name_field, 3, 1)
        grid_layout.addWidget(QLabel("Father's Name:"), 3, 2)
        grid_layout.addWidget(self.father_name_field, 3, 3)
        grid_layout.addWidget(QLabel("Address:"), 4, 0)
        grid_layout.addWidget(self.address_field, 4, 1)
        #########################################################
        grid_layout.addWidget(QLabel("Phone Number"),4,2)
        grid_layout.addWidget(self.customer_phone_number1,4,3)
        ###################################################
        grid_layout.addWidget(QLabel("CNIC No:"), 5, 0)
        grid_layout.addWidget(self.cnic_field, 5, 1)
        # Step (4e): Add image label with border and fixed size
        #####################################################################
        self.image_label = QLabel(self)
        self.image_label.setFixedSize(field_width,130)  # Fixed size for image
        self.image_label.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")  # Add border to the image label
        self.image_label.setAlignment(Qt.AlignLeft)
        self.image_label.setStyleSheet("border:2px solid black;border-radius:5px;")
        ##############################################################################
        grid_layout.addWidget(QLabel("Purchaser Image:"), 5, 2)
        grid_layout.addWidget(self.image_label, 5, 3)
        # Step (5): Add Vehicle Description section
        ###################################################################################################################################################################################################
        purchase_title_label=QLabel("Add Purchase Details")
        purchase_title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 20px; color: #333;")
        grid_layout.addWidget(purchase_title_label, 6, 0, 1, 4)
        ######################################################################3
        self.purchase_car_names=QComboBox()
        self.purchase_car_names.setFixedHeight(field_height)
        self.purchase_car_names.setFixedWidth(field_width)
        self.purchase_car_names.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ##############################################
        self.fetch_latest_car_from_firestore()
        #######################################################
        self.purchase_car_price=QLineEdit(self)
        self.purchase_car_price.setFixedWidth(field_width)
        self.purchase_car_price.setFixedHeight(field_height)
        self.purchase_car_price.setAlignment(Qt.AlignLeft)
        self.purchase_car_price.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ################################################################
        grid_layout.addWidget(QLabel("Purchase Car Name:"),7,0)
        grid_layout.addWidget(self.purchase_car_names,7,1)
        ##############################################################################
        grid_layout.addWidget(QLabel("Purchase Car Price:"),7,2)
        grid_layout.addWidget(self.purchase_car_price,7,3)
        #######################################################3
        self.purchase_car_names.currentIndexChanged.connect(self.on_purchase_selection_changed)
        self.populate_purchase_combo_box()
        ###################################################################################################################################################################################################
        ########################################################################################3####################################################################
        vehicle_title_label = QLabel("Description of Vehicle")
        vehicle_title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 20px; color: #333;")
        grid_layout.addWidget(vehicle_title_label,8, 0, 1, 4)
        # Add vehicle details fields
        #####################################################################
        self.registration_field=QLineEdit()
        self.registration_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.registration_field.setAlignment(Qt.AlignLeft)
        self.registration_field.setFixedWidth(field_width)
        self.registration_field.setFixedHeight(field_height)
        #########################################################3
        self.chassis_field=QLineEdit()
        self.chassis_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.chassis_field.setFixedWidth(field_width)
        self.chassis_field.setFixedHeight(field_height)
        self.chassis_field.setAlignment(Qt.AlignLeft)
        ######################################################
        self.engine_field=QLineEdit()
        self.engine_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.engine_field.setFixedWidth(field_width)
        self.engine_field.setFixedHeight(field_height)
        self.engine_field.setAlignment(Qt.AlignLeft)
        #########################################################
        self.make_field=QLineEdit()
        self.make_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.make_field.setFixedWidth(field_width)
        self.make_field.setFixedHeight(field_height)
        self.make_field.setAlignment(Qt.AlignLeft)
        ########################################################
        self.horsepower_field=QLineEdit()
        self.horsepower_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.horsepower_field.setFixedWidth(field_width)
        self.horsepower_field.setFixedHeight(field_height)
        self.horsepower_field.setAlignment(Qt.AlignLeft)
        ##################################################
        self.model_field=QLineEdit()
        self.model_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.model_field.setFixedWidth(field_width)
        self.model_field.setFixedHeight(field_height)
        self.model_field.setAlignment(Qt.AlignLeft)
        ################################################
        self.color_field=QLineEdit()
        self.color_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.color_field.setFixedWidth(field_width)
        self.color_field.setFixedHeight(field_height)
        self.color_field.setAlignment(Qt.AlignLeft)
        ##########################################################################
        # self.sale_import_field=QLineEdit()
        # self.sale_import_field.setStyleSheet("""
        #     QLineEdit{
        #         background-color:#ECDFCC;
        #         color:black;
        #         font-weight:bold;
        #         border-radius:10px;
        #     }
        #     QLineEdit::Placeholder{
        #         color:white;
        #         font-weight:bold;
        #         text-align:center;
        #     }
        # """)
        # self.sale_import_field.setFixedWidth(field_width)
        # self.sale_import_field.setFixedHeight(field_height)
        # self.sale_import_field.setAlignment(Qt.AlignLeft)
        ##########################################################
        # Add vehicle details fields
        grid_layout.addWidget(QLabel("Registration #:"),9,0)
        grid_layout.addWidget(self.registration_field,9,1)
        grid_layout.addWidget(QLabel("Chassis #:"),9,2)
        grid_layout.addWidget(self.chassis_field,9,3)
        grid_layout.addWidget(QLabel("Engine #:"),10,0)
        grid_layout.addWidget(self.engine_field,10,1)
        grid_layout.addWidget(QLabel("Make:"),10,2)
        grid_layout.addWidget(self.make_field,10,3)
        grid_layout.addWidget(QLabel("Horse Power:"),11,0)
        grid_layout.addWidget(self.horsepower_field,11,1)
        grid_layout.addWidget(QLabel("Model #:"),11,2)
        grid_layout.addWidget(self.model_field,11,3)
        grid_layout.addWidget(QLabel("Color:"),12,0)
        grid_layout.addWidget(self.color_field,12,1)
        ########################################################
        # grid_layout.addWidget(QLabel("Sale Import:"),12,2)
        # grid_layout.addWidget(self.sale_import_field,12,3)
        #########################################################################################################################################
        self.Sale_Price_field=QLineEdit()
        self.Sale_Price_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Sale_Price_field.setFixedWidth(field_width)
        self.Sale_Price_field.setFixedHeight(field_height)
        self.Sale_Price_field.setAlignment(Qt.AlignLeft)
        ################################################################
        self.Sale_Price_Inwords=QLineEdit()
        self.Sale_Price_Inwords.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ###########################################
        self.Sale_Price_Inwords.setFixedWidth(field_width)
        self.Sale_Price_Inwords.setFixedHeight(field_height)
        self.Sale_Price_Inwords.setAlignment(Qt.AlignLeft)
        ##################################################################################################
        grid_layout.addWidget(QLabel("Car Sale Price:"),13,0)
        grid_layout.addWidget(self.Sale_Price_field,13,1)
        ##########################################
        grid_layout.addWidget(QLabel("Sale Price(InWords):"),13,2)
        grid_layout.addWidget(self.Sale_Price_Inwords,13,3)
        ##############################################################################################################################################
        #################(one time)##################################################
        ##############################################################################################################################
        self.sale_Payment_date1=QDateEdit()
        self.sale_Payment_date1.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.sale_Payment_date1.setFixedWidth(field_width)
        self.sale_Payment_date1.setFixedHeight(field_height)
        self.sale_Payment_date1.setDate(QDate.currentDate())
        self.sale_Payment_date1.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Sale Payment Date:"),14,0)
        grid_layout.addWidget(self.sale_Payment_date1,14,1)
        ######################################################################################################################
        self.Sale_Mode_of_Payment1 = QComboBox()
        self.Sale_Mode_of_Payment1.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
                padding: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #ECDFCC;
                selection-background-color: #D1BBA6;
            }
        """)
        self.Sale_Mode_of_Payment1.setFixedWidth(field_width)
        self.Sale_Mode_of_Payment1.setFixedHeight(field_height)
        self.Sale_Mode_of_Payment1.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Sale_Mode_of_Payment1.setEditable(True)
        self.Sale_Mode_of_Payment1.lineEdit().setAlignment(Qt.AlignLeft)
        self.Sale_Mode_of_Payment1.setEditable(False)
        grid_layout.addWidget(QLabel("Mode of Payment:"),14,2)
        grid_layout.addWidget(self.Sale_Mode_of_Payment1,14,3)
        #############################################################################################################################
        self.sale_priceIn_check1=QLineEdit()
        self.sale_priceIn_check1.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.sale_priceIn_check1.setFixedWidth(field_width)
        self.sale_priceIn_check1.setFixedHeight(field_height)
        self.sale_priceIn_check1.setAlignment(Qt.AlignLeft)
        grid_layout.addWidget(QLabel("Amount of Payment:"),14,4)
        grid_layout.addWidget(self.sale_priceIn_check1,14,5)
        #################(repeat instances)##################################################
        #################(Two time)##################################################
        ##############################################################################################################################
        self.sale_Payment_date2=QDateEdit()
        self.sale_Payment_date2.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.sale_Payment_date2.setFixedWidth(field_width)
        self.sale_Payment_date2.setFixedHeight(field_height)
        self.sale_Payment_date2.setDate(QDate.currentDate())
        self.sale_Payment_date2.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Sale Payment Date:"),15,0)
        grid_layout.addWidget(self.sale_Payment_date2,15,1)
        ######################################################################################################################
        self.Sale_Mode_of_Payment2= QComboBox()
        self.Sale_Mode_of_Payment2.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
                padding: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #ECDFCC;
                selection-background-color: #D1BBA6;
            }
        """)
        self.Sale_Mode_of_Payment2.setFixedWidth(field_width)
        self.Sale_Mode_of_Payment2.setFixedHeight(field_height)
        self.Sale_Mode_of_Payment2.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Sale_Mode_of_Payment2.setEditable(True)
        self.Sale_Mode_of_Payment2.lineEdit().setAlignment(Qt.AlignLeft)
        self.Sale_Mode_of_Payment2.setEditable(False)
        grid_layout.addWidget(QLabel("Mode of Payment:"),15,2)
        grid_layout.addWidget(self.Sale_Mode_of_Payment2,15,3)
        #############################################################################################################################
        self.sale_priceIn_check2=QLineEdit()
        self.sale_priceIn_check2.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.sale_priceIn_check2.setFixedWidth(field_width)
        self.sale_priceIn_check2.setFixedHeight(field_height)
        self.sale_priceIn_check2.setAlignment(Qt.AlignLeft)
        grid_layout.addWidget(QLabel("Amount of Payment:"),15,4)
        grid_layout.addWidget(self.sale_priceIn_check2,15,5)
        ###########################################################
        #################(Three time)##################################################
        ##############################################################################################################################
        self.sale_Payment_date3=QDateEdit()
        self.sale_Payment_date3.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
        """)
        self.sale_Payment_date3.setFixedWidth(field_width)
        self.sale_Payment_date3.setFixedHeight(field_height)
        self.sale_Payment_date3.setDate(QDate.currentDate())
        self.sale_Payment_date3.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Sale Payment Date:"),16,0)
        grid_layout.addWidget(self.sale_Payment_date3,16,1)
        ######################################################################################################################
        self.Sale_Mode_of_Payment3= QComboBox()
        self.Sale_Mode_of_Payment3.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                padding: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #ECDFCC;
                selection-background-color: #D1BBA6;
            }
        """)
        self.Sale_Mode_of_Payment3.setFixedWidth(field_width)
        self.Sale_Mode_of_Payment3.setFixedHeight(field_height)
        self.Sale_Mode_of_Payment3.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Sale_Mode_of_Payment3.setEditable(True)
        self.Sale_Mode_of_Payment3.lineEdit().setAlignment(Qt.AlignLeft)
        self.Sale_Mode_of_Payment3.setEditable(False)
        grid_layout.addWidget(QLabel("Mode of Payment:"),16,2)
        grid_layout.addWidget(self.Sale_Mode_of_Payment3,16,3)
        #############################################################################################################################
        self.sale_priceIn_check3=QLineEdit()
        self.sale_priceIn_check3.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.sale_priceIn_check3.setFixedWidth(field_width)
        self.sale_priceIn_check3.setFixedHeight(field_height)
        self.sale_priceIn_check3.setAlignment(Qt.AlignLeft)
        grid_layout.addWidget(QLabel("Amount of Payment:"),16,4)
        grid_layout.addWidget(self.sale_priceIn_check3,16,5)
        #################(four time)##################################################
        ##############################################################################################################################
        self.sale_Payment_date4=QDateEdit()
        self.sale_Payment_date4.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.sale_Payment_date4.setFixedWidth(field_width)
        self.sale_Payment_date4.setFixedHeight(field_height)
        self.sale_Payment_date4.setDate(QDate.currentDate())
        self.sale_Payment_date4.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Sale Payment Date:"),17,0)
        grid_layout.addWidget(self.sale_Payment_date4,17,1)
        ######################################################################################################################
        self.Sale_Mode_of_Payment4= QComboBox()
        self.Sale_Mode_of_Payment4.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
                padding: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #ECDFCC;
                selection-background-color: #D1BBA6;
            }
        """)
        self.Sale_Mode_of_Payment4.setFixedWidth(field_width)
        self.Sale_Mode_of_Payment4.setFixedHeight(field_height)
        self.Sale_Mode_of_Payment4.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Sale_Mode_of_Payment4.setEditable(True)
        self.Sale_Mode_of_Payment4.lineEdit().setAlignment(Qt.AlignLeft)
        self.Sale_Mode_of_Payment4.setEditable(False)
        grid_layout.addWidget(QLabel("Mode of Payment:"),17,2)
        grid_layout.addWidget(self.Sale_Mode_of_Payment4,17,3)
        #############################################################################################################################
        self.sale_priceIn_check4=QLineEdit()
        self.sale_priceIn_check4.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.sale_priceIn_check4.setFixedWidth(field_width)
        self.sale_priceIn_check4.setFixedHeight(field_height)
        self.sale_priceIn_check4.setAlignment(Qt.AlignLeft)
        grid_layout.addWidget(QLabel("Amount of Payment:"),17,4)
        grid_layout.addWidget(self.sale_priceIn_check4,17,5)
        #################(five time)##################################################
        ##############################################################################################################################
        self.sale_Payment_date5=QDateEdit()
        self.sale_Payment_date5.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.sale_Payment_date5.setFixedWidth(field_width)
        self.sale_Payment_date5.setFixedHeight(field_height)
        self.sale_Payment_date5.setDate(QDate.currentDate())
        self.sale_Payment_date5.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Sale Payment Date:"),18,0)
        grid_layout.addWidget(self.sale_Payment_date5,18,1)
        ######################################################################################################################
        self.Sale_Mode_of_Payment5= QComboBox()
        self.Sale_Mode_of_Payment5.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
                padding: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #ECDFCC;
                selection-background-color: #D1BBA6;
            }
        """)
        self.Sale_Mode_of_Payment5.setFixedWidth(field_width)
        self.Sale_Mode_of_Payment5.setFixedHeight(field_height)
        self.Sale_Mode_of_Payment5.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Sale_Mode_of_Payment5.setEditable(True)
        self.Sale_Mode_of_Payment5.lineEdit().setAlignment(Qt.AlignLeft)
        self.Sale_Mode_of_Payment5.setEditable(False)
        grid_layout.addWidget(QLabel("Mode of Payment:"),18,2)
        grid_layout.addWidget(self.Sale_Mode_of_Payment5,18,3)
        #############################################################################################################################
        self.sale_priceIn_check5=QLineEdit()
        self.sale_priceIn_check5.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.sale_priceIn_check5.setFixedWidth(field_width)
        self.sale_priceIn_check5.setFixedHeight(field_height)
        self.sale_priceIn_check5.setAlignment(Qt.AlignLeft)
        grid_layout.addWidget(QLabel("Amount of Payment:"),18,4)
        grid_layout.addWidget(self.sale_priceIn_check5,18,5)
        ############################(Check(2)#################################################################################################
        #############################################################################################################################
        #################################################################################################################
        self.sale_Remarks=QTextEdit()
        self.sale_Remarks.setStyleSheet("""
            QTextEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QTextEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ##############################################################
        self.sale_Remarks.setFixedWidth(340)
        self.sale_Remarks.setFixedHeight(120)
        self.sale_Remarks.setAlignment(Qt.AlignLeft)
        ################################################################
        grid_layout.addWidget(QLabel("Payment Details:"),19,0)
        grid_layout.addWidget(self.sale_Remarks,19,1)
        #############################################################################################################################################################
        #########################################################################################################################################################################
        #Step (6): Add Document attachment section with upload buttons and image preview
        document_title_label = QLabel("Documents Attached")
        document_title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin-top:20px; color: #333;")
        grid_layout.addWidget(document_title_label,20,0)
        #############################################################################
        cnic_upload_button=QPushButton("Upload Cnic Image")
        attachment_button=QPushButton("Upload Other Document")
        ###################################################################
        cnic_upload_button.setStyleSheet("""
            QPushButton {
                background-color:#40A2E3;  /* Modern red color */
                color: white;
                border: none;
                border-radius:5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color:#40A2E3;
            }
            QPushButton:pressed {
                background-color:#40A2E3;
            }
        """)
        cnic_upload_button.setFixedWidth(150)
        cnic_upload_button.setFixedHeight(35)
        #######################################################
        attachment_button.setStyleSheet("""
            QPushButton {
                background-color:#40A2E3;  /* Modern red color */
                color: white;
                border: none;
                border-radius:5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color:#40A2E3;
            }
            QPushButton:pressed {
                background-color:#40A2E3;
            }
        """)
        attachment_button.setFixedWidth(170)
        attachment_button.setFixedHeight(35)
        ########################################################################
        #Connect the upload button signals to methods for handling file uploads
        cnic_upload_button.clicked.connect(lambda: self.upload_image(self.cnic_upload_label))
        ################################################################################################
        attachment_button.clicked.connect(lambda: self.upload_image(self.attachment_label))
        attachment_button.clicked.connect(lambda:self.upload_images_to_label(self.attachment_label))
        ######################################################################################
        #Image labels with border and fixed size
        self.cnic_upload_label = QLabel("No image uploaded")
        self.attachment_label = QLabel("No image uploaded")
        self.cnic_upload_label.setFixedSize(250, 130)  # Fixed size for uploaded images
        self.attachment_label.setFixedSize(250, 130)
        self.cnic_upload_label.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")  # Add border to uploaded image labels
        self.attachment_label.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")
        self.cnic_upload_label.setAlignment(Qt.AlignCenter)
        self.attachment_label.setAlignment(Qt.AlignCenter)
        ##################################################################################
        self.My_other1=QLabel("No image uploaded")
        self.My_other2=QLabel("No image uploaded")
        self.My_other3=QLabel("No image uploaded")
        self.My_other4=QLabel("No image uploaded")
        self.My_other1.setFixedSize(250, 130)  # Fixed size for uploaded images
        self.My_other2.setFixedSize(250, 130)
        self.My_other3.setFixedSize(250, 130)  # Fixed size for uploaded images
        self.My_other4.setFixedSize(250, 130)
        ####################################
        self.My_other1.setAlignment(Qt.AlignCenter)
        self.My_other2.setAlignment(Qt.AlignCenter)
        self.My_other3.setAlignment(Qt.AlignCenter)
        self.My_other4.setAlignment(Qt.AlignCenter)
        ####################################################
        self.My_other1.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")
        self.My_other2.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")
        self.My_other3.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")
        self.My_other4.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")
        ###############################################################################
        # Add upload buttons and labels
        grid_layout.addWidget(cnic_upload_button,21,0)
        grid_layout.addWidget(self.cnic_upload_label,21,1)
        grid_layout.addWidget(attachment_button,21,2)
        grid_layout.addWidget(self.attachment_label,21,3)
        ############################################################
        grid_layout.addWidget(self.My_other1,22,0)
        grid_layout.addWidget(self.My_other2,22,1)
        grid_layout.addWidget(self.My_other3,22,2)
        grid_layout.addWidget(self.My_other4,22,3)
        #Set up event filter to detect clicks on QLabel
        self.My_other1.installEventFilter(self)
        self.My_other2.installEventFilter(self)
        self.My_other3.installEventFilter(self)
        self.My_other4.installEventFilter(self)
        #############################################################################
        self.Document_Note=QTextEdit()
        self.Document_Note.setStyleSheet("""
            QTextEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QTextEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Document_Note.setFixedWidth(340)
        self.Document_Note.setFixedHeight(120)
        grid_layout.addWidget(QLabel("Document Note:"),23,0)
        grid_layout.addWidget(self.Document_Note,23,1)
        ######################################################################################################
        # Step (7): Add Save, Clear, and Generate Report buttons with extra spacing and center-aligned
        l1=QLabel("____________________________________________________________________________________________________________________________________________")
        l1.setStyleSheet("color:blue;height:20px;")
        label_layout= QHBoxLayout()
        label_layout.setContentsMargins(0,40,0,0)
        label_layout.addWidget(l1)
        label_layout.setAlignment(Qt.AlignCenter)
        #############################################################
        save_button=QPushButton("Save")
        clear_button=QPushButton("Clear")
        report_button=QPushButton("Generate Report")
        search_button=QPushButton("Search Records")  # New button
        update_record=QPushButton("Update Records")
        #################################################################
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #018749;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0D7C66;
            }
            QPushButton:pressed {
                background-color: #018749;
            }
        """)
        #######################################################
        clear_button.setStyleSheet("""
            QPushButton {
                background-color:#624E88;/* Modern blue color */
                color: white;
                border: none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color:#8967B3;
            }
            QPushButton:pressed {
                background-color:#CB80AB;
            }
        """)
        ####################################################
        search_button.setStyleSheet("""
            QPushButton {
                background-color:#181C14;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#3C3D37;
            }
            QPushButton:pressed{
                background-color:#697565;
            }
        """)
        ######################################
        report_button.setStyleSheet("""
            QPushButton {
                background-color:#2D3250;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#424769;
            }
            QPushButton:pressed{
                background-color:#7077A1;
            }
        """)
        ####################################
        update_record.setStyleSheet("""
            QPushButton {
                background-color:#9D5C0D;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#E5890A;
            }
            QPushButton:pressed{
                background-color:#C36A2D;
            }
        """)
        ###########################################################################
        save_button.setFixedSize(button_width,30)
        clear_button.setFixedSize(button_width,30)
        report_button.setFixedSize(button_width,30)
        search_button.setFixedSize(button_width,30)
        update_record.setFixedSize(button_width,30)
        #############################################################
        save_button.clicked.connect(self.save_sale_note)
        clear_button.clicked.connect(self.clear_sale)
        report_button.clicked.connect(self.create_report)
        search_button.clicked.connect(self.search_record)
        update_record.clicked.connect(self.update_data)
        ######################################################################################################
        # Add spacing above the buttons
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0,80, 0, 0)  # Add top margin to create spacing above buttons
        #################################################################################
        # Add buttons to layout
        button_layout.addWidget(save_button)
        button_layout.addWidget(clear_button)
        button_layout.addWidget(report_button)
        button_layout.addWidget(search_button)  # Add the new button here
        button_layout.addWidget(update_record)
        ############################################################################################################
        label_layout.setSpacing(40)
        grid_layout.addLayout(label_layout,24,0, 1, 4)
        form_widget.setLayout(grid_layout)
        ########################################################
        button_layout.setSpacing(10)
        button_layout.setAlignment(Qt.AlignCenter)  # Center the buttons horizontally
        grid_layout.addLayout(button_layout,25, 0, 1, 4)
        form_widget.setLayout(grid_layout)
        #######################################################################################################################
        # Add the form to the tab
        tab_widget.addTab(form_widget, "Add Sale Car Details")
##########################################33333##########################################################################################
###############################################################################(View Selling Car Details)################################
##########################################################################################################################33#############
        #Create the view_widget
        view_widget = QWidget()
        view_layout = QVBoxLayout(view_widget)
        view_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)  # Align top and center horizontally
        ###################################################################################################################################################
        # Add the title of the tab
        view_title_label = QLabel("View Selling Car Details")
        view_title_label.setStyleSheet("font-size: 24px; font-weight: bold; margin-bottom: 20px; color: #333;")
        view_layout.addWidget(view_title_label, alignment=Qt.AlignLeft)  # Center horizontally
        #################################################################################################################
        #Add Search Records section
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0,0,650,0)
        search_label = QLabel("Search Records:")
        search_label.setStyleSheet("""
            QLabel{
                color:black;
                font-weight:bold;
                border-radius:10px;
                font-size:18px;
            }
            QLabel::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)

        self.shams_search_field1=QLineEdit()
        self.shams_search_field1.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                font-size:16px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.shams_search_field1.setFixedHeight(50)
        self.shams_search_field1.setFixedWidth(550)
        ###################################################################################################################
        search_button = QPushButton("Search Records")
        search_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;  /* Modern teal color */
                color: white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        search_button.setFixedWidth(150)
        search_button.setFixedHeight(50)
        search_button.clicked.connect(self.search_records)
        #############################################################################################################################################
        print_button=QPushButton("Print Reports")
        print_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;  /* Modern teal color */
                color: white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        print_button.setFixedWidth(150)
        print_button.setFixedHeight(50)
        print_button.clicked.connect(self.print_record)
        ###################################################################################################################################################
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.shams_search_field1)
        search_layout.addWidget(search_button)
        search_layout.addWidget(print_button)
        search_layout.setAlignment(Qt.AlignLeft | Qt.AlignTop)  # Center horizontally and align top
        view_layout.addLayout(search_layout)
        ##########################################################################################################################################################
        
        ###################################################################################################################################################
        #Add a table to display records
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(11)
        self.table_widget.setHorizontalHeaderLabels([
            "Invoice No", "Customer Name", "Father Name", "CNIC NO", 
            "Registration No", "Model", "Color", "Sale_Price", "Sale Date", "Chesis No", "Document ID"
        ])
        self.table_widget.setFixedWidth(1070)
        self.table_widget.setFixedHeight(300)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
                padding: 0;
            }
            QHeaderView::section {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
            }
        """)
        self.table_widget.horizontalHeader().setStretchLastSection(True)

        column_width=250
        for i in range(11):
            self.table_widget.setColumnWidth(i, column_width)
        view_layout.addWidget(self.table_widget, alignment=Qt.AlignLeft|Qt.AlignTop)  # Center horizontally
        #############################################################################################################################################################
        # Add Save, Clear, and Generate Report buttons with extra spacing and center-aligned
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0,0,650,0)
        delete_button = QPushButton("Delete Records")
        delete_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size:16px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        #############################################################################
        load_data = QPushButton("Load Data")
        load_data.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size:16px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        ###################################################################################
        
        ###############################################################################################################
        button_width = 250
        button_height = 35
        delete_button.setFixedSize(button_width, button_height)
        load_data.setFixedSize(button_width, button_height)
        delete_button.clicked.connect(self.delete_sale)
        load_data.clicked.connect(self.load_data)
        # Create a label for the total sale field
        total_sale_label = QLabel("Total Sale:")
        total_sale_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                font-size: 16px;
                color: #333;
                margin-right: 10px;
            }
        """)

        self.total_sale_field = QLineEdit()
        self.total_sale_field.setPlaceholderText("$0.00")
        self.total_sale_field.setReadOnly(True)
        self.total_sale_field.setStyleSheet("""
            QLineEdit {
                font-weight: bold;
                font-size: 16px;
                color: #333;
                background-color: #ECDFCC;
                border-radius: 10px;
                padding: 5px;
            }
        """)
        self.total_sale_field.setFixedWidth(200)
        self.total_sale_field.setFixedHeight(50)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(load_data)
        button_layout.addWidget(total_sale_label)
        button_layout.addWidget(self.total_sale_field)
        button_layout.setSpacing(10)
        button_layout.setAlignment(Qt.AlignHCenter | Qt.AlignTop)  # Center the buttons horizontally
        view_layout.addLayout(button_layout)
        ####################################################################################################################################
        # Add the "View Sale Car Details" tab to the tab widget
        tab_widget.addTab(view_widget, "View Sale Car Details")
        # Load the selling car records after setting up the table
        self.load_selling_car_records()
        self.Sale_Price_Inwords.textChanged.connect(self.calculate_price_to_words)
        # Inside your main window's constructor, add the tab widget to the layout
        layout.addWidget(tab_widget)
        return sales_management_widget
###########################################################################################################################################
#####################################################(Create Report Menu for sale delivery)################################################################
    def print_record(self):
        """Generate a PDF report for the car sales data in the table."""
        try:
            # Open a file save dialog to allow user to select save location
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            if not file_path:  # If no file is selected, cancel the operation
                return

            # Ensure the file extension is .pdf
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Create PDF file with landscape orientation
            pdf_canvas = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf_canvas.setTitle("Car Sales Report")

            # Get the page width for centering content
            page_width, _ = landscape(A4)

            # Set up header information centered
            pdf_canvas.setFont("Helvetica-Bold", 14)
            company_text = "CAR EXPERTS Sector-A, Bankers Town Ring Road, Near State Life Society Lahore"
            company_text_width = pdf_canvas.stringWidth(company_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - company_text_width) / 2, 570, company_text)

            # Center the report date
            report_date_text = f"Car Sales Report - {datetime.datetime.now().strftime('%Y-%m-%d')}"
            report_date_width = pdf_canvas.stringWidth(report_date_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - report_date_width) / 2, 550, report_date_text)

            # Adding report title centered with minimal margin at the top
            pdf_canvas.setFont("Helvetica-Bold", 16)
            report_title = "Car Sales Report"
            report_title_width = pdf_canvas.stringWidth(report_title, "Helvetica-Bold", 16)
            pdf_canvas.drawString((page_width - report_title_width) / 2, 510, report_title)

            # Setting up table headers and data
            headers = ["Invoice No", "Customer Name", "Father Name", "CNIC", "Registration", "Model", "Color", "Sale Price", "Date", "Make"]
            data = [headers]

            # Fetch table data and add to data list (excluding the Document ID column)
            for row in range(self.table_widget.rowCount()):
                row_data = []
                for col in range(self.table_widget.columnCount() - 1):  # Exclude the last column (Document ID)
                    item = self.table_widget.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Define table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all cells to the center
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Padding for the header cells
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Alternate row color
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Table border/grid lines
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Regular font for table content
                ('FONTSIZE', (0, 1), (-1, -1), 9),  # Adjust font size for table content
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),  # Padding for content cells
                ('TOPPADDING', (0, 1), (-1, -1), 5)  # Padding for content cells
            ])

            # Set column widths (adjusted to fit the new number of columns)
            col_widths = [
                0.8 * inch,   # Invoice No
                1.2 * inch,   # Supplier Name
                1.2 * inch,   # Father Name
                1.0 * inch,   # CNIC (increased width)
                1.2 * inch,   # Registration No (increased width)
                1.0 * inch,   # Model
                1.0 * inch,   # Color
                1.2 * inch,   # Chassis
                1.0 * inch,   # Purchase Date
                1.8 * inch    # Purchase Price (increased width)
            ]
            # Create table and set style
            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)

            # Calculate table dimensions
            table_width, table_height = table.wrap(0, 0)
            # Calculate the position for the table to be centered horizontally and below the title
            x_position = (page_width - table_width) / 2  # Center horizontally
            y_position = 470  # Position the table just below the title

            # Draw the table on the PDF
            table.drawOn(pdf_canvas, x_position, y_position - table_height)

            # Add total sale value below the table
            pdf_canvas.setFont("Helvetica-Bold", 10)
            pdf_canvas.drawString(x_position, y_position - table_height - 20, f"Total Sales Value: {self.total_sale_field.text()}")

            # Add page number
            pdf_canvas.setFont("Helvetica", 8)
            pdf_canvas.drawString(750, 30, f"Page {pdf_canvas.getPageNumber()}")

            # Save the PDF file
            pdf_canvas.save()

            # Notify the user that the PDF was successfully created
            self.show_custom_message(f"PDF report generated: {file_path}", "Success", is_success=True)

        except Exception as e:
            # Handle any errors that occur during the PDF generation
            self.show_custom_message(f"Error generating PDF report: {e}", "Error", is_success=False)
###########################################################################################################################################
    def get_latest_sales_id_from_db(self):
        #Reference to the 'Customers' collection in Firestore
        sale_ref=db.collection('Car_Selling_Note')
        shams_sales=sale_ref.stream()  # Get all customer documents
        selling_ids=[]
        for selling in shams_sales:
            data=selling.to_dict()  # Convert Firestore document to dictionary
            selling_ids.append(int(data.get('invoice_number', 0)))  # Fetch the 'customer_id' field value
        if selling_ids:
            # If customers exist, find the maximum customer_id and add 1
            max_id=max(selling_ids)
            return max_id+1
        else:
            # If there are no customers, start from 1
            return 1  # This will trigger '00001' when formatted
    ########################################################################
    def generate_dynamic_sales_id(self):
        # Generate the next ID with leading zeros (e.g., 00001)
        new_invoice_number=str(self.current_id2).zfill(6)  # Generate a new random number
        self.invoice_number_field.setText(new_invoice_number)  # Update the QLineEdit field
        #Increment the ID for the next call
        self.current_id2+=1
    #######################################################
    def fetch_latest_car_from_firestore(self):
        # Initialize Firestore client
        db = firestore.Client()
        try:
            # Fetch the most recent car document from the "Car_Selling_Note" collection
            cars_ref = (
                db.collection("Car_Selling_Note")
                .order_by("added_date", direction=firestore.Query.DESCENDING)
                .limit(1)
            )
            latest_car_doc = cars_ref.stream()

            # If a document exists, update the combo box with the car name
            for car in latest_car_doc:
                latest_car_name = car.get("purchase_car_name")  # Assuming "purchase_car_name" exists
                if latest_car_name:
                    self.purchase_car_names.clear()  # Clear existing items if any
                    self.purchase_car_names.addItem(latest_car_name)  # Add the latest car name
                    self.purchase_car_names.setCurrentText(latest_car_name)  # Set it as selected
                else:
                    self.purchase_car_names.clear()  # Clear existing items if any
                    self.purchase_car_names.setCurrentText("Add Purchase Details")  # Set it as selected
        except Exception as e:
            print(f"Error fetching latest car: {e}")
##############################################################################################################################################################
#############################################################(Create Report Menu for sale delivery)################################################################
    def calculate_price_to_words(self):
        #Get the text from the Sale_Price_field
        sale_price_text = self.Sale_Price_field.text()
        if sale_price_text.strip() == "":  # Handle empty input
            self.Sale_Price_Inwords.setText("")
            return
        try:
            # Convert the text to a number (e.g., int)
            sale_price = int(sale_price_text)
            # Convert the number to words using the custom number_to_words function
            sale_in_words =self.number_to_words(sale_price)
            # Update the Sale_Price_Inwords field with the converted words
            self.Sale_Price_Inwords.setText(sale_in_words)
        except ValueError:
            # Handle the case where the conversion to int fails
            self.Sale_Price_Inwords.setText("Invalid price format")
    #################################################################################
    def number_to_words(self,num):
        if num < 0:
            return "minus " + self.number_to_words(-num)
        if num == 0:
            return "zero"
        units = [
            "", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
            "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen", 
            "seventeen", "eighteen", "nineteen"
        ]
        tens = [
            "", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", 
            "ninety"
        ]
        thousands = [
            "", "thousand", "million", "billion", "trillion", "quadrillion", 
            "quintillion", "sextillion", "septillion", "octillion", 
            "nonillion", "decillion", "undecillion", "duodecillion", 
            "tredecillion", "quattuordecillion", "quindecillion", 
            "sexdecillion", "septendecillion", "octodecillion", 
            "novemdecillion", "vigintillion"
        ]
        def helper(n):
            if n < 20:
                return units[n]
            elif n < 100:
                return tens[n // 10] + ('' if n % 10 == 0 else ' ' + units[n % 10])
            elif n < 1000:
                return units[n // 100] + " hundred" + ('' if n % 100 == 0 else ' and ' + helper(n % 100))
            else:
                for i, word in enumerate(thousands):
                    if n < 1000 ** (i + 1):
                        return helper(n // (1000 ** i)) + ' ' + thousands[i] + ('' if n % (1000 ** i) == 0 else ' ' + helper(n % (1000 ** i)))

        return helper(num).strip()
################################################################################################################################################################
############################################################################################################################################################
    def header(self, pdf):
        # Add centered image at the top
        self.add_centered_image(pdf, 'report_header.jpg', y=5)
        # First number (Cell: 0322-999066) aligned with the image on the right
        pdf.set_font('Arial', 'B', 12)  # Regular font for the phone number
        pdf.set_xy(pdf.w - 20, 15)  # Adjusted position to ensure the cell fits within the page
        pdf.cell(0, 5, '0321-4946671', ln=True, align='R')
        # Second number (042-35222655) directly below the first number
        pdf.set_xy(pdf.w - 40, pdf.get_y())  # No extra gap, same x, new y directly below the first number
        pdf.cell(0, 5, '0321-8822086', ln=True, align='R')
        # Space after the image and phone numbers
        pdf.ln(0)
        # Centered address
        pdf.set_font('Arial', 'B', 12)  # Bold font for the address
        pdf.cell(0, 10,'CAR EXPERTS Sector-A,Bankers Town Ring Road,Near State Life Society Lahore',align='C', ln=True)
        # Add space before the title
        pdf.ln(1)
        # Sale Receipt/Delivery Note Title centered
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Delivery Note', align='C', ln=True)
        pdf.ln(0)
    ##########################################################################################################################
    def add_centered_image(self, pdf, image_name, y):
        # Load the image dynamically
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path = os.path.join(base_path, f'dashboard_images/{image_name}')  # Dynamically construct the image path
        # Define the width of the image
        page_width=pdf.w
        image_width=90  # Image width (can be adjusted)
        # Calculate the center X position to center the image
        x = (page_width - image_width) / 2
        # Add the image to the PDF at the given position
        try:
            pdf.image(image_path, x=x, y=y, w=image_width)
        except RuntimeError as e:
            print(f"Error adding image: {e}")
    ########################################################################################################################
    def get_image_from_label999(self, label):
        pixmap = label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a temporary file to save the image
            with NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
                temp_filename = temp_file.name
                # Save the QImage as a PNG file
                image.save(temp_filename, "PNG")
            return temp_filename  # Return the path to the temporary image file
        return None
    #########################################################################################
    def sale_info(self, pdf):
        # Fetching data from PyQt5 form fields
        ##############(Content1)##################
        date = self.date_field.date().toString("yyyy-MM-dd")
        time = self.time_field.time().toString("HH:mm")
        day = self.day_field.text()
        invoice_number = self.invoice_number_field.text()
        # Set font for labels
        pdf.set_font('Arial', '', 10)
        # Prepare the labels and values
        label_date = f'Date: '
        label_time = f'Time: '
        label_day = f'Day: '
        label_invoice_number = f'Invoice Number: '
        # Define the widths of labels and values
        label_widths = [
            pdf.get_string_width(label_date),
            pdf.get_string_width(label_time),
            pdf.get_string_width(label_day),
            pdf.get_string_width(label_invoice_number)
        ]
        value_widths = [
            pdf.get_string_width(date),
            pdf.get_string_width(time),
            pdf.get_string_width(day),
            pdf.get_string_width(invoice_number)
        ]
        # Define horizontal spacing between each section
        spacing=2  # Adjust spacing between sections
        # Calculate the total width needed
        total_width = sum(label_widths) + sum(value_widths) + spacing * (len(label_widths) + len(value_widths) - 1)
        margin = (pdf.w - total_width) / 5  # Center the text horizontally
        # Starting Y position for Content1
        y_position = 50  # Adjust as needed
        # Set starting position
        pdf.set_xy(margin, y_position)
        # Print each label and value with underlining for values only
        pdf.set_font('Arial', '', 10)
        pdf.cell(label_widths[0], 10, label_date, ln=False, align='L')
        # Underline value part
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[0], 10, date, ln=False, align='L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(spacing, 10, '', ln=False)
        pdf.cell(label_widths[1], 10, label_time, ln=False, align='L')
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[1], 10, time, ln=False, align='L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(spacing, 10, '', ln=False)
        pdf.cell(label_widths[2], 10, label_day, ln=False, align='L')
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[2], 10, day, ln=False, align='L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(15, 10, '', ln=False)
        pdf.cell(label_widths[3], 10, label_invoice_number, ln=False, align='L')
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[3], 10, invoice_number, ln=True, align='L')
        # Reset font to regular
        pdf.set_font('Arial', '', 10)
        #########################################################################################################################################
        #######################################adding the table#################################################
        #########################################################################################################################################
        #Set table Y position immediately after Content1
        table_y_position = pdf.get_y() + 2  # Reduce spacing by using a small value like 2
        #Define margins, table dimensions, and cell content
        ########################################################################################################################################################
        num_cols = 2
        margin = 10
        label_width=30  # Minimized width for labels (e.g., "Customer Name")
        value_width=120  # Maximized width for values (e.g., "ALI RAFAY MUSHTAQ")
        cell_height=7
        #############################################################################################################################################
        # Image file path and positioning details
        image_path=self.get_image_from_label999(self.image_label)
        #####################################################################3######################################################################################################################
        #image_path = 'Images/circle_loading.jpg'  # Adjust the path based on your setup
        image_width=45
        image_height=35
        # Calculate available space for the table and image
        image_x_position = pdf.w -4- image_width  # Position image to the far right
        table_x_position = margin  # Start the table from the left margin
        ############################################################################################################################################################
        my_customer_name=self.customer_name_field.currentText()
        my_father_name=self.father_name_field.text()
        my_address_name=self.address_field.text()
        my_cnic_name=self.cnic_field.text()
        ###########################################
        sale_customer_phone=self.customer_phone_number1.text()
        ######################################################################################################################################################################
        table_data=[
                ["Purchaser Name:",my_customer_name],
                ["Father Name:",my_father_name],
                ["Address:",my_address_name],
                ["Mobile #:",sale_customer_phone],
                ["CNIC:",my_cnic_name]
        ]
        ###########################################################################################################################################################
        # Set position for the table
        pdf.set_xy(table_x_position, table_y_position)
        # Draw table rows
        for row in range(0, len(table_data)):
            pdf.set_font('Arial', 'B', 10)  # Bold font for labels
            pdf.cell(label_width, cell_height, table_data[row][0], border=1, ln=False, align='L')
            pdf.set_font('Arial', '',10)  # Regular font for values
            pdf.cell(value_width, cell_height, table_data[row][1], border=1, ln=False, align='L')
            pdf.ln(cell_height)  # Move to the next line
        #Add the customer photo to the right of the table
        pdf.image(image_path, x=image_x_position, y=table_y_position, w=image_width, h=image_height)
        # Add a border around the image
        pdf.rect(x=image_x_position, y=table_y_position, w=image_width, h=image_height)
        #########################################################################################################################################
        #######################################adding the table###########################################################
        my_s1=self.registration_field.text()
        my_s2=self.chassis_field.text()
        my_s3=self.engine_field.text()
        my_s4=self.make_field.text()
        my_s5=self.horsepower_field.text()
        my_s6=self.model_field.text()
        my_s7=self.color_field.text()
        # my_s8=self.sale_import_field.text()
        ##########################################(3rd content)############################################################
        #Define the data for the third content (Vehicle Description Table)
        vehicle_description_data=[
                ["Registration #", my_s1, "Make", my_s4],
                ["Chassis #", my_s2, "Horse Power", my_s5],
                ["Engine #", my_s3, "Model #", my_s6],
                ["Color", my_s7, "", ""]  # Adjusted row after removing "Import"
        ]
        #################################################################################################################################
        #######################################################################################################################################
        #Add header for vehicle description
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 10, "Description of Vehicle", ln=True, align='L')
        # Set font for table content
        pdf.set_font('Arial', '', 10)
        # Define column widths
        col1_width = 40  # First column (e.g., "Registration #")
        col2_width = 40  # Second column (e.g., "ARG-945")
        col3_width = 35  # Third column (e.g., "Make")
        col4_width = 40  # Fourth column (e.g., "TOYOTA CROSS")
        row_height = 7  # Height of each ro
        # Draw the table rows for vehicle description
        for row in vehicle_description_data:
            pdf.set_font('Arial', 'B', 10)  # Bold font for first and third columns
            pdf.cell(col1_width, row_height, row[0], border=1, ln=False, align='L')
            pdf.set_font('Arial', '', 10)  # Regular font for second and fourth columns
            pdf.cell(col2_width, row_height, row[1], border=1, ln=False, align='L')
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(col3_width, row_height, row[2], border=1, ln=False, align='L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(col4_width, row_height, row[3], border=1, ln=True, align='L')
        # Add space before the next section
        pdf.ln(0)
        # Add the "Documents Attached" section
        ##############################################################################################################
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, row_height, "Documents Attached:", ln=True, align='L')
        # # Add the documents attached data
        #######################################################################################################################
        ###############################################################################################################################
        #Fetch all document data
        all_document_data = self.Document_Note.toPlainText()
        # Define initial rectangle dimensions
        rect_x = 10  # X-coordinate of the rectangle's top-left corner
        rect_y = pdf.get_y()  # Y position after the previous content
        rect_width = 195  # Width of the rectangle
        padding=1  # Padding inside the rectangle
        # Set font for the text
        pdf.set_font('Arial', '', 8)
        # Calculate the available width for text and line height
        text_width = rect_width-2*padding  # Use the full rectangle width minus padding
        line_height=4  # Height of each line
        # Wrap the text to determine the total height required for the rectangle
        wrapped_lines = pdf.multi_cell(
            text_width, 
            line_height, 
            all_document_data, 
            border=0, 
            align='L', 
            split_only=True
        )
        rect_height = line_height * len(wrapped_lines)+1 * padding  # Total height of the rectangle
        # Check if the rectangle height exceeds the page's remaining height
        if rect_y + rect_height > pdf.h - 20:  # Account for bottom margin
            pdf.add_page()  # Add a new page
            rect_y = 10  # Reset Y position for the new page
        # Draw the rectangle
        pdf.rect(rect_x, rect_y, rect_width, rect_height)
        # Add the text inside the rectangle
        pdf.set_xy(rect_x + padding, rect_y + padding)
        pdf.multi_cell(
            text_width, 
            line_height, 
            all_document_data, 
            border=0, 
            align='L'
        )  # Display the text
        # Move the cursor below the rectangle for further content
        pdf.set_y(rect_y + rect_height+0)  # Add spacing for the next section
        ######################################################################################################################
        #########################################################################################################################################
        total_price_figures1=self.Sale_Price_field.text()
        total_price_words1=self.Sale_Price_Inwords.text()
        ##################################################################################
        #Total Price Figures, Total Price Words, and Remarks
        pdf.set_font('Arial', 'B', 10)  # Set font to bold
        pdf.cell(45, 6, "Total Price Figures:", ln=False, align='L')  # Bold label
        pdf.set_font('Arial', '', 10)  # Regular font for value
        pdf.cell(25, 6, total_price_figures1, ln=False, align='L')  # Value with minimal spacing
        ##################################################
        pdf.set_font('Arial', 'B', 10)  # Bold label
        pdf.cell(45, 6, "Total Price Words:", ln=False, align='L')  # Bold label
        pdf.set_font('Arial', '', 10)  # Regular font for value
        pdf.cell(25, 6, total_price_words1, ln=False, align='L')  # Value with minimal spacing
        pdf.ln(5)
        ###############################################################################################################################################################################
        sale_cheque_payment1 = self.sale_priceIn_check1.text()
        sale_cheque_payment2 = self.sale_priceIn_check2.text()
        sale_cheque_payment3 = self.sale_priceIn_check3.text()
        sale_cheque_payment4 = self.sale_priceIn_check4.text()
        sale_cheque_payment5 = self.sale_priceIn_check5.text()
        #########################################################
        remarks1 = self.sale_Remarks.toPlainText()
        #################################################
        #Set title for the section
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 7, "Payment Details:", ln=True, align='L')
        # Define rectangle dimensions and initial position
        rect_x = 10  # X position of the rectangle
        rect_y = pdf.get_y()  # Y position after the title
        rect_width = 195  # Width of the rectangle
        padding = 1  # Padding inside the rectangle
        # Combine payment amounts and remarks
        Sale_payments = [sale_cheque_payment1, sale_cheque_payment2, sale_cheque_payment3, sale_cheque_payment4, sale_cheque_payment5]
        payments_text1 = " , ".join(map(str, Sale_payments))  # Convert payments to strings and join with commas
        combined_text = f"{payments_text1} | {remarks1}"  # Combine payments and remarks
        # Set font and calculate text wrapping
        pdf.set_font('Arial', '', 8)  # Set font for text
        line_height = 4  # Line height
        text_width = rect_width - 2 * padding  # Available width for text inside the rectangle

        # Wrap text to determine height
        wrapped_lines = pdf.multi_cell(text_width, line_height, combined_text, border=0, align='L', split_only=True)
        rect_height = line_height * len(wrapped_lines) + 2 * padding  # Height of the rectangle based on wrapped lines

        # Check if rectangle height exceeds page's remaining height
        if rect_y + rect_height > pdf.h - 20:  # Account for bottom margin
            pdf.add_page()  # Add a new page
            rect_y = 10  # Reset Y position for the new page

        # Draw the rectangle
        pdf.rect(rect_x, rect_y, rect_width, rect_height)

        # Add text inside the rectangle
        pdf.set_xy(rect_x + padding, rect_y + padding)
        pdf.multi_cell(text_width, line_height, combined_text, border=0, align='L')  # Display text

        # Move the cursor below the rectangle for further content
        pdf.set_y(rect_y + rect_height + 1)  # Add vertical spacing for next section

        ######################################################################################################Adding the final paragraph#######################################
        pdf.set_font('Arial', '',8)
        #####################################################################################Adding the final paragraph#######################################
        base_path= getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path= os.path.join(base_path, 'card_images/desc.PNG')  # Construct the path to the image
        # Specify the image dimensions
        image_width=200  # Width of the image
        image_height=15  # Height of the image
        # Calculate the center position on the PDF page
        page_width = pdf.w  # Get the page width from the PDF
        center_x = (page_width - image_width) / 2  # Calculate the center X position
        # Add the image to the PDF at the calculated center position
        # If y is None, the image will be placed at the current position
        try:
            pdf.image(image_path, x=center_x, y=None, w=image_width, h=image_height)
        except Exception as e:
            print(f"Error loading image: {e}")
        pdf.ln(1)
        ##################################################################
        #################################################################################################################
        #################################################################################################################
        ##################################(4th content)##################################################################
        pdf.set_font('Arial', '', 10)
        # Center and insert the image
        ##################################################################################
        # Set the base path to dynamically load the image
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path = os.path.join(base_path, 'card_images/my_bottom1.PNG')  # Construct the path to the image
        # Specify the image dimensions
        image_width=200  # Width of the image
        image_height=51  # Height of the image
        # Calculate the center position on the PDF page
        page_width = pdf.w  # Get the page width from the PDF
        center_x = (page_width - image_width) / 2  # Calculate the center X position
        # Add the image to the PDF at the calculated center position
        # If y is None, the image will be placed at the current position
        try:
            pdf.image(image_path, x=center_x, y=None, w=image_width, h=image_height)
        except Exception as e:
            print(f"Error loading image: {e}")
        ######################################################################
        pdf.ln(1)
        #########################################################################
        # Set font
        pdf.set_font('Arial', '',7)
        # Load the image dynamically
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path = os.path.join(base_path, 'card_images/my_bottom.PNG')  # Dynamically construct the image path
        # Specify the image dimensions
        image_width=200  # Width of the image
        image_height=24  # Height of the image
        # Calculate the center position on the PDF page
        page_width = pdf.w  # Get the width of the page
        center_x = (page_width - image_width) / 2  # Calculate the center X position
        # Add the image to the PDF at the calculated center position
        try:
            pdf.image(image_path, x=center_x, y=None, w=image_width, h=image_height)
        except Exception as e:
            print(f"Error loading image: {e}")
        ######################################################################################################################################
    def save_image_from_label(label, file_path):
        pixmap = label.pixmap()
        if pixmap:
            pixmap.save(file_path, "PNG")
    ########################################################################################################################################################
    ###################################################3#################################################################################################################
    def generate_report(self, save_path):
        pdf=FPDF('P','mm','A4')  # Page format
        pdf.add_page()
        self.header(pdf)
        self.sale_info(pdf)  # Call the method to add sale info to PDF
        try:
            pdf.output(save_path)
            return True
        except RuntimeError as e:
            print(f"Error generating PDF: {e}")
            return False
    ##############################################################################
    def create_report(self):
        try:
            save_path = self.save_report_dialog()
            if save_path:
                if self.generate_report(save_path):
                    self.show_custom_message("Report generated successfully.", "Success", is_success=True)
                else:
                    self.show_custom_message("Failed to generate report.", "Error", is_success=False)
            else:
                self.show_custom_message("No file selected for saving.", "Error", is_success=False)

        except Exception as e:
            print("Error in create_report:", e)
            self.show_custom_message(f"An error occurred: {e}", "Error", is_success=False)
    ####################################################################################################
    def save_report_dialog(self):
        # Open a file save dialog to choose the location to save the report
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self,  # Parent widget
            "Save Report",  # Dialog title
            "",  # Default directory (empty means the user's home directory)
            "PDF Files (*.pdf);;All Files (*)",  # Filter options
            options=options
        )
        return file_path if file_path else None
    ###############################################################################################################################################################################3
    ##################################################################################################################################################################################
    def load_data(self):
        self.show_custom_message("Sales records is Loaded successfully!", "Success", is_success=True)
        self.load_sales_data()
        self.shams_search_field1.clear()
    ######################################################################################################################################################
    ####################################################################################################################################################
    def search_records(self):
        """Search for records based on the entered text and display all matching records."""
        shams_search_text1 = self.shams_search_field1.text().strip().lower()  # Strip spaces and convert to lowercase
        # Check if the search field is empty
        if not shams_search_text1:
            self.show_custom_message("Please enter a value to search Records.", "Warning", is_success=False)
            return
        try:
            # Fetch a broader set of records for partial matching
            all_records = db.collection('Car_Selling_Note').get()
            # Filter records locally for partial matches
            matching_records = []
            for record in all_records:
                sale_data = record.to_dict()
                if any(shams_search_text1 in str(sale_data.get(field, '')).lower() for field in 
                    ['invoice_number', 'registration', 'chassis']):
                    matching_records.append(record)

            # Clear the table before displaying search results
            self.table_widget.setRowCount(0)
            
            # Initialize total sale value
            total_sale_value = 0
            
            # Check if there are no matching records
            if not matching_records:
                self.show_custom_message("No records found for the given search criteria.", "Info", is_success=True)
                # Reset total sale field if no records found
                self.total_sale_field.setText('0')
                return
            
            # Loop through the matching records and display them in the table
            for record in matching_records:
                sale_data = record.to_dict()
                row_position = self.table_widget.rowCount()
                self.table_widget.insertRow(row_position)
                # Fill in the table with the matching record's data
                self.table_widget.setItem(row_position, 0, QTableWidgetItem(sale_data.get('invoice_number', '')))
                self.table_widget.setItem(row_position, 1, QTableWidgetItem(sale_data.get('customer_name', '')))
                self.table_widget.setItem(row_position, 2, QTableWidgetItem(sale_data.get('father_name', '')))
                self.table_widget.setItem(row_position, 3, QTableWidgetItem(sale_data.get('cnic', '')))
                self.table_widget.setItem(row_position, 4, QTableWidgetItem(sale_data.get('registration', '')))
                self.table_widget.setItem(row_position, 5, QTableWidgetItem(sale_data.get('model', '')))
                self.table_widget.setItem(row_position, 6, QTableWidgetItem(sale_data.get('color', '')))
                
                # Directly retrieve and format sale_price as an integer string
                sale_price = sale_data.get('sale_price', 0)
                self.table_widget.setItem(row_position, 7, QTableWidgetItem(str(int(sale_price))))  # Ensure it's an integer and then convert to string
                self.table_widget.setItem(row_position, 8, QTableWidgetItem(sale_data.get('date', '')))
                self.table_widget.setItem(row_position, 9, QTableWidgetItem(sale_data.get('make', '')))
                self.table_widget.setItem(row_position, 10, QTableWidgetItem(record.id))  # Store document ID
                
                # Accumulate total sale value
                total_sale_value += sale_price
            
            # Update the total sale field with the calculated total
            self.total_sale_field.setText(str(total_sale_value))  # Convert total to string for the text field
            
            # Resize columns to fit contents
            self.table_widget.resizeColumnsToContents()
        
        except Exception as e:
            self.show_custom_message(f"Error searching for records: {e}", "Error", is_success=False)
    ############################################################################################################################################   
    def load_sales_data(self):
        """Load existing selling car records from Firebase Firestore and display them in the table."""
        try:
            # Fetch records from Firestore
            sales_records = db.collection('Car_Selling_Note').get()
            # Clear existing rows in the table before loading new data
            self.table_widget.setRowCount(0)
            
            total_sale_value = 0  # Initialize total sale value

            # Loop through each record and insert it into the table
            for record in sales_records:
                sale_data = record.to_dict()
                # Insert a new row in the table
                row_position = self.table_widget.rowCount()
                self.table_widget.insertRow(row_position)
                # Fill in the columns with data from Firestore
                self.table_widget.setItem(row_position, 0, QTableWidgetItem(sale_data.get('invoice_number', '')))
                self.table_widget.setItem(row_position, 1, QTableWidgetItem(sale_data.get('customer_name', '')))
                self.table_widget.setItem(row_position, 2, QTableWidgetItem(sale_data.get('father_name', '')))
                self.table_widget.setItem(row_position, 3, QTableWidgetItem(sale_data.get('cnic', '')))
                self.table_widget.setItem(row_position, 4, QTableWidgetItem(sale_data.get('registration', '')))
                self.table_widget.setItem(row_position, 5, QTableWidgetItem(sale_data.get('model', '')))
                self.table_widget.setItem(row_position, 6, QTableWidgetItem(sale_data.get('color', '')))
                
                # Replace chassis with sale_price (integer)
                sale_price = sale_data.get('sale_price', 0)  # Default to 0 if sale_price is missing
                self.table_widget.setItem(row_position, 7, QTableWidgetItem(str(sale_price)))  # Convert integer to string

                total_sale_value += sale_price  # Accumulate total sale value

                self.table_widget.setItem(row_position, 8, QTableWidgetItem(sale_data.get('date', '')))
                self.table_widget.setItem(row_position, 9, QTableWidgetItem(sale_data.get('chassis', '')))
                self.table_widget.setItem(row_position, 10, QTableWidgetItem(record.id))  # Store document ID
                
            # Update the total sale field with the calculated total
            self.total_sale_field.setText(str(total_sale_value))  # Convert total to string for the text field

            # Optionally, resize columns to fit contents
            self.table_widget.resizeColumnsToContents()
        except Exception as e:
            self.show_custom_message(f"Error loading sale records: {e}", "Error", is_success=False)
            #pass
        ###########################################################################################################################################
    def load_selling_car_records(self):
        """Load existing selling car records from Firebase Firestore and display them in the table."""
        try:
            # Fetch records from Firestore
            sales_records = db.collection('Car_Selling_Note').get()
            # Clear existing rows in the table before loading new data
            self.table_widget.setRowCount(0)

            total_sale_value = 0  # Initialize total sale value

            # Loop through each record and insert it into the table
            for record in sales_records:
                sale_data = record.to_dict()
                # Insert a new row in the table
                row_position = self.table_widget.rowCount()
                self.table_widget.insertRow(row_position)
                # Fill in the columns with data from Firestore
                self.table_widget.setItem(row_position, 0, QTableWidgetItem(sale_data.get('invoice_number', '')))
                self.table_widget.setItem(row_position, 1, QTableWidgetItem(sale_data.get('customer_name', '')))
                self.table_widget.setItem(row_position, 2, QTableWidgetItem(sale_data.get('father_name', '')))
                self.table_widget.setItem(row_position, 3, QTableWidgetItem(sale_data.get('cnic', '')))
                self.table_widget.setItem(row_position, 4, QTableWidgetItem(sale_data.get('registration', '')))
                self.table_widget.setItem(row_position, 5, QTableWidgetItem(sale_data.get('model', '')))
                self.table_widget.setItem(row_position, 6, QTableWidgetItem(sale_data.get('color', '')))

                # Replace chassis with sale_price (integer)
                sale_price = sale_data.get('sale_price', 0)  # Default to 0 if sale_price is missing
                self.table_widget.setItem(row_position, 7, QTableWidgetItem(str(sale_price)))  # Convert integer to string

                total_sale_value += sale_price  # Accumulate total sale value

                self.table_widget.setItem(row_position, 8, QTableWidgetItem(sale_data.get('date', '')))
                self.table_widget.setItem(row_position, 9, QTableWidgetItem(sale_data.get('chassis', '')))
                self.table_widget.setItem(row_position, 10, QTableWidgetItem(record.id))  # Store document ID

            # Update the total sale field with the calculated total
            self.total_sale_field.setText(str(total_sale_value))  # Convert total to string for the text field

            # Optionally, resize columns to fit contents
            self.table_widget.resizeColumnsToContents()
        except Exception as e:
            self.show_custom_message(f"Error loading sale records: {e}", "Error", is_success=False)
            #pass
        ##########################################################################################################################################3
        #################################################################################
    def delete_record_from_firestore1(self,document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Car_Selling_Note').document(document_id)
            # Delete the document from Firestore
            doc_ref.delete()
            print(f"Document {document_id} deleted successfully.")
            self.load_sale_Summary_records()
            self.current_id2=self.get_latest_sales_id_from_db()
            self.invoice_number_field.setText(str(self.current_id2).zfill(6))
        except Exception as e:
            print(f"An error occurred while deleting the document: {e}")
            # Optionally raise the exception to propagate the error further
            raise
        ################################################################################
    def delete_sale(self):
        """Delete the customer record based on the selected row in the table."""
        # Get the selected items in the table
        selected_items=self.table_widget.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this customer?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 11th column (index 10)
                document_id =self.table_widget.item(row, 10).text()

                # Retrieve the sale price from the 8th column (index 7)
                sale_price = float(self.table_widget.item(row, 7).text())

                # Call the function to delete the document from Firestore
                self.delete_record_from_firestore1(document_id)
                # Remove the row from the table
                self.table_widget.removeRow(row)

                # Update total sale value
                current_total = float(self.total_sale_field.text())  # Get current total from the text field
                new_total = current_total - sale_price  # Calculate new total
                self.total_sale_field.setText(str(new_total))  # Update the total sale field

                # Show success message
                self.show_custom_message("Sales record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # If any error occurs, display it
                #self.show_custom_message(f"Error deleting record: {e}", "Error", is_success=False)
                pass
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    #########################################################################################################################################################
    def update_data(self):
        """Handle saving or updating sale note details with validation and Firebase storage."""
        # Retrieve inputs
        date = self.date_field.date().toString("yyyy-MM-dd")
        time = self.time_field.time().toString("HH:mm")
        day = self.day_field.text()
        invoice_number = self.invoice_number_field.text()
        customer_name = self.customer_name_field.currentText()
        father_name = self.father_name_field.text()
        address = self.address_field.text()
        cnic = self.cnic_field.text()
        registration = self.registration_field.text()
        chassis = self.chassis_field.text()
        engine = self.engine_field.text()
        make = self.make_field.text()
        horsepower = self.horsepower_field.text()
        model = self.model_field.text()
        color = self.color_field.text()
        document_note = self.Document_Note.toPlainText()
        #####################################################################################################################
        # Retrieve new additional fields
        purchase_car_name=self.purchase_car_names.currentText()  # ComboBox for car names
        # sale_import = self.sale_import_field.text()
        customer_phone_number1 = self.customer_phone_number1.text()
        sale_price_in_words = self.Sale_Price_Inwords.text()
        sale_remarks=self.sale_Remarks.toPlainText()
        ############################################################################################################################
        # Convert purchase_car_price to integer
        try:
            purchase_car_price = int(self.purchase_car_price.text())  # Convert to integer
        except ValueError:
            self.show_custom_message("Purchase car price must be a valid number.", "Error", is_success=False)
            return
        ##############################################################################################################################
        #Gather payment modes
        sale_payment_modes = [
            self.Sale_Mode_of_Payment1.currentText(),
            self.Sale_Mode_of_Payment2.currentText(),
            self.Sale_Mode_of_Payment3.currentText(),
            self.Sale_Mode_of_Payment4.currentText(),
            self.Sale_Mode_of_Payment5.currentText()
        ]

        # Gather payment dates
        sale_payment_dates = [
            self.sale_Payment_date1.date().toPyDate(),
            self.sale_Payment_date2.date().toPyDate(),
            self.sale_Payment_date3.date().toPyDate(),
            self.sale_Payment_date4.date().toPyDate(),
            self.sale_Payment_date5.date().toPyDate()
        ]

        try:
            sale_price = int(self.Sale_Price_field.text())
            #########################################################
            cheque_data= {
                f"price_in_cheque{i + 1}": int(field.text().replace(',', ''))
                for i, field in enumerate([
                    self.sale_priceIn_check1,
                    self.sale_priceIn_check2,
                    self.sale_priceIn_check3,
                    self.sale_priceIn_check4,
                    self.sale_priceIn_check5
                ]) if field.text().strip()
            }
        except ValueError:
            self.show_custom_message("Sale price and cheque values must be valid numbers.", "Error", is_success=False)
            return
        ##################################################################################################
        cnic_regex = r'^\d{5}-\d{7}-\d{1}$|^\d{13}$'
        # Check for empty fields (include new fields in the validation)
        # if not date or not time or not day or not invoice_number or not customer_name or not father_name or not address or not registration or not chassis or not engine or not make or not horsepower or not model or not color or not sale_price or not sale_price_in_words or not sale_remarks or not purchase_car_name or not purchase_car_price or not sale_import or not customer_phone_number1 or not document_note:
        #     self.show_custom_message("All required fields must be filled.", "Error", is_success=False)
        #     return

        # Validate CNIC format
        # if not re.match(cnic_regex, cnic):
        #     self.show_custom_message("CNIC must be in the format XXXXX-YYYYYYY-Z or a 13-digit number.", "Error", is_success=False)
        #     return

        # Check for missing images
        #############################################################################################################
        self.customer_image_data = self.get_image_from_label(self.image_label)
        # if not self.customer_image_data:
        #     self.show_custom_message("Please upload the customer image.", "Error", is_success=False)
        #     return
        self.cnic_image_data = self.get_image_data_from_label1()
        # if not self.cnic_image_data:
        #     self.show_custom_message("Please upload the CNIC image.", "Error", is_success=False)
        #     return
        self.other_image_data = self.get_image_data_from_label2()
        # if not self.other_image_data:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        self.my_other_img1 = self.my_otherLabel1()
        # if not self.my_other_img1:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        self.my_other_img2 = self.my_otherLabel2()
        # if not self.my_other_img2:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        self.my_other_img3 = self.my_otherLabel3()
        # if not self.my_other_img3:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        self.my_other_img4 = self.my_otherLabel4()
        # if not self.my_other_img4:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        ##########################################################################################################################
        #Prepare sale data
        ##############################################################################################################################################
        # Prepare sale data
        sale_data = {
            "date": date,
            "time": time,
            "day": day,
            "invoice_number": invoice_number,
            "customer_name": customer_name,
            "father_name": father_name,
            "address": address,
            "cnic": cnic,
            "customer_image": self.customer_image_data,
            "registration": registration,
            "chassis": chassis,
            "engine": engine,
            "make": make,
            "horsepower": horsepower,
            "model": model,
            "color": color,
            "cnic_image": self.cnic_image_data,
            "other_image": self.other_image_data,
            "sale_price": sale_price,  # Stored as integer
            "sale_price_in_words": sale_price_in_words,
            #####################################################################################
            #################(one time1)##################################
            "sale_payment_modes":sale_payment_modes,
            "sale_payment_dates":[date.isoformat() for date in sale_payment_dates],
            "cheque_data":cheque_data,
            #########################################################################################
            "sale_remarks":sale_remarks,
            # Adding the new fields to the sale data
            ################################################################################
            "purchase_car_name": purchase_car_name,
            "purchase_car_price": purchase_car_price,  # Stored as integer
            # "sale_import": sale_import,
            "customer_phone_number1": customer_phone_number1,
            ######################################################################
            "other_image1":self.my_other_img1,
            "other_image2":self.my_other_img2,
            "other_image3":self.my_other_img3,
            "other_image4":self.my_other_img4,
            "Document_note":document_note
        }
        ############################################################################################################################
        # Check if sale record already exists (based on invoice number)
        existing_sale = db.collection('Car_Selling_Note').where('invoice_number', '==', invoice_number).get()
        if existing_sale:
            # Update existing record
            try:
                for doc in existing_sale:
                    doc_ref = db.collection('Car_Selling_Note').document(doc.id)
                    doc_ref.update(sale_data)
                self.show_custom_message("Sale note updated successfully!", "Success", is_success=True)
                # Clear form fields
                self.clear_sale()
                # Reload the table to reflect updated data
                self.load_sales_data()
                #########################################################
                self.current_id2=self.get_latest_sales_id_from_db()
                self.invoice_number_field.setText(str(self.current_id2).zfill(6))
                self.fetch_latest_car_from_firestore()
                ################################################################################
            except Exception as e:
                self.show_custom_message(f"Error updating sale record: {e}", "Error", is_success=False)
        else:
            self.show_custom_message("No sale record found with the given invoice number.", "Error", is_success=False)
    ##########################################################################################################################################################
    def get_image_from_label(self, label):
        pixmap = label.pixmap()
        if pixmap:
            image_data = pixmap.toImage().bits().asstring(pixmap.toImage().byteCount())
            base64_data = base64.b64encode(image_data).decode('utf-8')
            return base64_data
        return None
        ######################################################################
    def get_image_data_from_label1(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap = self.cnic_upload_label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
        #############################################################################
    def get_image_data_from_label2(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap = self.attachment_label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    #############################################################################################################################################
    #############################################################################################################################################
    def my_otherLabel1(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.My_other1.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def my_otherLabel2(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.My_other2.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def my_otherLabel3(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.My_other3.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def my_otherLabel4(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.My_other4.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    #############################################################################################################################################
    ##################################################################################################################################################
    def save_sale_note(self):
        """Handle saving sale note details with validation and Firebase storage."""
        # Retrieve inputs
        date = self.date_field.date().toString("yyyy-MM-dd")
        time = self.time_field.time().toString("HH:mm")
        day = self.day_field.text()
        ######################################################################
        invoice_number = self.invoice_number_field.text()
        customer_name = self.customer_name_field.currentText()
        father_name = self.father_name_field.text()
        address = self.address_field.text()
        cnic = self.cnic_field.text()
        #####################################################################
        registration = self.registration_field.text()
        chassis = self.chassis_field.text()
        engine = self.engine_field.text()
        make = self.make_field.text()
        #################################################################
        horsepower = self.horsepower_field.text()
        model = self.model_field.text()
        color = self.color_field.text()
        document_note = self.Document_Note.toPlainText()
        ##############################################################
        #Gather payment modes
        sale_payment_modes = [
            self.Sale_Mode_of_Payment1.currentText(),
            self.Sale_Mode_of_Payment2.currentText(),
            self.Sale_Mode_of_Payment3.currentText(),
            self.Sale_Mode_of_Payment4.currentText(),
            self.Sale_Mode_of_Payment5.currentText()
        ]

        # Gather payment dates
        sale_payment_dates = [
            self.sale_Payment_date1.date().toPyDate(),
            self.sale_Payment_date2.date().toPyDate(),
            self.sale_Payment_date3.date().toPyDate(),
            self.sale_Payment_date4.date().toPyDate(),
            self.sale_Payment_date5.date().toPyDate()
        ]

        try:
            sale_price = int(self.Sale_Price_field.text())
            cheque_data= {
                f"price_in_cheque{i + 1}": int(field.text().replace(',', ''))
                for i, field in enumerate([
                    self.sale_priceIn_check1,
                    self.sale_priceIn_check2,
                    self.sale_priceIn_check3,
                    self.sale_priceIn_check4,
                    self.sale_priceIn_check5
                ]) if field.text().strip()
            }
        except ValueError:
            self.show_custom_message("Sale price and cheque values must be valid numbers.", "Error", is_success=False)
            return
        ####################################################################################################################
        sale_price_in_words=self.Sale_Price_Inwords.text()
        sale_remarks=self.sale_Remarks.toPlainText()
        # Validation regex
        cnic_regex = r'^\d{5}-\d{7}-\d{1}$|^\d{13}$'
        # Retrieve additional fields
        purchase_car_name = self.purchase_car_names.currentText()  # Assuming QComboBox
        # sale_import = self.sale_import_field.text()
        customer_phone_number1 = self.customer_phone_number1.text()
        ####################################################################################
        # Add conversion for purchase_car_price
        try:
            purchase_car_price = int(self.purchase_car_price.text())  # Convert to integer
        except ValueError:
            self.show_custom_message("Purchase car price must be a valid number.", "Error", is_success=False)
            return     
        #############################################################################################################################
        # Check for empty fields
        # if not date or not time or not day or not invoice_number or not customer_name or not father_name or not address or not registration or not chassis or not engine or not make or not horsepower or not model or not color or not sale_price or not sale_price_in_words or not sale_remarks or not purchase_car_name or not purchase_car_price or not sale_import or not customer_phone_number1 or not document_note:
        #     self.show_custom_message("All required fields must be filled.", "Error", is_success=False)
        #     return

        # Validate CNIC format
        # if not re.match(cnic_regex, cnic):
        #     self.show_custom_message("CNIC must be in the format XXXXX-YYYYYYY-Z or a 13-digit number.", "Error", is_success=False)
        #     return

        # Check for missing images
        self.customer_image_data = self.get_image_from_label(self.image_label)
        # if not self.customer_image_data:
        #     self.show_custom_message("Please upload the customer image.", "Error", is_success=False)
        #     return
        self.cnic_image_data = self.get_image_data_from_label1()
        # if not self.cnic_image_data:
        #     self.show_custom_message("Please upload the CNIC image.", "Error", is_success=False)
        #     return
        self.other_image_data = self.get_image_data_from_label2()
        # if not self.other_image_data:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        # ############################################################################################################################################################
        self.my_other_img1=self.my_otherLabel1()
        # if not self.my_other_img1:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        self.my_other_img2=self.my_otherLabel2()
        # if not self.my_other_img2:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        self.my_other_img3=self.my_otherLabel3()
        # if not self.my_other_img3:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        self.my_other_img4=self.my_otherLabel4()
        # if not self.my_other_img4:
        #     self.show_custom_message("Please upload other relevant images.", "Error", is_success=False)
        #     return
        ##############################################################################################################################################
        # Prepare sale data
        sale_data = {
            "date": date,
            "time": time,
            "day": day,
            "invoice_number": invoice_number,
            "customer_name": customer_name,
            "father_name": father_name,
            "address": address,
            "cnic": cnic,
            "customer_image": self.customer_image_data,
            "registration": registration,
            "chassis": chassis,
            "engine": engine,
            "make": make,
            "horsepower": horsepower,
            "model": model,
            "color": color,
            "cnic_image": self.cnic_image_data,
            "other_image": self.other_image_data,
            "sale_price": sale_price,  # Stored as integer
            "sale_price_in_words": sale_price_in_words,
            #####################################################################################
            #################(one time1)##################################
            "sale_payment_modes":sale_payment_modes,
            "sale_payment_dates":[date.isoformat() for date in sale_payment_dates],
            "cheque_data":cheque_data,
            #########################################################################################
            "sale_remarks":sale_remarks,
            # Adding the new fields to the sale data
            ################################################################################
            "purchase_car_name": purchase_car_name,
            "purchase_car_price": purchase_car_price,  # Stored as integer
            # "sale_import": sale_import,
            "customer_phone_number1": customer_phone_number1,
            ######################################################################
            "other_image1":self.my_other_img1,
            "other_image2":self.my_other_img2,
            "other_image3":self.my_other_img3,
            "other_image4":self.my_other_img4,
            "Document_note":document_note
        }
        # Check if sale record already exists (based on registration number or other unique identifier)
        existing_sale = db.collection('Car_Selling_Note').where('registration', '==', registration).get()
        if existing_sale:
            self.show_custom_message("Sale record already exists! Please check the vehicle registration.", "Error", is_success=False)
            return
        # Try saving the data to Firebase Firestore
        try:
            doc_ref = db.collection('Car_Selling_Note').add(sale_data)
            document_id = doc_ref[1].id
            db.collection('Car_Selling_Note').document(document_id).update({'document_id': document_id})
            self.show_custom_message("Delivery Sale Note Record is added successfully!", "Success", is_success=True)
            self.clear_sale()
            self.load_selling_car_records()
            self.current_id2=self.get_latest_sales_id_from_db()
            self.invoice_number_field.setText(str(self.current_id2).zfill(6))
            self.fetch_latest_car_from_firestore()
        except Exception as e:
            self.show_custom_message("Please first connect to the internet before using the system.", "Error", is_success=False)
    #######################################################################################################################################################
    ######################################################################################################################################################
    def eventFilter(self, source, event):
        # Check if the source is the QLabel and the event is a mouse button press
        if source == self.My_other1 and event.type()==event.MouseButtonPress:
            self.upload_images_to_label(self.My_other1)
        if source == self.My_other2 and event.type()==event.MouseButtonPress:
            self.upload_images_to_label(self.My_other2)
        if source == self.My_other3 and event.type()==event.MouseButtonPress:
            self.upload_images_to_label(self.My_other3)
        if source == self.My_other4 and event.type()==event.MouseButtonPress:
            self.upload_images_to_label(self.My_other4)
        return super().eventFilter(source, event)
    ####################################################################################################################################################
    #####################################################################################################################################################
    def upload_images_to_label(self, label: QLabel, fill_label: bool = False):  # Accept QLabel as parameter
        # Function to handle multiple image uploads
        options = QFileDialog.Options()
        file_names, _ = QFileDialog.getOpenFileNames(self, "Select Images", "", "Images (*.png *.jpg *.bmp);;All Files (*)", options=options)
        
        for file_name in file_names:
            if file_name:
                pixmap = QPixmap(file_name)
                # Scale the pixmap to fit within the label while maintaining aspect ratio
                pixmap = pixmap.scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                label.setPixmap(pixmap)
                label.setAlignment(Qt.AlignCenter)  # Center the image in the label
    ######################################################################################################################################################
    ###########################################################################################################################################################
    def upload_image(self, label: QLabel, fill_label: bool = False):
        # Function to handle image upload
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Images (*.png *.jpg *.bmp);;All Files (*)", options=options)
        if file_name:
            pixmap = QPixmap(file_name)
            if fill_label:
                # Scale the pixmap to fill the entire label area
                pixmap = pixmap.scaled(label.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation)
            else:
                # Scale the pixmap to fit within the label while maintaining aspect ratio
                pixmap = pixmap.scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            label.setPixmap(pixmap)
            label.setAlignment(Qt.AlignCenter)  # Center the image in the label
    ###########################################################################################################################################
    #######################################################################################
    def search_record(self):
        """Handle searching and displaying a single customer record based on the selected address."""
        #Retrieve customer address and name from the search fields
        self.address = self.address_field.text()
        self.purchase_cars_name=self.purchase_car_names.currentText()  # ComboBox for car names
        
        # Check which field is provided and search accordingly
        if not self.address and not self.purchase_cars_name:
            self.show_custom_message("Please enter an address or customer name to search.", "Error", is_success=False)
            return
        # Initialize the query based on the input field
        if self.address:
            query = db.collection('Car_Selling_Note').where('address', '==', self.address).limit(1)
        elif self.purchase_cars_name:
            query = db.collection('Car_Selling_Note').where('purchase_car_name','==',self.purchase_cars_name).limit(1)
        # Execute the query
        records = query.get()
        if len(records) == 0:
            self.show_custom_message("No records found for the given search criteria.", "Information", is_success=False)
            return
        # Display the first record
        first_record = records[0]  # Get the first record
        data = first_record.to_dict()
        # Display the first record
        # Fetch date, time, and day from the Firestore record
        self.date_field.setDate(QDate.fromString(data.get('date', ''), "yyyy-MM-dd"))
        self.time_field.setTime(QTime.fromString(data.get('time', ''), "HH:mm"))
        self.day_field.setText(data.get('day', ''))
        # Fetch other fields (not date/time/day) from Firebase Storage
        self.invoice_number_field.setText(data.get('invoice_number', ''))
        self.customer_name_field.setCurrentText(data.get('customer_name', ''))
        self.father_name_field.setText(data.get('father_name', ''))
        self.address_field.setText(data.get('address', ''))
        self.cnic_field.setText(data.get('cnic', ''))
        self.registration_field.setText(data.get('registration', ''))
        self.chassis_field.setText(data.get('chassis', ''))
        self.engine_field.setText(data.get('engine', ''))
        self.make_field.setText(data.get('make', ''))
        self.horsepower_field.setText(data.get('horsepower', ''))
        self.model_field.setText(data.get('model', ''))
        self.color_field.setText(data.get('color', ''))
        # Handle image data display from Firebase Storage
        cnic_image_data = data.get('cnic_image', '')
        other_image_data = data.get('other_image', '')
        other_image1_data = data.get('other_image1', '')
        other_image2_data = data.get('other_image2', '')
        other_image3_data = data.get('other_image3', '')
        other_image4_data = data.get('other_image4', '')
        # Display images using the display_image function
        self.display_image(self.cnic_upload_label, cnic_image_data)
        self.display_image(self.attachment_label, other_image_data)
        self.display_image(self.My_other1, other_image1_data)
        self.display_image(self.My_other2, other_image2_data)
        self.display_image(self.My_other3, other_image3_data)
        self.display_image(self.My_other4, other_image4_data)
        # Fetch and display sale-related fields
        #################################################################
        sale_price = data.get('sale_price', 0)  # Integer
        sale_price_in_words = data.get('sale_price_in_words', '')  # String
        ######################################################################################
        sale_remarks = data.get('sale_remarks', '')  # String
        # Display the sale-related fields
        self.Sale_Price_field.setText(str(sale_price))
        self.Sale_Price_Inwords.setText(sale_price_in_words)
        ###########################################################################################
        self.sale_Remarks.setPlainText(sale_remarks)
        # Fetch and display purchase car details, including the purchase car price
        self.purchase_car_names.setCurrentText(data.get('purchase_car_name', ''))
        # Convert purchase_car_price back to string to display in the field
        purchase_car_price = data.get('purchase_car_price', 0)
        self.purchase_car_price.setText(str(purchase_car_price))
        # Fetch and display other purchase-related fields
        self.customer_phone_number1.setText(data.get('customer_phone_number1', ''))
        # self.sale_import_field.setText(data.get('sale_import', ''))
        # Fetch and display the document note
        document_note = data.get('Document_note', '')
        self.Document_Note.setPlainText(document_note)
        ##########################################################################################
        # Display payment dates and modes
        my_sale_payment_dates=data.get('sale_payment_dates', [])
        for i, date_str in enumerate(my_sale_payment_dates):
            date = QDate.fromString(date_str, 'yyyy-MM-dd')
            if i == 0: self.sale_Payment_date1.setDate(date)
            elif i == 1: self.sale_Payment_date2.setDate(date)
            elif i == 2: self.sale_Payment_date3.setDate(date)
            elif i == 3: self.sale_Payment_date4.setDate(date)
            elif i == 4: self.sale_Payment_date5.setDate(date)
        #########################################################################
        sale_payment_modes = data.get('sale_payment_modes', [])
        if sale_payment_modes:
            self.Sale_Mode_of_Payment1.setCurrentText(sale_payment_modes[0] if len(sale_payment_modes) > 0 else '')
            self.Sale_Mode_of_Payment2.setCurrentText(sale_payment_modes[1] if len(sale_payment_modes) > 1 else '')
            self.Sale_Mode_of_Payment3.setCurrentText(sale_payment_modes[2] if len(sale_payment_modes) > 2 else '')
            self.Sale_Mode_of_Payment4.setCurrentText(sale_payment_modes[3] if len(sale_payment_modes) > 3 else '')
            self.Sale_Mode_of_Payment5.setCurrentText(sale_payment_modes[4] if len(sale_payment_modes) > 4 else '')
        ########################################################################################################
        #Display cheque data
        cheque_data = data.get('cheque_data', {})
        self.sale_priceIn_check1.setText(str(cheque_data.get('price_in_cheque1', '')))
        self.sale_priceIn_check2.setText(str(cheque_data.get('price_in_cheque2', '')))
        self.sale_priceIn_check3.setText(str(cheque_data.get('price_in_cheque3', '')))
        self.sale_priceIn_check4.setText(str(cheque_data.get('price_in_cheque4', '')))
        self.sale_priceIn_check5.setText(str(cheque_data.get('price_in_cheque5', '')))
    ################################################################################################################
    def display_image(self, label: QLabel, image_data: str):
        """Display an image in the given QLabel.
        :param label: QLabel where the image will be displayed.
        :param image_data: Base64 encoded image data or raw binary image data.
        """
        try:
            # Check if image_data is base64 encoded
            if isinstance(image_data, str) and image_data.startswith('data:image'):
                # Extract base64 data from the data URL
                base64_data = image_data.split(',')[1]
                # Decode base64 data to binary
                image_data = base64.b64decode(base64_data)
            else:
                # Handle if the image is stored as base64 without data URL structure
                image_data = base64.b64decode(image_data)
            # Create an in-memory binary stream from the image data
            image_stream = io.BytesIO(image_data)
            # Load the image from the stream
            pixmap = QPixmap()
            pixmap.loadFromData(image_stream.read(), 'PNG')  # Use 'PNG' for loading PNG images
            # Set the QPixmap to the QLabel, scale the image to fit and maintain aspect ratio
            label.setPixmap(pixmap.scaled(310, 200, Qt.KeepAspectRatio))
            label.setScaledContents(True)  # Ensure scaling of image to fit the label
        except Exception as e:
            # Handle exceptions (e.g., log error, display a placeholder image, etc.)
            print(f"Error displaying image: {e}")
            label.clear()  # Clear the label if image display fails
        #######################################################################################################
    def clear_sale(self):
        """Clear all sale form fields."""
        self.address_field.clear()
        #self.invoice_number_field.clear()
        self.customer_name_field.setCurrentIndex(0)
        self.father_name_field.clear()
        self.cnic_field.clear()
        self.registration_field.clear()
        self.chassis_field.clear()
        self.engine_field.clear()
        self.make_field.clear()
        self.horsepower_field.clear()
        self.model_field.clear()
        self.color_field.clear()
        self.Sale_Price_field.clear()
        self.Sale_Price_Inwords.clear()
        ###########################################################################
        self.Sale_Mode_of_Payment1.setCurrentIndex(0)  # Corrected here
        self.Sale_Mode_of_Payment2.setCurrentIndex(0)  # Corrected here
        self.Sale_Mode_of_Payment3.setCurrentIndex(0)  # Corrected here
        self.Sale_Mode_of_Payment4.setCurrentIndex(0)  # Corrected here
        self.Sale_Mode_of_Payment5.setCurrentIndex(0)  # Corrected here
        ####################################################
        self.sale_priceIn_check1.clear()
        self.sale_priceIn_check2.clear()
        self.sale_priceIn_check3.clear()
        self.sale_priceIn_check4.clear()
        self.sale_priceIn_check5.clear()
        ###################################################
        self.sale_Payment_date1.setDate(QDate.currentDate())  # Sets the date to today
        self.sale_Payment_date2.setDate(QDate.currentDate())  # Sets the date to today
        self.sale_Payment_date3.setDate(QDate.currentDate())  # Sets the date to today
        self.sale_Payment_date4.setDate(QDate.currentDate())  # Sets the date to today
        self.sale_Payment_date5.setDate(QDate.currentDate())  # Sets the date to today
        ################################################################
        self.sale_Remarks.clear()
        #####################################################################
        self.purchase_car_names.setCurrentIndex(0)
        self.purchase_car_price.clear()
        # self.sale_import_field.clear()
        self.customer_phone_number1.clear()
        #################################################
        self.image_label.setText("No image uploaded")
        self.attachment_label.setText("No image uploaded")
        self.cnic_upload_label.setText("No image uploaded")
        #####################################################
        self.My_other1.setText("No image uploaded")
        self.My_other2.setText("No image uploaded")
        self.My_other3.setText("No image uploaded")
        self.My_other4.setText("No image uploaded")
        self.Document_Note.clear()
#######################################################################################################################################################
#########################################################################################################################################################
###########################################################################################################################################################
#################################################################################################################################################################
################################################################################################################################################################
###########################################################################(Create_inventory_Store_management_view)##############################################
    def create_inventory_store_management_view(self):
        """Create the inventory store management view with a modern look."""
        inventory_store_widget = QWidget()
        ##################################################
        inventory_store_widget.setStyleSheet("background-color:white;")
        #################################################
        layout = QVBoxLayout(inventory_store_widget)
        # Create QTabWidget with modern styling
        tab_widget = QTabWidget()
        tab_widget.setStyleSheet("""
            QTabWidget::pane { /* The tab widget frame */
                border-top: 1px solid #C0C0C0;
                background-color:white;
                border-radius:4px;
            }
            QTabBar::tab {
                background: #E0E0E0;
                color: #333;
                border: 1px solid #C0C0C0;
                padding: 10px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
                min-width:200px;
                                 
            }
            QTabBar::tab:hover {
                background: #B0BEC5;
            }
            QTabBar::tab:selected {
                background-color: #42A5F5;
                color: white;
                font-weight: bold;
            }
            QTabBar::tab:!selected {
                margin-top: 2px;
            }
        """)
        
        # Add Customer Details Tab
        customer_tab =self.create_customer_details_tab()
        tab_widget.addTab(customer_tab, "Add Purchaser Details")
        ###############################################################################
        # Add Supplier Details Tab
        supplier_tab=self.create_supplier_details_tab()
        tab_widget.addTab(supplier_tab, "Add Seller Details")
        ########################################################################################
        # Add Stock Details Tab (Placeholder for stock display)
        stock_tab = self.create_stock_tab()
        tab_widget.addTab(stock_tab, "Show Complete Stock")
        ######################################################################################
        sale_and_purchase=self.create_sale_purchase_stock_tab()
        tab_widget.addTab(sale_and_purchase, "Sales and Purchases Summary")
        #####################################################################################
        # Add the QTabWidget to the layout
        layout.addWidget(tab_widget)
        return inventory_store_widget
########################################################################################################################################
########################################################################################################################################
###############################################(it is Customer Tab)####################################################################
    def create_customer_details_tab(self):
        """Create the tab for adding customer details with a modern table and show already registered customers."""
        customer_widget = QWidget()
        layout = QVBoxLayout(customer_widget)
        # Create a form layout for fields (2 fields per row)
        form_layout = QFormLayout()
        form_layout.setSpacing(20)  # Spacing between rows
        ##############################################################################################
        #Customer Fields
        self.customer_id=QLineEdit()
        self.customer_id.setPlaceholderText("Purchaser ID")
        self.current_id = self.get_latest_customer_id_from_db()
        self.customer_id.setText(str(self.current_id).zfill(6))
        ###########################################################
        self.customer_id.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ################################################################
        self.customer_id.setFixedWidth(300)  # Fixed width for fields
        # Use a lambda function to call the method without issues
        self.customer_id.mousePressEvent = lambda event: self.generate_dynamic_customer_id()
        # Assuming this variable is initialized somewhere in your class
        ##########################################################################################################################################
        self.customer_type=QComboBox()
        self.customer_type.addItems(["Select Purchaser_Type","Dealer/Showroom", "Individual"])
        self.customer_type.setFixedWidth(300)
        self.customer_type.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ###############################################################################################################################
        self.customer_name = QLineEdit()
        self.customer_name.setPlaceholderText("Purchaser Name")
        self.customer_name.setFixedWidth(300)
        self.customer_name.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ###################################################################
        self.phone = QLineEdit()
        self.phone.setPlaceholderText("CNIC NO")
        self.phone.setFixedWidth(300)
        self.phone.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ########################################################################
        self.customer_email = QLineEdit()
        self.customer_email.setPlaceholderText("Email")
        self.customer_email.setFixedWidth(300)
        self.customer_email.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #########################################################################
        self.address = QLineEdit()
        self.address.setPlaceholderText("Purchaser Address")
        self.address.setFixedWidth(300)
        self.address.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-size:14px;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #########################################################################
        self.customer_phone=QLineEdit()
        self.customer_phone.setPlaceholderText("Purchaser Phone")
        self.customer_phone.setFixedWidth(300)
        self.customer_phone.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ################################################################
        self.father_name = QLineEdit()
        self.father_name.setPlaceholderText("Father Name")
        self.father_name.setFixedWidth(300)
        self.father_name.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #############################################################
        # Adding fields in pairs to the form layout
        form_layout.addRow(self.create_bold_label("Purchaser ID"),self.customer_id)
        form_layout.addRow(self.create_bold_label("Purchaser Type"),self.customer_type)
        form_layout.addRow(self.create_bold_label("Purchaser Name"),self.customer_name)
        form_layout.addRow(self.create_bold_label("Father Name"),self.father_name)
        form_layout.addRow(self.create_bold_label("CNIC"), self.phone)
        form_layout.addRow(self.create_bold_label("Email"),self.customer_email)
        form_layout.addRow(self.create_bold_label("Phone Number"),self.customer_phone)
        form_layout.addRow(self.create_bold_label("Address"),self.address)
        # Upload Customer Image
        upload_image_button = QPushButton("Upload Purchaser Image")
        upload_image_button.setFixedWidth(200)
        ##############################################################
        upload_image_button.setStyleSheet("""
            QPushButton {
                background-color:#5F6F65;/*Modern grey color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#808D7C;
            }
            QPushButton:pressed{
                background-color:#9CA986;
                color:black;
            }
        """)
        ############################################################
        upload_image_button.clicked.connect(self.upload_customer_image)
        # Define self.customer_image_label as an instance variable for displaying the image
        self.customer_image_label = QLabel("No image uploaded")  # Instance variable
        # Set a fixed size for the image label and add a border
        self.customer_image_label.setFixedSize(300, 200)
        self.customer_image_label.setStyleSheet("""
            QLabel{
                border: 2px solid black;
                padding: 5px;
                background-color: white;
                text-align: center;
            }
        """)
        self.customer_image_label.setAlignment(Qt.AlignCenter)
        form_layout.addRow(upload_image_button, self.customer_image_label)
        # Add Form Layout to Main Layout
        layout.addLayout(form_layout)
        # Buttons for Save, Search, and Clear
        button_layout = QHBoxLayout()
        save_button = QPushButton("Save Purchaser")
        save_button.clicked.connect(self.save_customer)
        search_button = QPushButton("Search Purchaser")
        search_button.clicked.connect(self.search_customer)  # Connect to search function
        clear_button = QPushButton("Clear Record")
        clear_button.clicked.connect(self.clear_form)
        ###################################################################################
        search_button.setStyleSheet("""
            QPushButton {
                background-color:#181C14;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#3C3D37;
            }
            QPushButton:pressed{
                background-color:#697565;
            }
        """)
        #####################################################################################################################
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #018749;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0D7C66;
            }
            QPushButton:pressed {
                background-color: #018749;
            }
        """)
        ############################################################################
        clear_button.setStyleSheet("""
            QPushButton {
                background-color:#624E88;/* Modern blue color */
                color: white;
                border: none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color:#8967B3;
            }
            QPushButton:pressed {
                background-color:#CB80AB;
            }
        """)
        #######################################################################################################################
        save_button.setFixedWidth(230)
        save_button.setFixedHeight(35)
        search_button.setFixedWidth(230)
        search_button.setFixedHeight(35)
        clear_button.setFixedWidth(230)
        clear_button.setFixedHeight(35)
        #####################################################################
        button_layout.addWidget(save_button)
        button_layout.addWidget(search_button)  # Add search button
        button_layout.addWidget(clear_button)
        #########################################################33################
        ###########################################################################################################################
        layout.addLayout(button_layout)
        ############################################################
        # Customer Table with modern design
        #########################################################################################################################################3
        button_layout_2=QHBoxLayout()
        self.update_button=QPushButton("Update Records")
        self.delete_button=QPushButton("Delete Records")
        ############################################################################################
        self.delete_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        ############################################################################
        self.update_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;  /* Modern blue color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0069d9;
            }
            QPushButton:pressed {
                background-color: #0056b3;
            }
        """)
        #################################################
        ######################################################
        self.update_button.setFixedWidth(150)
        self.update_button.setFixedHeight(35)
        ######################################################
        self.delete_button.setFixedWidth(150)
        self.delete_button.setFixedHeight(35)
        ############################################
        ###############################################
        self.update_button.clicked.connect(self.update_customer)  # Connect to update function
        self.delete_button.clicked.connect(self.delete_customer)  # Connect to delete function
        #########################################################
        button_layout_2.addWidget(self.update_button)
        button_layout_2.addWidget(self.delete_button)
        layout.addLayout(button_layout_2)
        button_layout_2.setSpacing(0)  # Set the spacing between widgets to 0
        button_layout_2.setContentsMargins(0,20,0,0)
        ####################################################################
        self.table1_layout=QHBoxLayout()
        self.customer_table = QTableWidget(0,7)  # 6 columns
        self.customer_table.setHorizontalHeaderLabels(["ID", "Name", "CNIC", "Email", "Address", "Father Name","Document ID"])
        self.customer_table.setStyleSheet("""
            QHeaderView::section {
                background-color:#825B32;
                color: white;
                font-weight: bold;
            }
            QTableWidget {
                background-color: #FAFAFA;
                alternate-background-color: #E0E0E0;
                gridline-color: #B0BEC5;
            }
            QTableWidget::item {
                text-align: center;
            }
        """)

        # Set the table to stretch the columns to fit the entire table's width
        header = self.customer_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)  # Stretch columns to fit the table width
        # Enable alternating row colors for better readability
        self.customer_table.setAlternatingRowColors(True)
        self.customer_table.setFixedHeight(250)  # Fixed height for table
        # Add the table widget to the layout with a stretch factor
        self.table1_layout.addWidget(self.customer_table)
        self.table1_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        layout.addLayout(self.table1_layout,1)
        # Buttons for Update and Delete
        ####################################################################################################################################################
        # Load sales table
        ############################
        #############################################################################################3
        self.load_customer_records()
        return customer_widget
################################################################################################################################################################################
#############################################################################################################################################################
    def create_bold_label(self, text):
        label = QLabel(text)
        label.setStyleSheet("""
            QLabel {
                font-size:14px;  /* Set font size */
                font-weight:bold;  /* Make the text bold */
                color:black;  /* Set text color */
            }
        """)
        return label
##########################################################################################################################################
    def get_latest_customer_id_from_db(self):
        #Reference to the 'Customers' collection in Firestore
        customers_ref = db.collection('Customers')
        customers = customers_ref.stream()  # Get all customer documents
        customer_ids = []
        for customer in customers:
            data = customer.to_dict()  # Convert Firestore document to dictionary
            customer_ids.append(int(data.get('customer_id', 0)))  # Fetch the 'customer_id' field value
        if customer_ids:
            # If customers exist, find the maximum customer_id and add 1
            max_id = max(customer_ids)
            return max_id + 1
        else:
            # If there are no customers, start from 1
            return 1  # This will trigger '00001' when formatted
    ####################################################################################
    def generate_dynamic_customer_id(self):
        # Generate the next ID with leading zeros (e.g., 000001)
        dynamic_customer_id = str(self.current_id).zfill(6)
        self.customer_id.setText(dynamic_customer_id)  # Update the QLineEdit field
        # Increment the ID for the next call
        self.current_id += 1
################################################################################################################################################
    def clear_form(self):
        """Clear the form fields and reset the image label."""
        #self.customer_id.clear()
        self.customer_name.clear()
        self.phone.clear()
        self.customer_email.clear()
        self.address.clear()
        self.father_name.clear()
        self.customer_phone.clear()
        self.customer_image_label.clear()
        self.customer_type.setCurrentIndex(0)
        self.customer_image_label.setText("No image uploaded")
        self.refresh_customer_table()
##############################################################
#################################################################################################################################
    def delete_record_from_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Customers').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
            self.populate_customer_combo_box()
            #Update the customer_id to the next available ID after saving
            self.current_id = self.get_latest_customer_id_from_db()
            self.customer_id.setText(str(self.current_id).zfill(6))
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
############################################################################################################################
    def delete_customer(self):
        """Delete the customer record based on the selected row in the table."""
        # Get the selected items in the table
        selected_items = self.customer_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return

        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this customer?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)

                # Get the row index of the first selected item
                row = selected_items[0].row()

                # Retrieve the document_id from the 6th column (index 5)
                document_id = self.customer_table.item(row,6).text()  # Changed index to 5

                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.delete_record_from_firestore(document_id)

                # Remove the row from the table
                self.customer_table.removeRow(row)

                # Show a success message to the user
                self.show_custom_message("Record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                #self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
                pass
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
#####################################################################################################################################33
#####################################################################################################################################
    def search_customer(self):
        """Handle searching and displaying a single customer record based on the selected criteria."""
        # Retrieve input values from the search fields
        customer_id_text = self.customer_id.text()
        customer_name_text = self.customer_name.text()
        cnic_text = self.phone.text()  # Assuming the field for CNIC is named 'phone'
        
        # Ensure at least one search field is provided
        if not (customer_id_text or customer_name_text or cnic_text):
            self.show_custom_message("Please enter Customer ID, Name, or CNIC to Search Records.", "Error", is_success=False)
            return
        
        # Initialize the query (assuming a Firestore collection named 'Customers')
        query = db.collection('Customers')  # Replace with the actual Firestore reference
        
        # Build the query with dynamic criteria based on what has been entered
        filters = []
        if customer_name_text:
            filters.append(('customer_name', '==', customer_name_text))
        if cnic_text:
            filters.append(('cnic', '==', cnic_text))
        
        for filter_condition in filters:
            query = query.where(*filter_condition)
        
        # Limit the query to one result for a single customer
        query = query.limit(1)
        
        # Execute the query
        try:
            records = query.get()
        except Exception as e:
            self.show_custom_message(f"Error while fetching records: {str(e)}", "Error", is_success=False)
            return
        
        if len(records) == 0:
            self.show_custom_message("No records found for the given criteria.", "Information", is_success=False)
            return
        
        # Display the first record
        first_record = records[0]
        data = first_record.to_dict()
        
        # Ensure `self.address` and other fields are valid widgets before using `setText`
        if hasattr(self.address, 'setText'):
            self.address.setText(data.get('address', ''))
        else:
            print("Warning: `self.address` is not a valid widget. Skipping setting text.")

        # Populate other fields safely
        self.customer_id.setText(data.get('customer_id', ''))
        self.customer_type.setCurrentText(data.get('customer_type', ''))
        self.customer_name.setText(data.get('customer_name', ''))
        self.father_name.setText(data.get('father_name', ''))
        self.phone.setText(data.get('cnic', ''))
        self.customer_email.setText(data.get('email', ''))
        self.customer_phone.setText(data.get('customer_phone', ''))
        
        # Handle image data display
        customer_image_data = data.get('customer_image', '')  # Assumes customer_image is a URL or base64 string
        self.display_image(self.customer_image_label, customer_image_data)
        
        # Clear existing data in the table
        self.customer_table.setRowCount(0)
        try:
            # Fetch the filtered data from Firestore
            customers_ref = query.stream()
            # Iterate over the records and insert them into the table
            for customer in customers_ref:
                customer_data = customer.to_dict()
                row_position = self.customer_table.rowCount()
                self.customer_table.insertRow(row_position)
                self.customer_table.setItem(row_position, 0, QTableWidgetItem(customer_data.get('customer_id', '')))
                self.customer_table.setItem(row_position, 1, QTableWidgetItem(customer_data.get('customer_name', '')))
                self.customer_table.setItem(row_position, 2, QTableWidgetItem(customer_data.get('cnic', '')))
                self.customer_table.setItem(row_position, 3, QTableWidgetItem(customer_data.get('email', '')))
                self.customer_table.setItem(row_position, 4, QTableWidgetItem(customer_data.get('address', '')))
                self.customer_table.setItem(row_position, 5, QTableWidgetItem(customer_data.get('father_name', '')))
                self.customer_table.setItem(row_position, 6, QTableWidgetItem(customer_data.get('document_id', '')))
        except Exception as e:
            self.show_custom_message(f"Error refreshing customer table: {e}", "Error", is_success=False)
        
        # Store the document ID for later use
        self.stored_customer_id = first_record.id
        return self.stored_customer_id
########################################################################################################################################
    def update_customer(self):
        """Update the selected customer record."""
        # Retrieve inputs, using local variables to avoid overwriting the original widgets
        customer_id = self.customer_id.text()
        my_customer_type=self.customer_type.currentText()
        customer_name = self.customer_name.text()
        phone = self.phone.text()
        email = self.customer_email.text()
        my_customer_phone=self.customer_phone.text()
        address = self.address.text()
        father_name = self.father_name.text()
        # Validation regex
        customer_name_regex = r'^[A-Za-z\s]{3,}$'
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{3,}$'
        phone_regex = r'^\d{5}-\d{7}-\d{1}$|^\d{13}$'
        # Check for empty fields
        # if not customer_id or not my_customer_type or not customer_name or not email or not address or not father_name or not my_customer_phone:
        #     self.show_custom_message("All fields are required.", "Error", is_success=False)
        #     return
        # Validate customer name
        # if not re.match(customer_name_regex, customer_name):
        #     self.show_custom_message("Purchaser Name must be at least 3 characters long and contain only letters.", "Error", is_success=False)
        #     return
        # # Validate phone number
        # if not re.match(phone_regex, phone):
        #     self.show_custom_message("CNIC must be in the format XXXXX-YYYYYYY-Z or a 13-digit number.", "Error", is_success=False)
        #     return
        # customer_phone_regex=r'^\d{11}$'
        # if not re.match(customer_phone_regex,my_customer_phone):
        #     self.show_custom_message("Phone number must be 11 digits.", "Error", is_success=False)
        #     return
        # # Validate email
        # if not re.match(email_regex, email):
        #     self.show_custom_message("Invalid email format.", "Error", is_success=False)
        #     return
        # Validate image
        image_data=self.get_image_data_from_label9()
        # Prepare data for Firebase
        updated_data={
            'customer_type':my_customer_type,
            'customer_name':customer_name,
            'cnic':phone,
            'email':email,
            'address':address,
            'customer_phone':my_customer_phone,
            'father_name':father_name,
            'customer_image':image_data  # Include the base64 image data
        }
        # Check if the customer ID exists in Firestore
        customer_ref = db.collection('Customers').where('customer_id', '==', customer_id).get()
        if not customer_ref:
            self.show_custom_message("Purchaser ID does not exist! Please check the ID and try again.", "Error", is_success=False)
            return
        # Try updating the data in Firebase Firestore
        try:
            for doc in customer_ref:
                doc_ref = db.collection('Customers').document(doc.id)
                doc_ref.update(updated_data)
            self.show_custom_message("Record updated successfully!", "Success", is_success=True)
            self.refresh_customer_table()
            self.load_customer_records1()
            self.clear_form()
            self.populate_customer_combo_box()
            self.current_id = self.get_latest_customer_id_from_db()
            self.customer_id.setText(str(self.current_id).zfill(6))
            # Clear the input fields after updating
        except Exception as e:
            self.show_custom_message("An error occurred while updating the record. Please try again.", "Error", is_success=False)   
            #pass         
###############################################################################################################################
    def refresh_customer_table(self):
        """Refresh the table to show the latest customer records."""
        # Clear existing data
        self.customer_table.setRowCount(0)
        # Fetch updated data from Firestore
        try:
            customers_ref = db.collection('Customers').stream()
            for customer in customers_ref:
                customer_data = customer.to_dict()
                # Insert the customer data into the table
                row_position = self.customer_table.rowCount()
                self.customer_table.insertRow(row_position)
                self.customer_table.setItem(row_position,0,QTableWidgetItem(customer_data.get('customer_id', '')))
                self.customer_table.setItem(row_position,1,QTableWidgetItem(customer_data.get('customer_name', '')))
                self.customer_table.setItem(row_position,2,QTableWidgetItem(customer_data.get('cnic', '')))
                self.customer_table.setItem(row_position,3,QTableWidgetItem(customer_data.get('email', '')))
                self.customer_table.setItem(row_position,4,QTableWidgetItem(customer_data.get('address', '')))
                self.customer_table.setItem(row_position,5,QTableWidgetItem(customer_data.get('father_name', '')))
                self.customer_table.setItem(row_position,6,QTableWidgetItem(customer_data.get('document_id', '')))
                # Handle image data if necessary
        except Exception as e:
            self.show_custom_message(f"Error refreshing customer table: {e}", "Error", is_success=False)
            #pass
######################################################################################################################################
    def display_image(self, label: QLabel, image_data: str):
        """Display an image in the given QLabel.
        :param label: QLabel where the image will be displayed.
        :param image_data: Base64 encoded image data or raw binary image data.
        """
        try:
            # Check if image_data is base64 encoded
            if isinstance(image_data, str) and image_data.startswith('data:image'):
                # Extract base64 data from the data URL
                base64_data = image_data.split(',')[1]
                # Decode base64 data to binary
                image_data = base64.b64decode(base64_data)
            else:
                # Handle if the image is stored as base64 without data URL structure
                image_data = base64.b64decode(image_data)

            # Create an in-memory binary stream from the image data
            image_stream = io.BytesIO(image_data)

            # Load the image from the stream
            pixmap = QPixmap()
            pixmap.loadFromData(image_stream.read(), 'PNG')  # Use 'PNG' for loading PNG images

            # Set the QPixmap to the QLabel, scale the image to fit and maintain aspect ratio
            label.setPixmap(pixmap.scaled(310, 200, Qt.KeepAspectRatio))
            label.setScaledContents(True)  # Ensure scaling of image to fit the label

        except Exception as e:
            # Handle exceptions (e.g., log error, display a placeholder image, etc.)
            print(f"Error displaying image: {e}")
            label.clear()  # Clear the label if image display fails
#############################################################################################################################3
##############################################################################################################################
###########################################################################################
    def load_customer_records(self):
        """Load existing customer records from Firebase Firestore and display them in the table."""
        try:
            customers = db.collection('Customers').get()
            for customer in customers:
                customer_data = customer.to_dict()
                # Insert data into the table
                row_position = self.customer_table.rowCount()
                self.customer_table.insertRow(row_position)
                self.customer_table.setItem(row_position,0,QTableWidgetItem(customer_data.get('customer_id', '')))
                self.customer_table.setItem(row_position,1,QTableWidgetItem(customer_data.get('customer_name', '')))
                self.customer_table.setItem(row_position,2,QTableWidgetItem(customer_data.get('cnic', '')))
                self.customer_table.setItem(row_position,3,QTableWidgetItem(customer_data.get('email', '')))
                self.customer_table.setItem(row_position,4,QTableWidgetItem(customer_data.get('address', '')))
                self.customer_table.setItem(row_position,5,QTableWidgetItem(customer_data.get('father_name', '')))
                self.customer_table.setItem(row_position,6,QTableWidgetItem(customer_data.get('document_id', '')))
        except Exception as e:
            self.show_custom_message(f"Error loading customers: {e}", "Error", is_success=False)
            #pass
#######################################################################################################################################
    def upload_customer_image(self):
        """Handle customer image upload and display it in the label."""
        file_name, _ = QFileDialog.getOpenFileName(None, "Upload Image", "", "Image Files (*.png *.jpg *.jpeg)")
        if file_name:
            #Load the image from the selected file
            pixmap = QPixmap(file_name)
            
            # Resize the pixmap to fit within the label dimensions (optional, if the image is too large)
            pixmap = pixmap.scaled(self.customer_image_label.size(), aspectRatioMode=1)  # Keep aspect ratio
            
            # Set the pixmap (image) onto the label
            self.customer_image_label.setPixmap(pixmap)
###########################################################################################################################################
    def get_image_data_from_label9(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap = self.customer_image_label.pixmap()
        if pixmap:
            image = pixmap.toImage()

            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)

            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')

            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
###########################################################################################
    def save_customer(self):
        """Handle saving customer details with validation and Firebase image storage."""
        #Retrieve inputs, using local variables to avoid overwriting the original widgets
        customer_id=self.customer_id.text()
        my_customer_type=self.customer_type.currentText()
        customer_name=self.customer_name.text()
        phone=self.phone.text()
        email=self.customer_email.text()
        my_customer_phone=self.customer_phone.text()
        address=self.address.text()
        father_name=self.father_name.text()
        # Validation regex
        customer_name_regex = r'^[A-Za-z\s]{3,}$'
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{3,}$'
        #phone_regex = r'^\d{5}-\d{7}-\d{1}$|^\d{13}$'
        phone_regex = r'^(\d{5}-\d{7}-\d{1}|\d{13}|0\d{2}\d{8}|\d{11}|\+92\d{10})$'
        image_data = self.get_image_data_from_label9()
        # Check for empty fields
        # if not customer_id or not my_customer_type or not customer_name or not email or not address or not father_name or not my_customer_phone:
        #     self.show_custom_message("All fields are required.", "Error", is_success=False)
        #     return

        # # Validate customer name
        # if not re.match(customer_name_regex, customer_name):
        #     self.show_custom_message("Customer Name must be at least 3 characters long and contain only letters.", "Error", is_success=False)
        #     return

        # # Validate phone number
        # if not re.match(phone_regex, phone):
        #     self.show_custom_message("CNIC must be in the format XXXXX-YYYYYYY-Z or a 13-digit number.", "Error", is_success=False)
        #     return
        # #########################################3
        # customer_phone_regex=r'^\d{11}$'
        # if not re.match(customer_phone_regex,my_customer_phone):
        #     self.show_custom_message("Phone number must be 11 digits.", "Error", is_success=False)
        #     return
        # # Validate email
        # if not re.match(email_regex, email):
        #     self.show_custom_message("Invalid email format.", "Error", is_success=False)
        #     return

        # Validate image
        # Prepare data for Firebase
        customer_data = {
            'customer_id':customer_id,
            'customer_type':my_customer_type,
            'customer_name':customer_name,
            'cnic': phone,
            'email': email,
            'address': address,
            'customer_phone':my_customer_phone,
            'father_name': father_name,
            'customer_image':image_data  # Include the base64 image data
        }
        #Check if the customer_id already exists in Firestore
        existing_customer_id = db.collection('Customers').where('customer_id', '==', customer_id).get()
        if existing_customer_id:
            self.show_custom_message("Purchaser ID already exists! Please use another customer ID.", "Error", is_success=False)
            return

        # Check if the customer already exists in Firestore
        existing_customer = db.collection('Customers').where('email', '==', email).get()
        if existing_customer:
            self.show_custom_message("Purchaser already exists! Please use another email address.", "Error", is_success=False)
            return
        # Try saving the data to Firebase Firestore
        try:
            doc_ref = db.collection('Customers').add(customer_data)
            document_id = doc_ref[1].id
            db.collection('Customers').document(document_id).update({'document_id': document_id})
            self.show_custom_message("Record is Added successfully!", "Success", is_success=True)
            
            # Add saved customer record to the table
            row_position = self.customer_table.rowCount()
            self.customer_table.insertRow(row_position)
            self.customer_table.setItem(row_position, 0, QTableWidgetItem(customer_id))
            self.customer_table.setItem(row_position, 1, QTableWidgetItem(customer_name))
            self.customer_table.setItem(row_position, 2, QTableWidgetItem(phone))
            self.customer_table.setItem(row_position, 3, QTableWidgetItem(email))
            self.customer_table.setItem(row_position, 4, QTableWidgetItem(address))
            self.customer_table.setItem(row_position, 5, QTableWidgetItem(father_name))
            self.customer_table.setItem(row_position, 6, QTableWidgetItem(document_id))
            self.populate_customer_combo_box()
            self.load_customer_records1()
            #####################################
            #Update the customer_id to the next available ID after saving
            self.current_id = self.get_latest_customer_id_from_db()
            self.customer_id.setText(str(self.current_id).zfill(6))
            # Clear the input fields after saving
            self.clear_form()
        except Exception as e:
            #self.show_custom_message("Please first connect to the internet before using the system", "Error", is_success=False)
            pass
#########################################################################################################################################################
#########################################################################################################################################################
    def fetch_customer_details(self):
        customers_ref = db.collection('Customers')
        docs = customers_ref.stream()
        customer_details = {}
        customer_names = []
        for doc in docs:
            data = doc.to_dict()
            customer_name = data.get('customer_name')
            if customer_name:
                customer_names.append(customer_name)
                customer_details[customer_name] = {
                    'father_name':data.get('father_name',''),
                    'address':data.get('address',''),
                    'customer_phone':data.get('customer_phone',''),
                    'cnic': data.get('cnic', ''),
                    'customer_image': data.get('customer_image', '')
                }
        return customer_details, customer_names
    ##################################################################
    def populate_customer_combo_box(self):
        self.customer_name_field.clear()
        self.customer_details, customer_names = self.fetch_customer_details()
        # Add "Please Select Customer" option and then the customer names
        self.customer_name_field.addItems(["Please Select Purchaser"] + customer_names)
    ############################################################################################
    def on_customer_selection_changed(self, index):
        #Ensure fields are cleared if no customer is selected
        if index == 0:
            if hasattr(self, 'father_name_field'):
                self.father_name_field.setText('')
            if hasattr(self, 'cnic_field'):
                self.cnic_field.setText('')
            if hasattr(self, 'address_field'):
                self.address_field.setText('')
            if hasattr(self,'customer_phone_number1'):
                self.customer_phone_number1.setText('')
            if hasattr(self, 'image_label'):
                self.image_label.clear()
        else:
            customer_name = self.customer_name_field.currentText()
            details = self.customer_details.get(customer_name, {})

            if hasattr(self, 'father_name_field'):
                self.father_name_field.setText(details.get('father_name', ''))
            #####################################################################################
            if hasattr(self, 'address_field'):
                self.address_field.setText(details.get('address', ''))
            ############################################################################################
            if hasattr(self,'customer_phone_number1'):
                self.customer_phone_number1.setText(details.get('customer_phone', ''))
            ########################################################################################
            if hasattr(self,'cnic_field'):
                self.cnic_field.setText(details.get('cnic', ''))
            ################################################################################################
            customer_image = details.get('customer_image', '')
            if hasattr(self, 'image_label') and customer_image:
                try:
                    image_data = base64.b64decode(customer_image)
                    pixmap = QPixmap()
                    pixmap.loadFromData(image_data, 'PNG')
                    self.image_label.setPixmap(pixmap.scaled(310, 200, Qt.KeepAspectRatio))
                except Exception as e:
                    print(f"Error decoding image: {e}")
                    self.image_label.clear()
            else:
                self.image_label.clear()
#########################################################################################################################################################
#########################################################################################################################################################
    def show_custom_message(parent, message, title, is_success=True):
        """Display a custom message dialog."""
        # Create a QDialog for custom layout
        dialog = QDialog(parent)
        dialog.setWindowTitle(title)

        # Set fixed height and width for the dialog
        dialog.setFixedHeight(120)
        dialog.setFixedWidth(480)

        # Set dialog layout
        layout = QVBoxLayout()

        # Add message label
        label = QLabel(message)
        layout.addWidget(label, alignment=Qt.AlignCenter)

        # Add OK button to the center
        button_box = QDialogButtonBox(QDialogButtonBox.Ok)

        # Use lambda to connect dialog.accept to the button click
        button_box.accepted.connect(dialog.accept)

        # Add the button to the layout and center it
        layout.addWidget(button_box, alignment=Qt.AlignCenter)

        # Set layout and custom style
        dialog.setLayout(layout)

        # Set custom styles
        dialog.setStyleSheet(f"""
            QDialog {{
                background-color: {"#e0f7fa" if is_success else "#ffebee"};
                border: 2px solid {"green" if is_success else "red"};
                border-radius: 10px;
            }}
            QLabel {{
                color: black;
                font-size: 16px;
                font-weight: bold;
            }}
            QPushButton {{
                background-color: {"#81c784" if is_success else "#e57373"};
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 14px;
            }}
        """)

        # Execute the dialog
        dialog.exec_()
############################################################################################################################################################# 
###################################################################################################################################################################
########################################################################################################################################################################3
    def create_purchase_management_view(self):
        #Create the main widget
        purchase_management_widget=QWidget()
        ###############################################################
        purchase_management_widget.setStyleSheet("background-color:white;")
        #####################################################################
        layout=QVBoxLayout(purchase_management_widget)
        purchase_management_widget.setStyleSheet("background-color:white;")
        # Step (1): Add Tab bar with name "Add Sale Car Details"
        tab_widget = QTabWidget()
        # Style for the tab bar
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #C0C0C0;
                border-radius:10px;
            }
            QTabBar::tab {
                background-color: #F1F1F1;
                border: 1px solid #C0C0C0;
                border-radius: 5px 5px 0 0;
                padding: 10px;
                min-width: 120px;
                font-weight: bold;
                color: #333;
            }
            QTabBar::tab:selected {
                background-color: #007BFF;
                color: white;
            }
        """)

        # Create the form widget inside the tab
        form_widget = QWidget()
        grid_layout = QGridLayout(form_widget)
        # Define fixed size for fields
        field_width = 270
        field_height = 35
        Purchase_button_width = 150
        Purchase_button_height = 40
        spacing = 10

        # General stylesheet for the form
        form_widget.setStyleSheet("""
            QWidget {
                font-family: Arial, sans-serif;
                font-size:12px;
            }
            QLabel {
                font-weight: bold;
                color: #333;
                font-size:15px;
                font-weight:bold;
            }
            QLineEdit, QDateEdit, QTimeEdit {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #F9F9F9;
                font-weight:bold;
                font-size:14px;
            }
            QPushButton {
                background-color: #007BFF;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #003d7a;
            }
            QLabel[style*="border"] {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #F9F9F9;
            }
            QVBoxLayout,QHBoxLayout {
                spacing:0px;
            }
        """)
        # Step (2): Add the Title of the form INSIDE the Tab bar
        title_label = QLabel("Purchase Receipt")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 24px; font-weight: bold; margin-bottom: 20px; color: #333;")
        grid_layout.addWidget(title_label,0,0,1,4)  # Span the title across 4 columns
        # Step (3): Add date, time, day, and invoice number
        self.Purchase_date_field=QDateEdit(QDate.currentDate())
        self.Purchase_date_field.setAlignment(Qt.AlignCenter)
        self.Purchase_date_field.setContentsMargins(0,0,0,30)
        # self.Purchase_date_field.setReadOnly(True)
        ###########################################################33
        self.Purchase_date_field.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QDateEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_date_field.setFixedWidth(field_width)
        self.Purchase_date_field.setFixedHeight(field_height)
        #########################################################
        self.Purchase_time_field=QTimeEdit(QTime.currentTime())
        self.Purchase_time_field.setAlignment(Qt.AlignCenter)
        self.Purchase_time_field.setReadOnly(True)  # Set the field to read-only
        #####################################################
        self.Purchase_time_field.setFixedWidth(field_width)
        self.Purchase_time_field.setFixedHeight(field_height)
        self.Purchase_time_field.setContentsMargins(0,0,0,30)
        self.Purchase_time_field.setStyleSheet("""
            QTimeEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QTimeEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        #################################################################33
        self.Purchase_day_field=QLineEdit(QDate.currentDate().toString("dddd"))  # Automatically fill day
        self.Purchase_day_field.setAlignment(Qt.AlignCenter)
        self.Purchase_day_field.setReadOnly(True)  # Set the field to read-only
        self.Purchase_day_field.setFixedWidth(field_width)
        self.Purchase_day_field.setFixedHeight(field_height)
        self.Purchase_day_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        #############################################################################################################3
        self.Purchase_invoice_number_field=QLineEdit()  # Random invoice number
        self.current_id3=self.get_latest_purchase_id_from_db()
        self.Purchase_invoice_number_field.setText(str(self.current_id3).zfill(6))
        ########################################################################################
        self.Purchase_invoice_number_field.setAlignment(Qt.AlignCenter)
        self.Purchase_invoice_number_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_invoice_number_field.setFixedWidth(field_width)
        self.Purchase_invoice_number_field.setFixedHeight(field_height)
        self.Purchase_invoice_number_field.mousePressEvent=lambda event:self.generate_new_invoice_number()
        ############################################################################################################
        # Add fields to grid layout
        grid_layout.addWidget(QLabel("Purchase Date:"), 1,0)
        grid_layout.addWidget(self.Purchase_date_field, 1, 1)
        grid_layout.addWidget(QLabel("Time:"), 1, 2)
        grid_layout.addWidget(self.Purchase_time_field, 1, 3)
        grid_layout.addWidget(QLabel("Day:"), 2, 0)
        grid_layout.addWidget(self.Purchase_day_field, 2, 1)
        grid_layout.addWidget(QLabel("Purchase Invoice No:"), 2, 2)
        grid_layout.addWidget(self.Purchase_invoice_number_field, 2, 3)
        ######################################################################################################################
        #Step (4): Add customer details
        self.Supplier_name_field=QComboBox()
        # Populate the combo box with example customer names
        #self.customer_name_field.addItems(["Select Customer Name"]+customer_names)
        #customer_name_field = QLineEdit()
        self.Supplier_name_field.setFixedHeight(field_height)
        self.Supplier_name_field.setFixedWidth(field_width)
        #self.customer_name_field.currentIndexChanged.connect(self.on_customer_selection_changed)
        #self.populate_customer_combo_box()
        self.populate_supplier_combo_box()
        #############################################################
        self.Supplier_name_field.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                font-weight:bold;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ####################################################
        self.Supplier_father_name_field=QLineEdit(self)
        self.Supplier_father_name_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Supplier_father_name_field.setFixedWidth(field_width)
        self.Supplier_father_name_field.setFixedHeight(field_height)
        self.Supplier_father_name_field.setAlignment(Qt.AlignCenter)
        ######################################################
        self.Supplier_address_field=QLineEdit(self)
        self.Supplier_address_field.setFixedWidth(field_width)
        self.Supplier_address_field.setFixedHeight(field_height)
        self.Supplier_address_field.setAlignment(Qt.AlignCenter)
        self.Supplier_address_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ##########################################################################################################
        self.my_Supplier_phonenumber=QLineEdit(self)
        self.my_Supplier_phonenumber.setFixedWidth(field_width)
        self.my_Supplier_phonenumber.setFixedHeight(field_height)
        self.my_Supplier_phonenumber.setAlignment(Qt.AlignCenter)
        self.my_Supplier_phonenumber.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        #######################################################################################################################
        self.Supplier_cnic_field = QLineEdit(self)
        self.Supplier_cnic_field.setFixedWidth(field_width)
        self.Supplier_cnic_field.setFixedHeight(field_height)
        self.Supplier_cnic_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Supplier_cnic_field.setAlignment(Qt.AlignCenter)
        ##########################################################
        # Add fields to grid layout
        grid_layout.addWidget(QLabel("Seller Name:"), 3, 0)
        grid_layout.addWidget(self.Supplier_name_field, 3, 1)
        grid_layout.addWidget(QLabel("Father's Name:"), 3, 2)
        grid_layout.addWidget(self.Supplier_father_name_field, 3, 3)
        grid_layout.addWidget(QLabel("Address:"), 4, 0)
        grid_layout.addWidget(self.Supplier_address_field, 4, 1, 1, 1)
        ##########################################################################
        grid_layout.addWidget(QLabel("Phone Number:"),4,2)
        grid_layout.addWidget(self.my_Supplier_phonenumber,4,3)
        ###############################################################################
        grid_layout.addWidget(QLabel("CNIC No:"), 5, 0)
        grid_layout.addWidget(self.Supplier_cnic_field, 5, 1)
        # Step (4e): Add image label with border and fixed size
        ################################################################
        #####################################################################
        self.supplier_image_label1=QLabel(self)
        self.supplier_image_label1.setFixedSize(field_width,130)  # Fixed size for image
        self.supplier_image_label1.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")  # Add border to the image label
        self.supplier_image_label1.setAlignment(Qt.AlignCenter)
        self.supplier_image_label1.setStyleSheet("border:2px solid black;border-radius:5px;")
        ##############################################################################
        grid_layout.addWidget(QLabel("Seller Image:"), 5, 2)
        grid_layout.addWidget(self.supplier_image_label1, 5, 3)
        # Step (5): Add Vehicle Description section
        ##################################################################################
        self.Supplier_name_field.currentIndexChanged.connect(self.on_supplier_selection_changed)
        ###############################################################################################
        vehicle_title_label = QLabel("Description of Vehicle")
        vehicle_title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 20px; color: #333;")
        grid_layout.addWidget(vehicle_title_label, 6, 0, 1, 4)
        # Add vehicle details fields
        #####################################################################
        self.Purchase_registration_field=QLineEdit()
        self.Purchase_registration_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_registration_field.setAlignment(Qt.AlignCenter)
        self.Purchase_registration_field.setFixedWidth(field_width)
        self.Purchase_registration_field.setFixedHeight(field_height)
        ######################################################################
        #########################################################3
        self.Purchase_chassis_field=QLineEdit()
        self.Purchase_chassis_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_chassis_field.setFixedWidth(field_width)
        self.Purchase_chassis_field.setFixedHeight(field_height)
        self.Purchase_chassis_field.setAlignment(Qt.AlignCenter)
        ######################################################
        self.Purchase_engine_field=QLineEdit()
        self.Purchase_engine_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_engine_field.setFixedWidth(field_width)
        self.Purchase_engine_field.setFixedHeight(field_height)
        self.Purchase_engine_field.setAlignment(Qt.AlignCenter)
        #########################################################
        self.Purchase_make_field=QLineEdit()
        self.Purchase_make_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_make_field.setFixedWidth(field_width)
        self.Purchase_make_field.setFixedHeight(field_height)
        self.Purchase_make_field.setAlignment(Qt.AlignCenter)
        ########################################################
        self.Purchase_horsepower_field=QLineEdit()
        self.Purchase_horsepower_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_horsepower_field.setFixedWidth(field_width)
        self.Purchase_horsepower_field.setFixedHeight(field_height)
        self.Purchase_horsepower_field.setAlignment(Qt.AlignCenter)
        ##################################################
        self.Purchase_model_field=QLineEdit()
        self.Purchase_model_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_model_field.setFixedWidth(field_width)
        self.Purchase_model_field.setFixedHeight(field_height)
        self.Purchase_model_field.setAlignment(Qt.AlignCenter)
        ################################################
        self.Purchase_color_field=QLineEdit()
        self.Purchase_color_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_color_field.setFixedWidth(field_width)
        self.Purchase_color_field.setFixedHeight(field_height)
        self.Purchase_color_field.setAlignment(Qt.AlignCenter)
        ##############################################################################################
        # self.Purchase_import_field=QLineEdit()
        # self.Purchase_import_field.setStyleSheet("""
        #     QLineEdit{
        #         background-color:#ECDFCC;
        #         color:black;
        #         font-weight:bold;
        #         border-radius:10px;
        #     }
        #     QLineEdit::Placeholder{
        #         color:white;
        #         font-weight:bold;
        #         text-align:center;
        #     }
        # """)
        # self.Purchase_import_field.setFixedWidth(field_width)
        # self.Purchase_import_field.setFixedHeight(field_height)
        # self.Purchase_import_field.setAlignment(Qt.AlignCenter)
        ##########################################################
        # Add vehicle details fields
        grid_layout.addWidget(QLabel("Registration #:"), 7, 0)
        grid_layout.addWidget(self.Purchase_registration_field, 7, 1)
        grid_layout.addWidget(QLabel("Chassis #:"), 7, 2)
        grid_layout.addWidget(self.Purchase_chassis_field, 7, 3)
        grid_layout.addWidget(QLabel("Engine #:"), 8, 0)
        grid_layout.addWidget(self.Purchase_engine_field, 8, 1)
        grid_layout.addWidget(QLabel("Make:"), 8, 2)
        grid_layout.addWidget(self.Purchase_make_field, 8, 3)
        grid_layout.addWidget(QLabel("Horse Power:"), 9, 0)
        grid_layout.addWidget(self.Purchase_horsepower_field, 9, 1)
        grid_layout.addWidget(QLabel("Model #:"),9,2)
        grid_layout.addWidget(self.Purchase_model_field, 9, 3)
        grid_layout.addWidget(QLabel("Color:"),10,0)
        grid_layout.addWidget(self.Purchase_color_field,10,1)
        #############################################################################################
        # grid_layout.addWidget(QLabel("Import:"),10,2)
        # grid_layout.addWidget(self.Purchase_import_field,10,3)
        ######################################################################################
        self.Purchase_Price_field=QLineEdit()
        self.Purchase_Price_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_Price_field.setFixedWidth(field_width)
        self.Purchase_Price_field.setFixedHeight(field_height)
        self.Purchase_Price_field.setAlignment(Qt.AlignCenter)
        ################################################################
        self.Purchase_InwWords_field=QLineEdit()
        self.Purchase_InwWords_field.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ###########################################
        self.Purchase_InwWords_field.setFixedWidth(field_width)
        self.Purchase_InwWords_field.setFixedHeight(field_height)
        self.Purchase_InwWords_field.setAlignment(Qt.AlignCenter)
        ##################################################################################################
        grid_layout.addWidget(QLabel("Purchase Price:"),11,0)
        grid_layout.addWidget(self.Purchase_Price_field,11,1)
        ##########################################
        grid_layout.addWidget(QLabel("Price InWords:"),11,2)
        grid_layout.addWidget(self.Purchase_InwWords_field,11,3)
    ##########################################################################################################################
    #####################################################################################################################################
        #################################(One Time (1)###########################################################
        self.Purchase_Payment_date1=QDateEdit()
        self.Purchase_Payment_date1.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.Purchase_Payment_date1.setFixedWidth(field_width)
        self.Purchase_Payment_date1.setFixedHeight(field_height)
        self.Purchase_Payment_date1.setDate(QDate.currentDate())
        self.Purchase_Payment_date1.setCalendarPopup(True)
        #####################################################################################
        grid_layout.addWidget(QLabel("Purchase Payment Date:"),12,0)
        grid_layout.addWidget(self.Purchase_Payment_date1,12,1)
        ###########################################################################################################################
        self.Purchase_ModeOf_payment_field1=QComboBox()
        self.Purchase_ModeOf_payment_field1.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                font-weight:bold;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_ModeOf_payment_field1.setFixedWidth(field_width)
        self.Purchase_ModeOf_payment_field1.setFixedHeight(field_height)
        self.Purchase_ModeOf_payment_field1.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Purchase_ModeOf_payment_field1.setEditable(True)
        self.Purchase_ModeOf_payment_field1.lineEdit().setAlignment(Qt.AlignLeft)
        self.Purchase_ModeOf_payment_field1.setEditable(False)
        grid_layout.addWidget(QLabel("Mode Of Payment:"),12,2)
        grid_layout.addWidget(self.Purchase_ModeOf_payment_field1,12,3)
        ###################################################################################################################
        self.Purchase_InCheck_field1=QLineEdit()
        self.Purchase_InCheck_field1.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_InCheck_field1.setFixedWidth(field_width)
        self.Purchase_InCheck_field1.setFixedHeight(field_height)
        self.Purchase_InCheck_field1.setAlignment(Qt.AlignCenter)
        grid_layout.addWidget(QLabel("Amount of Payment:"),12,4)
        grid_layout.addWidget(self.Purchase_InCheck_field1,12,5)
        ##########################################################################################################################
        ############################################(One Time (2)###########################################################
        self.Purchase_Payment_date2=QDateEdit()
        self.Purchase_Payment_date2.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.Purchase_Payment_date2.setFixedWidth(field_width)
        self.Purchase_Payment_date2.setFixedHeight(field_height)
        self.Purchase_Payment_date2.setDate(QDate.currentDate())
        self.Purchase_Payment_date2.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Purchase Payment Date:"),13,0)
        grid_layout.addWidget(self.Purchase_Payment_date2,13,1)
        ###########################################################################################################################
        self.Purchase_ModeOf_payment_field2=QComboBox()
        self.Purchase_ModeOf_payment_field2.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                font-weight:bold;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_ModeOf_payment_field2.setFixedWidth(field_width)
        self.Purchase_ModeOf_payment_field2.setFixedHeight(field_height)
        self.Purchase_ModeOf_payment_field2.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Purchase_ModeOf_payment_field2.setEditable(True)
        self.Purchase_ModeOf_payment_field2.lineEdit().setAlignment(Qt.AlignLeft)
        self.Purchase_ModeOf_payment_field2.setEditable(False)
        grid_layout.addWidget(QLabel("Mode Of Payment:"),13,2)
        grid_layout.addWidget(self.Purchase_ModeOf_payment_field2,13,3)
        ###################################################################################################################
        self.Purchase_InCheck_field2=QLineEdit()
        self.Purchase_InCheck_field2.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_InCheck_field2.setFixedWidth(field_width)
        self.Purchase_InCheck_field2.setFixedHeight(field_height)
        self.Purchase_InCheck_field2.setAlignment(Qt.AlignCenter)
        grid_layout.addWidget(QLabel("Amount of Payment:"),13,4)
        grid_layout.addWidget(self.Purchase_InCheck_field2,13,5)
        ###########################################################################################################################
        ###############################################(One Time (3)###########################################################
        self.Purchase_Payment_date3=QDateEdit()
        self.Purchase_Payment_date3.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.Purchase_Payment_date3.setFixedWidth(field_width)
        self.Purchase_Payment_date3.setFixedHeight(field_height)
        self.Purchase_Payment_date3.setDate(QDate.currentDate())
        self.Purchase_Payment_date3.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Purchase Payment Date:"),14,0)
        grid_layout.addWidget(self.Purchase_Payment_date3,14,1)
        ###########################################################################################################################
        self.Purchase_ModeOf_payment_field3=QComboBox()
        self.Purchase_ModeOf_payment_field3.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-size:14px;
                font-weight:bold;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_ModeOf_payment_field3.setFixedWidth(field_width)
        self.Purchase_ModeOf_payment_field3.setFixedHeight(field_height)
        self.Purchase_ModeOf_payment_field3.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Purchase_ModeOf_payment_field3.setEditable(True)
        self.Purchase_ModeOf_payment_field3.lineEdit().setAlignment(Qt.AlignLeft)
        self.Purchase_ModeOf_payment_field3.setEditable(False)
        grid_layout.addWidget(QLabel("Mode Of Payment:"),14,2)
        grid_layout.addWidget(self.Purchase_ModeOf_payment_field3,14,3)
        ###################################################################################################################
        self.Purchase_InCheck_field3=QLineEdit()
        self.Purchase_InCheck_field3.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_InCheck_field3.setFixedWidth(field_width)
        self.Purchase_InCheck_field3.setFixedHeight(field_height)
        self.Purchase_InCheck_field3.setAlignment(Qt.AlignCenter)
        grid_layout.addWidget(QLabel("Amount of Payment:"),14,4)
        grid_layout.addWidget(self.Purchase_InCheck_field3,14,5)
        ##################################################################################################################
        ###############################################(One Time (4)###########################################################
        self.Purchase_Payment_date4=QDateEdit()
        self.Purchase_Payment_date4.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.Purchase_Payment_date4.setFixedWidth(field_width)
        self.Purchase_Payment_date4.setFixedHeight(field_height)
        self.Purchase_Payment_date4.setDate(QDate.currentDate())
        self.Purchase_Payment_date4.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Purchase Payment Date:"),15,0)
        grid_layout.addWidget(self.Purchase_Payment_date4,15,1)
        ###########################################################################################################################
        self.Purchase_ModeOf_payment_field4=QComboBox()
        self.Purchase_ModeOf_payment_field4.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_ModeOf_payment_field4.setFixedWidth(field_width)
        self.Purchase_ModeOf_payment_field4.setFixedHeight(field_height)
        self.Purchase_ModeOf_payment_field4.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Purchase_ModeOf_payment_field4.setEditable(True)
        self.Purchase_ModeOf_payment_field4.lineEdit().setAlignment(Qt.AlignLeft)
        self.Purchase_ModeOf_payment_field4.setEditable(False)
        grid_layout.addWidget(QLabel("Mode Of Payment:"),15,2)
        grid_layout.addWidget(self.Purchase_ModeOf_payment_field4,15,3)
        ###################################################################################################################
        self.Purchase_InCheck_field4=QLineEdit()
        self.Purchase_InCheck_field4.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_InCheck_field4.setFixedWidth(field_width)
        self.Purchase_InCheck_field4.setFixedHeight(field_height)
        self.Purchase_InCheck_field4.setAlignment(Qt.AlignCenter)
        grid_layout.addWidget(QLabel("Amount of Payment:"),15,4)
        grid_layout.addWidget(self.Purchase_InCheck_field4,15,5)
    ##################################################################################################################
    ###############################################(One Time (5)###########################################################
        self.Purchase_Payment_date5=QDateEdit()
        self.Purchase_Payment_date5.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
        """)
        self.Purchase_Payment_date5.setFixedWidth(field_width)
        self.Purchase_Payment_date5.setFixedHeight(field_height)
        self.Purchase_Payment_date5.setDate(QDate.currentDate())
        self.Purchase_Payment_date5.setCalendarPopup(True)
        grid_layout.addWidget(QLabel("Purchase Payment Date:"),16,0)
        grid_layout.addWidget(self.Purchase_Payment_date5,16,1)
        ###########################################################################################################################
        self.Purchase_ModeOf_payment_field5=QComboBox()
        self.Purchase_ModeOf_payment_field5.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_ModeOf_payment_field5.setFixedWidth(field_width)
        self.Purchase_ModeOf_payment_field5.setFixedHeight(field_height)
        self.Purchase_ModeOf_payment_field5.addItems(["","Select Mode of Payment", "Check", "Cash", "PayOrder", "Deposit"])
        self.Purchase_ModeOf_payment_field5.setEditable(True)
        self.Purchase_ModeOf_payment_field5.lineEdit().setAlignment(Qt.AlignLeft)
        self.Purchase_ModeOf_payment_field5.setEditable(False)
        grid_layout.addWidget(QLabel("Mode Of Payment:"),16,2)
        grid_layout.addWidget(self.Purchase_ModeOf_payment_field5,16,3)
        ###################################################################################################################
        self.Purchase_InCheck_field5=QLineEdit()
        self.Purchase_InCheck_field5.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.Purchase_InCheck_field5.setFixedWidth(field_width)
        self.Purchase_InCheck_field5.setFixedHeight(field_height)
        self.Purchase_InCheck_field5.setAlignment(Qt.AlignCenter)
        grid_layout.addWidget(QLabel("Amount of Payment:"),16,4)
        grid_layout.addWidget(self.Purchase_InCheck_field5,16,5)
    #########################################################################################################################
    ######################################################################################################################################
    ###################################################################################################################################
        self.Purchase_remarks_field=QTextEdit()
        self.Purchase_remarks_field.setStyleSheet("""
            QTextEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QTextEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        ##############################################################
        self.Purchase_remarks_field.setFixedWidth(340)
        self.Purchase_remarks_field.setFixedHeight(120)
        #self.Purchase_remarks_field.setAlignment(Qt.AlignCenter)
        ################################################################
        grid_layout.addWidget(QLabel("Payment Details:"),17,0)
        grid_layout.addWidget(self.Purchase_remarks_field,17,1)
#############################################################################################################################################
#############################################################################################################################################
        #####################################################################################################
        # Step (6): Add Document attachment section with upload buttons and image preview
        Purchase_document_title_label = QLabel("Documents Attached")
        Purchase_document_title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 20px; color: #333;")
        grid_layout.addWidget(Purchase_document_title_label,18,0)
        #############################################################################
        Purchase_cnic_upload_button=QPushButton("Upload Cnic Image")
        Purchase_other_upload_button=QPushButton("Upload Other Document")
        ###################################################################
        Purchase_cnic_upload_button.setStyleSheet("""
            QPushButton {
                background-color:#40A2E3;  /* Modern red color */
                color: white;
                border: none;
                border-radius:5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color:#40A2E3;
            }
            QPushButton:pressed {
                background-color:#40A2E3;
            }
        """)
        Purchase_cnic_upload_button.setFixedWidth(150)
        Purchase_cnic_upload_button.setFixedHeight(35)
        #######################################################
        Purchase_other_upload_button.setStyleSheet("""
            QPushButton {
                background-color:#40A2E3;  /* Modern red color */
                color: white;
                border: none;
                border-radius:5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color:#40A2E3;
            }
            QPushButton:pressed {
                background-color:#40A2E3;
            }
        """)
        Purchase_other_upload_button.setFixedWidth(170)
        Purchase_other_upload_button.setFixedHeight(35)
        ########################################################################
        #Connect the upload button signals to methods for handling file uploads
        Purchase_cnic_upload_button.clicked.connect(lambda:self.Purchase_upload_image(self.Purchase_cnic_upload_label))
        Purchase_other_upload_button.clicked.connect(lambda:self.Purchase_upload_image(self.Purchase_attachment_label))
        ######################################################################################
        self.Purchase_cnic_upload_label=QLabel()
        self.Purchase_attachment_label=QLabel()
        style="border: 1px solid #C0C0C0; border-radius: 5px;"
        self.Purchase_cnic_upload_label.setStyleSheet(style)
        self.Purchase_attachment_label.setStyleSheet(style)
        self.Purchase_cnic_upload_label.setFixedSize(250, 130)
        self.Purchase_attachment_label.setFixedSize(250, 130)
        grid_layout.addWidget(Purchase_cnic_upload_button,19,0)
        grid_layout.addWidget(self.Purchase_cnic_upload_label,19,1)
        grid_layout.addWidget(Purchase_other_upload_button,19,2)
        grid_layout.addWidget(self.Purchase_attachment_label,19,3)
        #########################################################################################################################
        # Set up QLabel for images
        self.purchase_other1 = QLabel("No image uploaded")
        self.purchase_other2 = QLabel("No image uploaded")
        self.purchase_other3 = QLabel("No image uploaded")
        self.purchase_other4 = QLabel("No image uploaded")
        # Fixed size for uploaded images
        self.purchase_other1.setFixedSize(250, 130)
        self.purchase_other2.setFixedSize(250, 130)
        self.purchase_other3.setFixedSize(250, 130)
        self.purchase_other4.setFixedSize(250, 130)

        # Center alignment for text in QLabel
        self.purchase_other1.setAlignment(Qt.AlignCenter)
        self.purchase_other2.setAlignment(Qt.AlignCenter)
        self.purchase_other3.setAlignment(Qt.AlignCenter)
        self.purchase_other4.setAlignment(Qt.AlignCenter)

        # Set styles for QLabel
        style = "border: 1px solid #C0C0C0; border-radius: 5px;"
        self.purchase_other1.setStyleSheet(style)
        self.purchase_other2.setStyleSheet(style)
        self.purchase_other3.setStyleSheet(style)
        self.purchase_other4.setStyleSheet(style)

        # Add QLabel to the grid layout
        grid_layout.addWidget(self.purchase_other1, 20, 0)
        grid_layout.addWidget(self.purchase_other2, 20, 1)
        grid_layout.addWidget(self.purchase_other3, 20, 2)
        grid_layout.addWidget(self.purchase_other4, 20, 3)

        # Install event filter to detect clicks on QLabel
        ##########################################################################################################
        # Set properties for each label
        for label in [self.purchase_other1, self.purchase_other2, self.purchase_other3, self.purchase_other4]:
            label.setFixedSize(250, 130)  # Fixed size for uploaded images
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("border: 1px solid #C0C0C0; border-radius: 5px;")
            label.mousePressEvent = lambda event, lbl=label: self.handle_label_click(event, lbl)
        ################################################################################################################################
        # Set up QTextEdit for purchase note
        self.purchase_Note = QTextEdit()
        self.purchase_Note.setStyleSheet("""
            QTextEdit{
                background-color: #ECDFCC;
                color: black;
                font-weight: bold;
                border-radius: 10px;
            }
            QTextEdit::Placeholder{
                color: white;
                font-weight: bold;
                text-align: center;
            }
        """)
        self.purchase_Note.setFixedWidth(340)
        self.purchase_Note.setFixedHeight(120)
        grid_layout.addWidget(QLabel("Document Note:"),21, 0)
        grid_layout.addWidget(self.purchase_Note, 21, 1)
        ############################################################################################################################################################
        #############################################################################################################################################################################################################
        #Step (7): Add Save, Clear, and Generate Report buttons with extra spacing and center-aligned
        label_layout= QHBoxLayout()
        l1=QLabel("_______________________________________________________________________________________________________________________________________________________")
        l1.setStyleSheet("color:blue;height:20px;")
        label_layout.setContentsMargins(0,50,0,0)
        label_layout.addWidget(l1)
        label_layout.setAlignment(Qt.AlignCenter)
        #############################################################
        Purchase_save_button=QPushButton("Save")
        Purchase_clear_button=QPushButton("Clear")
        Purchase_report_button=QPushButton("Generate Report")
        Purchase_search_button=QPushButton("Search Records")  # New button
        Purchase_update_record=QPushButton("Update Records")
        #################################################################
        Purchase_save_button.setStyleSheet("""
            QPushButton {
                background-color: #018749;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0D7C66;
            }
            QPushButton:pressed {
                background-color: #018749;
            }
        """)
        #######################################################
        Purchase_clear_button.setStyleSheet("""
            QPushButton {
                background-color:#624E88;/* Modern blue color */
                color: white;
                border: none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color:#8967B3;
            }
            QPushButton:pressed {
                background-color:#CB80AB;
            }
        """)
        ####################################################
        Purchase_search_button.setStyleSheet("""
            QPushButton {
                background-color:#181C14;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#3C3D37;
            }
            QPushButton:pressed{
                background-color:#697565;
            }
        """)
        ######################################
        Purchase_report_button.setStyleSheet("""
            QPushButton {
                background-color:#2D3250;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#424769;
            }
            QPushButton:pressed{
                background-color:#7077A1;
            }
        """)
        ####################################
        Purchase_update_record.setStyleSheet("""
            QPushButton {
                background-color:#9D5C0D;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#E5890A;
            }
            QPushButton:pressed{
                background-color:#C36A2D;
            }
        """)
        ###########################################################################
        Purchase_save_button.setFixedSize(Purchase_button_width,30)
        Purchase_clear_button.setFixedSize(Purchase_button_width,30)
        Purchase_report_button.setFixedSize(Purchase_button_width,30)
        Purchase_search_button.setFixedSize(Purchase_button_width,30)
        Purchase_update_record.setFixedSize(Purchase_button_width,30)
        #############################################################
        Purchase_save_button.clicked.connect(self.save_purchase_note)
        Purchase_clear_button.clicked.connect(self.clear_purchase)
        Purchase_report_button.clicked.connect(self.Purchase_create_report)
        Purchase_search_button.clicked.connect(self.search_purchase)
        Purchase_update_record.clicked.connect(self.Purchase_update_data)
        ######################################################################################################
        # Add spacing above the buttons
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0,40, 0, 0)  # Add top margin to create spacing above buttons
        #################################################################################
        # Add buttons to layout
        button_layout.addWidget(Purchase_save_button)
        button_layout.addWidget(Purchase_clear_button)
        button_layout.addWidget(Purchase_report_button)
        button_layout.addWidget(Purchase_search_button)  # Add the new button here
        button_layout.addWidget(Purchase_update_record)
        ############################################################################################################
        label_layout.setSpacing(80)
        grid_layout.addLayout(label_layout,22, 0, 1, 4)
        form_widget.setLayout(grid_layout)
        ########################################################
        button_layout.setSpacing(10)
        button_layout.setAlignment(Qt.AlignCenter)  # Center the buttons horizontally
        grid_layout.addLayout(button_layout,23, 0, 1, 4)
        form_widget.setLayout(grid_layout)
        #######################################################################################################################
        # Add the form to the tab
        tab_widget.addTab(form_widget, "Add Purchase Car Details")
##########################################33333##########################################################################################
###############################################################################(View Selling Car Details)################################
##########################################################################################################################33#############
        #Create the view_widget
        view_widget = QWidget()
        view_layout = QVBoxLayout(view_widget)
        view_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        # Add the title of the tab
        view_title_label = QLabel("View Purchase Car Details")
        view_title_label.setStyleSheet("font-size: 24px; font-weight: bold; margin-bottom: 20px; color: #333;")
        view_layout.addWidget(view_title_label)
        ###################Add Search Records section#############
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0,0,650,0)
        Purchase_search_label = QLabel("Search Records:")
        Purchase_search_label.setStyleSheet("""
            QLabel{
                color:black;
                font-weight:bold;
                border-radius:10px;
                font-size:18px;
            }
            QLabel::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        #########################################################
        self.shams_search_field2=QLineEdit()
        self.shams_search_field2.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                font-size:16px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
                text-align:center;
            }
        """)
        self.shams_search_field2.setFixedHeight(50)
        self.shams_search_field2.setFixedWidth(500)
        ################################################################
        Purchase_search_button = QPushButton("Search Records")
        Purchase_search_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;  /* Modern teal color */
                color: white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        Purchase_search_button.setFixedWidth(150)
        Purchase_search_button.setFixedHeight(50)
        ###################################################################################################################
        ################################################################
        Purchase_print_button = QPushButton("Print Reports")
        Purchase_print_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;  /* Modern teal color */
                color: white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        Purchase_print_button.setFixedWidth(150)
        Purchase_print_button.setFixedHeight(50)
        Purchase_print_button.clicked.connect(self.print_record1)
        ################################################################################################################################
        search_layout.addWidget(Purchase_search_label)
        search_layout.addWidget(self.shams_search_field2)
        search_layout.addWidget(Purchase_search_button)
        search_layout.addWidget(Purchase_print_button)
        search_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        view_layout.addLayout(search_layout)
        view_layout.setSpacing(10)
        # Add Radio Buttons for search methods
        ########################################################################################################################
        #################################################################################################################################################
        ##search_button.clicked.connect(self.my_search)
        Purchase_search_button.clicked.connect(self.purchase_search_records)
        #######################################################################################################################
        ############################################################################################################################################
        line_layout=QHBoxLayout()
        self.line33=QLabel("__________________________________________________________________________________________________________________________________________________________________________________")
        self.line33.setFixedWidth(1100)
        line_layout.addWidget(self.line33)
        line_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        view_layout.addLayout(line_layout)
        ##################################################################################################################################################################
        ################################################################################################################################################################################
        # Add Save, Clear, and Generate Report buttons with extra spacing and center-aligned
        button_layout = QHBoxLayout()
        Purchase_delete_button = QPushButton("Delete Records")
        Purchase_delete_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size:16px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        Purchase_button_width=250
        Purchase_button_height=35
        ##################################################################
        Purchase_delete_button.clicked.connect(self.Purchase_delete)
        #######################################################################################
        Purchase_load_data=QPushButton("Load Data")
        Purchase_load_data.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size:16px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        ###########################################################
        Purchase_delete_button.setFixedSize(Purchase_button_width,Purchase_button_height)
        Purchase_load_data.setFixedSize(Purchase_button_width,Purchase_button_height)
        #############################################################
        Purchase_load_data.clicked.connect(self.purchase_load)
        ########################################################
        button_layout.addWidget(Purchase_delete_button)
        button_layout.addWidget(Purchase_load_data)
        ############################################
        button_layout.setSpacing(10)
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)  # Center the buttons horizontally
        view_layout.addLayout(button_layout)
        ##################################################################################################################################################
        # Add a table to display records
        my_purchase_layout=QHBoxLayout()
        self.Purchase_table_widget=QTableWidget()  # Make table_widget an instance attribute
        self.Purchase_table_widget.setColumnCount(11)  # 11 columns including 'Document ID'
        self.Purchase_table_widget.setHorizontalHeaderLabels(["Invoice No", "Supplier Name", "Father Name", "CNIC NO", 
                                                    "Registration No", "Model", "Color", "Chassis No", "Purchase Date","Purchase_Price","Document ID"])
        self.Purchase_table_widget.setFixedWidth(1070)
        self.Purchase_table_widget.setFixedHeight(400)  # Set a fixed height for the table
        self.Purchase_table_widget.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
                padding: 0;  /* Remove internal padding to match the frame */
            }
            QHeaderView::section {
                background-color: #007BFF;  /* Professional blue color */
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;  /* Center align text */
            }
        """)
        ################################################################
        # Enable alternating row colors for better readability
        self.Purchase_table_widget.setAlternatingRowColors(True)
        self.Purchase_table_widget.horizontalHeader().setStretchLastSection(True)
        #Option to let some columns automatically resize
        self.Purchase_table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Set fixed column widths for specific columns if needed (e.g., for the first column)
        self.Purchase_table_widget.setColumnWidth(0,320)  # For Invoice No
        # Other columns will adjust dynamically due to the Stretch mode
        ##############################################################
        my_purchase_layout.addWidget(self.Purchase_table_widget)
        #####################################################
        my_purchase_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        view_layout.addLayout(my_purchase_layout)
        ################################################################################################################################################
        ####################################################################################################################################################################################
        # Add the "View Sale Car Details" tab to the tab widget
        tab_widget.addTab(view_widget, "View Purchase Car Details")
        self.Purchase_load_data()
        self.Purchase_Price_field.textChanged.connect(self.convert_price_to_words)
        layout.addWidget(tab_widget)
        ############################################################################################################
        return purchase_management_widget
################################################################################################################################################################################################################
###############################################################################################################################################################
    #######################################################################################################
    def print_record1(self):
        """Generate a PDF report for the purchase data in the table."""
        try:
            # Open a file save dialog to allow user to select save location
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            if not file_path:  # If no file is selected, cancel the operation
                return

            # Ensure the file extension is .pdf
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Create PDF file with landscape orientation
            pdf_canvas = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf_canvas.setTitle("Purchase Report")

            # Get the page width for centering content
            page_width, _ = landscape(A4)

            # Set up header information centered
            pdf_canvas.setFont("Helvetica-Bold", 14)

            # Center the "Company Name or Logo"
            company_text = "CAR EXPERTS Sector-A, Bankers Town Ring Road, Near State Life Society Lahore"
            company_text_width = pdf_canvas.stringWidth(company_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - company_text_width) / 2, 570, company_text)

            # Center the report date
            report_date_text = f"Purchase Report - {datetime.datetime.now().strftime('%Y-%m-%d')}"
            report_date_width = pdf_canvas.stringWidth(report_date_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - report_date_width) / 2, 550, report_date_text)

            # Adding report title centered with minimal margin at the top
            pdf_canvas.setFont("Helvetica-Bold", 16)
            report_title = "Purchase Report"
            report_title_width = pdf_canvas.stringWidth(report_title, "Helvetica-Bold", 16)
            pdf_canvas.drawString((page_width - report_title_width) / 2, 510, report_title)

            # Setting up table headers and data
            headers = ["Invoice No", "Supplier Name", "Father Name", "CNIC", "Registration", "Model", "Color", "Chassis", "Date", "Purchase Price"]
            data = [headers]

            # Fetch table data and add to data list (excluding the Document ID column)
            for row in range(self.Purchase_table_widget.rowCount()):
                row_data = []
                for col in range(self.Purchase_table_widget.columnCount() - 1):  # Exclude the last column (Document ID)
                    item = self.Purchase_table_widget.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Define table style with adjusted font size and padding
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all cells to the center
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Increased padding for the header cells
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Alternate row color
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Table border/grid lines
                ('FONTSIZE', (0, 1), (-1, -1), 9),  # Set font size for the data rows
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),  # Padding for data cells
            ])

            # Set column widths (adjusted to fit the new number of columns)
            col_widths = [
                0.8 * inch,   # Invoice No
                1.2 * inch,   # Supplier Name
                1.2 * inch,   # Father Name
                1.0 * inch,   # CNIC (increased width)
                1.2 * inch,   # Registration No (increased width)
                1.0 * inch,   # Model
                1.0 * inch,   # Color
                1.2 * inch,   # Chassis
                1.0 * inch,   # Purchase Date
                1.8 * inch    # Purchase Price (increased width)
            ]
            
            # Create table and set style
            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)

            # Calculate table dimensions
            table_width, table_height = table.wrap(0, 0)
            # Calculate the position for the table to be centered horizontally and below the title
            x_position = (page_width - table_width) / 2  # Center horizontally
            y_position = 470  # Position the table just below the title (adjust as needed)

            # Draw the table on the PDF
            table.drawOn(pdf_canvas, x_position, y_position - table_height)

            # Calculate the total purchase value (assuming you have a way to get this data)
            total_purchase_value = self.calculate_total_purchase_value()  # Replace with your method to calculate the total

            # Add total purchase value below the table
            pdf_canvas.setFont("Helvetica-Bold", 10)
            pdf_canvas.drawString(x_position, y_position - table_height - 20, f"Total Purchase Value: {total_purchase_value}")

            # Add page number
            pdf_canvas.setFont("Helvetica", 8)
            pdf_canvas.drawString(750, 30, f"Page {pdf_canvas.getPageNumber()}")

            # Save the PDF file
            pdf_canvas.save()

            # Notify the user that the PDF was successfully created using QMessageBox
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setText(f"PDF report generated successfully: {file_path}")
            msg_box.setWindowTitle("Success")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec_()

        except Exception as e:
            # Handle any errors in the PDF generation process using QMessageBox
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Critical)
            msg_box.setText(f"Error generating PDF report: {str(e)}")
            msg_box.setWindowTitle("Error")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec_()
    ###########################################################################
    def calculate_total_purchase_value(self):
        """Calculate the total purchase value based on the table data."""
        total = 0.0
        for row in range(self.Purchase_table_widget.rowCount()):
            item = self.Purchase_table_widget.item(row, 9)  # Assuming column 9 is Purchase Price
            if item and item.text():
                try:
                    total += float(item.text())
                except ValueError:
                    continue  # Skip if conversion fails
        return total
###########################################################################################################################################
    def purchase_load(self):
        self.Purchase_load_data()
        self.shams_search_field2.clear()
#########################################################################################################################################################
    def get_latest_purchase_id_from_db(self):
        purchase_ref=db.collection("Purchases")
        shams_purchase=purchase_ref.stream()
        purchase_ids=[]
        for my_purchase in shams_purchase:
            data=my_purchase.to_dict()
            purchase_ids.append(int(data.get('invoice_number',0)))
        if purchase_ids:
            max_id=max(purchase_ids)
            return max_id+1
        else:
            return 1
    ####################################################################
    #Function to generate a new random invoice number when clicked
    def generate_new_invoice_number(self):
        ##########(Generate ID for Purchase)#################
        new_invoice_number=str(self.current_id3).zfill(6)# Generate a new random number
        self.Purchase_invoice_number_field.setText(new_invoice_number)  # Update the QLineEdit field
        #Increment the ID for the next call
        self.current_id3+=1
######################################################################
    def Purchase_delete(self):
        """Delete the suppliers record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.Purchase_table_widget.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Supplier?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.Purchase_table_widget.item(row,10).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.delete_purchase_from_firestore(document_id)
                # Remove the row from the table
                self.Purchase_table_widget.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Purchase record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                #self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
                pass
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    ########################################################################################################################
    def delete_purchase_from_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Purchases').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
            self.Purchase_load_Summary_data()
            self.current_id3=self.get_latest_purchase_id_from_db()
            self.Purchase_invoice_number_field.setText(str(self.current_id3).zfill(6))
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
#######################################################################################################################################################################################
    def get_supplier_image_from_label_in_purchase(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.supplier_image_label1.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def get_purchase_upload_cnic_Image(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.Purchase_cnic_upload_label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def get_purchase_upload_other_Image(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.Purchase_attachment_label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    ##################################################################################################################################################################################################
    #########################################################################################################################################################
    def Purchase_update_data(self):
        #Retrieve inputs
        purchase_date = self.Purchase_date_field.text()
        purchase_time = self.Purchase_time_field.text()
        purchase_day = self.Purchase_day_field.text()
        invoice_number = self.Purchase_invoice_number_field.text()
        # Supplier details
        supplier_name = self.Supplier_name_field.currentText()
        supplier_father_name = self.Supplier_father_name_field.text()
        supplier_address = self.Supplier_address_field.text()
        supplier_cnic = self.Supplier_cnic_field.text()
        supplier_phone_number = self.my_Supplier_phonenumber.text()
        # Car details
        purchase_registration = self.Purchase_registration_field.text()
        purchase_chassis = self.Purchase_chassis_field.text()
        purchase_engine = self.Purchase_engine_field.text()
        purchase_make = self.Purchase_make_field.text()
        purchase_horsepower = self.Purchase_horsepower_field.text()
        purchase_model = self.Purchase_model_field.text()
        purchase_color = self.Purchase_color_field.text()
        # Payment details
        purchase_price = self.Purchase_Price_field.text()
        purchase_in_words = self.Purchase_InwWords_field.text()
        purchase_remarks=self.Purchase_remarks_field.toPlainText()
        purchase_note = self.purchase_Note.toPlainText()
        # import_parts = self.Purchase_import_field.text()

        # Payment dates and modes
        payment_dates = [
            self.Purchase_Payment_date1.date().toPyDate(),
            self.Purchase_Payment_date2.date().toPyDate(),
            self.Purchase_Payment_date3.date().toPyDate(),
            self.Purchase_Payment_date4.date().toPyDate(),
            self.Purchase_Payment_date5.date().toPyDate()
        ]
        payment_modes = [
            self.Purchase_ModeOf_payment_field1.currentText(),
            self.Purchase_ModeOf_payment_field2.currentText(),
            self.Purchase_ModeOf_payment_field3.currentText(),
            self.Purchase_ModeOf_payment_field4.currentText(),
            self.Purchase_ModeOf_payment_field5.currentText()
        ]

        # Cheque details
        try:
            cheque_data = {
                f"purchase_in_cheque{i + 1}": int(field.text().replace(',', ''))
                for i, field in enumerate([
                    self.Purchase_InCheck_field1,
                    self.Purchase_InCheck_field2,
                    self.Purchase_InCheck_field3,
                    self.Purchase_InCheck_field4,
                    self.Purchase_InCheck_field5
                ]) if field.text().strip()
            }
        except ValueError:
            self.show_custom_message("Please enter valid numbers for cheque amounts.", "Error", is_success=False)
            return

        # # Validate mandatory fields
        # if not all([purchase_date, purchase_time, invoice_number, supplier_name, supplier_cnic, supplier_phone_number,purchase_note, purchase_remarks]):
        #     self.show_custom_message("Please fill in all mandatory fields.", "Error", is_success=False)
        #     return
        # # Validate CNIC and phone number formats
        # if not re.match(r'^\d{5}-\d{7}-\d{1}$|^\d{13}$', supplier_cnic):
        #     self.show_custom_message("CNIC must follow XXXXX-YYYYYYY-Z or be a 13-digit number.", "Error", is_success=False)
        #     return
        # if not re.match(r'^\d{11}$', supplier_phone_number):
        #     self.show_custom_message("Phone number must be 11 digits.", "Error", is_success=False)
        #     return

        # # Convert purchase price to integer
        # try:
        #     purchase_price = int(purchase_price.replace(',', '').strip())
        # except ValueError:
        #     self.show_custom_message("Purchase price must be a valid integer.", "Error", is_success=False)
        #     return

        # Retrieve image data
        images = {
            'supplier_image': self.get_supplier_image_from_label_in_purchase(),
            'purchase_cnic_image': self.get_purchase_upload_cnic_Image(),
            'purchase_attachment_image': self.get_purchase_upload_other_Image(),
            'purchase_other_image1': self.purchase_my_otherLabel1(),
            'purchase_other_image2': self.purchase_my_otherLabel2(),
            'purchase_other_image3': self.purchase_my_otherLabel3(),
            'purchase_other_image4': self.purchase_my_otherLabel4()
        }

        # Check for all images
        # if any(img is None for img in images.values()):
        #     self.show_custom_message("Please upload all necessary images.", "Error", is_success=False)
        #     return

        # Prepare data for Firebase
        purchase_data = {
            'purchase_date': purchase_date,
            'purchase_time': purchase_time,
            'purchase_day': purchase_day,
            'invoice_number': invoice_number,
            'supplier_name': supplier_name,
            'supplier_father_name': supplier_father_name,
            'supplier_address': supplier_address,
            'supplier_phone_number': supplier_phone_number,
            'supplier_cnic': supplier_cnic,
            **images,  # Include image data
            'purchase_registration': purchase_registration,
            'purchase_chassis': purchase_chassis,
            'purchase_engine': purchase_engine,
            'purchase_make': purchase_make,
            'purchase_horsepower': purchase_horsepower,
            'purchase_model': purchase_model,
            'purchase_color': purchase_color,
            # 'import_parts': import_parts,
            'purchase_price': purchase_price,
            'purchase_in_words': purchase_in_words,
            'cheque_data': cheque_data,
            'payment_dates': [date.isoformat() for date in payment_dates],
            'payment_modes': payment_modes,
            'purchase_remarks': purchase_remarks,
            'purchase_note': purchase_note
        }

        # Check if the invoice number already exists in Firestore
        existing_invoice = db.collection('Purchases').where('invoice_number', '==', invoice_number).get()
        if existing_invoice:
            # Update the existing record
            try:
                for doc in existing_invoice:
                    doc_ref = db.collection('Purchases').document(doc.id)
                    doc_ref.update(purchase_data)
                self.show_custom_message("Purchase record updated successfully!", "Success", is_success=True)
                # Refresh the table to reflect updated data
                self.Purchase_load_data()
                self.My_Purchase_load_data1()
                self.populate_purchase_combo_box()
                # Clear the input fields after updating
                self.current_id3=self.get_latest_purchase_id_from_db()
                self.Purchase_invoice_number_field.setText(str(self.current_id3).zfill(6))
                self.clear_purchase()
            except Exception as e:
                print(f"Error details: {e}")  # Add detailed logging
                self.show_custom_message("An error occurred while updating the record. Please try again.", "Error", is_success=False)
        else:
            self.show_custom_message("No record found with the provided invoice number.", "Error", is_success=False)
    #################################################################################################################################################
    ##################################################################################################################################################################################################
    #########################################################################################################################################################
    def Purchase_load_data(self):
        """Load existing purchase records from Firebase Firestore and display them in the table."""
        try:
            # Fetch purchase records from Firebase Firestore
            purchase_records = db.collection('Purchases').get()
            
            # Clear existing rows in the table before loading new data
            self.Purchase_table_widget.setRowCount(0)
            
            # Loop through each record and insert it into the table
            for record in purchase_records:
                purchase_data = record.to_dict()
                
                # Extract relevant fields for the table
                invoice_number = purchase_data.get('invoice_number', '')
                supplier_name = purchase_data.get('supplier_name', '')
                supplier_father_name = purchase_data.get('supplier_father_name', '')
                supplier_cnic = purchase_data.get('supplier_cnic', '')
                purchase_registration = purchase_data.get('purchase_registration', '')
                purchase_model = purchase_data.get('purchase_model', '')
                purchase_color = purchase_data.get('purchase_color', '')
                purchase_chassis = purchase_data.get('purchase_chassis', '')
                purchase_date = purchase_data.get('purchase_date', '')
                
                # Directly retrieve and format purchase_price as an integer string
                purchase_price = purchase_data.get('purchase_price', 0)
                formatted_purchase_price = str(int(purchase_price))  # Ensure it's an integer and then convert to string
                
                document_id = record.id  # Use Firestore's document ID for "Document ID"
                
                # Insert a new row in the table
                row_position = self.Purchase_table_widget.rowCount()
                self.Purchase_table_widget.insertRow(row_position)
                
                # Fill in the columns with data from the data source
                self.Purchase_table_widget.setItem(row_position, 0, QTableWidgetItem(invoice_number))
                self.Purchase_table_widget.setItem(row_position, 1, QTableWidgetItem(supplier_name))
                self.Purchase_table_widget.setItem(row_position, 2, QTableWidgetItem(supplier_father_name))
                self.Purchase_table_widget.setItem(row_position, 3, QTableWidgetItem(supplier_cnic))
                self.Purchase_table_widget.setItem(row_position, 4, QTableWidgetItem(purchase_registration))
                self.Purchase_table_widget.setItem(row_position, 5, QTableWidgetItem(purchase_model))
                self.Purchase_table_widget.setItem(row_position, 6, QTableWidgetItem(purchase_color))
                self.Purchase_table_widget.setItem(row_position, 7, QTableWidgetItem(purchase_chassis))
                self.Purchase_table_widget.setItem(row_position, 8, QTableWidgetItem(purchase_date))
                self.Purchase_table_widget.setItem(row_position, 9, QTableWidgetItem(formatted_purchase_price))  # Display as integer
                self.Purchase_table_widget.setItem(row_position, 10, QTableWidgetItem(document_id))
            
            # Optionally, resize columns to fit contents
        except Exception as e:
            self.show_custom_message(f"Error loading purchase records: {e}", "Error", is_success=False)
            #pass
#################################################################################3
    def convert_price_to_words(self):
        #Get the text from the Purchase_Price_field
        purchase_price_text = self.Purchase_Price_field.text()
        if purchase_price_text.strip() == "":  # Handle empty input
            self.Purchase_InwWords_field.setText("")
            return
        try:
            # Convert the text to a number (e.g., int)
            purchase_price = int(purchase_price_text)
            # Convert the number to words using the custom number_to_words function
            purchase_in_words =self.number_to_words(purchase_price)
            # Update the Purchase_InwWords_field with the converted words
            self.Purchase_InwWords_field.setText(purchase_in_words)
        except ValueError:
            # Handle the case where the conversion to int fails
            self.Purchase_InwWords_field.setText("Invalid price format")
    ##################################################################################
    def number_to_words(self,num):
        if num < 0:
            return "minus " + self.number_to_words(-num)
        if num == 0:
            return "zero"
        units = [
            "", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
            "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen", 
            "seventeen", "eighteen", "nineteen"
        ]
        tens = [
            "", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", 
            "ninety"
        ]
        thousands = [
            "", "thousand", "million", "billion", "trillion", "quadrillion", 
            "quintillion", "sextillion", "septillion", "octillion", 
            "nonillion", "decillion", "undecillion", "duodecillion", 
            "tredecillion", "quattuordecillion", "quindecillion", 
            "sexdecillion", "septendecillion", "octodecillion", 
            "novemdecillion", "vigintillion"
        ]
        def helper(n):
            if n < 20:
                return units[n]
            elif n < 100:
                return tens[n // 10] + ('' if n % 10 == 0 else ' ' + units[n % 10])
            elif n < 1000:
                return units[n // 100] + " hundred" + ('' if n % 100 == 0 else ' and ' + helper(n % 100))
            else:
                for i, word in enumerate(thousands):
                    if n < 1000 ** (i + 1):
                        return helper(n // (1000 ** i)) + ' ' + thousands[i] + ('' if n % (1000 ** i) == 0 else ' ' + helper(n % (1000 ** i)))
        return helper(num).strip()
#############################################################################################################################################################
#####################################################################################################################################################
    def purchase_my_otherLabel1(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.purchase_other1.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def purchase_my_otherLabel2(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.purchase_other2.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def purchase_my_otherLabel3(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.purchase_other3.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    def purchase_my_otherLabel4(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.purchase_other4.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
#################################################################################################################################################################
###################################################################################################################################################################################
##############################################################################################################################################################
    def save_purchase_note(self):
        #Retrieve inputs
        purchase_date = self.Purchase_date_field.text()
        purchase_time = self.Purchase_time_field.text()
        purchase_day = self.Purchase_day_field.text()
        invoice_number = self.Purchase_invoice_number_field.text()
        # Supplier details
        supplier_name = self.Supplier_name_field.currentText()
        supplier_father_name = self.Supplier_father_name_field.text()
        supplier_address = self.Supplier_address_field.text()
        supplier_cnic = self.Supplier_cnic_field.text()
        supplier_phone_number = self.my_Supplier_phonenumber.text()
        # Car details
        purchase_registration = self.Purchase_registration_field.text()
        purchase_chassis = self.Purchase_chassis_field.text()
        purchase_engine = self.Purchase_engine_field.text()
        purchase_make = self.Purchase_make_field.text()
        purchase_horsepower = self.Purchase_horsepower_field.text()
        purchase_model = self.Purchase_model_field.text()
        purchase_color = self.Purchase_color_field.text()
        # Payment details
        purchase_price = self.Purchase_Price_field.text()
        purchase_in_words = self.Purchase_InwWords_field.text()
        purchase_remarks = self.Purchase_remarks_field.toPlainText()
        purchase_note = self.purchase_Note.toPlainText()
        # import_parts = self.Purchase_import_field.text()
        # Payment dates and modes
        payment_dates = [
            self.Purchase_Payment_date1.date().toPyDate(),
            self.Purchase_Payment_date2.date().toPyDate(),
            self.Purchase_Payment_date3.date().toPyDate(),
            self.Purchase_Payment_date4.date().toPyDate(),
            self.Purchase_Payment_date5.date().toPyDate()
        ]
        payment_modes = [
            self.Purchase_ModeOf_payment_field1.currentText(),
            self.Purchase_ModeOf_payment_field2.currentText(),
            self.Purchase_ModeOf_payment_field3.currentText(),
            self.Purchase_ModeOf_payment_field4.currentText(),
            self.Purchase_ModeOf_payment_field5.currentText()
        ]
        # Cheque details
        try:
            cheque_data = {
                f"purchase_in_cheque{i + 1}": int(field.text().replace(',', ''))
                for i, field in enumerate([
                    self.Purchase_InCheck_field1,
                    self.Purchase_InCheck_field2,
                    self.Purchase_InCheck_field3,
                    self.Purchase_InCheck_field4,
                    self.Purchase_InCheck_field5
                ]) if field.text().strip()
            }
        except ValueError:
            self.show_custom_message("Please enter valid numbers for cheque amounts.", "Error", is_success=False)
            return

        # Validate mandatory fields
        # if not all([purchase_date, purchase_time, invoice_number, supplier_name, supplier_cnic, supplier_phone_number, purchase_note, purchase_remarks]):
        #     self.show_custom_message("Please fill in all mandatory fields.", "Error", is_success=False)
        #     return

        # # Validate CNIC and phone number formats
        # if not re.match(r'^\d{5}-\d{7}-\d{1}$|^\d{13}$', supplier_cnic):
        #     self.show_custom_message("CNIC must follow XXXXX-YYYYYYY-Z or be a 13-digit number.", "Error", is_success=False)
        #     return
        # if not re.match(r'^\d{11}$', supplier_phone_number):
        #     self.show_custom_message("Phone number must be 11 digits.", "Error", is_success=False)
        #     return

        # Convert purchase price to integer
        try:
            purchase_price = int(purchase_price.replace(',', '').strip())
        except ValueError:
            self.show_custom_message("Purchase price must be a valid integer.", "Error", is_success=False)
            return

        # Retrieve image data
        images = {
            'supplier_image': self.get_supplier_image_from_label_in_purchase(),
            'purchase_cnic_image': self.get_purchase_upload_cnic_Image(),
            'purchase_attachment_image': self.get_purchase_upload_other_Image(),
            'purchase_other_image1': self.purchase_my_otherLabel1(),
            'purchase_other_image2': self.purchase_my_otherLabel2(),
            'purchase_other_image3': self.purchase_my_otherLabel3(),
            'purchase_other_image4': self.purchase_my_otherLabel4()
        }

        # Check for all images
        #if any(img is None for img in images.values()):
        #    self.show_custom_message("Please upload all necessary images.", "Error", is_success=False)
        #    return

        # Prepare data for Firebase
        purchase_data = {
            'purchase_date': purchase_date,
            'purchase_time': purchase_time,
            'purchase_day': purchase_day,
            'invoice_number': invoice_number,
            'supplier_name': supplier_name,
            'supplier_father_name': supplier_father_name,
            'supplier_address': supplier_address,
            'supplier_phone_number': supplier_phone_number,
            'supplier_cnic': supplier_cnic,
            **images,  # Include image data
            'purchase_registration': purchase_registration,
            'purchase_chassis': purchase_chassis,
            'purchase_engine': purchase_engine,
            'purchase_make': purchase_make,
            'purchase_horsepower': purchase_horsepower,
            'purchase_model': purchase_model,
            'purchase_color': purchase_color,
            # 'import_parts': import_parts,
            'purchase_price': purchase_price,
            'purchase_in_words': purchase_in_words,
            'cheque_data': cheque_data,
            'payment_dates': [date.isoformat() for date in payment_dates],
            'payment_modes': payment_modes,
            'purchase_remarks': purchase_remarks,
            'purchase_note': purchase_note
        }
        # Check if the invoice number already exists in Firestore
        try:
            existing_invoice = db.collection('Purchases').where('invoice_number', '==', invoice_number).get()
            if existing_invoice:
                self.show_custom_message("Invoice number already exists! Please use a different invoice number.", "Error", is_success=False)
                return
            # Save the data to Firestore
            doc_ref = db.collection('Purchases').add(purchase_data)
            document_id = doc_ref[1].id
            db.collection('Purchases').document(document_id).update({'document_id': document_id})
            # Success message
            self.show_custom_message("Purchase record added successfully!", "Success", is_success=True)
            # Load and clear UI data
            self.Purchase_load_data()
            self.My_Purchase_load_data1()
            self.populate_purchase_combo_box()
            self.current_id3=self.get_latest_purchase_id_from_db()
            self.Purchase_invoice_number_field.setText(str(self.current_id3).zfill(6))
            self.clear_purchase()
        except Exception as e:
            print(f"Error details: {e}")  # Add detailed logging
            self.show_custom_message("An error occurred while saving the record. Please check your internet connection and try again.", "Error", is_success=False)
    ######################################################################################
    def fetch_purchase_details(self):
        purchase_ref=db.collection('Purchases')
        docs=purchase_ref.stream()
        purchase_details={}
        purchase_names=[]
        for doc in docs:
            data = doc.to_dict()
            purchase_name = data.get('purchase_make')
            if purchase_name:
                purchase_names.append(purchase_name)
                purchase_details[purchase_name] = {
                    'purchase_price':data.get('purchase_price',''),
                    'purchase_registration':data.get('purchase_registration',''),
                    'purchase_chassis':data.get('purchase_chassis',''),
                    'purchase_engine':data.get('purchase_engine', ''),
                    'purchase_make': data.get('purchase_make', ''),
                    #############################################################
                    'purchase_horsepower':data.get('purchase_horsepower', ''),
                    'purchase_model':data.get('purchase_model', ''),
                    'purchase_color':data.get('purchase_color', ''),
                    # 'import_parts':data.get('import_parts', '')
                }
        return purchase_details,purchase_names
    ##################################################################
    def populate_purchase_combo_box(self):
        self.purchase_car_names.clear()
        self.purchase_details,purchase_names=self.fetch_purchase_details()
        # Add "Please Select Customer" option and then the customer names
        self.purchase_car_names.addItems(["Select Purchase Car_Names"]+purchase_names)
    ############################################################################################
    def on_purchase_selection_changed(self, index):
        # Ensure fields are cleared if no customer is selected
        if index == 0:
            if hasattr(self, 'purchase_car_price'):
                self.purchase_car_price.setText('')
            # if hasattr(self, 'sale_import_field'):
            #     self.sale_import_field.setText('')
            if hasattr(self, 'registration_field'):
                self.registration_field.setText('')
            if hasattr(self, 'chassis_field'):
                self.chassis_field.setText('')
            if hasattr(self, 'engine_field'):
                self.engine_field.clear()
            if hasattr(self, 'make_field'):
                self.make_field.clear()
            if hasattr(self, 'horsepower_field'):
                self.horsepower_field.clear()
            if hasattr(self, 'engine_field'):
                self.engine_field.clear()
            if hasattr(self, 'model_field'):
                self.model_field.clear()
            if hasattr(self, 'color_field'):
                self.color_field.clear()
        ######################################################################################################################################################
        else:
            purchase_name = self.purchase_car_names.currentText()
            details = self.purchase_details.get(purchase_name, {})
            
            if hasattr(self, 'purchase_car_price'):
                self.purchase_car_price.setText(str(details.get('purchase_price', '')))
            
            if hasattr(self, 'registration_field'):
                self.registration_field.setText(str(details.get('purchase_registration', '')))
            
            if hasattr(self, 'chassis_field'):
                self.chassis_field.setText(str(details.get('purchase_chassis', '')))
            
            if hasattr(self, 'engine_field'):
                self.engine_field.setText(str(details.get('purchase_engine', '')))
            
            if hasattr(self, 'make_field'):
                self.make_field.setText(str(details.get('purchase_make', '')))
            
            if hasattr(self, 'horsepower_field'):
                self.horsepower_field.setText(str(details.get('purchase_horsepower', '')))
            
            if hasattr(self, 'model_field'):
                self.model_field.setText(str(details.get('purchase_model', '')))
            
            if hasattr(self, 'color_field'):
                self.color_field.setText(str(details.get('purchase_color', '')))
            
            # # if hasattr(self, 'sale_import_field'):
            # #     # self.sale_import_field.setText(str(details.get('import_parts', '')))
            #     pass
######################################################################################################################################################
##############################################################################################################################################################
    def display_image(self, label: QLabel, image_data: str):
        """Display an image in the given QLabel.
        :param label: QLabel where the image will be displayed.
        :param image_data: Base64 encoded image data or raw binary image data.
        """
        try:
            # Check if image_data is base64 encoded
            if isinstance(image_data, str) and image_data.startswith('data:image'):
                # Extract base64 data from the data URL
                base64_data = image_data.split(',')[1]
                # Decode base64 data to binary
                image_data = base64.b64decode(base64_data)
            else:
                # Handle if the image is stored as base64 without data URL structure
                image_data = base64.b64decode(image_data)

            # Create an in-memory binary stream from the image data
            image_stream = io.BytesIO(image_data)

            # Load the image from the stream
            pixmap = QPixmap()
            pixmap.loadFromData(image_stream.read(), 'PNG')  # Use 'PNG' for loading PNG images

            # Set the QPixmap to the QLabel, scale the image to fit and maintain aspect ratio
            label.setPixmap(pixmap.scaled(310, 200, Qt.KeepAspectRatio))
            label.setScaledContents(True)  # Ensure scaling of image to fit the label

        except Exception as e:
            # Handle exceptions (e.g., log error, display a placeholder image, etc.)
            print(f"Error displaying image: {e}")
            label.clear()  # Clear the label if image display fails
    ###################################################################################################################################################################33
    def search_purchase(self):
        """Handle searching and displaying a single purchase record based on the selected address."""
        # Retrieve address from the search field
        self.cnic=self.Supplier_cnic_field.text()
        if not self.cnic:
            self.show_custom_message("Please enter Data to search Records.", "Error", is_success=False)
            return
        try:
            # Search for records in Firebase Firestore
            records = db.collection('Purchases').where('supplier_cnic','==',self.cnic).limit(1).get()
            if len(records) == 0:
                self.show_custom_message("No records were found.", "Information", is_success=False)
                return
            
            # Clear previous data in the form
            self.clear_purchase()
            
            # Display the first record
            first_record = records[0]
            data = first_record.to_dict()
            
            # Parse date and time from the data
            purchase_date_str = data.get('purchase_date', '')
            purchase_time_str = data.get('purchase_time', '')
            
            purchase_date = QDate.fromString(purchase_date_str, 'yyyy-MM-dd') if purchase_date_str else QDate.currentDate()
            purchase_time = QTime.fromString(purchase_time_str, 'HH:mm:ss') if purchase_time_str else QTime.currentTime()
            # Update the UI components with parsed data
            self.Purchase_date_field.setDate(purchase_date)
            self.Purchase_time_field.setTime(purchase_time)
            # self.Purchase_day_field.setText(purchase_date.toString('dddd'))
            self.Purchase_day_field.setText(data.get('purchase_day', ''))
            
            # Populate other fields from Firestore
            self.Purchase_invoice_number_field.setText(data.get('invoice_number', ''))
            self.Supplier_name_field.setCurrentText(data.get('supplier_name', ''))
            self.Supplier_father_name_field.setText(data.get('supplier_father_name', ''))
            self.Supplier_address_field.setText(data.get('supplier_address', ''))
            self.my_Supplier_phonenumber.setText(data.get('supplier_phone_number', ''))
            self.Supplier_cnic_field.setText(data.get('supplier_cnic', ''))
            self.Purchase_registration_field.setText(data.get('purchase_registration', ''))
            self.Purchase_chassis_field.setText(data.get('purchase_chassis', ''))
            self.Purchase_engine_field.setText(data.get('purchase_engine', ''))
            self.Purchase_make_field.setText(data.get('purchase_make', ''))
            self.Purchase_horsepower_field.setText(data.get('purchase_horsepower', ''))
            self.Purchase_model_field.setText(data.get('purchase_model', ''))
            self.Purchase_color_field.setText(data.get('purchase_color', ''))
            # self.Purchase_import_field.setText(data.get('import_parts', ''))
            self.Purchase_Price_field.setText(str(data.get('purchase_price', '')))
            self.Purchase_InwWords_field.setText(data.get('purchase_in_words', ''))
            self.Purchase_remarks_field.setPlainText(data.get('purchase_remarks', ''))
            self.purchase_Note.setPlainText(data.get('purchase_note', ''))
            
            # Display payment dates and modes
            payment_dates = data.get('payment_dates', [])
            for i, date_str in enumerate(payment_dates):
                date = QDate.fromString(date_str, 'yyyy-MM-dd')
                if i == 0: self.Purchase_Payment_date1.setDate(date)
                elif i == 1: self.Purchase_Payment_date2.setDate(date)
                elif i == 2: self.Purchase_Payment_date3.setDate(date)
                elif i == 3: self.Purchase_Payment_date4.setDate(date)
                elif i == 4: self.Purchase_Payment_date5.setDate(date)
            
            payment_modes = data.get('payment_modes', [])
            if payment_modes:
                self.Purchase_ModeOf_payment_field1.setCurrentText(payment_modes[0] if len(payment_modes) > 0 else '')
                self.Purchase_ModeOf_payment_field2.setCurrentText(payment_modes[1] if len(payment_modes) > 1 else '')
                self.Purchase_ModeOf_payment_field3.setCurrentText(payment_modes[2] if len(payment_modes) > 2 else '')
                self.Purchase_ModeOf_payment_field4.setCurrentText(payment_modes[3] if len(payment_modes) > 3 else '')
                self.Purchase_ModeOf_payment_field5.setCurrentText(payment_modes[4] if len(payment_modes) > 4 else '')
            
            # Display cheque data
            cheque_data = data.get('cheque_data', {})
            self.Purchase_InCheck_field1.setText(str(cheque_data.get('purchase_in_cheque1', '')))
            self.Purchase_InCheck_field2.setText(str(cheque_data.get('purchase_in_cheque2', '')))
            self.Purchase_InCheck_field3.setText(str(cheque_data.get('purchase_in_cheque3', '')))
            self.Purchase_InCheck_field4.setText(str(cheque_data.get('purchase_in_cheque4', '')))
            self.Purchase_InCheck_field5.setText(str(cheque_data.get('purchase_in_cheque5', '')))
            
            # Display image data
            self.display_image(self.supplier_image_label1, data.get('supplier_image', ''))
            self.display_image(self.Purchase_cnic_upload_label, data.get('purchase_cnic_image', ''))
            self.display_image(self.Purchase_attachment_label, data.get('purchase_attachment_image', ''))
            self.display_image(self.purchase_other1, data.get('purchase_other_image1', ''))
            self.display_image(self.purchase_other2, data.get('purchase_other_image2', ''))
            self.display_image(self.purchase_other3, data.get('purchase_other_image3', ''))
            self.display_image(self.purchase_other4, data.get('purchase_other_image4', ''))
            
            # Make certain fields read-only to prevent accidental modification
            self.Supplier_address_field.setReadOnly(True)
            self.Supplier_father_name_field.setReadOnly(True)
            self.Supplier_cnic_field.setReadOnly(True)
            self.my_Supplier_phonenumber.setReadOnly(True)

        except Exception as e:
            self.show_custom_message(f"An error occurred while searching: {str(e)}", "Error", is_success=False)
    ####################################################################################################################################################################
    def clear_purchase(self):
        self.Supplier_name_field.setCurrentIndex(0)  # Reset ComboBox to first item
        self.Supplier_father_name_field.clear()  # Clear QLineEdit field
        self.Supplier_address_field.clear()  # Clear QLineEdit field
        self.Supplier_cnic_field.clear()  # Clear QLineEdit field
        self.supplier_image_label1.clear()
        self.my_Supplier_phonenumber.clear()
        ##################################################################
        #Clear Purchase Vehicle fields
        self.Purchase_registration_field.clear()  # Clear QLineEdit field
        self.Purchase_chassis_field.clear()  # Clear QLineEdit field
        self.Purchase_engine_field.clear()  # Clear QLineEdit field
        self.Purchase_make_field.clear()  # Clear QLineEdit field
        self.Purchase_horsepower_field.clear()  # Clear QLineEdit field
        self.Purchase_model_field.clear()  # Clear QLineEdit field
        self.Purchase_color_field.clear()  # Clear QLineEdit field
        # self.Purchase_import_field.clear()
        ##############################################################
        #Clear Purchase details fields
        self.Purchase_Price_field.clear()  # Clear QLineEdit field
        self.Purchase_InwWords_field.clear()  # Clear QLineEdit field
        #################################################################################
        self.Purchase_ModeOf_payment_field1.setCurrentIndex(0)
        self.Purchase_ModeOf_payment_field2.setCurrentIndex(0)
        self.Purchase_ModeOf_payment_field3.setCurrentIndex(0)
        self.Purchase_ModeOf_payment_field4.setCurrentIndex(0)
        self.Purchase_ModeOf_payment_field5.setCurrentIndex(0)
        #################################################################
        self.Purchase_InCheck_field1.clear()  # Clear QLineEdit field
        self.Purchase_InCheck_field2.clear()  # Clear QLineEdit field
        self.Purchase_InCheck_field3.clear()  # Clear QLineEdit field
        self.Purchase_InCheck_field4.clear()  # Clear QLineEdit field
        self.Purchase_InCheck_field5.clear()  # Clear QLineEdit field
        ####################################################################################################
        self.Purchase_remarks_field.clear()  # Clear QLineEdit field
        ########################################################################################
        self.Purchase_cnic_upload_label.clear()
        self.Purchase_attachment_label.clear()
        self.Purchase_cnic_upload_label.setText("No image uploaded")
        self.Purchase_attachment_label.setText("No image uploaded")
        ###################################################
        self.purchase_other1.clear()
        self.purchase_other2.clear()
        self.purchase_other3.clear()
        self.purchase_other4.clear()
        self.purchase_other1.setText("No image uploaded")
        self.purchase_other2.setText("No image uploaded")
        self.purchase_other3.setText("No image uploaded")
        self.purchase_other4.setText("No image uploaded")
        self.purchase_Note.clear()
##################################################################################################################################
    def Purchase_upload_image(self, label: QLabel, fill_label: bool = False):
        # Function to handle image upload
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Images (*.png *.jpg *.bmp);;All Files (*)", options=options)
        if file_name:
            pixmap = QPixmap(file_name)
            if fill_label:
                # Scale the pixmap to fill the entire label area
                pixmap = pixmap.scaled(label.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation)
            else:
                # Scale the pixmap to fit within the label while maintaining aspect ratio
                pixmap = pixmap.scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            label.setPixmap(pixmap)
            label.setAlignment(Qt.AlignCenter)  # Center the image in the label
    ##################################################################################################################
    # New method to handle label clicks
    def handle_label_click(self, event, label):
        self.Purchase_other_upload_image(label)
    # Method to handle image upload
    ######################################################
    def Purchase_other_upload_image(self, label):
        # Logic to upload the image
        image_path, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Image Files (*.png *.jpg *.jpeg *.bmp)")
        if image_path:  # Check if an image was selected
            label.setPixmap(QPixmap(image_path).scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
            label.setText("")  # Clear the text if an image is uploaded
    #################################################################################################################################################################
    def purchase_search_records(self):
        """Search for purchase records based on the entered text."""
        shams_search_text2 = self.shams_search_field2.text().strip().lower()  # Strip spaces and convert to lowercase
        
        # Check if the search field is empty
        if not shams_search_text2:
            self.show_custom_message("Please enter a value to search Records.", "Warning", is_success=False)
            return
        
        try:
            # Fetch a broader set of records for partial matching
            all_records = db.collection('Purchases').get()
            
            # Filter records locally for partial matches
            matching_records = []
            for record in all_records:
                purchase_data = record.to_dict()
                if any(shams_search_text2 in str(purchase_data.get(field, '')).lower() for field in 
                    ['invoice_number', 'purchase_registration', 'purchase_chassis']):
                    matching_records.append(record)
            
            # Clear the table before displaying search results
            self.Purchase_table_widget.setRowCount(0)
            
            # Check if there are no matching records
            if not matching_records:
                self.show_custom_message("No records found for the given search criteria.", "Info", is_success=True)
                return
            
            # Loop through the matching records and display them in the table
            for record in matching_records:
                purchase_data = record.to_dict()
                row_position = self.Purchase_table_widget.rowCount()
                self.Purchase_table_widget.insertRow(row_position)
                
                # Fill in the table with the matching record's data
                self.Purchase_table_widget.setItem(row_position, 0, QTableWidgetItem(purchase_data.get('invoice_number', '')))
                self.Purchase_table_widget.setItem(row_position, 1, QTableWidgetItem(purchase_data.get('supplier_name', '')))
                self.Purchase_table_widget.setItem(row_position, 2, QTableWidgetItem(purchase_data.get('supplier_father_name', '')))
                self.Purchase_table_widget.setItem(row_position, 3, QTableWidgetItem(purchase_data.get('supplier_cnic', '')))
                self.Purchase_table_widget.setItem(row_position, 4, QTableWidgetItem(purchase_data.get('purchase_registration', '')))
                self.Purchase_table_widget.setItem(row_position, 5, QTableWidgetItem(purchase_data.get('purchase_model', '')))
                self.Purchase_table_widget.setItem(row_position, 6, QTableWidgetItem(purchase_data.get('purchase_color', '')))
                self.Purchase_table_widget.setItem(row_position, 7, QTableWidgetItem(purchase_data.get('purchase_chassis', '')))
                self.Purchase_table_widget.setItem(row_position, 8, QTableWidgetItem(purchase_data.get('purchase_date', '')))
                
                # Directly retrieve and format purchase_price as an integer string
                purchase_price = purchase_data.get('purchase_price', 0)
                formatted_purchase_price = str(int(purchase_price))  # Ensure it's an integer and then convert to string
                self.Purchase_table_widget.setItem(row_position, 9, QTableWidgetItem(formatted_purchase_price))
                self.Purchase_table_widget.setItem(row_position, 10, QTableWidgetItem(record.id))  # Store document ID
            
            # Resize columns to fit contents
            self.Purchase_table_widget.resizeColumnsToContents()
            
        except Exception as e:
            self.show_custom_message(f"Error searching purchase records: {e}", "Error", is_success=False)
#######################################################################################################################################################################
    def Purchase_header(self, pdf):
        # Add centered image at the top
        self.Purchase_add_centered_image1(pdf, 'report_header.jpg', y=5)
        # First number (Cell: 0322-999066) aligned with the image on the right
        pdf.set_font('Arial', 'B', 12)  # Regular font for the phone number
        pdf.set_xy(pdf.w - 20, 15)  # Adjusted position to ensure the cell fits within the page
        pdf.cell(0, 5, '0321-4946671', ln=True, align='R')
        # Second number (042-35222655) directly below the first number
        pdf.set_xy(pdf.w - 40, pdf.get_y())  # No extra gap, same x, new y directly below the first number
        pdf.cell(0, 5, '0321-8822086', ln=True, align='R')
        # Space after the image and phone numbers
        pdf.ln(0)
        # Centered address
        pdf.set_font('Arial', 'B', 12)  # Bold font for the address
        pdf.cell(0, 10,'CAR EXPERTS Sector-A,Bankers Town Ring Road,Near State Life Society Lahore',align='C', ln=True)
        # Add space before the title
        pdf.ln(3)
        # Sale Receipt/Delivery Note Title centered
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Purchase Receipt', align='C', ln=True)
        pdf.ln(0)
##########################################################################################################################################
    def Purchase_add_centered_image1(self, pdf,image_name, y):
        # Load the image dynamically
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path = os.path.join(base_path, f'dashboard_images/{image_name}')  # Dynamically construct the image path
        # Define the width of the image
        page_width = pdf.w
        image_width = 90  # Image width (can be adjusted)
        # Calculate the center X position to center the image
        x = (page_width - image_width) / 2
        # Add the image to the PDF at the given position
        try:
            pdf.image(image_path, x=x, y=y, w=image_width)
        except RuntimeError as e:
            print(f"Error adding image: {e}")
###################################################################################################
    def get_image_from_label888(self, label):
        pixmap = label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a temporary file to save the image
            with NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
                temp_filename = temp_file.name
                # Save the QImage as a PNG file
                image.save(temp_filename, "PNG")
            return temp_filename  # Return the path to the temporary image file
        return None
#########################################################################################
    def Purchase_sale_info1(self, pdf):
        # Fetching data from PyQt5 form fields
        ##############(Content1)##################
        date = self.date_field.date().toString("yyyy-MM-dd")
        time = self.time_field.time().toString("HH:mm")
        day = self.day_field.text()
        invoice_number = self.invoice_number_field.text()
        # Set font for labels
        pdf.set_font('Arial', '', 10)
        # Prepare the labels and values
        label_date = f'Date: '
        label_time = f'Time: '
        label_day = f'Day: '
        label_invoice_number = f'Invoice Number: '
        # Define the widths of labels and values
        label_widths = [
            pdf.get_string_width(label_date),
            pdf.get_string_width(label_time),
            pdf.get_string_width(label_day),
            pdf.get_string_width(label_invoice_number)
        ]
        value_widths = [
            pdf.get_string_width(date),
            pdf.get_string_width(time),
            pdf.get_string_width(day),
            pdf.get_string_width(invoice_number)
        ]
        # Define horizontal spacing between each section
        spacing = 5  # Adjust spacing between sections
        # Calculate the total width needed
        total_width = sum(label_widths) + sum(value_widths) + spacing * (len(label_widths) + len(value_widths) - 1)
        margin = (pdf.w - total_width) / 5  # Center the text horizontally
        # Starting Y position for Content1
        y_position = 50  # Adjust as needed
        # Set starting position
        pdf.set_xy(margin, y_position)
        # Print each label and value with underlining for values only
        pdf.set_font('Arial', '', 10)
        pdf.cell(label_widths[0], 10, label_date, ln=False, align='L')
        # Underline value part
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[0], 10, date, ln=False, align='L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(spacing, 10, '', ln=False)
        pdf.cell(label_widths[1], 10, label_time, ln=False, align='L')
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[1], 10, time, ln=False, align='L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(spacing, 10, '', ln=False)
        pdf.cell(label_widths[2], 10, label_day, ln=False, align='L')
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[2], 10, day, ln=False, align='L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(15, 10, '', ln=False)
        pdf.cell(label_widths[3], 10, label_invoice_number, ln=False, align='L')
        pdf.set_font('Arial', 'U', 10)
        pdf.cell(value_widths[3], 10, invoice_number, ln=True, align='L')
        # Reset font to regular
        pdf.set_font('Arial', '', 10)
        #########################################################################################################################################
        #######################################adding the table#################################################
        #########################################################################################################################################
        # Set table Y position immediately after Content1
        table_y_position = pdf.get_y() + 2  # Reduce spacing by using a small value like 2
        #Define margins, table dimensions, and cell content
        num_cols = 2
        margin = 10
        label_width=30  # Minimized width for labels (e.g., "Customer Name")
        value_width=120  # Maximized width for values (e.g., "ALI RAFAY MUSHTAQ")
        cell_height=7
        ##############################################################################################
        image_path=self.get_image_from_label888(self.supplier_image_label1)
        #####################################################################################################
        # Image file path and positioning details
        ##image_path = 'Images/circle_loading.jpg'  # Adjust the path based on your setup
        image_width=45
        image_height=35
        # Calculate available space for the table and image
        image_x_position = pdf.w -4- image_width  # Position image to the far right
        table_x_position = margin  # Start the table from the left margin
        ######################################################################
        my_supplier_name=self.Supplier_name_field.currentText()
        my_supplier_father=self.Supplier_father_name_field.text()
        my_supplier_address=self.Supplier_address_field.text()
        my_supplier_phone_no=self.my_Supplier_phonenumber.text()
        my_supplier_cnic=self.Supplier_cnic_field.text()
        ################################################################################################################################
        table_data = [
            ["Seller Name:",my_supplier_name],
            ["Father Name:",my_supplier_father],
            ["Address:",my_supplier_address],
            ["Mobile #:",my_supplier_phone_no],
            ["CNIC:",my_supplier_cnic]
        ]
        ###################################################################################################################################
        # Content to be added to the table
        ###################################################################################################################
        # Set position for the table
        pdf.set_xy(table_x_position, table_y_position)
        # Draw table rows
        for row in range(0, len(table_data)):
            pdf.set_font('Arial', 'B', 10)  # Bold font for labels
            pdf.cell(label_width, cell_height, table_data[row][0], border=1, ln=False, align='L')
            pdf.set_font('Arial', '',10)  # Regular font for values
            pdf.cell(value_width, cell_height, table_data[row][1], border=1, ln=False, align='L')
            pdf.ln(cell_height)  # Move to the next line

        #Add the customer photo to the right of the table
        pdf.image(image_path, x=image_x_position, y=table_y_position, w=image_width, h=image_height)
        # Add a border around the image
        pdf.rect(x=image_x_position, y=table_y_position, w=image_width, h=image_height)
        #######################################adding the table###########################################################
        purchase_s1=self.Purchase_registration_field.text()
        purchase_s2=self.Purchase_make_field.text()
        purchase_s3=self.Purchase_chassis_field.text()
        purchase_s4=self.Purchase_horsepower_field.text()
        purchase_s5=self.Purchase_engine_field.text()
        purchase_s6=self.Purchase_model_field.text()
        # purchase_s8=self.Purchase_import_field.text()
        purchase_s7=self.Purchase_color_field.text()
        ##########################################(3rd content)############################################################
        # Define the data for the third content (Vehicle Description Table)
        #Define the data for the third content (Vehicle Description Table)
        vehicle_description_data=[
                ["Registration #",purchase_s1, "Make",purchase_s2],
                ["Chassis #",purchase_s3, "Horse Power",purchase_s4],
                ["Engine #",purchase_s5, "Model #",purchase_s6],
                ["Color",purchase_s7, "", ""]  # Adjusted row after removing "Import"
        ]
        #################################################################################################################################
        #Add header for vehicle description
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 10, "Description of Vehicle", ln=True, align='L')
        # Set font for table content
        pdf.set_font('Arial', '', 10)
        # Define column widths
        col1_width = 40  # First column (e.g., "Registration #")
        col2_width = 40  # Second column (e.g., "ARG-945")
        col3_width = 35  # Third column (e.g., "Make")
        col4_width = 40  # Fourth column (e.g., "TOYOTA CROSS")
        row_height = 7  # Height of each ro
        # Draw the table rows for vehicle description
        for row in vehicle_description_data:
            pdf.set_font('Arial', 'B', 10)  # Bold font for first and third columns
            pdf.cell(col1_width, row_height, row[0], border=1, ln=False, align='L')
            pdf.set_font('Arial', '', 10)  # Regular font for second and fourth columns
            pdf.cell(col2_width, row_height, row[1], border=1, ln=False, align='L')
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(col3_width, row_height, row[2], border=1, ln=False, align='L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(col4_width, row_height, row[3], border=1, ln=True, align='L')
        # Add space before the next section
        pdf.ln(0)
        ######################################################################################
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, row_height, "Documents Attached:", ln=True, align='L')
        #########################################################################
        # Fetch all document data
        purchase_all_document_data = self.purchase_Note.toPlainText()

        # Define the initial rectangle dimensions
        rect_x = 10  # X-coordinate of the rectangle's top-left corner
        rect_y = pdf.get_y()  # Y position after the previous content
        rect_width = 195  # Width of the rectangle
        padding = 1  # Padding inside the rectangle
        line_height = 4  # Height of each line in the text
        line_spacing = 5  # Line spacing for the text

        # Set font for the text
        pdf.set_font('Arial', '', 8)

        # Calculate the available width for text and height for wrapping
        text_width = rect_width - 2 * padding  # Use the full rectangle width minus padding
        wrapped_lines = pdf.multi_cell(
            text_width, 
            line_height, 
            purchase_all_document_data, 
            border=0, 
            align='L', 
            split_only=True
        )

        # Calculate the total height of the rectangle based on the wrapped text
        rect_height = line_height * len(wrapped_lines) + padding * 2  # Total height of the rectangle

        # Check if the rectangle height exceeds the page's remaining height
        if rect_y + rect_height > pdf.h - 20:  # Account for bottom margin
            pdf.add_page()  # Add a new page
            rect_y = 10  # Reset Y position for the new page

        # Draw the rectangle
        pdf.rect(rect_x, rect_y, rect_width, rect_height)

        # Add the wrapped text inside the rectangle
        pdf.set_xy(rect_x + padding, rect_y + padding)
        pdf.multi_cell(
            text_width, 
            line_height, 
            purchase_all_document_data, 
            border=0, 
            align='L'
        )
        # Move the cursor below the rectangle for further content
        pdf.set_y(rect_y + rect_height + 0)  # Add spacing for the next section
        #######################################################################################################
        cash_payment=self.Purchase_ModeOf_payment_field1.currentText()
        cash_payment1=self.Purchase_ModeOf_payment_field2.currentText()
        cash_payment2=self.Purchase_ModeOf_payment_field3.currentText()
        cash_payment3=self.Purchase_ModeOf_payment_field4.currentText()
        cash_payment4=self.Purchase_ModeOf_payment_field5.currentText()
        #################################################################################
        cheque_payment=self.Purchase_InCheck_field1.text()
        cheque_payment1=self.Purchase_InCheck_field2.text()
        cheque_payment2=self.Purchase_InCheck_field3.text()
        cheque_payment3=self.Purchase_InCheck_field4.text()
        cheque_payment4=self.Purchase_InCheck_field5.text()
        ##########################################################################################################
        total_price_figures=self.Purchase_Price_field.text()
        total_price_words=self.Purchase_InwWords_field.text()
        ##################################################################################
        #Total Price Figures, Total Price Words, and Remarks
        pdf.set_font('Arial', 'B', 10)  # Set font to bold
        pdf.cell(45, 6, "Total Price Figures:", ln=False, align='L')  # Bold label
        pdf.set_font('Arial', '', 10)  # Regular font for value
        pdf.cell(25, 6, total_price_figures, ln=False, align='L')  # Value with minimal spacing
        ##################################################
        pdf.set_font('Arial', 'B', 10)  # Bold label
        pdf.cell(45, 6, "Total Price Words:", ln=False, align='L')  # Bold label
        pdf.set_font('Arial', '', 10)  # Regular font for value
        pdf.cell(25, 6, total_price_words, ln=False, align='L')  # Value with minimal spacing
        pdf.ln(5)
        ###################################################################################################
        ###################################################################################################
        remarks = self.Purchase_remarks_field.toPlainText()
        #############################################################################
        # Set title for the section
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 7, "Payment Details:", ln=True, align='L')
        #####################################################################################
        # Define rectangle dimensions and text to display
        rect_x = 10  # X position of the rectangle
        rect_y = pdf.get_y()  # Y position after the title
        rect_width = 195  # Width of the rectangle
        padding = 2  # Padding for text within the rectangle
        # Combine payment amounts and remarks
        purchase_payments=[cheque_payment,cheque_payment1,cheque_payment2,cheque_payment3,cheque_payment4]
        payments_text1 = " , ".join(map(str,purchase_payments))  # Convert to string and join
        combined_text = f"{payments_text1} | {remarks}"  # Combine payments and remarks
        # Calculate text height based on the rectangle width and font size
        pdf.set_font('Arial', '', 8)  # Set font for text
        line_height = 5  # Height of each line
        text_width = rect_width - 2 * padding  # Width available for text inside the rectangle
        lines = pdf.multi_cell(text_width, line_height, combined_text, border=0, align='L', split_only=True)  # Split text into lines
        rect_height = line_height * len(lines) + 2 * padding  # Adjust rectangle height based on line count
        # Check if rectangle height exceeds page's remaining height
        if rect_y + rect_height > pdf.h - 20:  # Account for bottom margin
            pdf.add_page()  # Add a new page
            rect_y = 10  # Reset Y position for the new page
        # Draw the rectangle
        pdf.rect(rect_x, rect_y, rect_width, rect_height)
        # Position and display text inside the rectangle
        pdf.set_xy(rect_x + padding, rect_y + padding)
        pdf.multi_cell(text_width, line_height, combined_text, border=0, align='L')  # Display text
        # Move the cursor below the rectangle for further content
        pdf.set_y(rect_y + rect_height + 3)  # Add vertical spacing for next section
        ###################################################################################################################################################
        #####################################################################################Adding the final paragraph#######################################
        base_path= getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path= os.path.join(base_path, 'card_images/desc.PNG')  # Construct the path to the image
        # Specify the image dimensions
        image_width=200  # Width of the image
        image_height=15  # Height of the image
        # Calculate the center position on the PDF page
        page_width = pdf.w  # Get the page width from the PDF
        center_x = (page_width - image_width) / 2  # Calculate the center X position
        # Add the image to the PDF at the calculated center position
        # If y is None, the image will be placed at the current position
        try:
            pdf.image(image_path, x=center_x, y=None, w=image_width, h=image_height)
        except Exception as e:
            print(f"Error loading image: {e}")
        pdf.ln(1)
        ##################################(4th content)##################################################################
        pdf.set_font('Arial', '', 10)
        # Center and insert the image
        ##################################################################################
        # Set the base path to dynamically load the image
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path = os.path.join(base_path, 'card_images/my_bottom1.PNG')  # Construct the path to the image
        # Specify the image dimensions
        image_width=200  # Width of the image
        image_height=51  # Height of the image
        # Calculate the center position on the PDF page
        page_width = pdf.w  # Get the page width from the PDF
        center_x = (page_width - image_width) / 2  # Calculate the center X position
        # Add the image to the PDF at the calculated center position
        # If y is None, the image will be placed at the current position
        try:
            pdf.image(image_path, x=center_x, y=None, w=image_width, h=image_height)
        except Exception as e:
            print(f"Error loading image: {e}")
        ######################################################################
        pdf.ln(1)
        #########################################################################
        pdf.set_font('Arial', '',7)
        # Center and insert the image
        # Load the image dynamically
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))  # Set the base path
        image_path = os.path.join(base_path, 'card_images/my_bottom.PNG')  # Dynamically construct the image path
        # Specify the image dimensions
        image_width=200  # Width of the image
        image_height=24  # Height of the image
        # Calculate the center position on the PDF page
        page_width = pdf.w  # Get the width of the page
        center_x = (page_width - image_width) / 2  # Calculate the center X position
        # Add the image to the PDF at the calculated center position
        try:
            pdf.image(image_path, x=center_x, y=None, w=image_width, h=image_height)
        except Exception as e:
            print(f"Error loading image: {e}")
######################################################################################################################################
    def save_image_from_label1(label, file_path):
        pixmap = label.pixmap()
        if pixmap:
            pixmap.save(file_path, "PNG")
###################################################3#################################################################################################################
    def Purchase_generate_report1(self, save_path):
        pdf=FPDF('P','mm','A4')  # Page format
        pdf.add_page()
        self.Purchase_header(pdf)
        self.Purchase_sale_info1(pdf)  # Call the method to add sale info to PDF
        try:
            pdf.output(save_path)
            return True
        except RuntimeError as e:
            print(f"Error generating PDF: {e}")
            return False
    ##################################################
    def Purchase_create_report(self):
        try:
            save_path = self.Purchase_save_report_dialog1()
            if save_path:
                if self.Purchase_generate_report1(save_path):
                    self.show_custom_message("Report generated successfully.", "Success", is_success=True)
                else:
                    self.show_custom_message("Failed to generate report.", "Error", is_success=False)
            else:
                self.show_custom_message("No file selected for saving.", "Error", is_success=False)

        except Exception as e:
            print("Error in create_report:", e)
            self.show_custom_message(f"An error occurred: {e}", "Error", is_success=False)
    ##########################################################################################################
    def Purchase_save_report_dialog1(self):
        # Open a file save dialog to choose the location to save the report
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self,  # Parent widget
            "Save Report",  # Dialog title
            "",  # Default directory (empty means the user's home directory)
            "PDF Files (*.pdf);;All Files (*)",  # Filter options
            options=options
        )
        return file_path if file_path else None
###########################################################################################################################################################################
######################################################################################################(It is Supplier Tab)#############################################################
######################################################################################################(It is Supplier Tab)#############################################################
######################################################################################################(It is Supplier Tab)#############################################################
######################################################################################################(It is Supplier Tab)#############################################################
######################################################################################################(It is Supplier Tab)#############################################################
    def create_supplier_details_tab(self):
        """Create the tab for adding supplier details with a modern table."""
        supplier_widget = QWidget()
        layout = QVBoxLayout(supplier_widget)
        ###################################################################
        # Create a form layout for fields (2 fields per row)
        form_layout = QFormLayout()
        form_layout.setSpacing(20)
        ################################################################
        # Define fixed width for fields
        field_width=300
        ####################################################################
        #Supplier Fields
        self.supplier_id=QLineEdit()
        ########################################
        self.current_id1=self.get_latest_supplier_id_from_db()
        self.supplier_id.setText(str(self.current_id1).zfill(6))
        ################################################################
        self.supplier_id.setPlaceholderText("Seller ID")
        self.supplier_id.setFixedWidth(field_width)
        self.supplier_id.setFixedHeight(35)
        self.supplier_id.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #################################################################################
        self.supplier_id.mousePressEvent=lambda event:self.generate_dynamic_supplier_Id()
        ##################################################################################################
        self.supplier_type=QComboBox()
        self.supplier_type.addItems(["Select_Seller_Type","Dealer/Showroom", "Individual"])
        self.supplier_type.setFixedWidth(field_width)
        self.supplier_type.setFixedHeight(35)
        self.supplier_type.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #########################################################################
        self.supplier_name = QLineEdit()
        self.supplier_name.setPlaceholderText("Seller Name")
        self.supplier_name.setFixedWidth(field_width)
        self.supplier_name.setFixedHeight(35)
        self.supplier_name.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #################################################################
        self.supplier_cnic=QLineEdit()
        self.supplier_cnic.setPlaceholderText("CNIC")
        self.supplier_cnic.setFixedWidth(field_width)
        self.supplier_cnic.setFixedHeight(35)
        self.supplier_cnic.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #################################################################################################
        self.supplier_phone_number=QLineEdit()
        self.supplier_phone_number.setPlaceholderText("Phone Number")
        self.supplier_phone_number.setFixedWidth(field_width)
        self.supplier_phone_number.setFixedHeight(35)
        self.supplier_phone_number.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ##########################################################################################################
        self.supplier_email=QLineEdit()
        self.supplier_email.setPlaceholderText("Email")
        self.supplier_email.setFixedWidth(field_width)
        self.supplier_email.setFixedHeight(35)
        self.supplier_email.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ##############################################################
        self.supplier_address=QLineEdit()
        self.supplier_address.setPlaceholderText("Seller Address")
        self.supplier_address.setFixedWidth(field_width)
        self.supplier_address.setFixedHeight(35)
        self.supplier_address.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ###################################################################
        self.supplier_fathername=QLineEdit()
        self.supplier_fathername.setPlaceholderText("Father_Name")
        self.supplier_fathername.setFixedWidth(field_width)
        self.supplier_fathername.setFixedHeight(35)
        self.supplier_fathername.setStyleSheet("""
            QLineEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                font-size:14px;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
                                               
        """)
        #######################################################
        # Adding fields in pairs to the form layout
        form_layout.addRow(self.create_bold_label("Seller ID"),self.supplier_id)
        form_layout.addRow(self.create_bold_label("Seller Type"),self.supplier_type)
        form_layout.addRow(self.create_bold_label("Seller Name"),self.supplier_name)
        form_layout.addRow(self.create_bold_label("Seller CNIC"),self.supplier_cnic)
        form_layout.addRow(self.create_bold_label("Seller Email"),self.supplier_email)
        ############################################################################
        form_layout.addRow(self.create_bold_label("Seller Phone Number"),self.supplier_phone_number)
        #############################################################################
        form_layout.addRow(self.create_bold_label("Seller Address"),self.supplier_address)
        form_layout.addRow(self.create_bold_label("Seller Father Name"),self.supplier_fathername)
        ###################################################################################################
        # Upload Supplier Image
        upload_image_button = QPushButton("Upload Image")
        upload_image_button.setFixedWidth(200)
        upload_image_button.setFixedHeight(40)
        upload_image_button.setStyleSheet("""
            QPushButton {
                background-color:#5F6F65;/*Modern grey color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#808D7C;
            }
            QPushButton:pressed{
                background-color:#9CA986;
                color:black;
            }
        """)
        upload_image_button.clicked.connect(self.upload_supplier_image)
        ###################################################################
        self.supplier_image_label=QLabel("No image uploaded")
        self.supplier_image_label.setFixedWidth(300)
        self.supplier_image_label.setFixedHeight(200)
        self.supplier_image_label.setAlignment(Qt.AlignCenter)
        self.supplier_image_label.setStyleSheet("""
            QLabel{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                border:2px solid black;
            }
            QLabel::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        form_layout.addRow(upload_image_button,self.supplier_image_label)
        form_layout.setSpacing(20)
        ###########################################################
        # Add Form Layout to Main Layout
        layout.addLayout(form_layout)
        layout.setSpacing(10)
        #####################################################################
        # Buttons for Save and Clear
        button_layout = QHBoxLayout()
        ##########################################################################
        supplier_search_button=QPushButton("Search Record")
        supplier_search_button.setFixedWidth(250)
        supplier_search_button.setFixedHeight(35)
        supplier_search_button.setStyleSheet("""
            QPushButton {
                background-color:#181C14;/*Modern red color */
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#3C3D37;
            }
            QPushButton:pressed{
                background-color:#697565;
            }
        """)
        ##########################################
        supplier_search_button.clicked.connect(self.search_suppllier)
        ########################################################################
        supplier_save_button = QPushButton("Save Records")
        supplier_save_button.setFixedWidth(250)
        supplier_save_button.setFixedHeight(35)
        supplier_save_button.setStyleSheet("""
            QPushButton {
                background-color: #018749;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0D7C66;
            }
            QPushButton:pressed {
                background-color: #018749;
            }
        """)
        #################################
        supplier_save_button.clicked.connect(self.save_suppllier)
        ###################################################################
        Supplier_clear_button = QPushButton("Clear Record")
        Supplier_clear_button.setFixedWidth(250)
        Supplier_clear_button.setFixedHeight(35)
        Supplier_clear_button.setStyleSheet("""
            QPushButton {
                background-color:#624E88;/* Modern blue color */
                color: white;
                border: none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color:#8967B3;
            }
            QPushButton:pressed {
                background-color:#CB80AB;
            }
        """)
        ########################################
        Supplier_clear_button.clicked.connect(self.clear_suppllier)
        ############################################################################################
        ####################################################33
        button_layout.addWidget(supplier_save_button)
        button_layout.addWidget(supplier_search_button)
        button_layout.addWidget(Supplier_clear_button)
        button_layout.setContentsMargins(0,0,0,0)
        layout.addLayout(button_layout)
        # Supplier Table with modern design
        #############################################################################################################################################
        #########################################################################################################################################################
        button_layout_2 = QHBoxLayout()
        ########################################################################################################
        ##################################################################################################################################################
        supplier_update_button=QPushButton("Update Records")
        supplier_update_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;  /* Modern blue color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0069d9;
            }
            QPushButton:pressed {
                background-color: #0056b3;
            }
        """)
        ##############################
        ######################################################
        supplier_update_button.clicked.connect(self.update_suppllier)
        ######################################################
        supplier_update_button.setFixedWidth(200)
        supplier_update_button.setFixedHeight(35)
        ########################################################
        supplier_delete_button = QPushButton("Delete Records")
        supplier_delete_button.setFixedWidth(200)
        supplier_delete_button.setFixedHeight(35)
        supplier_delete_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;  /* Modern red color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        ###############################################################################
        supplier_delete_button.clicked.connect(self.delete_suppllier)
        ###########################################################################
        button_layout_2.addWidget(supplier_update_button)
        button_layout_2.addWidget(supplier_delete_button)
        button_layout_2.setContentsMargins(0,20,0,0)
        layout.addLayout(button_layout_2)
        ########################################################################
        self.table2_layout=QHBoxLayout()
        self.supplier_table = QTableWidget()  # Create a QTableWidget
        self.supplier_table.setColumnCount(6)  # Set the number of columns
        self.supplier_table.setHorizontalHeaderLabels(["Seller ID","Seller Name","CNIC","Email","Address","Document ID"])  # Set headers
        # Modern table styling with background and text color
        self.supplier_table.setStyleSheet("""
            QHeaderView::section {
                background-color:#825B32;
                color: white;
                font-weight: bold;
            }
            QTableWidget {
                background-color: #FAFAFA;
                alternate-background-color: #E0E0E0;
                gridline-color: #B0BEC5;
            }
            QTableWidget::item {
                text-align: center;
            }
        """)
        ################################################################################
        header = self.supplier_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)  # Stretch columns to fit the table width
        # Enable alternating row colors for better readability
        self.supplier_table.setAlternatingRowColors(True)
        self.supplier_table.setFixedHeight(250)  # Fixed height for table
        # Add the table widget to the layout with a stretch factor
        ############################################################################
        self.table2_layout.addWidget(self.supplier_table)
        self.table2_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        layout.addLayout(self.table2_layout,1) 
        #############################################################
        self.load_supplier_records()
        ############################################
        return supplier_widget
######################################################################################################################################################################
##################################################################################################################################
    def get_latest_supplier_id_from_db(self):
        supplier_ref=db.collection('Suppliers')
        my_supplier=supplier_ref.stream() #Get all Supplier
        supplier_ids=[]
        for suplier in my_supplier:
            supplier_data=suplier.to_dict() #convert into dectionary
            supplier_ids.append(int(supplier_data.get('supplier_id',0)))
        if supplier_ids:
           #if supplier exists then increase by one
           supplier_max_id=max(supplier_ids)
           return supplier_max_id+1
        else:
            return 1
    #################################################################################################
    def generate_dynamic_supplier_Id(self):
         # Generate the next ID with leading zeros (e.g., 00001)
         dynamic_supplier_id=str(self.current_id1).zfill(6)
         self.supplier_id.setText(dynamic_supplier_id)
         self.current_id1+=1
###########################################################################################################
    def upload_supplier_image(self):
        """Handle supplier image upload."""
        file_name, _ = QFileDialog.getOpenFileName(None, "Upload Image", "", "Image Files (*.png *.jpg *.jpeg)")
        if file_name:
            #Load the image from the selected file
            pixmap = QPixmap(file_name)
            # Resize the pixmap to fit within the label dimensions (optional, if the image is too large)
            pixmap = pixmap.scaled(self.supplier_image_label.size(), aspectRatioMode=1)  # Keep aspect ratio
            # Set the pixmap (image) onto the label
            self.supplier_image_label.setPixmap(pixmap)
########################################################################################################################################
#############################################################################################################################
    def load_supplier_records(self):
        """Load existing supplier records from Firebase Firestore and display them in the table."""
        try:
            # Fetch supplier records from Firebase Firestore
            suppliers = db.collection('Suppliers').get()
            # Clear existing rows in the supplier table before loading new data
            self.supplier_table.setRowCount(0)
            for supplier in suppliers:
                supplier_data = supplier.to_dict()
                # Insert supplier data into the table
                row_position = self.supplier_table.rowCount()
                self.supplier_table.insertRow(row_position)
                self.supplier_table.setItem(row_position, 0, QTableWidgetItem(supplier_data.get('supplier_id', '')))
                self.supplier_table.setItem(row_position, 1, QTableWidgetItem(supplier_data.get('supplier_name', '')))
                self.supplier_table.setItem(row_position, 2, QTableWidgetItem(supplier_data.get('cnic', '')))
                self.supplier_table.setItem(row_position, 3, QTableWidgetItem(supplier_data.get('email', '')))
                self.supplier_table.setItem(row_position, 4, QTableWidgetItem(supplier_data.get('address', '')))
                self.supplier_table.setItem(row_position, 5, QTableWidgetItem(supplier_data.get('document_id', '')))
        except Exception as e:
            # Show a custom error message if there's an issue loading the supplier records
            self.show_custom_message(f"Error loading suppliers: {e}", "Error", is_success=False)
    ######################################################################################################
    def search_suppllier(self):
        """Handle searching and displaying a single supplier record based on the given criteria."""
        # Retrieve search inputs
        supplier_id_txt = self.supplier_id.text()
        shams_supplier_name = self.supplier_name.text()
        shams_supplier_cnic = self.supplier_cnic.text()
        # Check if any input is provided
        if not (supplier_id_txt or shams_supplier_name or shams_supplier_cnic):
            self.show_custom_message("Please enter Supplier Name or Supplier CNIC.", "Error", is_success=False)
            return
        # Initialize Firestore query
        supplier_query = db.collection('Suppliers')
        # Apply filters based on input fields
        if shams_supplier_name:
            supplier_query = supplier_query.where('supplier_name', '==', shams_supplier_name)
        if shams_supplier_cnic:
            supplier_query = supplier_query.where('cnic', '==', shams_supplier_cnic)
        # Limit the query to one result for a single supplier
        supplier_query = supplier_query.limit(1)
        # Execute the query and handle errors
        try:
            supplier_records = supplier_query.get()
        except Exception as e:
            self.show_custom_message(f"Error while fetching records: {str(e)}", "Error", is_success=False)
            return
        # Check if any records are found
        if len(supplier_records) == 0:
            self.show_custom_message("No records found for the given criteria.", "Information", is_success=False)
            return
        # Process the first record
        first_record = supplier_records[0]
        data = first_record.to_dict()
        # Fill in the fields with the fetched data
        self.supplier_id.setText(data.get('supplier_id', ''))
        self.supplier_type.setCurrentText(data.get('supplier_type', ''))
        self.supplier_name.setText(data.get('supplier_name', ''))
        self.supplier_fathername.setText(data.get('father_name', ''))
        self.supplier_cnic.setText(data.get('cnic', ''))
        self.supplier_phone_number.setText(data.get('supplier_phone', ''))
        self.supplier_email.setText(data.get('email', ''))
        self.supplier_address.setText(data.get('address', ''))
        # Handle image data display from Firebase Storage
        supplier_image_data = data.get('supplier_image', '')  # Assumes supplier_image is a URL or base64 string
        self.display_image(self.supplier_image_label, supplier_image_data)

        # Store the document ID for later use
        self.stored_supplier_id = first_record.id

        # Update supplier table
        self.supplier_table.setRowCount(0)
        try:
            # Populate the table with the single record found
            row_position = self.supplier_table.rowCount()
            self.supplier_table.insertRow(row_position)
            self.supplier_table.setItem(row_position, 0, QTableWidgetItem(data.get('supplier_id', '')))
            self.supplier_table.setItem(row_position, 1, QTableWidgetItem(data.get('supplier_name', '')))
            self.supplier_table.setItem(row_position, 2, QTableWidgetItem(data.get('cnic', '')))
            self.supplier_table.setItem(row_position, 3, QTableWidgetItem(data.get('email', '')))
            self.supplier_table.setItem(row_position, 4, QTableWidgetItem(data.get('address', '')))
            self.supplier_table.setItem(row_position, 5, QTableWidgetItem(first_record.id))
        except Exception as e:
            self.show_custom_message(f"Error updating supplier table: {e}", "Error", is_success=False)

        return self.stored_supplier_id
    #########################################################################
    def display_image(self, label: QLabel, image_data: str):
        """Display an image in the given QLabel.
        :param label: QLabel where the image will be displayed.
        :param image_data: Base64 encoded image data or raw binary image data.
        """
        try:
            # Check if image_data is base64 encoded
            if isinstance(image_data, str) and image_data.startswith('data:image'):
                # Extract base64 data from the data URL
                base64_data = image_data.split(',')[1]
                # Decode base64 data to binary
                image_data = base64.b64decode(base64_data)
            else:
                # Handle if the image is stored as base64 without data URL structure
                image_data = base64.b64decode(image_data)

            # Create an in-memory binary stream from the image data
            image_stream = io.BytesIO(image_data)

            # Load the image from the stream
            pixmap = QPixmap()
            pixmap.loadFromData(image_stream.read(), 'PNG')  # Use 'PNG' for loading PNG images

            # Set the QPixmap to the QLabel, scale the image to fit and maintain aspect ratio
            label.setPixmap(pixmap.scaled(310, 200, Qt.KeepAspectRatio))
            label.setScaledContents(True)  # Ensure scaling of image to fit the label

        except Exception as e:
            # Handle exceptions (e.g., log error, display a placeholder image, etc.)
            print(f"Error displaying image: {e}")
            label.clear()  # Clear the label if image display fails
    #################################################################################################
    #############################################################################
    def get_supplier_image_data_from_label(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.supplier_image_label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    ################################################
    def save_suppllier(self):
        #Retrieve inputs, using local variables to avoid overwriting the original widgets
        supplier_id=self.supplier_id.text()
        my_supplier_type=self.supplier_type.currentText()
        supplier_name=self.supplier_name.text()
        cnic = self.supplier_cnic.text()
        email = self.supplier_email.text()
        supplier_phone=self.supplier_phone_number.text()
        address =self.supplier_address.text()
        father_name = self.supplier_fathername.text()
        # Validation regex
        supplier_name_regex = r'^[A-Za-z\s]{3,}$'
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{3,}$'
        cnic_regex = r'^\d{5}-\d{7}-\d{1}$|^\d{13}$'
        suplier_phone_regex = r'^(\d{5}-\d{7}-\d{1}|\d{13}|0\d{2}\d{8}|\d{11}|\+92\d{10})$'
        # Check for empty fields
        # if not supplier_id or not my_supplier_type or not supplier_name or not email or not address or not father_name or not supplier_phone:
        #     self.show_custom_message("All fields are required.", "Error", is_success=False)
        #     return
        # # Validate supplier name
        # if not re.match(supplier_name_regex, supplier_name):
        #     self.show_custom_message("Supplier Name must be at least 3 characters long and contain only letters.", "Error", is_success=False)
        #     return
        # # Validate CNIC
        # #####################################################################################
        # if not re.match(suplier_phone_regex,supplier_phone):
        #     self.show_custom_message("Phone number must be 11 digits.", "Error", is_success=False)
        #     return
        # #########################################################
        # if not re.match(cnic_regex, cnic):
        #     self.show_custom_message("CNIC must be in the format XXXXX-YYYYYYY-Z or a 13-digit number.", "Error", is_success=False)
        #     return
        # # Validate email
        # if not re.match(email_regex, email):
        #     self.show_custom_message("Invalid email format.", "Error", is_success=False)
        #     return
        supplier_image=self.get_supplier_image_data_from_label()
        # if not supplier_image:
        #     self.show_custom_message("Please Upload a Supplier Image.", "Error", is_success=False)
        # Prepare data for Firebase
        supplier_data={
            'supplier_id':supplier_id,
            'supplier_type':my_supplier_type,
            'supplier_name':supplier_name,
            'cnic':cnic,
            'email':email,
            'supplier_phone':supplier_phone,
            'address':address,
            'father_name':father_name,
            'supplier_image':supplier_image  # Include the base64 image data
        }
        # Check if the supplier_id already exists in Firestore
        existing_supplier_id = db.collection('Suppliers').where('supplier_id', '==', supplier_id).get()
        if existing_supplier_id:
            self.show_custom_message("Seller ID already exists! Please use another supplier ID.", "Error", is_success=False)
            return
        # Check if the supplier already exists in Firestore
        existing_supplier = db.collection('Suppliers').where('email', '==', email).get()
        if existing_supplier:
            self.show_custom_message("Seller already exists! Please use another email address.", "Error", is_success=False)
            return
        # Try saving the data to Firebase Firestore
        try:
            doc_ref = db.collection('Suppliers').add(supplier_data)
            document_id = doc_ref[1].id
            db.collection('Suppliers').document(document_id).update({'document_id': document_id})
            self.show_custom_message("Seller Record is Added successfully!", "Success", is_success=True)
            # Add saved supplier record to the table
            self.populate_supplier_combo_box()
            self.load_supplier_records()
            ##self.populate_supplier_combo_box()
            #Clear the input fields after saving
            self.clear_suppllier()
            self.load_supplier_records1()
            self.current_id1=self.get_latest_supplier_id_from_db()
            self.supplier_id.setText(str(self.current_id1).zfill(6))
            ###########################################################
        except Exception as e:
            self.show_custom_message("Please first connect to the internet before using the system", "Error", is_success=False)
    ######################################################
    def update_suppllier(self):
        """Update the selected supplier record."""
        # Retrieve inputs, using local variables to avoid overwriting the original widgets
        supplier_id = self.supplier_id.text()
        my_supplier_type=self.supplier_type.currentText()
        supplier_name = self.supplier_name.text()
        cnic = self.supplier_cnic.text()
        email = self.supplier_email.text()
        supplier_phone=self.supplier_phone_number.text()
        address = self.supplier_address.text()
        father_name = self.supplier_fathername.text()
        
        # Validation regex
        supplier_name_regex = r'^[A-Za-z\s]{3,}$'
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{3,}$'
        cnic_regex = r'^\d{5}-\d{7}-\d{1}$|^\d{13}$'
        
        # Check for empty fields
        # if not supplier_id or not my_supplier_type or not supplier_name or not email or not address or not father_name or not supplier_phone:
        #     self.show_custom_message("All fields are required.", "Error", is_success=False)
        #     return
        
        # # Validate supplier name
        # if not re.match(supplier_name_regex, supplier_name):
        #     self.show_custom_message("Seller Name must be at least 3 characters long and contain only letters.", "Error", is_success=False)
        #     return
        
        # # Validate CNIC
        # if not re.match(cnic_regex, cnic):
        #     self.show_custom_message("CNIC must be in the format XXXXX-YYYYYYY-Z or a 13-digit number.", "Error", is_success=False)
        #     return
        
        # # Validate email
        # if not re.match(email_regex, email):
        #     self.show_custom_message("Invalid email format.", "Error", is_success=False)
        #     return
        
        # Validate supplier image
        supplier_image1= self.get_supplier_image_data_from_label()
        # if not supplier_image1:
        #     self.show_custom_message("Please Upload a Supplier Image.", "Error", is_success=False)
        #     return
        
        # Prepare data for Firebase
        updated_data = {
            'supplier_type':my_supplier_type,
            'supplier_name': supplier_name,
            'cnic': cnic,
            'email': email,
            'supplier_phone':supplier_phone,
            'address': address,
            'father_name': father_name,
            'supplier_image': supplier_image1  # Include the base64 image data
        }
        
        # Check if the supplier ID exists in Firestore
        supplier_ref = db.collection('Suppliers').where('supplier_id', '==', supplier_id).get()
        if not supplier_ref:
            self.show_custom_message("Seller ID does not exist! Please check the ID and try again.", "Error", is_success=False)
            return
        
        # Try updating the data in Firebase Firestore
        try:
            for doc in supplier_ref:
                doc_ref = db.collection('Suppliers').document(doc.id)
                doc_ref.update(updated_data)
            self.show_custom_message("Seller Record updated successfully!", "Success", is_success=True)
            self.load_supplier_records()  # Refresh the table to reflect updated data
            self.clear_suppllier()  # Clear the input fields after updating
            self.load_supplier_records1()
            self.populate_supplier_combo_box()
            self.current_id1=self.get_latest_supplier_id_from_db()
            self.supplier_id.setText(str(self.current_id1).zfill(6))
            ###############################################################################
        except Exception as e:
            self.show_custom_message("An error occurred while updating the record. Please try again.", "Error", is_success=False)
    ###########################################################################################################################
    def delete_supplier_from_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Suppliers').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
            self.populate_supplier_combo_box()
            self.current_id1=self.get_latest_supplier_id_from_db()
            self.supplier_id.setText(str(self.current_id1).zfill(6))
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
    ########################################################################################################################
    def delete_suppllier(self):
        """Delete the suppliers record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.supplier_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Supplier?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.supplier_table.item(row,5).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.delete_supplier_from_firestore(document_id)
                # Remove the row from the table
                self.supplier_table.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Seller record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    ################################################################################################################################
    def clear_suppllier(self):
        ##self.supplier_id.clear()
        self.supplier_name.clear()
        self.supplier_address.clear()
        self.supplier_cnic.clear()
        self.supplier_phone_number.clear()
        self.supplier_fathername.clear()
        self.supplier_email.clear()
        self.supplier_type.setCurrentIndex(0) 
        self.supplier_image_label.setText('No image uploaded')
        self.load_supplier_records()
    #########################################################################################################
    # Function to fetch supplier details
    def fetch_supplier_details(self):
        suppliers_ref = db.collection('Suppliers')
        docs = suppliers_ref.stream()
        suppliers_details = {}
        suppliers_names = []
        for doc in docs:
            data = doc.to_dict()
            supplier_name = data.get('supplier_name')  # Assuming you want to use supplier_name for selection
            if supplier_name:
                suppliers_names.append(supplier_name)
                suppliers_details[supplier_name] = {
                    'father_name': data.get('father_name', ''),
                    'address': data.get('address', ''),
                    'supplier_phone': data.get('supplier_phone', ''),
                    'cnic': data.get('cnic', ''),
                    'supplier_image': data.get('supplier_image', '')  # Image stored as base64
                }
        return suppliers_details, suppliers_names
    ########################################################
    #Function to populate ComboBox with supplier names
    def populate_supplier_combo_box(self):
        self.Supplier_name_field.clear()
        self.suppliers_details,suppliers_names = self.fetch_supplier_details()
        # Add "Please Select Supplier" option and then the supplier names
        self.Supplier_name_field.addItems(["Select Seller Name"] + suppliers_names)
    ############################################################################################
    #Function to handle supplier selection change
    def on_supplier_selection_changed(self, index):
        # Ensure fields are cleared if no supplier is selected
        if index == 0:
            if hasattr(self, 'Supplier_father_name_field'):
                self.Supplier_father_name_field.setText('')
            if hasattr(self, 'Supplier_cnic_field'):
                self.Supplier_cnic_field.setText('')
            if hasattr(self, 'Supplier_address_field'):
                self.Supplier_address_field.setText('')
            #################################################################################
            if hasattr(self,'my_Supplier_phonenumber'):
                self.my_Supplier_phonenumber.setText('')
            #######################################################################
            if hasattr(self, 'supplier_image_label'):
                self.supplier_image_label1.clear()
        else:
            supplier_name = self.Supplier_name_field.currentText()
            details = self.suppliers_details.get(supplier_name, {})

            if hasattr(self, 'Supplier_father_name_field'):
                self.Supplier_father_name_field.setText(details.get('father_name', ''))
            if hasattr(self, 'Supplier_address_field'):
                self.Supplier_address_field.setText(details.get('address', ''))
            if hasattr(self, 'Supplier_cnic_field'):
                self.Supplier_cnic_field.setText(details.get('cnic', ''))
            ###########################(for phone number)#####################################################
            ###############################################################################################
            if hasattr(self,'my_Supplier_phonenumber'):
                self.my_Supplier_phonenumber.setText(details.get('supplier_phone',''))
            #################################################################################################
            # Handling the image
            supplier_image = details.get('supplier_image', '')
            if hasattr(self, 'supplier_image_label1') and supplier_image:
                try:
                    # Decode the image from base64 and display it
                    supplier_image_data = base64.b64decode(supplier_image)
                    pixmap = QPixmap()
                    pixmap.loadFromData(supplier_image_data, 'PNG')
                    self.supplier_image_label1.setPixmap(pixmap.scaled(310, 200, Qt.KeepAspectRatio))
                except Exception as e:
                    print(f"Error decoding image: {e}")
                    self.supplier_image_label1.clear()
            else:
                self.supplier_image_label1.clear()
#############################################################################################################################################################################
###################################################################################################################################################################################################################################
######################################################################################################(It is Supplier Tab)#############################################################
    def create_employee_management_view(self):
        """Create the inventory store management view with a modern look."""
        employee_management_widget=QWidget()
        ##################################################
        employee_management_widget.setStyleSheet("background-color:white;")
        #################################################
        layout=QVBoxLayout(employee_management_widget)
        # Create QTabWidget with modern styling
        employee_tab_widget=QTabWidget()
        employee_tab_widget.setStyleSheet("""
            QTabWidget::pane { /* The tab widget frame */
                border-top: 1px solid #C0C0C0;
                background-color:white;
                border-radius:4px;
            }
            QTabBar::tab {
                background: #E0E0E0;
                color: #333;
                border: 1px solid #C0C0C0;
                padding: 10px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
                min-width: 150px;                
            }
            QTabBar::tab:hover {
                background: #B0BEC5;
            }
            QTabBar::tab:selected {
                background-color: #42A5F5;
                color: white;
                font-weight: bold;
            }
            QTabBar::tab:!selected {
                margin-top: 2px;
            }
        """)
        # Add Customer Details Tab
        add_employee=self.create_employee_details_tab()
        employee_tab_widget.addTab(add_employee,"Add Employee Details")
        # Add Supplier Details Tab#########################
        # Add Stock Details Tab (Placeholder for stock display)
        payment_employee=self.payment_employee_details_tab()
        employee_tab_widget.addTab(payment_employee,"Employee Payment Details")
        # Add the QTabWidget to the layout
        layout.addWidget(employee_tab_widget)
        return employee_management_widget
###########################################################################################################################################################################
###################################################################################################################################################################
###################################################################################################################################################################
#################################(employee_Payment Module)##################################################################################################################################
###################################################################################################################################################################
###################################################################################################################################################################
############################################################################################################################################################################################
    def payment_employee_details_tab(self):
        #Create the Employee Payment Details tab
        employee_payment_widget = QWidget()
        main_layout = QVBoxLayout(employee_payment_widget)
        main_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        # Title Label
        title_label = QLabel("Add Employee Salaries Details")
        title_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        title_label.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                margin-bottom: 10px;
            }
        """)
        main_layout.addWidget(title_label)
        self.line = QLabel("_______________________________________________________________________________________________________________________________________________________________")
        self.line.setFixedWidth(1080)
        main_layout.addWidget(self.line)
        # Global styles for QLabel and QLineEdit
        employee_payment_widget.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
            }
            QLineEdit {
                font-size: 14px;
                font-weight: bold;
                background-color:#ECDFCC;
                color: black;
                border-radius: 10px;
            }
            QComboBox {
                background-color:#ECDFCC;
                color: black;
                font-weight: bold;
                border-radius: 10px;
            }
        """)

        # Employee Details Layout
        ##########################################################
        employee_details_layout = QGridLayout()
        employee_details_layout.setSpacing(20)
        # Fields without labels
        self.employee_id = QComboBox()
        self.employee_id.setFixedWidth(230)
        self.employee_id.currentIndexChanged.connect(self.on_employee_selection_changed)

        self.date_of_joining = QLineEdit()
        self.date_of_joining.setPlaceholderText("Employee Join Date")
        self.date_of_joining.setFixedWidth(250)

        self.gender_combo = QLineEdit()
        self.gender_combo.setPlaceholderText("Employee Gender")
        self.gender_combo.setFixedWidth(250)

        self.contact_number = QLineEdit()
        self.contact_number.setPlaceholderText("Contact Number")
        self.contact_number.setFixedWidth(250)

        self.designation = QLineEdit()
        self.designation.setPlaceholderText("Employee Designation")
        self.designation.setFixedWidth(250)

        self.employee_image_label = QLabel("No image uploaded")
        self.employee_image_label.setFixedSize(230, 150)
        self.employee_image_label.setAlignment(Qt.AlignCenter)
        self.employee_image_label.setStyleSheet("""
            QLabel {
                background-color: white;
                color: black;
                border: 2px solid black;
                border-radius: 2px;
            }
        """)

        # Adding elements in a grid without labels
        employee_details_layout.addWidget(self.employee_id, 0, 0)
        employee_details_layout.addWidget(self.date_of_joining, 0, 1)
        employee_details_layout.addWidget(self.gender_combo, 1, 0)
        employee_details_layout.addWidget(self.contact_number, 1, 1)
        employee_details_layout.addWidget(self.designation, 2, 0)
        employee_details_layout.addWidget(self.employee_image_label, 2, 1)
        ############################################################################
        employee_details_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        ##################################################
        self.populate_employee_combo_box()
        main_layout.addLayout(employee_details_layout)
        main_layout.setSpacing(10)
        ###########################################################################################################################################################################################
        ###################################################################################################################################################################
        # Salary Details Layout
        salary_details_layout = QGridLayout()
        salary_details_layout.setSpacing(20)
        # Fields without labels
        self.basic_salary = QLineEdit()
        self.basic_salary.setPlaceholderText("Basic Salary")
        self.basic_salary.setFixedWidth(250)
        self.allowances_combo = QComboBox()
        self.allowances_combo.addItems(["Housing", "Transportation", "Other"])
        self.allowances_combo.setFixedWidth(230)
        self.bonuses = QLineEdit()
        self.bonuses.setPlaceholderText("Bonuses/Commission")
        self.bonuses.setFixedWidth(250)
        self.deductions = QLineEdit()
        self.deductions.setPlaceholderText("Deductions")
        self.deductions.setFixedWidth(250)
        self.net_salary = QLineEdit()
        self.net_salary.setPlaceholderText("Net Salary")
        self.net_salary.setFixedWidth(250)
        self.net_salary.setReadOnly(True)
        self.bank_account = QLineEdit()
        self.bank_account.setPlaceholderText("Bank Account Number")
        self.bank_account.setFixedWidth(250)
        # Adding elements in a grid without labels
        salary_details_layout.addWidget(self.basic_salary, 0, 0)
        salary_details_layout.addWidget(self.allowances_combo, 0, 1)
        salary_details_layout.addWidget(self.bonuses, 1, 0)
        salary_details_layout.addWidget(self.deductions, 1, 1)
        salary_details_layout.addWidget(self.net_salary, 2, 0)
        salary_details_layout.addWidget(self.bank_account, 2, 1)
        ######################################################################
        ###############################################################################
        # Calculate Net Salary Button
        calculate_button = QPushButton("Calculate Net Salary")
        calculate_button.setFixedWidth(200)
        calculate_button.setFixedHeight(40)
        calculate_button.clicked.connect(self.calculate_net_salary)
        calculate_button.setStyleSheet("""
            QPushButton {
                background-color: #A28B55;
                color: white;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #86AB89;
            }
            QPushButton:pressed {
                background-color: #CBE2B5;
            }
        """)
        salary_details_layout.addWidget(calculate_button, 3, 0, 1, 2, Qt.AlignCenter)
        main_layout.addLayout(salary_details_layout)
        # Set the layout to the employee_payment_widget
        salary_details_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        ####################################################################################################################################################
        # Save and Clear Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Save Records")
        save_button.setFixedWidth(200)
        save_button.setFixedHeight(35)
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #018749;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0D7C66;
            }
            QPushButton:pressed {
                background-color: #018749;
            }
        """)
        save_button.clicked.connect(self.save_employee_salary)

        clear_button = QPushButton("Clear Records")
        clear_button.setFixedWidth(200)
        clear_button.setFixedHeight(35)
        clear_button.setStyleSheet("""
            QPushButton {
                background-color:#624E88;
                color: white;
                border: none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color:#8967B3;
            }
            QPushButton:pressed {
                background-color:#CB80AB;
            }
        """)
        clear_button.clicked.connect(self.clear_employee_salary)
        button_layout.addWidget(save_button)
        button_layout.addWidget(clear_button)
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(button_layout)
        ############################################################################################################################################################
        self.line_value=QLabel("_______________________________________________________________________________________________________________________________________________________________")
        self.line_value.setFixedWidth(1080)
        main_layout.addWidget(self.line_value)
        #############################################################################################################################################
        #########################################################################################################
        # Buttons for Update, Delete, and Search
        button_layout_2 = QHBoxLayout()
        update_button = QPushButton("Update Records")
        update_button.setFixedWidth(200)
        update_button.setFixedHeight(35)
        update_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius:5px;
                padding:10px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0069d9;
            }
            QPushButton:pressed {
                background-color: #0056b3;
            }
        """)
        update_button.clicked.connect(self.Salary_update_employee)
        delete_button = QPushButton("Delete Records")
        delete_button.setFixedWidth(200)
        delete_button.setFixedHeight(35)
        delete_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        delete_button.clicked.connect(self.Salary_delete_employee)
        search_button = QPushButton("Search Records")
        search_button.setFixedWidth(200)
        search_button.setFixedHeight(35)
        search_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        search_button.clicked.connect(self.Salary_search_employee)
        button_layout_2.addWidget(update_button)
        button_layout_2.addWidget(delete_button)
        button_layout_2.addWidget(search_button)
        button_layout_2.setAlignment(Qt.AlignCenter|Qt.AlignTop)
        main_layout.addLayout(button_layout_2)
        #######################################################################################################################################################
        # Table for Employee Records
        table_layout=QHBoxLayout()
        self.employee_salary_table = QTableWidget()
        self.employee_salary_table.setFixedSize(1040, 300)
        self.employee_salary_table.setColumnCount(6)
        self.employee_salary_table.setHorizontalHeaderLabels(
            ["Employee Name","Contact", "Gender", "Net Salary", "Account NO", "Document ID"]
        )
        self.employee_salary_table.setStyleSheet("""
            QHeaderView::section {
                background-color: #825B32;
                color: white;
                font-weight: bold;
            }
            QTableWidget {
                background-color: #FAFAFA;
                alternate-background-color: #E0E0E0;
                gridline-color: #B0BEC5;
            }
            QTableWidget::item {
                text-align: center;
            }
        """)

        header = self.employee_salary_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Fixed)
        for i in range(7):
            self.employee_salary_table.setColumnWidth(i, 150)
        ##############################################################
        table_layout.addWidget(self.employee_salary_table)
        table_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        #################################################################
        main_layout.addLayout(table_layout)
        self.load_employee_Salary_data()  #########loading data#############
        ####################################################################################################################################################
        employee_payment_widget.setLayout(main_layout)
        # Return or display employee_payment_widget in your main window
        return employee_payment_widget
####################################################################################################################################################################
#############################################################################################################################################################
    def Salary_update_employee(self):
        """Handle updating employee salary details with validation and Firebase storage."""
        # Retrieve input values
        employee_id = self.employee_id.currentText()
        date_of_joining = self.date_of_joining.text()
        gender = self.gender_combo.text()
        contact_number = self.contact_number.text()
        designation = self.designation.text()
        basic_salary = self.basic_salary.text()
        allowances = self.allowances_combo.currentText()
        bonuses = self.bonuses.text()
        deductions = self.deductions.text()
        net_salary = self.net_salary.text()
        bank_account = self.bank_account.text()
        
        # # Validation: Check if required fields are filled
        # if not employee_id or employee_id == "Select Employee ID" or not date_of_joining or not gender or not contact_number or not designation or not basic_salary or not net_salary or not bank_account:
        #     self.show_custom_message("All required fields must be filled.", "Error", is_success=False)
        #     return

        # Salary validation: Make sure salary-related fields are numbers
        try:
            basic_salary = float(basic_salary)
            bonuses = float(bonuses) if bonuses else 0
            deductions = float(deductions) if deductions else 0
            net_salary = float(net_salary)
        except ValueError:
            self.show_custom_message("Salary fields must be valid numbers.", "Error", is_success=False)
            return

        # Prepare the employee salary data
        salary_data = {
            "employee_name":employee_id,
            "date_of_joining": date_of_joining,
            "gender": gender,
            "contact_number": contact_number,
            "designation": designation,
            "basic_salary": basic_salary,
            "allowances": allowances,
            "bonuses": bonuses,
            "deductions": deductions,
            "net_salary": net_salary,
            "bank_account": bank_account
        }
        # Check if a salary record for this employee exists
        existing_salary = db.collection('Employee_Salary').where('employee_name','==', employee_id).get()
        if not existing_salary:
            self.show_custom_message("Salary record not found for this employee ID.", "Error", is_success=False)
            return
        # Try updating the data in Firebase Firestore
        try:
            for doc in existing_salary:
                doc_ref = db.collection('Employee_Salary').document(doc.id)
                doc_ref.update(salary_data)
            self.show_custom_message("Employee salary record updated successfully!", "Success", is_success=True)
            # Clear the form after successful update
            self.clear_employee_salary()
            self.load_employee_Salary_data()
            self.load_emplpyee_payment_report11()
        except Exception as e:
            self.show_custom_message("Failed to update record. Please check your internet connection.", "Error", is_success=False)
            print(f"Error: {e}")
    ############################################################################
    def Salary_search_employee(self):
        """Search for an employee by Employee ID and populate the fields."""
        employee_id = self.employee_id.currentText()  # Get the selected employee ID
        if not employee_id or employee_id == "Select Employee Name":
            self.show_custom_message("Employee Name is required for searching.", "Error", is_success=False)
            return
        # Search for the employee in Firestore
        try:
            # Query Firestore for documents where 'employee_id' matches the given ID
            results = db.collection('Employee_Salary').where('employee_name', '==', employee_id).stream()
            # Check if any results were found
            results_list = list(results)
            if not results_list:
                self.show_custom_message("No records found for this Employee ID.", "Error", is_success=False)
                return
            # Use the first result (if multiple results are possible, handle accordingly)
            employee_data = results_list[0].to_dict()
            # Populate the fields with employee data, ensuring no None values are passed
            self.basic_salary.setText(str(employee_data.get('basic_salary', '') or ''))
            self.allowances_combo.setCurrentText(str(employee_data.get('allowances', '') or ''))
            self.bonuses.setText(str(employee_data.get('bonuses', '') or ''))
            self.deductions.setText(str(employee_data.get('deductions', '') or ''))
            self.net_salary.setText(str(employee_data.get('net_salary', '') or ''))
            self.bank_account.setText(str(employee_data.get('bank_account', '') or ''))
            # Load employee image if available
        except Exception as e:
            self.show_custom_message(f"Error fetching employee details: {str(e)}", "Error", is_success=False)
    #################################################################################################################################################################
    def Salary_employee_delete_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Employee_Salary').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
    ######################################################################################################################################
    #########################################################################################################################################
    #################################################################################################################################################################
    def Salary_delete_employee(self):
        """Delete the customer record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.employee_salary_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self,"Confirm Deletion",
                                    "Are you sure you want to delete this customer?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.employee_salary_table.item(row,5).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.Salary_employee_delete_firestore(document_id)
                # Remove the row from the table
                self.employee_salary_table.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Employees record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    #################################################################################################################################################################
    #################################################################################################################################################################
    def get_image_data_from_label(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.employee_image_label.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
    ##############################################################################################################################################################
    #########################################################################################################################################################
    def clear_employee_salary(self):
        """Clear all employee salary form fields."""
        # Clear all the input fields
        self.employee_id.setCurrentIndex(0)
        self.date_of_joining.clear()
        self.gender_combo.clear()
        self.contact_number.clear()
        self.designation.clear()
        self.basic_salary.clear()
        self.allowances_combo.setCurrentIndex(0)
        self.bonuses.clear()
        self.deductions.clear()
        self.net_salary.clear()
        self.bank_account.clear()
        self.employee_image_label.setText("No image uploaded")  # Clear image label if necessary
    ###########################################################################################################################################
    #######################################################################################################################################
    def load_employee_Salary_data(self):
            """Load existing employee records from Firebase Firestore and display them in the table."""
            try:
                # Fetch employee salary records from Firebase Firestore
                employee_records = db.collection('Employee_Salary').get()
                # Clear existing rows in the table before loading new data
                self.employee_salary_table.setRowCount(0)
                # Loop through each record and insert it into the table
                for record in employee_records:
                    employee_data = record.to_dict()
                    # Insert a new row in the table
                    row_position = self.employee_salary_table.rowCount()
                    self.employee_salary_table.insertRow(row_position)
                    # Fill in the columns with data from the data source
                    self.employee_salary_table.setItem(row_position, 0, QTableWidgetItem(str(employee_data.get('employee_name', ''))))
                    self.employee_salary_table.setItem(row_position, 1, QTableWidgetItem(employee_data.get('contact_number', '')))
                    self.employee_salary_table.setItem(row_position, 2, QTableWidgetItem(employee_data.get('gender', '')))
                    self.employee_salary_table.setItem(row_position, 3, QTableWidgetItem(str(employee_data.get('net_salary', ''))))
                    self.employee_salary_table.setItem(row_position, 4, QTableWidgetItem(employee_data.get('bank_account', '')))
                    self.employee_salary_table.setItem(row_position, 5, QTableWidgetItem(employee_data.get('document_id', '')))
                # Optionally, resize columns to fit contents
            except Exception as e:
                self.show_custom_message(f"Error loading employee records: {e}", "Error", is_success=False)
    ############################################################################################################################################################
    #################################################################################################################################################
    def save_employee_salary(self):
        """Handle saving employee salary details with validation and Firebase storage."""
        # Retrieve input values
        employee_id = self.employee_id.currentText()
        date_of_joining = self.date_of_joining.text()
        gender = self.gender_combo.text()
        contact_number = self.contact_number.text()
        designation = self.designation.text()
        basic_salary = self.basic_salary.text()
        allowances = self.allowances_combo.currentText()
        bonuses = self.bonuses.text()
        deductions = self.deductions.text()
        net_salary = self.net_salary.text()
        bank_account = self.bank_account.text()
        # Validation: Check if required fields are filled
        # if not employee_id or employee_id == "Select Employee Name" or not date_of_joining or not gender or not contact_number or not designation or not basic_salary or not net_salary or not bank_account:
        #     self.show_custom_message("All required fields must be filled.", "Error", is_success=False)
        #     return
        # Salary validation: Make sure salary-related fields are numbers
        employe_salary_image_data = self.get_image_data_from_label()
        try:
            basic_salary = float(basic_salary)
            bonuses = float(bonuses) if bonuses else 0
            deductions = float(deductions) if deductions else 0
            net_salary = float(net_salary)
        except ValueError:
            self.show_custom_message("Salary fields must be valid numbers.", "Error", is_success=False)
            return
        # Prepare the employee salary data
        salary_data = {
            "employee_name": employee_id,
            "date_of_joining": date_of_joining,
            "gender": gender,
            "contact_number": contact_number,
            "designation": designation,
            "employee_image": employe_salary_image_data,
            "basic_salary": basic_salary,
            "allowances": allowances,
            "bonuses": bonuses,
            "deductions": deductions,
            "net_salary": net_salary,
            "bank_account": bank_account
        }
        # Check if a salary record for this employee already exists
        existing_salary = db.collection('Employee_Salary').where('employee_name', '==', employee_id).get()
        if existing_salary:
            self.show_custom_message(f"Salary record for {employee_id} already exists.", "Error", is_success=False)
            return
        # Try saving the data to Firebase Firestore
        try:
            doc_ref = db.collection('Employee_Salary').add(salary_data)
            document_id=doc_ref[1].id  # Getting the document ID
            db.collection('Employee_Salary').document(document_id).update({'document_id': document_id})
            self.show_custom_message("Employee salary record saved successfully!", "Success", is_success=True)
            # Clear the form after successful save
            self.clear_employee_salary()
            self.load_employee_Salary_data()
            self.load_emplpyee_payment_report11()
        except Exception as e:
            self.show_custom_message("Failed to save record. Please check your internet connection.", "Error", is_success=False)
            print(f"Error: {e}")
    ##################################################################################################################################################################
    ############################################################################################################################################################
    def calculate_net_salary(self):
        """Calculate the net salary based on basic salary, allowances, bonuses, and deductions."""
        basic_salary=self.basic_salary.text()
        bonuses=self.bonuses.text()
        deductions=self.deductions.text()
        if not basic_salary or not bonuses or not deductions:
            self.show_custom_message("All fields are required.", "Error", is_success=False)
            return
        else:
            try:
                basic_salary = float(self.basic_salary.text())
                bonuses = float(self.bonuses.text())
                deductions = float(self.deductions.text())
                # Calculate net salary
                net_salary = basic_salary + bonuses - deductions
                self.net_salary.setText(f"{net_salary:.2f}")
            except ValueError:
                self.show_custom_message(self,"Please enter valid numeric values for salary fields.","Error", is_success=False)
##################################################################################################################################################################################################################################################
###################################################################################################################################################################################################
###################################################################################################################################################################################################
#############################################################################################################################################################################################
###################################################(Add Employee Module)###############################################################################################################################
###################################################(Add Employee Module)##########################################################################################################################
###################################################(Add Employee Module)##########################################################################################################################
###################################################(Add Employee Module)########################################################################################################################
    def create_employee_details_tab(self):
        """Create the Add Employee Details tab."""
        employee_widget = QWidget()
        layout = QVBoxLayout(employee_widget)
        layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        # Title Label
        title_label = QLabel("Add Employee Details")
        title_label.setAlignment(Qt.AlignLeft)
        title_label.setStyleSheet("""
            QLabel {
                font-size:20px;
                font-weight: bold;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(title_label)
        # Form Layout for employee fields (2 fields per row)
        ###################################################################
        form_layout = QFormLayout()
        form_layout.setSpacing(20)
        # Define fixed width for fields
        field_width=300
        field_height=35
        # Employee ID
        self.employee_id1=QLineEdit()
        self.employee_id1.setPlaceholderText("Employee ID")
        #########################################################
        self.employee_id1.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ########################################################
        self.employee_id1.setFixedWidth(field_width)
        self.employee_id1.setFixedHeight(field_height)
        self.employee_id1.setAlignment(Qt.AlignLeft)
        ###################################################################
        form_layout.addRow(QLabel("Employee ID:"), self.employee_id1)
        #####################################################################################################
        ###################Employee Name################################
        self.employee_name1= QLineEdit()
        self.employee_name1.setPlaceholderText("Employee Name")
        self.employee_name1.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #################################################################
        self.employee_name1.setAlignment(Qt.AlignLeft)
        ###########################################################
        self.employee_name1.setFixedWidth(field_width)
        self.employee_name1.setFixedHeight(field_height)
        form_layout.addRow(QLabel("Employee Name:"),self.employee_name1)
        ###################################################################
        # Date of Joining
        self.date_of_joining1=QDateEdit()
        #self.date_of_joining.setPlaceholderText("Date of Joining")
        self.date_of_joining1.setCalendarPopup(True)  # Enable the calendar popup view
        ###################################################
        self.date_of_joining1.setStyleSheet("""
            QDateEdit{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QDateEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #############################################################
        # Optionally, set a default date
        self.date_of_joining1.setDate(QDate.currentDate())
        ##########################################################
        self.date_of_joining1.setFixedWidth(field_width)
        self.date_of_joining1.setFixedHeight(field_height)
        #######################################################
        # Add the date picker to the form layout
        form_layout.addRow(QLabel("Date of Joining:"), self.date_of_joining1)
        ##################################################################
        # Gender
        self.gender_combo1=QComboBox()
        self.gender_combo1.addItems(["Select Gender","Male", "Female", "Other"])
        #######################################
        self.gender_combo1.setStyleSheet("""
            QComboBox{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QComboBox::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ##################################################################
        self.gender_combo1.setFixedWidth(field_width)
        self.gender_combo1.setFixedHeight(field_height)
        #################################################################
        form_layout.addRow(QLabel("Gender:"),self.gender_combo1)
        #################################################################
        # Contact Number
        self.contact_number1=QLineEdit()
        self.contact_number1.setPlaceholderText("Contact Number")
        #################################################
        self.contact_number1.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        #############################################################
        self.contact_number1.setFixedWidth(field_width)
        self.contact_number1.setFixedHeight(field_height)
        ##############################################################
        form_layout.addRow(QLabel("Contact Number:"),self.contact_number1)
        #################################################################
        # Email Address
        self.email_address1=QLineEdit()
        self.email_address1.setPlaceholderText("Email Address")
        #################################################################
        self.email_address1.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        ###################################################################
        self.email_address1.setFixedWidth(field_width)
        self.email_address1.setFixedHeight(field_height)
        ###########################################################
        form_layout.addRow(QLabel("Email Address:"), self.email_address1)
        #################################################################
        # Employee Designation
        self.designation1=QLineEdit()
        self.designation1.setPlaceholderText("Employee Designation")
        self.designation1.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
            }
            QLineEdit::Placeholder{
                color:white;
                font-weight:bold;
            }
        """)
        self.designation1.setFixedWidth(field_width)
        self.designation1.setFixedHeight(field_height)
        ########################################################
        form_layout.addRow(QLabel("Employee Designation:"),self.designation1)
        ########################################################################
        # Employee Image
        self.upload_image_button=QPushButton("Upload Employee Image")
        self.upload_image_button.setFixedWidth(200)
        self.upload_image_button.setFixedHeight(40)
        self.upload_image_button.setStyleSheet("""
            QPushButton {
                background-color:#5F6F65;
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#808D7C;
            }
            QPushButton:pressed{
                background-color:#9CA986;
                color:black;
            }
        """)
        self.upload_image_button.clicked.connect(self.upload_employee_image)

        self.employee_image_label1= QLabel("No image uploaded")
        self.employee_image_label1.setFixedWidth(300)
        self.employee_image_label1.setFixedHeight(200)
        self.employee_image_label1.setAlignment(Qt.AlignCenter)
        self.employee_image_label1.setStyleSheet("""
            QLabel{
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                border:2px solid black;
            }
        """)

        form_layout.addRow(self.upload_image_button, self.employee_image_label1)
        form_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        # Add Form Layout to Main Layout
        layout.addLayout(form_layout)
        layout.setSpacing(10)
##########################################################################
        # Buttons for Save, Search, and Clear
        button_layout = QHBoxLayout()
        save_button = QPushButton("Save Records")
        save_button.setFixedWidth(200)
        save_button.setFixedHeight(35)
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #018749;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0D7C66;
            }
            QPushButton:pressed {
                background-color: #018749;
            }
        """)
        ###########################################
        save_button.clicked.connect(self.save_employee)
        ###########################################################3
        search_button=QPushButton("Search Records")
        search_button.setFixedWidth(200)
        search_button.setFixedHeight(35)
        search_button.setStyleSheet("""
            QPushButton {
                background-color:#181C14;
                color:white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover{
                background-color:#3C3D37;
            }
            QPushButton:pressed{
                background-color:#697565;
            }
        """)
        ########################################################
        search_button.clicked.connect(self.search_employee)
        ######################################################################
        clear_button=QPushButton("Clear Records")
        clear_button.setFixedWidth(200)
        clear_button.setFixedHeight(35)
        clear_button.setStyleSheet("""
            QPushButton {
                background-color:#624E88;
                color: white;
                border: none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color:#8967B3;
            }
            QPushButton:pressed {
                background-color:#CB80AB;
            }
        """)
        #################################
        clear_button.clicked.connect(self.clear_employee)
        ###############################################################
        button_layout.addWidget(save_button)
        button_layout.addWidget(search_button)
        button_layout.addWidget(clear_button)
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        layout.addLayout(button_layout)
        #######################################################
        line_layout_2=QHBoxLayout()
        self.line=QLabel("____________________________________________________________________________________________________________________________________________________________________________")
        self.line.setFixedWidth(1080)
        line_layout_2.addWidget(self.line)
        line_layout_2.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        layout.addLayout(line_layout_2)
        #######################################################################
        # Buttons for Update and Delete
        button_layout_2 = QHBoxLayout()
        update_button = QPushButton("Update Records")
        update_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius:5px;
                padding:10px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0069d9;
            }
            QPushButton:pressed {
                background-color: #0056b3;
            }
        """)
        update_button.setFixedWidth(200)
        update_button.setFixedHeight(35)
        update_button.clicked.connect(self.update_employee)
        ###################################################################################
        delete_button = QPushButton("Delete Records")
        delete_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        delete_button.setFixedWidth(200)
        delete_button.setFixedHeight(35)
        ####################################################
        delete_button.clicked.connect(self.delete_employee)
        ###################################################################################################
        button_layout_2.addWidget(update_button)
        button_layout_2.addWidget(delete_button)
        button_layout_2.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        ###############################################################################
        layout.addLayout(button_layout_2)
        layout.setSpacing(10)
        ###################################################################################################################################################
        self.my_table_data=QHBoxLayout()
        # Employee Table
        self.employee_table=QTableWidget(0,8)  # 8 columns
        self.employee_table.setHorizontalHeaderLabels([
            "Employee ID","Employee Name","Joining Date","Gender","Contact", 
            "Email", "Designation","Document ID"
        ])
        self.employee_table.setStyleSheet("""
            QHeaderView::section {
                background-color:#825B32;
                color: white;
                font-weight: bold;
            }
            QTableWidget {
                background-color: #FAFAFA;
                alternate-background-color: #E0E0E0;
                gridline-color: #B0BEC5;
            }
            QTableWidget::item {
                text-align: center;
            }
        """)

        header = self.employee_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)  # Stretch columns to fit the table width
        self.employee_table.setAlternatingRowColors(True)
        self.employee_table.setFixedHeight(250)  # Fixed height for table
        self.my_table_data.addWidget(self.employee_table)
        #################################################################
        self.my_table_data.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        layout.addLayout(self.my_table_data,1)  # Stretch factor 1
        #############################################################################################
        ##########################################################################
        self.load_employee_data()
        ###################################################################
        return employee_widget
#############################################################################################################################33####
###################################################################################################################################
    def employee_delete_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Employees').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
            self.populate_employee_combo_box()
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
##################################################################################################################################
    def delete_employee(self):
        """Delete the customer record based on the selected row in the table."""
        # Get the selected items in the table
        selected_items=self.employee_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this customer?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.employee_table.item(row,7).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.employee_delete_firestore(document_id)
                # Remove the row from the table
                self.employee_table.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Employees record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
#################################################################3####################3##########################################3
    def load_employee_data(self):
        """Load existing employee records from a data source and display them in the table."""
        try:
            # Fetch records from the data source (replace with your actual data fetching code)
            employee_records = db.collection('Employees').get()
            # Clear existing rows in the table before loading new data
            self.employee_table.setRowCount(0)
            # Loop through each record and insert it into the table
            for record in employee_records:
                employee_data = record.to_dict()
                # Insert a new row in the table
                row_position = self.employee_table.rowCount()
                self.employee_table.insertRow(row_position)
                # Fill in the columns with data from the data source
                self.employee_table.setItem(row_position, 0, QTableWidgetItem(employee_data.get('employee_id', '')))
                self.employee_table.setItem(row_position, 1, QTableWidgetItem(employee_data.get('employee_name', '')))
                self.employee_table.setItem(row_position, 2, QTableWidgetItem(employee_data.get('date_of_joining', '')))
                self.employee_table.setItem(row_position, 3, QTableWidgetItem(employee_data.get('gender', '')))
                self.employee_table.setItem(row_position, 4, QTableWidgetItem(employee_data.get('contact_number', '')))
                self.employee_table.setItem(row_position, 5, QTableWidgetItem(employee_data.get('email_address', '')))
                self.employee_table.setItem(row_position, 6, QTableWidgetItem(employee_data.get('designation', '')))
                self.employee_table.setItem(row_position, 7, QTableWidgetItem(employee_data.get('document_id', '')))
            # Optionally, resize columns to fit contents
            self.employee_table.resizeColumnsToContents()
        except Exception as e:
            self.show_custom_message(f"Error loading employee records: {e}", "Error", is_success=False)
#######################################################3#############################################################
    def clear_employee(self):
        """Clear all the input fields and the table."""
        self.employee_id1.clear()
        self.employee_name1.clear()
        self.contact_number1.clear()
        self.email_address1.clear()
        self.designation1.clear()
        self.date_of_joining1.setDate(QDate.currentDate())
        self.gender_combo1.setCurrentIndex(0)
        self.employee_image_label1.clear()
        self.employee_image_label1.setText("No image uploaded")
###################################################################################################################
    def set_image_to_label(self, image_data):
        """Set the image data (base64) to the label widget."""
        if not image_data:
            self.employee_image_label1.setText("No image uploaded")
            return
        # Decode the base64 image data
        try:
            image_bytes = base64.b64decode(image_data)
            image=QImage.fromData(image_bytes)
            if image.isNull():
                self.show_custom_message("Invalid image data.", "Error", is_success=False)
                self.employee_image_label1.setText("No image uploaded")
            else:
                # Convert QImage to QPixmap and set it on the label
                pixmap = QPixmap.fromImage(image)
                self.employee_image_label1.setPixmap(pixmap)
                self.employee_image_label1.setScaledContents(True)  # Ensure image fits within label bounds
        except Exception as e:
            self.show_custom_message("Error loading image.", "Error", is_success=False)
            self.employee_image_label1.setText("No image uploaded")
##########################################################################################################################################
    def update_employee(self):
        """Handle updating employee details with validation and Firebase storage."""
        #Retrieve inputs
        employee_id =self.employee_id1.text()
        employee_name = self.employee_name1.text()
        contact_number = self.contact_number1.text()
        email_address = self.email_address1.text()
        designation = self.designation1.text()
        date_of_joining = self.date_of_joining1.date().toString("yyyy-MM-dd")
        gender = self.gender_combo1.currentText()

        # Validation regex
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{3,}$'
        contact_number_regex = r'^\d{11}$'  # Example for 11-digit contact number

        # # Check for empty fields
        # if not employee_id or not employee_name or not contact_number or not email_address or not designation:
        #     self.show_custom_message("All fields are required.", "Error", is_success=False)
        #     return

        # # Validate contact number
        # if not re.match(contact_number_regex, contact_number):
        #     self.show_custom_message("Contact Number must be an 11-digit number.", "Error", is_success=False)
        #     return

        # # Validate email
        # if not re.match(email_regex, email_address):
        #     self.show_custom_message("Invalid email format.", "Error", is_success=False)
        #     return

        # Validate image
        image_data=self.get_image_data_from_label()
        # if not image_data:
        #     self.show_custom_message("Please upload an employee image.", "Error", is_success=False)
        #     return

        # Prepare employee data
        employee_data = {
            'employee_name': employee_id,
            'employee_name': employee_name,
            'contact_number': contact_number,
            'email_address': email_address,
            'designation': designation,
            'date_of_joining': date_of_joining,
            'gender': gender,
            'employee_image': image_data,  # Include the base64 image data
        }

        # Check if employee record already exists (based on employee_id)
        existing_employee=db.collection('Employees').where('employee_id','==', employee_id).get()
        if existing_employee:
            # Update existing record
            try:
                for doc in existing_employee:
                    doc_ref = db.collection('Employees').document(doc.id)
                    doc_ref.update(employee_data)
                self.show_custom_message("Employee record updated successfully!", "Success", is_success=True)
                # Reload the table to reflect updated data
                self.clear_employee()
                self.load_employee_data()
                self.load_employee_records1()
                #self.employee_id.currentIndexChanged.connect(self.on_employee_selection_changed)
                self.populate_employee_combo_box()
            except Exception as e:
                self.show_custom_message(f"Error updating employee record: {e}", "Error", is_success=False)
        else:
            self.show_custom_message("Employee ID not found! Please check the ID.", "Error", is_success=False)
###########################################################################################################################################
    def search_employee(self):
        """Search for an employee by Employee ID and populate the fields."""
        employee_id = self.employee_id1.text()
        if not employee_id:
            self.show_custom_message("Employee ID is required for searching.", "Error", is_success=False)
            return
        # Search for the employee in Firestore
        try:
            results = db.collection('Employees').where('employee_id', '==', employee_id).get()
            if not results:
                self.show_custom_message("No records found for this Employee ID.", "Error", is_success=False)
                return
            # Assuming there is only one result for a given employee_id
            employee_data=results[0].to_dict()
            self.employee_name1.setText(employee_data.get('employee_name', ''))
            self.contact_number1.setText(employee_data.get('contact_number', ''))
            self.email_address1.setText(employee_data.get('email_address', ''))
            self.designation1.setText(employee_data.get('designation', ''))
            self.date_of_joining1.setDate(QDate.fromString(employee_data.get('date_of_joining', ''), "yyyy-MM-dd"))
            self.gender_combo1.setCurrentText(employee_data.get('gender', ''))
            # Load image if available
            if 'employee_image' in employee_data:
                self.set_image_to_label(employee_data['employee_image'])
            else:
                self.employee_image_label1.setText("No image uploaded")
        except Exception as e:
            self.show_custom_message("Error fetching employee details. Please try again.", "Error", is_success=False)
###########################################################################################################################################
    def get_image_data_from_label(self):
        """Convert the image from QLabel to base64-encoded string."""
        pixmap=self.employee_image_label1.pixmap()
        if pixmap:
            image = pixmap.toImage()
            # Create a QBuffer to hold the image data
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the image as PNG into the buffer
            image.save(buffer, 'PNG')
            # Encode the image data to base64
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            return image_data
        return None
#########################################################################################################################################
#########################################################################################################################################
    def save_employee(self):
        """Handle saving employee details with validation and Firebase image storage."""
        # Retrieve inputs
        employee_id = self.employee_id1.text()
        employee_name = self.employee_name1.text()
        contact_number = self.contact_number1.text()
        email_address = self.email_address1.text()
        designation = self.designation1.text()
        date_of_joining = self.date_of_joining1.date().toString("yyyy-MM-dd")
        gender = self.gender_combo1.currentText()

        # Validation regex
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{3,}$'
        contact_number_regex = r'^\d{11}$'  # Example for 10-digit contact number

        # Check for empty fields
        # if not employee_id or not employee_name or not contact_number or not email_address or not designation:
        #     self.show_custom_message("All fields are required.", "Error", is_success=False)
        #     return

        # # Validate contact number
        # if not re.match(contact_number_regex, contact_number):
        #     self.show_custom_message("Contact Number must be a 10-digit number.", "Error", is_success=False)
        #     return
        # # Validate email
        # if not re.match(email_regex, email_address):
        #     self.show_custom_message("Invalid email format.", "Error", is_success=False)
        #     return

        # Validate image
        image_data=self.get_image_data_from_label()
        # if not image_data:
        #     self.show_custom_message("Please upload an employee image.", "Error", is_success=False)
        #     return

        # Prepare data for Firebase
        employee_data = {
            'employee_id': employee_id,
            'employee_name': employee_name,
            'contact_number': contact_number,
            'email_address': email_address,
            'designation': designation,
            'date_of_joining': date_of_joining,
            'gender': gender,
            'employee_image': image_data,  # Include the base64 image data
        }
        #######################################################
        # Check if sale record already exists (based on registration number or other unique identifier)
        ###################
        # Check if the employee_id already exists in Firestore
        existing_employee_id=db.collection('Employees').where('employee_name', '==', employee_id).get()
        if existing_employee_id:
            self.show_custom_message("Employee ID already exists! Please use another employee ID.", "Error", is_success=False)
            return
        # Try saving the data to Firebase Firestore
        try:
            doc_ref = db.collection('Employees').add(employee_data)
            document_id=doc_ref[1].id
            db.collection('Employees').document(document_id).update({'document_id': document_id})
            self.show_custom_message("Employee Record is Added successfully!", "Success", is_success=True)
            # Add saved employee record to the table
            self.load_employee_data()
            self.load_employee_records1()
            #Call fetch_employee_details to refresh the employee details after saving
            ##employee_details=self.fetch_employee_details()
            ##self.employee_id=list(employee_details.keys())
            self.populate_employee_combo_box()
            # Clear the input fields after saving
            self.clear_employee()
        except Exception as e:
            self.show_custom_message("Please first connect to the internet before using the system", "Error", is_success=False)
###############################################################
    def fetch_employee_details(self):
        employee_ref = db.collection('Employees')
        docs = employee_ref.stream()
        employee_details = {}
        employee_ids = []
        for doc in docs:
            data = doc.to_dict()
            employee_id = data.get('employee_name')
            if employee_id:
                employee_ids.append(employee_id)
                employee_details[employee_id]={
                    'date_of_joining': data.get('date_of_joining', ''),
                    'gender': data.get('gender', ''),
                    'contact_number': data.get('contact_number', ''),
                    'designation': data.get('designation', ''),
                    'employee_image': data.get('employee_image', '')
                }
        return employee_details, employee_ids
##########################################################################################################################################
    def populate_employee_combo_box(self):
        self.employee_id.clear()
        self.employee_details,employee_ids=self.fetch_employee_details()
        # Employee ID and Employee Name (First row
        self.employee_id.addItems(["Select Employee Name"]+employee_ids)
#####################################################################################################################################3
    def on_employee_selection_changed(self, index):
        if index == 0:  # Check if the selected index is the "Select Employee ID" option
            # Clear the fields if the default option is selected
            self.date_of_joining.setText('')
            self.gender_combo.setText('')  # Use setCurrentText for combo boxes
            self.contact_number.setText('')
            self.designation.setText('')
            self.employee_image_label.clear()  # Clear the image if index 0 is selected
        else:
            # Get the selected employee ID from the combo box
            employee_id = self.employee_id.currentText()
            # Fetch details from self.employee_details dictionary
            details = self.employee_details.get(employee_id, {})
            # Update the UI fields with employee details
            self.date_of_joining.setText(details.get('date_of_joining', ''))
            self.gender_combo.setText(details.get('gender', ''))  # Use setCurrentText for combo boxes
            self.contact_number.setText(details.get('contact_number', ''))
            self.designation.setText(details.get('designation', ''))

            # Handle image display
            image_data = details.get('employee_image', '')
            if image_data:
                image_data = base64.b64decode(image_data)
                pixmap = QPixmap()
                pixmap.loadFromData(image_data, 'PNG')
                self.employee_image_label.setPixmap(pixmap.scaled(230, 230, Qt.KeepAspectRatio))
                self.employee_image_label.setFixedWidth(230)
                self.employee_image_label.setFixedHeight(150)
                self.employee_image_label.setAlignment(Qt.AlignCenter)
            else:
                self.employee_image_label.clear()  # Clear image if no image data is available
###################################3###################################################################################################    
    def upload_employee_image(self):
        """Handle customer image upload and display it in the label."""
        file_name, _ = QFileDialog.getOpenFileName(None, "Upload Image", "", "Image Files (*.png *.jpg *.jpeg)")
        if file_name:
            #Load the image from the selected file
            pixmap = QPixmap(file_name)
            # Resize the pixmap to fit within the label dimensions (optional, if the image is too large)
            pixmap = pixmap.scaled(self.employee_image_label1.size(), aspectRatioMode=1)  # Keep aspect ratio
            # Set the pixmap (image) onto the label
            self.employee_image_label1.setPixmap(pixmap)  
###########################################################################################################################################################################################################################################################
#######################################################################################################################################################################################
###########################################(Report Management)#########################################################################################################################################
###########################################(Report Management)#############################################################################################################################################
    def create_reports_management_view(self):
        """Create the inventory store management view with a modern look."""
        self.Complete_report_Views = QWidget()
        ##################################################
        self.Complete_report_Views.setStyleSheet("background-color:white;")
        #################################################
        self.report_layout = QVBoxLayout(self.Complete_report_Views)
        # Create QTabWidget with modern styling
        self.Report_tab_widget = QTabWidget()
        self.Report_tab_widget.setStyleSheet("""
            QTabWidget::pane { /* The tab widget frame */
                border-top: 1px solid #C0C0C0;
                background-color:white;
                border-radius:4px;
            }
            QTabBar::tab {
                background: #E0E0E0;
                color: #333;
                border: 1px solid #C0C0C0;
                padding: 10px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
                min-width:80px;
            }
            QTabBar::tab:hover {
                background: #B0BEC5;
            }
            QTabBar::tab:selected {
                background-color: #42A5F5;
                color: white;
                font-weight: bold;
            }
            QTabBar::tab:!selected {
                margin-top: 2px;
            }
        """)

        # Add Customer Details Tab
        self.customer_report = self.Customer_report_tab()
        self.Report_tab_widget.addTab(self.customer_report,"Purchaser Report")
        
        # Add Supplier Details Tab
        self.supplier_report = self.create_supplier_report_tab()
        self.Report_tab_widget.addTab(self.supplier_report,"Seller Report")
        
        # Add Employee Details Tab
        self.employee_report = self.create_employee_report_tab()
        self.Report_tab_widget.addTab(self.employee_report,"Employee Report")
         
        # Add Employee Payment Details Tab
        self.Employee_Payment_report = self.Employee_Payment_report_tab()
        self.Report_tab_widget.addTab(self.Employee_Payment_report,"Employee_Payment Report")
        # Add the QTabWidget to the layout
        self.report_layout.addWidget(self.Report_tab_widget)
        return self.Complete_report_Views
##############################################################################################################################################################################################
##############################################################################################################################################################################################
    def Customer_report_tab(self):
        tab_widget1 = QWidget()
        main_layout = QVBoxLayout(tab_widget1)
        main_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        #####################################################################
        # Title Label
        self.title_layout=QHBoxLayout()
        title_label = QLabel("Purchaser Report")
        title_label.setAlignment(Qt.AlignLeft)
        title_label.setStyleSheet("""
            font-size: 26px;
            font-weight: bold;
            color: #2E86C1;
            margin-bottom: 20px;
        """)
        ###############
        self.title_layout.addWidget(title_label)
        self.title_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.title_layout)
        ##########################################################################
        # Search Layout
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0,0,830,0)
        search_label = QLabel("Search Records:")
        search_label.setStyleSheet("""
            color:black;
            font-weight:bold;
            font-size:18px;
        """)

        # Search Field
        self.search_field1 = QLineEdit()
        self.search_field1.setPlaceholderText("Enter search term")
        self.search_field1.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                font-size:16px;
            }
        """)
        self.search_field1.setFixedHeight(50)
        self.search_field1.setFixedWidth(450)
        #################################################################3
        # Search Button
        search_button1 = QPushButton("Search Records")
        search_button1.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        search_button1.setFixedSize(150,50)
        search_button1.clicked.connect(self.rapid_search1)
        ######################################################################
        # Add widgets to the search layout
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_field1)
        search_layout.addWidget(search_button1)
        search_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(search_layout)
        main_layout.setSpacing(10)
        ###############################################################################################
        # Button Layout
        button_layout = QHBoxLayout()
        button_style = """
            background-color: #2980B9;
            color: white;
            border-radius: 5px;
            padding: 10px 20px;
            font-size: 14px;
            font-weight: bold;
        """
        
        delete_button = QPushButton("Delete")
        refresh_button = QPushButton("Refresh Records")
        pdf_button = QPushButton("Generate PDF Report")
        excel_button = QPushButton("Generate Excel Report")
        refresh_button.clicked.connect(self.refresh_1)
        delete_button.clicked.connect(self.delete_report)
        pdf_button.clicked.connect(self.Pdf_report1)
        excel_button.clicked.connect(self.excel_report1)
        for button in [delete_button, refresh_button, pdf_button, excel_button]:
            button.setStyleSheet(button_style)
            button.setCursor(Qt.PointingHandCursor)
            button.setMinimumWidth(150)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(refresh_button)
        button_layout.addWidget(pdf_button)
        button_layout.addWidget(excel_button)
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        
        main_layout.addLayout(button_layout)
        ###########################################################
        # Table for Customer Reports
        self.table_layout=QHBoxLayout()
        self.customer_report_table = QTableWidget()
        self.customer_report_table.setColumnCount(7)
        self.customer_report_table.setHorizontalHeaderLabels(["Purchaser ID", "Purchase Name", "CNIC", "Email", "Address", "Father Name", "Document ID"])
        self.customer_report_table.setFixedSize(1050, 500)
        self.customer_report_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
                padding: 0;
            }
            QHeaderView::section {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
            }
        """)

        # Set column widths
        self.customer_report_table.setColumnWidth(0, 50)   # ID
        self.customer_report_table.setColumnWidth(1, 180)  # Customer Name
        self.customer_report_table.setColumnWidth(2, 140)  # CNIC
        self.customer_report_table.setColumnWidth(3, 180)  # Email
        self.customer_report_table.setColumnWidth(4, 200)  # Address
        self.customer_report_table.setColumnWidth(5, 180)  # Father Name
        self.customer_report_table.setColumnWidth(6, 100)  # Document ID
        self.table_layout.addWidget(self.customer_report_table)
        self.table_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.table_layout)
        #############################################################################################################################################################
        
        #######################################
        self.load_customer_records1()
        ###################################################################
        return tab_widget1
##################################################################################################################################################################################################
#################################################################################################################################################################################
    def rapid_search1(self):
        search_term = self.search_field1.text().lower()
        # Check if the search term is empty
        if not search_term:
            self.show_custom_message("Please enter a search term.", "Info", is_success=False)
            return
        # Clear the table before loading search results
        self.customer_report_table.setRowCount(0)
        try:
            # Get all customers from Firestore
            customers = db.collection('Customers').get()
            found_records = False  # Flag to track if any records are found
            # Filter based on the search term
            for customer in customers:
                customer_data = customer.to_dict()
                # Convert relevant fields to lowercase for comparison
                customer_name = customer_data.get('customer_name', '').lower()
                cnic = customer_data.get('cnic', '').lower()
                email = customer_data.get('email', '').lower()
                # Check if the search term matches customer_name, CNIC, or email
                if search_term in customer_name or search_term in cnic or search_term in email:
                    found_records = True  # Set flag to True since a matching record is found
                    # Add matching records to the table
                    row_position = self.customer_report_table.rowCount()
                    self.customer_report_table.insertRow(row_position)
                    self.customer_report_table.setItem(row_position, 0, QTableWidgetItem(customer_data.get('customer_id', '')))
                    self.customer_report_table.setItem(row_position, 1, QTableWidgetItem(customer_data.get('customer_name', '')))
                    self.customer_report_table.setItem(row_position, 2, QTableWidgetItem(customer_data.get('cnic', '')))
                    self.customer_report_table.setItem(row_position, 3, QTableWidgetItem(customer_data.get('email', '')))
                    self.customer_report_table.setItem(row_position, 4, QTableWidgetItem(customer_data.get('address', '')))
                    self.customer_report_table.setItem(row_position, 5, QTableWidgetItem(customer_data.get('father_name', '')))
                    self.customer_report_table.setItem(row_position, 6, QTableWidgetItem(customer_data.get('document_id', '')))
            # If no records were found, show a message
            if not found_records:
                self.show_custom_message("No matching records found.", "Info", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error while searching: {e}", "Error", is_success=False)
    ###################################################################################################################################################
    #####################################################################################################################################################
    def excel_report1(self):
        try:
            # Create Excel Workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Customer Report"

            # Report title (centered across all columns)
            report_title = "Customer Report"
            sheet.merge_cells('A1:G1')  # Merge cells A1 to G1 for the title
            sheet["A1"] = report_title  # Set the title in the first cell
            sheet["A1"].font = openpyxl.styles.Font(size=14, bold=True)  # Apply styling to the title
            sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")  # Center alignment

            # Table header
            headers = ["ID", "Customer Name", "CNIC", "Email", "Address", "Father Name", "Document ID"]
            sheet.append(headers)

            # Table data
            for row in range(self.customer_report_table.rowCount()):
                row_data = []
                for col in range(self.customer_report_table.columnCount()):
                    item = self.customer_report_table.item(row, col).text()
                    row_data.append(item)
                sheet.append(row_data)

            # Open a Save File dialog to select the path and file name
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_dialog = QFileDialog()
            file_dialog.setAcceptMode(QFileDialog.AcceptSave)
            file_dialog.setDefaultSuffix("xlsx")  # Default file extension
            save_path, _ = file_dialog.getSaveFileName(self, "Save Excel Report", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

            # If the user cancels, save_path will be empty
            if save_path:
                # Save the Excel file
                workbook.save(save_path)
                self.show_custom_message(f"Excel report generated: {save_path}", "Success", is_success=True)
            else:
                self.show_custom_message("Excel report generation was canceled.", "Canceled", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error generating Excel report: {e}", "Error", is_success=False)
    ###########################################################################################################################################
    ###########################################################################################################################################
    def Pdf_report1(self):
        try:
            # Open a file save dialog to allow user to select save location
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            if not file_path:  # If no file is selected, cancel the operation
                return

            # Ensure the file extension is .pdf
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Create PDF file with landscape orientation
            pdf_canvas = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf_canvas.setTitle("Customer Report")

            # Get the page width for centering content
            page_width, _ = landscape(A4)

            # Set up header information centered
            pdf_canvas.setFont("Helvetica-Bold", 14)

            # Center the "Company Name or Logo"
            company_text = "CAR EXPERTS Sector-A,Bankers Town Ring Road,Near State Life Society Lahore"
            company_text_width = pdf_canvas.stringWidth(company_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - company_text_width) / 2, 570, company_text)

            # Center the report date
            report_date_text = f"Customer Report - {datetime.datetime.now().strftime('%Y-%m-%d')}"
            report_date_width = pdf_canvas.stringWidth(report_date_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - report_date_width) / 2, 550, report_date_text)

            # Adding report title centered with minimal margin at the top
            pdf_canvas.setFont("Helvetica-Bold", 16)
            report_title = "Customer Report"
            report_title_width = pdf_canvas.stringWidth(report_title, "Helvetica-Bold", 16)
            pdf_canvas.drawString((page_width - report_title_width) / 2, 510, report_title)

            #################################################################################
            # Setting up table headers and data, excluding "Document ID"
            headers = ["ID", "Customer Name", "CNIC", "Email", "Address", "Father Name"]
            data = [headers]

            # Fetch table data and add to data list (excluding the last column "Document ID")
            for row in range(self.customer_report_table.rowCount()):
                row_data = []
                for col in range(self.customer_report_table.columnCount() - 1):  # Exclude the last column
                    item = self.customer_report_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Define table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all cells to the center
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Padding for the header cells
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Alternate row color
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Table border/grid lines
            ])

            # Set column widths, adjusting for the removed column
            col_widths = [0.9 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 2.0 * inch, 1.5 * inch]

            # Create table and set style
            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)

            # Calculate table dimensions
            table_width, table_height = table.wrap(0, 0)

            # Calculate the position for the table to be centered horizontally and below the title
            page_width, page_height = landscape(A4)
            x_position = (page_width - table_width) / 2  # Center horizontally
            y_position = 470  # Position the table just below the title (adjust as needed)

            # Draw the table on the PDF
            table.drawOn(pdf_canvas, x_position, y_position - table_height)

            # Add total number of customers below the table
            pdf_canvas.setFont("Helvetica-Bold", 10)
            total_customers = self.customer_report_table.rowCount()
            pdf_canvas.drawString(x_position, y_position - table_height - 20, f"Total Customers: {total_customers}")

            # Add page number
            pdf_canvas.setFont("Helvetica", 8)
            pdf_canvas.drawString(750, 30, f"Page {pdf_canvas.getPageNumber()}")

            # Save the PDF file
            pdf_canvas.save()

            # Notify the user that the PDF was successfully created
            self.show_custom_message(f"PDF report generated: {file_path}", "Success", is_success=True)

        except Exception as e:
            # Handle any errors that occur during the PDF generation
            self.show_custom_message(f"Error generating PDF report: {e}", "Error", is_success=False)
    #################################################################################################################################################3
    #################################################################################################################################################3
    def delete_customer_report_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Customers').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
    ################################################################################################################################################
    ################################################################################################################################################
    def delete_report(self):
        """Delete the suppliers record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.customer_report_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Supplier?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.customer_report_table.item(row,6).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.delete_customer_report_firestore(document_id)
                # Remove the row from the table
                self.customer_report_table.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Customer record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    ################################################################################################################################################
    ################################################################################################################################################
    def refresh_1(self):
        self.load_customer_records1()
    ########################################################################################################################################################
    ################################################################################################################################################
    def load_customer_records1(self):
        """Load existing customer records from Firebase Firestore and display them in the table."""
        try:
            # Clear the table before loading new data
            self.customer_report_table.setRowCount(0)

            customers = db.collection('Customers').get()
            for customer in customers:
                customer_data = customer.to_dict()
                # Insert data into the table
                row_position = self.customer_report_table.rowCount()
                self.customer_report_table.insertRow(row_position)
                self.customer_report_table.setItem(row_position, 0, QTableWidgetItem(customer_data.get('customer_id', '')))
                self.customer_report_table.setItem(row_position, 1, QTableWidgetItem(customer_data.get('customer_name', '')))
                self.customer_report_table.setItem(row_position, 2, QTableWidgetItem(customer_data.get('cnic', '')))
                self.customer_report_table.setItem(row_position, 3, QTableWidgetItem(customer_data.get('email', '')))
                self.customer_report_table.setItem(row_position, 4, QTableWidgetItem(customer_data.get('address', '')))
                self.customer_report_table.setItem(row_position, 5, QTableWidgetItem(customer_data.get('father_name', '')))
                self.customer_report_table.setItem(row_position, 6, QTableWidgetItem(customer_data.get('document_id', '')))
        except Exception as e:
            self.show_custom_message(f"Error loading customers: {e}", "Error", is_success=False)
#################################################(Ending of Customer Report)#############################################################################################################################################
#################################################(Ending of Customer Report)##################################################################################################################
    def create_supplier_report_tab(self):
        tab_widget2 = QWidget()
        main_layout = QVBoxLayout(tab_widget2)
        main_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        # Centered Title at the top
        #####################################################
        self.title_layout=QHBoxLayout()
        title_label = QLabel("Seller Report")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 26px;
            font-weight: bold;
            color: #2E86C1;
            margin-bottom: 20px;  /* Adds spacing below the title */
        """)
        #########################
        ###############
        self.title_layout.addWidget(title_label)
        self.title_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.title_layout)
        #########################################################
        # Search bar layout from left to right
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0,0,830,0)
        search_label = QLabel("Search Records:")
        search_label.setStyleSheet("""
            color: black;
            font-weight: bold;
            font-size: 18px;
        """)

        self.search_field2 = QLineEdit()
        self.search_field2.setPlaceholderText("Enter search term")
        self.search_field2.setStyleSheet("""
            QLineEdit {
                background-color: #ECDFCC;
                color: black;
                font-weight: bold;
                border-radius: 10px;
                font-size: 16px;
            }
        """)
        self.search_field2.setFixedHeight(50)
        self.search_field2.setFixedWidth(450)
        # Search Button
        search_button2=QPushButton("Search Records")
        search_button2.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;  /* Modern teal color */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        search_button2.setFixedSize(150, 50)
        search_button2.clicked.connect(self.rapid_search2)
        # Adding widgets horizontally in one line
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_field2)
        search_layout.addWidget(search_button2)
        search_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        # Add search layout to main layout
        main_layout.addLayout(search_layout)
        #########################################
        main_layout.setSpacing(10)
        ##################################################################################################################################################################
        # Button layout
        button_layout = QHBoxLayout()
        button_style = """
            QPushButton {
                background-color: #2980B9;  /* Modern blue color */
                color: white;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2471A3;
            }
            QPushButton:pressed {
                background-color: #1F618D;
            }
        """

        delete_button = QPushButton("Delete Records")
        refresh_button = QPushButton("Refresh Records")
        pdf_button = QPushButton("Generate PDF Report")
        excel_button = QPushButton("Generate Excel Report")
        refresh_button.clicked.connect(self.refresh_2)
        delete_button.clicked.connect(self.delete_report1)
        pdf_button.clicked.connect(self.Pdf_report2)
        excel_button.clicked.connect(self.excel_report2)
        # Set button styles
        for button in [delete_button, refresh_button, pdf_button, excel_button]:
            button.setStyleSheet(button_style)
            button.setCursor(Qt.PointingHandCursor)
            button.setMinimumWidth(150)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(refresh_button)
        button_layout.addWidget(pdf_button)
        button_layout.addWidget(excel_button)
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
         # Add button layout to the main layout
        main_layout.addLayout(button_layout)
        #########################################################################################################################################################
        # Table configuration
        # Table for Customer Reports
        self.supplier_layout=QHBoxLayout()
        self.supplier_report_table = QTableWidget(0, 6)
        self.supplier_report_table.setHorizontalHeaderLabels(["Seller ID", "Seller Name", "Cnic", "Email", "Address", "Document ID"])
        self.supplier_report_table.setFixedSize(1050, 500)
        self.supplier_report_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
            }
            QHeaderView::section {
                background-color: #007BFF;  /* Professional blue color */
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;  /* Center align text */
            }
        """)

        # Set fixed width for columns
        column_widths = [100, 200, 150, 200, 250, 150]
        for i, width in enumerate(column_widths):
            self.supplier_report_table.setColumnWidth(i, width)

        self.supplier_layout.addWidget(self.supplier_report_table)
        self.supplier_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.supplier_layout)
        #########################################################
        self.load_supplier_records1()
       

        return tab_widget2
########################################################################################################################################################################
######################################################################################################################################################################################
    def refresh_2(self):
        self.load_supplier_records1()
    ###############################################################################################################################################
    #####################################################################################################################################################################
    def rapid_search2(self):
        search_term = self.search_field2.text().lower()
        # Check if the search term is empty
        if not search_term:
            self.show_custom_message("Please enter a search term.", "Info", is_success=False)
            return
        # Clear the table before loading search results
        self.supplier_report_table.setRowCount(0)
        try:
            # Get all suppliers from Firestore
            suppliers = db.collection('Suppliers').get()
            found_records = False  # Flag to track if any records are found
            # Filter based on the search term
            for supplier in suppliers:
                supplier_data = supplier.to_dict()
                # Convert relevant fields to lowercase for comparison
                supplier_name = supplier_data.get('supplier_name', '').lower()
                cnic = supplier_data.get('cnic', '').lower()
                email = supplier_data.get('email', '').lower()
                # Check if the search term matches supplier_name, CNIC, or email
                if search_term in supplier_name or search_term in cnic or search_term in email:
                    found_records = True  # Set flag to True since a matching record is found
                    # Add matching records to the table
                    row_position = self.supplier_report_table.rowCount()
                    self.supplier_report_table.insertRow(row_position)
                    self.supplier_report_table.setItem(row_position, 0, QTableWidgetItem(supplier_data.get('supplier_id', '')))
                    self.supplier_report_table.setItem(row_position, 1, QTableWidgetItem(supplier_data.get('supplier_name', '')))
                    self.supplier_report_table.setItem(row_position, 2, QTableWidgetItem(supplier_data.get('cnic', '')))
                    self.supplier_report_table.setItem(row_position, 3, QTableWidgetItem(supplier_data.get('email', '')))
                    self.supplier_report_table.setItem(row_position, 4, QTableWidgetItem(supplier_data.get('address', '')))
                    self.supplier_report_table.setItem(row_position, 5, QTableWidgetItem(supplier_data.get('document_id', '')))
            # If no records were found, show a message
            if not found_records:
                self.show_custom_message("No matching records found.", "Info", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error while searching: {e}", "Error", is_success=False)
    ##########################################################################################################################################################
    ################################################################################################################################################################
    def delete_report1(self):
        """Delete the suppliers record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.supplier_report_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Supplier?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.supplier_report_table.item(row,5).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.delete_supplier_report_firestore(document_id)
                # Remove the row from the table
                self.supplier_report_table.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Supplier record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    #######################################################3###################################################################################################################
    #########################################################################################################################################################
    def delete_supplier_report_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Suppliers').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
    #######################################################################################################################################################        
    #####################################################################################################################################################################
    def Pdf_report2(self):
        try:
            # Open a file save dialog to allow user to select save location
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            if not file_path:  # If no file is selected, cancel the operation
                return

            # Ensure the file extension is .pdf
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Create PDF file with landscape orientation
            pdf_canvas = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf_canvas.setTitle("Supplier Report")

            # Get the page width for centering content
            page_width, _ = landscape(A4)

            # Set up header information centered
            pdf_canvas.setFont("Helvetica-Bold", 14)

            # Center the "Company Name or Logo"
            company_text = "CAR EXPERTS Sector-A, Bankers Town Ring Road, Near State Life Society Lahore"
            company_text_width = pdf_canvas.stringWidth(company_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - company_text_width) / 2, 570, company_text)

            # Center the report date
            report_date_text = f"Supplier Report - {datetime.datetime.now().strftime('%Y-%m-%d')}"
            report_date_width = pdf_canvas.stringWidth(report_date_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - report_date_width) / 2, 550, report_date_text)

            # Adding report title centered with minimal margin at the top
            pdf_canvas.setFont("Helvetica-Bold", 16)
            report_title = "Supplier Report"
            report_title_width = pdf_canvas.stringWidth(report_title, "Helvetica-Bold", 16)
            pdf_canvas.drawString((page_width - report_title_width) / 2, 510, report_title)

            # Setting up table headers and data without 'Document ID'
            headers = ["ID", "Supplier Name", "CNIC", "Email", "Address"]
            data = [headers]

            # Fetch table data and add to data list, excluding the 'Document ID' column (last column)
            for row in range(self.supplier_report_table.rowCount()):
                row_data = []
                for col in range(self.supplier_report_table.columnCount() - 1):  # Exclude the last column (Document ID)
                    item = self.supplier_report_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Define table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all cells to the center
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Padding for the header cells
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Alternate row color
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Table border/grid lines
            ])

            # Set column widths to fit the landscape orientation (adjusted to 5 columns)
            col_widths = [0.9 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 2.0 * inch]

            # Create table and set style
            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)

            # Calculate table dimensions
            table_width, table_height = table.wrap(0, 0)

            # Calculate the position for the table to be centered horizontally and below the title
            page_width, page_height = landscape(A4)
            x_position = (page_width - table_width) / 2  # Center horizontally
            y_position = 470  # Position the table just below the title (adjust as needed)

            # Draw the table on the PDF
            table.drawOn(pdf_canvas, x_position, y_position - table_height)

            # Add total number of suppliers below the table
            pdf_canvas.setFont("Helvetica-Bold", 10)
            total_suppliers = self.supplier_report_table.rowCount()
            pdf_canvas.drawString(x_position, y_position - table_height - 20, f"Total Suppliers: {total_suppliers}")

            # Add page number
            pdf_canvas.setFont("Helvetica", 8)
            pdf_canvas.drawString(750, 30, f"Page {pdf_canvas.getPageNumber()}")

            # Save the PDF file
            pdf_canvas.save()

            # Notify the user that the PDF was successfully created
            self.show_custom_message(f"PDF report generated: {file_path}", "Success", is_success=True)

        except Exception as e:
            # Handle any errors that occur during the PDF generation
            self.show_custom_message(f"Error generating PDF report: {e}", "Error", is_success=False)
    ##############################################################################################################################################################
    #############################################################################################################################################################
    def excel_report2(self):
        try:
            # Create Excel Workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Supplier Report"

            # Report title (centered across all columns)
            report_title = "Supplier Report"
            sheet.merge_cells('A1:F1')  # Merge cells A1 to F1 for the title
            sheet["A1"] = report_title  # Set the title in the first cell
            sheet["A1"].font = openpyxl.styles.Font(size=14, bold=True)  # Apply styling to the title
            sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")  # Center alignment

            # Table header
            headers = ["Supplier ID", "Supplier Name", "CNIC", "Email", "Address", "Document ID"]
            sheet.append(headers)

            # Table data
            for row in range(self.supplier_report_table.rowCount()):
                row_data = []
                for col in range(self.supplier_report_table.columnCount()):
                    item = self.supplier_report_table.item(row, col).text()
                    row_data.append(item)
                sheet.append(row_data)

            # Open a Save File dialog to select the path and file name
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_dialog = QFileDialog()
            file_dialog.setAcceptMode(QFileDialog.AcceptSave)
            file_dialog.setDefaultSuffix("xlsx")  # Default file extension
            save_path, _ = file_dialog.getSaveFileName(self, "Save Excel Report", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

            # If the user cancels, save_path will be empty
            if save_path:
                # Save the Excel file
                workbook.save(save_path)
                self.show_custom_message(f"Excel report generated: {save_path}", "Success", is_success=True)
            else:
                self.show_custom_message("Excel report generation was canceled.", "Canceled", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error generating Excel report: {e}", "Error", is_success=False)
    #################################################################################################
    ############################################################################################
    def load_supplier_records1(self):
        """Load existing customer records from Firebase Firestore and display them in the table."""
        try:
            # Clear the table before loading new data
            self.supplier_report_table.setRowCount(0)

            customers = db.collection('Suppliers').get()
            for customer in customers:
                customer_data = customer.to_dict()
                # Insert data into the table
                row_position = self.supplier_report_table.rowCount()
                self.supplier_report_table.insertRow(row_position)
                self.supplier_report_table.setItem(row_position, 0, QTableWidgetItem(customer_data.get('supplier_id', '')))
                self.supplier_report_table.setItem(row_position, 1, QTableWidgetItem(customer_data.get('supplier_name', '')))
                self.supplier_report_table.setItem(row_position, 2, QTableWidgetItem(customer_data.get('cnic', '')))
                self.supplier_report_table.setItem(row_position, 3, QTableWidgetItem(customer_data.get('email', '')))
                self.supplier_report_table.setItem(row_position, 4, QTableWidgetItem(customer_data.get('address', '')))
                self.supplier_report_table.setItem(row_position, 5, QTableWidgetItem(customer_data.get('document_id', '')))
        except Exception as e:
            self.show_custom_message(f"Error loading customers: {e}", "Error", is_success=False)
##############################################################################################################################################################################################
##############################################################################################################################################################################################
    def create_employee_report_tab(self):
        tab_widget3 = QWidget()
        main_layout = QVBoxLayout(tab_widget3)
        main_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        ##########################################################################################################
        # Title Label
        self.title_layout=QHBoxLayout()
        title_label = QLabel("Employee Report")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 26px;
            font-weight: bold;
            color: #2E86C1;
            margin-bottom: 20px;  /* Consistent margin */
        """)
        ###############################
        self.title_layout.addWidget(title_label)
        self.title_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.title_layout)
        ##########################################################################
        # Search Layout
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0,0,830,0)
        search_label = QLabel("Search Records:")
        search_label.setStyleSheet("""
            color:black;
            font-weight:bold;
            font-size:18px;
        """)

        # Search Field
        self.search_field3 = QLineEdit()
        self.search_field3.setPlaceholderText("Enter search term")
        self.search_field3.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                font-size:16px;
            }
        """)
        self.search_field3.setFixedHeight(50)
        self.search_field3.setFixedWidth(450)
        ####################################################
        # Search Button
        search_button3=QPushButton("Search Records")
        search_button3.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        search_button3.setFixedSize(150, 50)
        search_button3.clicked.connect(self.rapid_search3)
        # Add widgets to the search layout
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_field3)
        search_layout.addWidget(search_button3)
        search_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(search_layout)
        main_layout.setSpacing(10)
        #########################################################################################################################################
        ################################################################################################################################################
        # Button Layout
        button_layout = QHBoxLayout()
        button_style = """
            background-color: #2980B9;
            color: white;
            border-radius: 5px;
            padding: 10px 20px;
            font-size: 14px;
            font-weight: bold;
        """
        delete_button = QPushButton("Delete")
        refresh_button = QPushButton("Refresh Records")
        pdf_button = QPushButton("Generate PDF Report")
        excel_button = QPushButton("Generate Excel Report")
        ###################################################################
        refresh_button.clicked.connect(self.refresh_3)
        delete_button.clicked.connect(self.delete_report2)
        pdf_button.clicked.connect(self.Pdf_report3)
        excel_button.clicked.connect(self.excel_report3)
        #########################################################
        for button in [delete_button, refresh_button, pdf_button, excel_button]:
            button.setStyleSheet(button_style)
            button.setCursor(Qt.PointingHandCursor)
            button.setMinimumWidth(150)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(refresh_button)
        button_layout.addWidget(pdf_button)
        button_layout.addWidget(excel_button)
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(button_layout)
        ##########################################################
        self.employee=QHBoxLayout()
        # Employee Report Table
        self.employee_report_table = QTableWidget(0,8)
        self.employee_report_table.setHorizontalHeaderLabels(
            ["Employee ID","Employee Name","Joining Date","Gender","Contact", "Email","Designation","Document ID"]
        )
        self.employee_report_table.setFixedSize(1050, 500)
        self.employee_report_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
                padding: 0;
            }
            QHeaderView::section {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
            }
        """)

        # Set column widths
        self.employee_report_table.setColumnWidth(0, 100)  # Employee ID
        self.employee_report_table.setColumnWidth(1, 180)  # Employee Name
        self.employee_report_table.setColumnWidth(2, 150)  # Contact
        self.employee_report_table.setColumnWidth(3, 100)  # Gender
        self.employee_report_table.setColumnWidth(4, 150)  # Net Salary
        self.employee_report_table.setColumnWidth(5, 150)  # Account NO
        self.employee_report_table.setColumnWidth(6, 150)  # Document ID
        self.employee_report_table.setColumnWidth(7, 150)  # Document ID
        #################################################
        self.employee.addWidget(self.employee_report_table)
        self.employee.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.employee)
        ##############################################################
        self.load_employee_records1()
        ##############################################
       
        return tab_widget3
#############################################################################################################################################
#############################################################################################################################################
    def refresh_3(self):
        self.load_employee_records1()
    #########################3###################################################################################################################
    #########################3###################################################################################################################
    def delete_employee_report_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Employees').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
    #########################3###################################################################################################################
    #########################3###################################################################################################################
    def delete_report2(self):
        """Delete the suppliers record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.employee_report_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Supplier?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.employee_report_table.item(row,5).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.delete_employee_report_firestore(document_id)
                # Remove the row from the table
                self.employee_report_table.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Employee record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    ################################################################################################################################################
    #########################3###################################################################################################################
    def Pdf_report3(self):
        try:
            # Open a file save dialog to allow user to select save location
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            if not file_path:  # If no file is selected, cancel the operation
                return

            # Ensure the file extension is .pdf
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Create PDF file with landscape orientation
            pdf_canvas = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf_canvas.setTitle("Employee Report")

            # Get the page width for centering content
            page_width, _ = landscape(A4)

            # Set up header information centered
            pdf_canvas.setFont("Helvetica-Bold", 14)

            # Center the "Company Name or Logo"
            company_text = "CAR EXPERTS Sector-A, Bankers Town Ring Road, Near State Life Society Lahore"
            company_text_width = pdf_canvas.stringWidth(company_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - company_text_width) / 2, 570, company_text)

            #Center the report date
            report_date_text = f"Employee Report - {datetime.datetime.now().strftime('%Y-%m-%d')}"
            report_date_width = pdf_canvas.stringWidth(report_date_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - report_date_width) / 2, 550, report_date_text)

            # Adding report title centered with minimal margin at the top
            pdf_canvas.setFont("Helvetica-Bold", 16)
            report_title = "Employee Report"
            report_title_width = pdf_canvas.stringWidth(report_title, "Helvetica-Bold", 16)
            pdf_canvas.drawString((page_width - report_title_width) / 2, 510, report_title)

            #################################################################################
            # Setting up table headers and data
            headers = ["ID", "Employee Name", "Date of Joining", "Gender", "Contact Number", "Email", "Designation"]
            data = [headers]

            # Fetch table data and add to data list, excluding the last column (Document ID)
            for row in range(self.employee_report_table.rowCount()):
                row_data = []
                for col in range(self.employee_report_table.columnCount() - 1):  # Exclude the last column
                    item = self.employee_report_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Define table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all cells to the center
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Padding for the header cells
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Alternate row color
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Table border/grid lines
            ])

            # Set column widths (adjust the widths as needed since the Document ID column is removed)
            col_widths = [0.7*inch, 1.5*inch, 1.4*inch, 1.0*inch, 1.5*inch, 1.8*inch, 1.4*inch]

            # Create table and set style
            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)

            # Calculate table dimensions
            table_width, table_height = table.wrap(0, 0)

            # Calculate the position for the table to be centered horizontally and below the title
            page_width, page_height = landscape(A4)
            x_position = (page_width - table_width) / 2  # Center horizontally
            y_position = 470  # Position the table just below the title (adjust as needed)

            # Draw the table on the PDF
            table.drawOn(pdf_canvas, x_position, y_position - table_height)

            # Add total number of employees below the table
            pdf_canvas.setFont("Helvetica-Bold", 10)
            total_employees = self.employee_report_table.rowCount()
            pdf_canvas.drawString(x_position, y_position - table_height - 20, f"Total Employees: {total_employees}")

            # Add page number
            pdf_canvas.setFont("Helvetica", 8)
            pdf_canvas.drawString(750, 30, f"Page {pdf_canvas.getPageNumber()}")

            # Save the PDF file
            pdf_canvas.save()

            # Notify the user that the PDF was successfully created
            self.show_custom_message(f"PDF report generated: {file_path}", "Success", is_success=True)

        except Exception as e:
            # Handle any errors that occur during the PDF generation
            self.show_custom_message(f"Error generating PDF report: {e}", "Error", is_success=False)
    ###########################################################################################################################################################
    #########################3###################################################################################################################
    def excel_report3(self):
        try:
            # Create Excel Workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Employee Report"

            # Report title (centered across all columns)
            report_title = "Employee Report"
            sheet.merge_cells('A1:H1')  # Merge cells A1 to H1 for the title
            sheet["A1"] = report_title  # Set the title in the first cell
            sheet["A1"].font = openpyxl.styles.Font(size=14, bold=True)  # Apply styling to the title
            sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")  # Center alignment

            # Table header
            headers = ["Employee ID", "Employee Name", "Date of Joining", "Gender", "Contact Number", "Email Address", "Designation", "Document ID"]
            sheet.append(headers)

            # Table data
            for row in range(self.employee_report_table.rowCount()):
                row_data = []
                for col in range(self.employee_report_table.columnCount()):
                    item = self.employee_report_table.item(row, col).text() if self.employee_report_table.item(row, col) else ""
                    row_data.append(item)
                sheet.append(row_data)

            # Open a Save File dialog to select the path and file name
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_dialog = QFileDialog()
            file_dialog.setAcceptMode(QFileDialog.AcceptSave)
            file_dialog.setDefaultSuffix("xlsx")  # Default file extension
            save_path, _ = file_dialog.getSaveFileName(self, "Save Excel Report", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

            # If the user cancels, save_path will be empty
            if save_path:
                # Save the Excel file
                workbook.save(save_path)
                self.show_custom_message(f"Excel report generated: {save_path}", "Success", is_success=True)
            else:
                self.show_custom_message("Excel report generation was canceled.", "Canceled", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error generating Excel report: {e}", "Error", is_success=False)
    ##################################################################################################################################################################
    #########################3###################################################################################################################
    def rapid_search3(self):
        search_term = self.search_field3.text().lower()  # Assuming you have a search field for employees
        # Check if the search term is empty
        if not search_term:
            self.show_custom_message("Please enter a search term.", "Info", is_success=False)
            return
        # Clear the table before loading search results
        self.employee_report_table.setRowCount(0)
        try:
            # Get all employees from Firestore
            employees = db.collection('Employees').get()
            found_records = False  # Flag to track if any records are found
            # Filter based on the search term
            for employee in employees:
                employee_data = employee.to_dict()
                # Convert relevant fields to lowercase for comparison
                employee_name = employee_data.get('employee_name', '').lower()
                employee_id = employee_data.get('employee_id', '').lower()
                email = employee_data.get('email_address', '').lower()
                # Check if the search term matches employee_name, employee_id, or email
                if search_term in employee_name or search_term in employee_id or search_term in email:
                    found_records = True  # Set flag to True since a matching record is found
                    # Add matching records to the table
                    row_position = self.employee_report_table.rowCount()
                    self.employee_report_table.insertRow(row_position)
                    self.employee_report_table.setItem(row_position, 0, QTableWidgetItem(employee_id))
                    self.employee_report_table.setItem(row_position, 1, QTableWidgetItem(employee_name))
                    self.employee_report_table.setItem(row_position, 2, QTableWidgetItem(employee_data.get('date_of_joining', '')))
                    self.employee_report_table.setItem(row_position, 3, QTableWidgetItem(employee_data.get('gender', '')))
                    self.employee_report_table.setItem(row_position, 4, QTableWidgetItem(employee_data.get('contact_number', '')))
                    self.employee_report_table.setItem(row_position, 5, QTableWidgetItem(email))
                    self.employee_report_table.setItem(row_position, 6, QTableWidgetItem(employee_data.get('designation', '')))
                    self.employee_report_table.setItem(row_position, 7, QTableWidgetItem(employee_data.get('document_id', '')))
            # If no records were found, show a message
            if not found_records:
                self.show_custom_message("No matching records found.", "Info", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error while searching: {e}", "Error", is_success=False)
    ##############################################################################################################################################################
    #########################3###################################################################################################################
    def load_employee_records1(self):
        """Load existing employee records from Firebase Firestore and display them in the table, returning the count of employees."""
        try:
            # Clear the table before loading new data
            if not hasattr(self, 'employee_report_table'):
                raise AttributeError("employee_report_table is not defined")
            self.employee_report_table.setRowCount(0)
            customers = db.collection('Employees').get()
            #total_employees = 0
            self.employee_count=0  # Initialize the count
            for customer in customers:
                customer_data = customer.to_dict()
                # Insert data into the table
                row_position = self.employee_report_table.rowCount()
                self.employee_report_table.insertRow(row_position)
                self.employee_report_table.setItem(row_position, 0, QTableWidgetItem(customer_data.get('employee_id', '')))
                self.employee_report_table.setItem(row_position, 1, QTableWidgetItem(customer_data.get('employee_name', '')))
                self.employee_report_table.setItem(row_position, 2, QTableWidgetItem(customer_data.get('date_of_joining', '')))
                self.employee_report_table.setItem(row_position, 3, QTableWidgetItem(customer_data.get('gender', '')))
                self.employee_report_table.setItem(row_position, 4, QTableWidgetItem(customer_data.get('contact_number', '')))
                self.employee_report_table.setItem(row_position, 5, QTableWidgetItem(customer_data.get('email_address', '')))
                self.employee_report_table.setItem(row_position, 6, QTableWidgetItem(customer_data.get('designation', '')))
                self.employee_report_table.setItem(row_position, 7, QTableWidgetItem(customer_data.get('document_id', '')))
                self.employee_count+=1  # Increment the count
            
            return self.employee_count  # Return the count of employees  
        except Exception as e:
            self.show_custom_message(f"Error loading Employee: {e}", "Error", is_success=False)
            return 0  # Return 0 in case of an error
#########################################################(End of Employee)#####################################################################################################################################
#########################################################(End of Employee)####################################################################################################################
    def Employee_Payment_report_tab(self):
        tab_widget6 = QWidget()
        main_layout = QVBoxLayout(tab_widget6)
        main_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        # Title Label for Employee Payment (centered)
        #########################################################################################################################
        self.title_layout=QHBoxLayout()
        title_label = QLabel("Employee Payment Report")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 26px;
            font-weight: bold;
            color: #2E86C1;
            margin-bottom: 20px;
        """)
        #############################################
        self.title_layout.addWidget(title_label)
        self.title_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.title_layout)
        #########################################################################################################################
        # Search Section for Employee Payment (aligned horizontally from left to right)
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0,0,830,0)
        search_label = QLabel("Search Records:")
        search_label.setStyleSheet("""
            color:black;
            font-weight:bold;
            font-size:18px;
        """)

        # Search Field
        self.search_field6 = QLineEdit()
        self.search_field6.setPlaceholderText("Enter search term")
        self.search_field6.setStyleSheet("""
            QLineEdit {
                background-color:#ECDFCC;
                color:black;
                font-weight:bold;
                border-radius:10px;
                font-size:16px;
            }
        """)
        self.search_field6.setFixedHeight(50)
        self.search_field6.setFixedWidth(450)
        # Search Button
        search_button = QPushButton("Search Records")
        search_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border:none;
                border-radius:5px;
                padding:5px;
                font-size:12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        search_button.setFixedSize(150, 50)
        search_button.clicked.connect(self.search_employee_salary)
        # Add widgets to the search layout
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_field6)
        search_layout.addWidget(search_button)
        search_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(search_layout)
        main_layout.setSpacing(10)
        #####################################################################################################################
        #########################################################################################################################
        # Button Layout with Consistent Design
        button_layout = QHBoxLayout()
        button_style = """
            background-color: #2980B9;
            color: white;
            border-radius: 5px;
            padding: 10px 20px;
            font-size: 14px;
            font-weight: bold;
        """
        delete_button=QPushButton("Delete")
        refresh_button=QPushButton("Refresh Records")
        pdf_button=QPushButton("Generate PDF Report")
        excel_button=QPushButton("Generate Excel Report")
        ##########################################################################
        delete_button.clicked.connect(self.delete_employee_salary)
        refresh_button.clicked.connect(self.refreshing_employee_salary)
        pdf_button.clicked.connect(self.pdf_employee_salary)
        excel_button.clicked.connect(self.excel_employee_salary)
        ####################################################
        for button in [delete_button, refresh_button, pdf_button, excel_button]:
            button.setStyleSheet(button_style)
            button.setCursor(Qt.PointingHandCursor)
            button.setMinimumWidth(150)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(refresh_button)
        button_layout.addWidget(pdf_button)
        button_layout.addWidget(excel_button)
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(button_layout)
        # Total Employee and Total Employee Salary Layout
        ##################################################################################################################################################
        # Employee Payment Table
        self.emplpyee_payment=QHBoxLayout()
        self.emplpyee_payment_report_table = QTableWidget()
        self.emplpyee_payment_report_table.setColumnCount(8)
        self.emplpyee_payment_report_table.setHorizontalHeaderLabels(
            ["Employee ID","Contact", "Gender", "Designation", "Net Salary", "Account NO", "Document ID"]
        )
        self.emplpyee_payment_report_table.setFixedSize(1050, 500)
        self.emplpyee_payment_report_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
                padding: 0;
            }
            QHeaderView::section {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
            }
        """)
        # Set column widths
        self.emplpyee_payment_report_table.setColumnWidth(0, 100)  # Employee ID
        self.emplpyee_payment_report_table.setColumnWidth(1, 150)  # Employee Name
        self.emplpyee_payment_report_table.setColumnWidth(2, 100)  # Contact
        self.emplpyee_payment_report_table.setColumnWidth(3, 150)   # Gender
        self.emplpyee_payment_report_table.setColumnWidth(4, 150)  # Net Salary
        self.emplpyee_payment_report_table.setColumnWidth(5, 150)  # Account NO
        self.emplpyee_payment_report_table.setColumnWidth(6, 150)  # Document ID
        self.emplpyee_payment_report_table.setColumnWidth(7, 100)  # Document ID
        self.emplpyee_payment_report_table.setColumnWidth(8, 150)  # Document ID
        #############################################################################
        self.emplpyee_payment.addWidget(self.emplpyee_payment_report_table)
        self.emplpyee_payment.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        main_layout.addLayout(self.emplpyee_payment)
        ##############################################################################################################################
        total_layout = QHBoxLayout()
        ######################################################################################
        # Total Employee Label and Field
        total_employee_label = QLabel("Total Employee:")
        total_employee_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        self.total_employee_field = QLineEdit()
        self.total_employee_field.setPlaceholderText("0")
        self.total_employee_field.setFixedWidth(300)
        self.total_employee_field.setStyleSheet("""
            QLineEdit {
                background-color:black;
                border:2px solid white;
                border-radius:5px;
                padding:5px;
                font-size:16px;
                color:white;
            }
        """)
        ##############################################
        self.total_employee_field.setAlignment(Qt.AlignCenter)
        ###############################################################################
        # Total Employee Salary Label and Field
        total_salary_label = QLabel("Total Employee Salary:")
        total_salary_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        self.total_salary_field = QLineEdit()
        self.total_salary_field.setPlaceholderText("0")
        self.total_salary_field.setFixedWidth(300)
        self.total_salary_field.setStyleSheet("""
            QLineEdit {
                background-color:black;
                border:2px solid white;
                border-radius:5px;
                padding:5px;
                font-size:16px;
                color:white;
            }
        """)
        ###################################################################################
        self.total_salary_field.setAlignment(Qt.AlignCenter)
        ############################################################################################
        # Add to the layout
        total_layout.addWidget(total_employee_label)
        total_layout.addWidget(self.total_employee_field)
        total_layout.addWidget(total_salary_label)
        total_layout.addWidget(self.total_salary_field)
        total_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        # Add the total layout to the main layout
        #############################################
        self.load_emplpyee_payment_report11()
        ###############################################################################
        main_layout.addLayout(total_layout)
        return tab_widget6
###########################################################################################################################################################################################################################################################
################################################################################################################################################################################################   
    def load_emplpyee_payment_report11(self):
        """Load existing customer records from Firebase Firestore and display them in the table."""
        try:
            # Clear the table before loading new data
            self.emplpyee_payment_report_table.setRowCount(0)
            # Initialize counters for total employee count and net salary
            total_employees = 0
            total_net_salary = 0.0
            # Fetch employee salary records from Firebase Firestore
            customers = db.collection('Employee_Salary').get()
            for customer in customers:
                customer_data = customer.to_dict()
                # Insert data into the table
                row_position = self.emplpyee_payment_report_table.rowCount()
                self.emplpyee_payment_report_table.insertRow(row_position)
                self.emplpyee_payment_report_table.setItem(row_position, 0, QTableWidgetItem(str(customer_data.get('employee_name', ''))))
                self.emplpyee_payment_report_table.setItem(row_position, 1, QTableWidgetItem(customer_data.get('contact_number', '')))
                self.emplpyee_payment_report_table.setItem(row_position, 2, QTableWidgetItem(customer_data.get('gender', '')))
                self.emplpyee_payment_report_table.setItem(row_position, 3, QTableWidgetItem(customer_data.get('designation', '')))
                self.emplpyee_payment_report_table.setItem(row_position, 4, QTableWidgetItem(str(customer_data.get('net_salary', ''))))
                self.emplpyee_payment_report_table.setItem(row_position, 5, QTableWidgetItem(customer_data.get('bank_account', '')))
                self.emplpyee_payment_report_table.setItem(row_position, 6, QTableWidgetItem(customer_data.get('document_id', '')))
                # Increment total employee count
                total_employees += 1
                # Sum up the total net salary (ensure net_salary is a valid number)
                try:
                    total_net_salary += float(customer_data.get('net_salary', 0))
                except ValueError:
                    pass  # Handle cases where net_salary is not a valid float
            # Update the total employee count and total salary fields
            self.total_employee_field.setText(str(total_employees))
            self.total_salary_field.setText(f"{total_net_salary:.2f}")
        except Exception as e:
            self.show_custom_message(f"Error loading customers: {e}", "Error", is_success=False)
    #######################################################################################################################################################
    #######################################################################################################################################################
    def delete_employee_salary(self):
        """Delete the employee salary record based on the selected row in the table."""
        # Get the selected items in the table
        selected_items = self.emplpyee_payment_report_table.selectedItems()
        
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Employee Salary record?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 8th column (index 7)
                document_id = self.emplpyee_payment_report_table.item(row, 7).text()
                # Call the delete_record_from_firestore function to delete the employee from Firestore
                self.delete_employee_salary_report_firestore(document_id)
                # Remove the row from the table
                self.emplpyee_payment_report_table.removeRow(row)
                
                # Recalculate the total employee count and net salary after deletion
                self.update_total_employee_and_salary_fields()
                
                # Show a success message to the user
                self.show_custom_message("Employee Salary record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting employee salary record: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    #########################################################################################################################################################
    #########################################################################################################################################################
    def delete_employee_salary_report_firestore(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Employee_Salary').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
    #########################################################################################################################################################
    #########################################################################################################################################################
    def update_total_employee_and_salary_fields(self):
        """Recalculate and update the total employee count and total salary fields."""
        total_employees = self.emplpyee_payment_report_table.rowCount()
        total_net_salary = 0.0
        
        # Sum up the net salary for all rows in the table
        for row in range(total_employees):
            try:
                net_salary = float(self.emplpyee_payment_report_table.item(row, 5).text())
                total_net_salary += net_salary
            except (ValueError, TypeError):
                pass  # Handle cases where net_salary is invalid or missing
        
        # Update the total employee count and total salary fields
        self.total_employee_field.setText(str(total_employees))
        self.total_salary_field.setText(f"{total_net_salary:.2f}")
    ##########################################################################################################################################################
    #########################################################################################################################################################
    def refreshing_employee_salary(self):
        self.load_emplpyee_payment_report11()
    ###############################################################################################################################################################
    #########################################################################################################################################################
    def search_employee_salary(self):
        search_term = self.search_field6.text().lower()
        # Check if the search term is empty
        if not search_term:
            self.show_custom_message("Please enter a search term.", "Info", is_success=False)
            return
        # Clear the table before loading search results
        self.emplpyee_payment_report_table.setRowCount(0)
        
        total_employees = 0  # Counter for total employees found
        total_net_salary = 0.0  # Counter for total net salary found
        
        try:
            # Get all employee salary records from Firestore
            employees = db.collection('Employee_Salary').get()
            found_records = False  # Flag to track if any records are found
            
            # Filter based on the search term
            for employee in employees:
                employee_data = employee.to_dict()
                # Convert relevant fields to lowercase for comparison
                employee_name = employee_data.get('employee_name', '').lower()
                contact_number = employee_data.get('contact_number', '').lower()
                bank_account = employee_data.get('bank_account', '').lower()
                
                # Check if the search term matches employee_name, contact number, or bank account
                if search_term in employee_name or search_term in contact_number or search_term in bank_account:
                    found_records = True  # Set flag to True since a matching record is found
                    # Add matching records to the table
                    row_position = self.emplpyee_payment_report_table.rowCount()
                    self.emplpyee_payment_report_table.insertRow(row_position)
                    self.emplpyee_payment_report_table.setItem(row_position, 0, QTableWidgetItem(str(employee_data.get('employee_id', ''))))
                    self.emplpyee_payment_report_table.setItem(row_position, 1, QTableWidgetItem(employee_data.get('employee_name', '')))
                    self.emplpyee_payment_report_table.setItem(row_position, 2, QTableWidgetItem(employee_data.get('contact_number', '')))
                    self.emplpyee_payment_report_table.setItem(row_position, 3, QTableWidgetItem(employee_data.get('gender', '')))
                    self.emplpyee_payment_report_table.setItem(row_position, 4, QTableWidgetItem(employee_data.get('designation', '')))
                    self.emplpyee_payment_report_table.setItem(row_position, 5, QTableWidgetItem(str(employee_data.get('net_salary', ''))))
                    self.emplpyee_payment_report_table.setItem(row_position, 6, QTableWidgetItem(employee_data.get('bank_account', '')))
                    self.emplpyee_payment_report_table.setItem(row_position, 7, QTableWidgetItem(employee_data.get('document_id', '')))
                    
                    # Increment total employee count and total net salary
                    total_employees += 1
                    try:
                        total_net_salary += float(employee_data.get('net_salary', 0))
                    except ValueError:
                        pass  # Handle cases where net_salary is not a valid float
            # If no records were found, clear the total fields and show a message
            if not found_records:
                self.show_custom_message("No matching records found.", "Info", is_success=False)
                self.total_employee_field.setText("")  # Clear the total employee field
                self.total_salary_field.setText("")  # Clear the total salary field
            else:
                # Update the total employee count and total salary fields
                self.total_employee_field.setText(str(total_employees))
                self.total_salary_field.setText(f"{total_net_salary:.2f}")
        except Exception as e:
            self.show_custom_message(f"Error while searching: {e}", "Error", is_success=False)
    #####################################################################################################################################################################################
    #####################################################################################################################################################################################
    def pdf_employee_salary(self):
        """Generate a PDF report for the employee salary data in the table."""
        try:
            # Open a file save dialog to allow user to select save location
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            if not file_path:  # If no file is selected, cancel the operation
                return

            # Ensure the file extension is .pdf
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Create PDF file with landscape orientation
            pdf_canvas = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf_canvas.setTitle("Employee Report")

            # Get the page width for centering content
            page_width, _ = landscape(A4)

            # Set up header information centered
            pdf_canvas.setFont("Helvetica-Bold", 14)

            # Center the "Company Name or Logo"
            company_text = "CAR EXPERTS Sector-A, Bankers Town Ring Road, Near State Life Society Lahore"
            company_text_width = pdf_canvas.stringWidth(company_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - company_text_width) / 2, 570, company_text)

            # Center the report date
            report_date_text = f"Employee Report Salary-{datetime.datetime.now().strftime('%Y-%m-%d')}"
            report_date_width = pdf_canvas.stringWidth(report_date_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - report_date_width) / 2, 550, report_date_text)

            # Adding report title centered with minimal margin at the top
            pdf_canvas.setFont("Helvetica-Bold", 16)
            report_title = "Employee Salary Report"
            report_title_width = pdf_canvas.stringWidth(report_title, "Helvetica-Bold", 16)
            pdf_canvas.drawString((page_width - report_title_width) / 2, 510, report_title)
            ############################################################################################################################################
            # Setting up table headers and data
            headers = ["Employee ID", "Employee Name", "Contact Number", "Gender", "Designation", "Net Salary", "Bank Account"]
            data = [headers]
            # Fetch table data and add to data list (excluding the Document ID column)
            for row in range(self.emplpyee_payment_report_table.rowCount()):
                row_data = []
                for col in range(self.emplpyee_payment_report_table.columnCount() - 1):  # Exclude the last column (Document ID)
                    item = self.emplpyee_payment_report_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Define table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all cells to the center
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Padding for the header cells
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Alternate row color
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Table border/grid lines
            ])

            # Set column widths (adjusted to fit the new number of columns)
            col_widths = [1.0*inch, 1.5*inch, 1.5*inch, 0.8*inch, 1.5*inch, 1.5*inch, 1.5*inch]
            # Create table and set style
            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)
            # Calculate table dimensions
            table_width, table_height = table.wrap(0, 0)
            # Calculate the position for the table to be centered horizontally and below the title
            x_position = (page_width - table_width) / 2  # Center horizontally
            y_position = 470  # Position the table just below the title (adjust as needed)
            # Draw the table on the PDF
            table.drawOn(pdf_canvas, x_position, y_position - table_height)
            # Add total number of employees below the table
            pdf_canvas.setFont("Helvetica-Bold", 10)
            total_employees = self.emplpyee_payment_report_table.rowCount()
            pdf_canvas.drawString(x_position, y_position - table_height - 20, f"Total Employees: {total_employees}")
            # Add total net salary in bold font
            total_net_salary = self.total_salary_field.text()  # Retrieve total salary
            pdf_canvas.drawString(x_position, y_position - table_height - 40, f"Total Net Salary: {total_net_salary}")
            # Add page number
            pdf_canvas.setFont("Helvetica", 8)
            pdf_canvas.drawString(750, 30, f"Page {pdf_canvas.getPageNumber()}")
            # Save the PDF file
            pdf_canvas.save()
            # Notify the user that the PDF was successfully created
            self.show_custom_message(f"PDF report generated: {file_path}", "Success", is_success=True)
        except Exception as e:
            # Handle any errors that occur during the PDF generation
            self.show_custom_message(f"Error generating PDF report: {e}", "Error", is_success=False)
    ########################################################################################################################################################################
    ########################################################################################################################################################################
    def excel_employee_salary(self):
        """Generate an Excel report for employee salaries."""
        try:
            # Create Excel Workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Employee Salary Report"

            # Report title (centered across all columns)
            report_title = "Employee Salary Report"
            sheet.merge_cells('A1:H1')  # Merge cells A1 to H1 for the title
            sheet["A1"] = report_title  # Set the title in the first cell
            sheet["A1"].font = openpyxl.styles.Font(size=14, bold=True)  # Apply styling to the title
            sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")  # Center alignment

            # Table header
            headers = ["Employee ID", "Employee Name", "Contact Number", "Gender", "Designation", "Net Salary", "Bank Account", "Document ID"]
            sheet.append(headers)

            # Table data
            for row in range(self.emplpyee_payment_report_table.rowCount()):
                row_data = []
                for col in range(self.emplpyee_payment_report_table.columnCount()):
                    item = self.emplpyee_payment_report_table.item(row, col).text() if self.emplpyee_payment_report_table.item(row, col) else ""
                    row_data.append(item)
                sheet.append(row_data)

            # Open a Save File dialog to select the path and file name
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_dialog = QFileDialog()
            file_dialog.setAcceptMode(QFileDialog.AcceptSave)
            file_dialog.setDefaultSuffix("xlsx")  # Default file extension
            save_path, _ = file_dialog.getSaveFileName(self, "Save Excel Report", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

            # If the user cancels, save_path will be empty
            if save_path:
                # Save the Excel file
                workbook.save(save_path)
                self.show_custom_message(f"Excel report generated: {save_path}", "Success", is_success=True)
            else:
                self.show_custom_message("Excel report generation was canceled.", "Canceled", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error generating Excel report: {e}", "Error", is_success=False)
####################################################(End of Report Code)#####################################################################################################
####################################################(End of Report Code)##################################################################################################################
####################################################(End of Report Code)#####################################################################################################
####################################################(End of Report Code)#####################################################################################################
    def create_logout_view(self):
        """Create the logout view with a modern look."""
        logout_widget = QWidget()
        ##################################################
        logout_widget.setStyleSheet("background-color:white;")
        #################################################
        layout = QVBoxLayout(logout_widget)
        # Create and style the logout button
        logout_button = QPushButton("Logout")
        logout_button.setStyleSheet("""
            QPushButton {
                background-color: #FF5252;
                color: white;
                padding: 10px;
                border-radius: 10px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF1744;
            }
            QPushButton:pressed {
                background-color: #D50000;
            }
        """)
        # Fix the width and height of the button (e.g., width = 200px, height = 50px)
        logout_button.setFixedSize(400,40)
        # Connect the button to the logout function
        logout_button.clicked.connect(self.handle_logout)
        # Add the logout button to the layout and align it at the top
        layout.addWidget(logout_button, alignment=Qt.AlignTop | Qt.AlignLeft)
        return logout_widget
    ##################################################################################################################
    # Inside Dashboard.py
    def handle_logout(self):
        from Login_Registration_Form import MainWindow,RegisterForm  # Delayed import to avoid circular reference
        self.hide()
        self.logout_window = MainWindow()
        self.logout_window.showMaximized()
        self.regis=RegisterForm()
        self.regis.clear_login_history()
        self.Firestore_delete_login_history()
        ######################
    ################################################################################################################################################
    def delayed_logout(self):
        # Set up a QTimer to trigger the delayed_logout function after 30 minutes (1800000 milliseconds)
        self.timer = QTimer(self)
        self.timer.setSingleShot(True)  # Run only once
        self.timer.timeout.connect(self.process_logout)  # Connect to the method that performs the logout operations
        self.timer.start(1800000)  # 30 minutes in milliseconds
    def process_logout(self):
        # Here, implement the actions needed to perform the logout
        from Login_Registration_Form import MainWindow, RegisterForm  # Delayed import to avoid circular reference
        self.hide()
        self.logout_window = MainWindow()
        self.logout_window.showMaximized()
        self.regis = RegisterForm()
        self.regis.clear_login_history()
        self.Firestore_delete_login_history()
##############################################################################################################################################################
#####################################################################################################################################################################################################################   
    def create_admin_user_management_view(self):
        #Create the main layout for the Admin User Management view
        main_layout = QVBoxLayout()
        # Create the QTabWidget
        tab_widget = QTabWidget()
        tab_widget.setStyleSheet("background-color:white;")
        # Apply modern styles to QTabWidget and tabs
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #A9A9A9;   /* Light gray border around tab pane */
                border-radius:5px;           /* Rounded corners for the tab pane */
                background-color:white;    /* Light blue background for the tab content area */
            }
            QTabBar::tab {
                background-color:white;    /* Light gray for inactive tabs */
                color: #333;                  /* Dark text color for inactive tabs */
                border: 1px solid #A9A9A9;    /* Light gray border */
                border-radius: 5px;           /* Rounded corners */
                padding: 10px 20px;           /* Padding inside the tabs for a modern look */
                font-size: 14px;              /* Slightly larger font size */
                margin-right: 2px;            /* Small spacing between tabs */
                width:200px;
            }
            QTabBar::tab:selected {
                background-color: #1E90FF;    /* Dodger blue for the active tab */
                color: white;                 /* White text for the active tab */
                border-bottom: 2px solid #FF9800; /* Orange underline for active tab */
            }
            QTabBar::tab:hover {
                background-color: #D3D3D3;    /* Slightly darker gray when hovered */
                color: #333;                  /* Keep the text dark */
            }
        """)
        # Title
        title = QLabel("Admin User Management")
        title.setFont(QFont("Algerian", 24))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("""
            QLabel {
                color: #1E90FF;  /* Dodger blue color for the title */
                margin-bottom: 20px;  /* Add some spacing below the title */
                font-size:22px;
                font-weight:bold;
            }
        """)
        main_layout.addWidget(title)
        # Tab 1: User Management Form
        user_management_tab = QWidget()
        form_layout = QFormLayout()
        # Input fields
        self.email = QLineEdit()
        self.old_password = QLineEdit()
        self.new_password = QLineEdit()
        self.confirm_password = QLineEdit()
        for field in [self.email, self.old_password, self.new_password, self.confirm_password]:
            field.setFixedSize(300, 40)
        self.email.setPlaceholderText("Enter User Email")
        self.old_password.setPlaceholderText("Enter Old Password")
        self.new_password.setPlaceholderText("Enter New Password")
        self.confirm_password.setPlaceholderText("Confirm New Password")
        self.new_password.setEchoMode(QLineEdit.Password)
        self.confirm_password.setEchoMode(QLineEdit.Password)
        self.email.setText(str(self.Login_email))
        for field in [self.email, self.old_password, self.new_password, self.confirm_password]:
            field.setStyleSheet("""
                QLineEdit {
                    border: 1px solid #A9A9A9;
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 16px;
                    background-color: #F0F8FF;
                    color: #333;
                }
                QLineEdit:focus {
                    border: 2px solid #1E90FF;
                }
            """)
        form_layout.addRow(QLabel("User Email:"), self.email)
        form_layout.addRow(QLabel("Old Password:"), self.old_password)
        form_layout.addRow(QLabel("New Password:"), self.new_password)
        form_layout.addRow(QLabel("Confirm New Password:"), self.confirm_password)
        # Buttons
        update_button = QPushButton("Update Record")
        search_button = QPushButton("Search Records")
        clear_button = QPushButton("Clear Records")
        ###############################################
        update_button.setFixedWidth(200)
        update_button.setFixedHeight(50)
        ###########################################
        search_button.setFixedWidth(200)
        search_button.setFixedHeight(50)
        ################################################
        clear_button.setFixedWidth(200)
        clear_button.setFixedHeight(50)
        #################################################
        update_button.clicked.connect(self.update_admin)
        search_button.clicked.connect(self.search_admin)
        clear_button.clicked.connect(self.clear_admin)
        for button in [update_button, search_button, clear_button]:
            button.setStyleSheet("""
                QPushButton {
                    background-color: #1E90FF;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 10px 20px;
                    font-size: 16px;
                    margin: 5px;
                }
                QPushButton:hover {
                    background-color: #4682B4;
                }
            """)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        button_layout.addWidget(update_button)
        button_layout.addWidget(search_button)
        button_layout.addWidget(clear_button)
        form_layout.addRow(button_layout)
        user_management_tab.setLayout(form_layout)
        #########################################################################################################################################
        #########################################################################################################################################
        #Tab3:Admin Users View
        admin_users_view_tab = QWidget()
        admin_users_view_layout = QVBoxLayout()
        # Title for the tab
        tab_title = QLabel("User Admin View Management")
        tab_title.setFont(QFont("Arial", 18))
        tab_title.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        tab_title.setStyleSheet("""
            QLabel {
                color: #1E90FF;
                margin-bottom: 15px;
                font-size:20px;
                font-weight:bold;
            }
        """) 
        ########################################################################################################
        # Table setup
        self.user_table=QTableWidget()
        self.user_table.setColumnCount(6)
        self.user_table.setHorizontalHeaderLabels([
            "Username", "Email_Address", "Password", "Confirm_Password", "Register_Date","Document_ID"
        ])
        # Set fixed column widths
        column_widths = [200,220,220,220,220,220]  # Example widths for each column
        for i, width in enumerate(column_widths):
            self.user_table.setColumnWidth(i, width)
        # Set a fixed row height
        row_height = 40  # Example row height
        for row in range(0,self.user_table.rowCount()):  # Assuming rows are already added
            self.user_table.setRowHeight(row, row_height)
        # Style the table
        self.user_table.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;  /* White background */
                border: 1px solid #A9A9A9;
                font-size: 14px;
                width:830px;
                height:400px;
            }
            QHeaderView::section {
                background-color: #1E90FF;  /* Dodger blue */
                color: white;
                font-weight: bold;
                border: 1px solid #A9A9A9;
                padding: 5px;
            }
        """)
        #############################################################################################################################
        # Add table and title to layout
        admin_users_view_layout.addWidget(tab_title)
        admin_users_view_layout.addWidget(self.user_table)
        ############################################################################
        admin_users_view_tab.setLayout(admin_users_view_layout)
        ##########################################################################################################################################
        #Tab2:Placeholder for other functionality
        other_tab = QWidget()
        other_tab_layout = QVBoxLayout()
        self.other_tab_label = QLabel("All Connected User")
        self.other_tab_label.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        self.other_tab_label.setStyleSheet("color:#1E90FF;font-size:20px;font-weight:bold;")
        other_tab_layout.addWidget(self.other_tab_label)
        # Add functionality to other_tab_label if needed (e.g., trigger action)
        ########################################################################################
        self.refresh_button=QPushButton("Reload_Connected User")
        self.refresh_button.clicked.connect(self.refresh_all_connected_users)
        self.refresh_button.setStyleSheet("""
                QPushButton {
                    background-color: #1E90FF;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 10px 20px;
                    font-size: 16px;
                    margin: 5px;
                }
                QPushButton:hover {
                    background-color: #4682B4;
                }
            """)
        self.refresh_button.setFixedWidth(250)
        self.refresh_button.setFixedHeight(50)
        ################################################################
        other_tab_layout.addWidget(self.refresh_button)
        ##########################################################################################
        self.Login_table=QTableWidget()
        self.Login_table.setColumnCount(4)
        self.Login_table.setHorizontalHeaderLabels([
            "Email_Address", "Password","Login Time","Document ID"
        ])
        # Set fixed column widths
        column1_widths= [250,250,250,250]  # Example widths for each column
        for i, width1 in enumerate(column1_widths):
            self.Login_table.setColumnWidth(i, width1)
        #Set a fixed row height
        row_height1=40  # Example row height
        for row1 in range(0,self.Login_table.rowCount()):  # Assuming rows are already added
            self.Login_table.setRowHeight(row1, row_height1)
        # Style the table
        self.Login_table.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;  /* White background */
                border: 1px solid #A9A9A9;
                font-size: 14px;
                width:830px;
                height:400px;
            }
            QHeaderView::section {
                background-color: #1E90FF;  /* Dodger blue */
                color: white;
                font-weight: bold;
                border: 1px solid #A9A9A9;
                padding: 5px;
            }
        """)
        other_tab_layout.addWidget(self.Login_table)
        ###################################################################################################################
        other_tab.setLayout(other_tab_layout)
        ###############################################################################################################################
        # Add tabs to QTabWidget
        self.load_user_details()
        self.my_all_connected_users()
        ###########################################################################################################################
        tab_widget.addTab(user_management_tab, "Admin_Users")
        tab_widget.addTab(admin_users_view_tab, "Admin_Users_View")
        tab_widget.addTab(other_tab, "Connected Users")
        # Add the QTabWidget to the main layout
        main_layout.addWidget(tab_widget)
        # Create a placeholder widget for the view
        placeholder_widget = QWidget()
        placeholder_widget.setLayout(main_layout)
        #####################################################################################
        return placeholder_widget
########################################################################################################################################################
################################################################################################################################################
    def refresh_all_connected_users(self):
        try:
            my_login_data = db.collection('Login').get()
            # Clear the table before loading new data
            self.Login_table.setRowCount(0)
            if not my_login_data:
                return
            # Loop through each record and insert it into the table
            for my_row_idx1, my_login in enumerate(my_login_data):
                Admin_login_data = my_login.to_dict()
                ###### Insert a new records row ##########
                self.Login_table.insertRow(my_row_idx1)
                # Populate each column with the respective data from Firestore
                email_item = QTableWidgetItem(Admin_login_data.get('email_id', ''))
                password_item = QTableWidgetItem('*' * len(Admin_login_data.get('password', '')))
                login_time_item = QTableWidgetItem(Admin_login_data.get('login_time', ''))
                document_id_item = QTableWidgetItem(Admin_login_data.get('document_id', ''))
                # Disable editing and interaction for all items
                email_item.setFlags(email_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                password_item.setFlags(password_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                login_time_item.setFlags(login_time_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                document_id_item.setFlags(document_id_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                # Add the items to the table
                self.Login_table.setItem(my_row_idx1, 0, email_item)
                self.Login_table.setItem(my_row_idx1, 1, password_item)
                self.Login_table.setItem(my_row_idx1, 2, login_time_item)
                self.Login_table.setItem(my_row_idx1, 3, document_id_item)
        except Exception as e:
            self.show_custom_message(f"An error occurred while loading login data: {e}", "Error", is_success=False)
    #########################################################################################################################
    def my_all_connected_users(self):
        try:
            my_login_data = db.collection('Login').get()
            # Clear the table before loading new data
            self.Login_table.setRowCount(0)
            if not my_login_data:
                return
            # Loop through each record and insert it into the table
            for my_row_idx1, my_login in enumerate(my_login_data):
                Admin_login_data = my_login.to_dict()
                ###### Insert a new records row ##########
                self.Login_table.insertRow(my_row_idx1)
                # Populate each column with the respective data from Firestore
                email_item = QTableWidgetItem(Admin_login_data.get('email_id', ''))
                password_item = QTableWidgetItem('*' * len(Admin_login_data.get('password', '')))
                login_time_item = QTableWidgetItem(Admin_login_data.get('login_time', ''))
                document_id_item = QTableWidgetItem(Admin_login_data.get('document_id', ''))
                # Disable editing and interaction for all items
                email_item.setFlags(email_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                password_item.setFlags(password_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                login_time_item.setFlags(login_time_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                document_id_item.setFlags(document_id_item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
                # Add the items to the table
                self.Login_table.setItem(my_row_idx1, 0, email_item)
                self.Login_table.setItem(my_row_idx1, 1, password_item)
                self.Login_table.setItem(my_row_idx1, 2, login_time_item)
                self.Login_table.setItem(my_row_idx1, 3, document_id_item)
        except Exception as e:
            self.show_custom_message(f"An error occurred while loading login data: {e}", "Error", is_success=False)
    ##########################################################################################################################
    def Firestore_delete_login_history(self):
        try:
            # Get the email from the instance variable
            login_email=self.Login_email  # Assuming Login_email is set elsewhere
            document_id = None
            # Query Firestore 'Login' collection to find document matching the email
            docs = db.collection('Login').where('email_id', '==', login_email).stream()
            #Debug:Check the number of documents found
            docs_list = list(docs)
            print(f"Documents found: {len(docs_list)}")
            if docs_list:
                # Loop through the documents and try to fetch the document_id
                for doc in docs_list:
                    document_id = doc.id  # This should give you the document ID directly
                    print(f"Found document with ID: {document_id}")
                    db.collection('Login').document(document_id).delete()
                    # Show success message
                    #self.show_custom_message("Login history record deleted successfully.", "Success", is_success=True)
                    break
            else:
                # Show error message if no matching document is found in Firestore
                #self.show_custom_message("No matching document found to delete.", "Error", is_success=False)
                return
        except Exception as e:
            #Handle errors while deleting data from Firestore
            #self.show_custom_message(f"Error deleting login history: {e}", "Error", is_success=False)
            print(e)
    ##################################################################################################################################
    def load_user_details(self):
        try:
            my_user_data = db.collection('Users').get()
            # Clear the table before loading new data
            self.user_table.setRowCount(0)
            if not my_user_data:
                # self.show_custom_message("No User Data found.", "Info", is_success=True)
                return
            # Loop through each record and insert it into the table
            for my_row_idx, my_users in enumerate(my_user_data):
                Admin_user_data = my_users.to_dict()
                # Insert a new row for each record
                self.user_table.insertRow(my_row_idx)
                # Populate each column with the respective data from Firestore
                full_name_item = QTableWidgetItem(Admin_user_data.get('full_name', ''))
                email_item = QTableWidgetItem(Admin_user_data.get('email_id', ''))
                password_item = QTableWidgetItem('*' * len(Admin_user_data.get('password', '')))
                confirm_password_item = QTableWidgetItem('*' * len(Admin_user_data.get('confirm_password', '')))
                registration_date_item = QTableWidgetItem(Admin_user_data.get('registration_date', ''))
                document_id_item = QTableWidgetItem(Admin_user_data.get('document_id', ''))
                # Disable editing for each cell
                for item in [full_name_item, email_item, password_item, confirm_password_item, registration_date_item, document_id_item]:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make item non-editable
                # Add items to the table
                self.user_table.setItem(my_row_idx, 0, full_name_item)
                self.user_table.setItem(my_row_idx, 1, email_item)
                self.user_table.setItem(my_row_idx, 2, password_item)
                self.user_table.setItem(my_row_idx, 3, confirm_password_item)
                self.user_table.setItem(my_row_idx, 4, registration_date_item)
                self.user_table.setItem(my_row_idx, 5, document_id_item)
        except Exception as e:
            self.show_custom_message("An error occurred while loading user data: {}".format(e), "Error", is_success=False)
    ########################################################################################################
    def update_admin(self):
        """Update the admin user record based on email address and change the password."""
        email_id = self.email.text()  # Assuming you use this field for the admin email.
        new_password = self.new_password.text()
        confirm_password = self.confirm_password.text()
        # Validation regex
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{3,}$'
        password_regex = r'^(?=.*[A-Za-z])(?=.*\d)[A-Za-z\d]{8,}$'
        # Check for empty fields
        if not email_id or not new_password or not confirm_password:
            self.show_custom_message("All fields are required.", "Error", is_success=False)
            return
        # Validate email
        if not re.match(email_regex, email_id):
            self.show_custom_message("Invalid Email ID! Please enter a valid email address.", "Error", is_success=False)
            return
        # Validate new password
        if not re.match(password_regex, new_password):
            self.show_custom_message("Password must be at least 8 characters long and contain both letters and numbers.", "Error", is_success=False)
            return
        # Check if new password and confirm password match
        if new_password != confirm_password:
            self.show_custom_message("Passwords do not match! Please ensure both password fields match.", "Error", is_success=False)
            return
        # Authenticate the admin and update the password
        try:
            # Check if the email exists in Firestore
            user_ref = db.collection('Users').where('email_id', '==', email_id).get()
            if not user_ref:
                self.show_custom_message("Email does not exist! Please check the email and try again.", "Error", is_success=False)
                return
            # Update the password for the found admin record
            for doc in user_ref:
                doc_ref = db.collection('Users').document(doc.id)
                doc_ref.update({'password': new_password})
            self.show_custom_message("Password updated successfully!", "Success", is_success=True)
            # Optional: Clear the input fields after updating
            self.clear_admin()
        except Exception as e:
            self.show_custom_message(f"An error occurred while updating the password: {e}", "Error", is_success=False)
    ########################################################################
    def search_admin(self):
        """Handle searching and displaying the admin's password based on the provided email."""
        # Retrieve the email from the search field
        admin_email_text=self.email.text()  # Assuming there is a field 'admin_email' to capture email input
        if not admin_email_text:
            self.show_custom_message("Please enter Email to search Admin Records.", "Error", is_success=False)
            return
        # Search for records in Firebase Firestore using the 'User' collection and the email
        try:
            records = db.collection('Users').where('email_id', '==', admin_email_text).limit(1).get()
        except Exception as e:
            self.show_custom_message(f"Error while fetching records: {str(e)}", "Error", is_success=False)
            return

        # If no records are found
        if len(records) == 0:
            self.show_custom_message("No records found for the given email.", "Information", is_success=False)
            return
        # Display the first record found
        first_record = records[0]
        data = first_record.to_dict()
        # Assuming you are fetching a password field from the database
        admin_password = data.get('password', '')
        # Fill in the admin password or display it as per your UI design
        self.old_password.setText(admin_password)  # Assuming there is a field 'admin_password' to display the password
        # Store the document ID for later use (if needed)
        self.stored_admin_id = first_record.id
        return self.stored_admin_id
    #######################################################################
    def clear_admin(self):
        self.old_password.clear()
        self.new_password.clear()
        self.confirm_password.clear()
###########################################################################################################################################################################
##########################################################################################################################################(it is Stock entry tap#############################################################
    def create_stock_tab(self):
        """Create the 'View Purchase Car Details' tab and add total employee fields."""
        # Create the view widget
        view_widget = QWidget()
        view_layout = QVBoxLayout(view_widget)
        view_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        ############################################################################################################################
        # Add the title of the tab
        label_layout=QHBoxLayout()
        view_title_label = QLabel("View Purchase Car Details")
        view_title_label.setStyleSheet("font-size: 24px; font-weight: bold; margin-bottom: 20px; color: #333;")
        label_layout.addWidget(view_title_label)
        label_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        view_layout.addLayout(label_layout)
        view_layout.setSpacing(5)
        #################################################################################### Add Search Records section #############
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0,0,750,0)
        Purchase_search_label = QLabel("Search Records:")
        Purchase_search_label.setStyleSheet("""
            QLabel {
                color: black;
                font-weight: bold;
                border-radius: 10px;
                font-size: 18px;
            }
            QLabel::Placeholder {
                color: white;
                font-weight: bold;
                text-align: center;
            }
        """)
        ############################################################
        self.Purchase_search_field = QLineEdit()
        self.Purchase_search_field.setStyleSheet("""
            QLineEdit {
                background-color: #ECDFCC;
                color: black;
                font-weight: bold;
                border-radius: 10px;
                font-size: 16px;
            }
            QLineEdit::Placeholder {
                color: white;
                font-weight: bold;
                text-align: center;
            }
        """)
        self.Purchase_search_field.setFixedHeight(50)
        self.Purchase_search_field.setFixedWidth(550)
        ################################################################
        Purchase_search_button = QPushButton("Search Records")
        Purchase_search_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        Purchase_search_button.setFixedWidth(150)
        Purchase_search_button.setFixedHeight(50)
        #############################################################################################
        Purchase_refresh_button = QPushButton("Refresh Table")
        Purchase_refresh_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        Purchase_refresh_button.setFixedWidth(150)
        Purchase_refresh_button.setFixedHeight(50)
        ##############################################################################################
        search_layout.addWidget(Purchase_search_label)
        search_layout.addWidget(self.Purchase_search_field)
        search_layout.addWidget(Purchase_search_button)
        search_layout.addWidget(Purchase_refresh_button)
        search_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        view_layout.addLayout(search_layout)
        view_layout.setSpacing(5)
        #############################################################################################################################
        #Connect search button to search function
        Purchase_search_button.clicked.connect(self.find_purchase_records)
        Purchase_refresh_button.clicked.connect(self.shams_refresh_table)
        #################################Add Table to display records#######################################################
        self.my_table_purchase_layout=QHBoxLayout()
        self.Purchase_table_widget12299=QTableWidget()  # Make table_widget an instance attribute
        self.Purchase_table_widget12299.setColumnCount(11)
        self.Purchase_table_widget12299.setHorizontalHeaderLabels([
            "Invoice No", "Supplier Name", "Father Name", "CNIC NO", "Registration No", 
            "Model", "Color", "Chassis No", "Purchase Date", "Purchase Price", "Document ID"
        ])
        self.Purchase_table_widget12299.setFixedWidth(1070)
        self.Purchase_table_widget12299.setFixedHeight(400)
        self.Purchase_table_widget12299.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
                padding: 0;
            }
            QHeaderView::section {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
            }
        """)
        self.Purchase_table_widget12299.setAlternatingRowColors(True)
        self.Purchase_table_widget12299.horizontalHeader().setStretchLastSection(True)
        self.Purchase_table_widget12299.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.Purchase_table_widget12299.setColumnWidth(0, 320)
        self.my_table_purchase_layout.addWidget(self.Purchase_table_widget12299)
        self.my_table_purchase_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        view_layout.addLayout(self.my_table_purchase_layout)
        view_layout.setSpacing(5)
        #################### Add Delete and Load buttons ####################
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0,0,550,0)
        Purchase_delete_button= QPushButton("Delete Records")
        Purchase_delete_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        Purchase_load_data = QPushButton("Load Data")
        Purchase_load_data.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:pressed {
                background-color: #bd2130;
            }
        """)
        Purchase_button_width = 250
        Purchase_button_height = 35
        Purchase_delete_button.setFixedSize(Purchase_button_width, Purchase_button_height)
        Purchase_load_data.setFixedSize(Purchase_button_width, Purchase_button_height)
        Purchase_delete_button.clicked.connect(self.Purchase_delete1)
        Purchase_load_data.clicked.connect(self.purchase_load1)
        ################################################################################################################################################
        self.total_purchase_label=QLabel("Total Purchase Price:")
        self.total_purchase_label.setStyleSheet("font-weight: bold; font-size: 16px;")
        #self.My_total_purchase_field22=QLineEdit()
        ######################################################################################################################
        self.My_total_purchase_field22.setStyleSheet("""
            QLineEdit {
                background-color:black;
                border:2px solid white;
                border-radius:5px;
                padding:5px;
                font-size:16px;
                color:white;
            }
        """)
        self.My_total_purchase_field22.setFixedSize(400,40)
        self.My_total_purchase_field22.setAlignment(Qt.AlignCenter)
        ######################################################################################################################
        button_layout.addWidget(Purchase_delete_button)
        button_layout.addWidget(Purchase_load_data)
        #####################################################################################################################################
        button_layout.addWidget(self.total_purchase_label)
        button_layout.addWidget(self.My_total_purchase_field22)
        #########################################################################################################################################
        button_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        view_layout.addLayout(button_layout)
        ########################################################################
        self.My_Purchase_load_data1()
        return view_widget
###################################################################################################################################################
########################################################################################################################################################
    def My_delete_purchase_from_firestore233(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('Purchases').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise
    ##############################################################################################
    def Purchase_delete1(self):
        """Delete the suppliers record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.Purchase_table_widget12299.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Supplier?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.Purchase_table_widget12299.item(row,10).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.My_delete_purchase_from_firestore233(document_id)
                # Remove the row from the table
                self.Purchase_table_widget12299.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("Purchase record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    #################################################################################################################################################
    def purchase_load1(self):
        self.My_Purchase_load_data1()
    def shams_refresh_table(self):
        self.My_Purchase_load_data1()
    ##############################################################################################################################################
    def My_Purchase_load_data1(self):
        """Load existing purchase records from Firebase Firestore and display them in the table."""
        try:
            # Clear the table before loading new data
            self.Purchase_table_widget12299.setRowCount(0)
            
            # Initialize total purchase price
            total_purchase_price = 0  # Use integer
            
            # Fetch purchase records from Firebase Firestore
            purchase_records = db.collection('Purchases').get()
            
            # Loop through each record and insert it into the table
            for record in purchase_records:
                purchase_data = record.to_dict()
                
                # Insert a new row in the table
                row_position = self.Purchase_table_widget12299.rowCount()
                self.Purchase_table_widget12299.insertRow(row_position)
                
                # Fill in the columns with data from Firestore
                self.Purchase_table_widget12299.setItem(row_position, 0, QTableWidgetItem(purchase_data.get('invoice_number', '')))
                self.Purchase_table_widget12299.setItem(row_position, 1, QTableWidgetItem(purchase_data.get('supplier_name', '')))
                self.Purchase_table_widget12299.setItem(row_position, 2, QTableWidgetItem(purchase_data.get('supplier_father_name', '')))
                self.Purchase_table_widget12299.setItem(row_position, 3, QTableWidgetItem(purchase_data.get('supplier_cnic', '')))
                self.Purchase_table_widget12299.setItem(row_position, 4, QTableWidgetItem(purchase_data.get('purchase_registration', '')))
                self.Purchase_table_widget12299.setItem(row_position, 5, QTableWidgetItem(purchase_data.get('purchase_model', '')))
                self.Purchase_table_widget12299.setItem(row_position, 6, QTableWidgetItem(purchase_data.get('purchase_color', '')))
                self.Purchase_table_widget12299.setItem(row_position, 7, QTableWidgetItem(purchase_data.get('purchase_chassis', '')))
                self.Purchase_table_widget12299.setItem(row_position, 8, QTableWidgetItem(purchase_data.get('purchase_date', '')))
                
                # Get purchase price and ensure it's an integer
                purchase_price = purchase_data.get('purchase_price', 0)
                try:
                    purchase_price = int(purchase_price)  # Ensure it's an integer
                except (ValueError, TypeError):
                    purchase_price = 0  # Default to 0 if conversion fails
                
                self.Purchase_table_widget12299.setItem(row_position, 9, QTableWidgetItem(str(purchase_price)))  # Convert to string
                
                # Accumulate total purchase price
                total_purchase_price += purchase_price
                
                self.Purchase_table_widget12299.setItem(row_position, 10, QTableWidgetItem(record.id))  # Firestore document ID
            
            # Update the total purchase price field
            self.My_total_purchase_field22.setText(f"{total_purchase_price}")  # Display total as an integer 
            # Optionally, resize columns to fit contents
            self.Purchase_table_widget12299.resizeColumnsToContents()

        except Exception as e:
            self.show_custom_message(f"Error loading purchase records: {e}", "Error", is_success=False)
    #################################################
    def find_purchase_records(self):
        """Search for purchase records based on the selected radio button and the entered text."""
        shams_purchase_data=self.Purchase_search_field.text().strip().lower()
        if not shams_purchase_data:
           self.show_custom_message("Please enter a value to search Records.", "Warning", is_success=False)
           return
        try:
            # Query Firestore for matching records
            purchase_all_records=db.collection('Purchases').get()
            # Filter records locally for partial matches
            matching_records = []
            for purchase_record in purchase_all_records:
                my_purchas_data=purchase_record.to_dict()
                if any(shams_purchase_data in str(my_purchas_data.get(field,'')).lower()for field in
                   ['invoice_number','purchase_registration','purchase_chassis']):
                   matching_records.append(purchase_record)
            ##################################################################################################
            # Clear the table before displaying search results
            self.Purchase_table_widget12299.setRowCount(0)
            # Initialize total purchase price
            total_purchase_price = 0  # Use integer
            # Check if records were found
            if not matching_records:
               self.show_custom_message("No records found for the given search criteria.", "Info", is_success=True)
               #Reset total sale field if no records found
               self.My_total_purchase_field22.setText('0')
            # Loop through the matching records and display them in the table
            for purchase_record in matching_records:
                purchase_data=purchase_record.to_dict()
                # Insert a new row in the table
                row_position = self.Purchase_table_widget12299.rowCount()
                self.Purchase_table_widget12299.insertRow(row_position)
                
                # Fill in the columns with data from Firestore
                self.Purchase_table_widget12299.setItem(row_position, 0, QTableWidgetItem(purchase_data.get('invoice_number', '')))
                self.Purchase_table_widget12299.setItem(row_position, 1, QTableWidgetItem(purchase_data.get('supplier_name', '')))
                self.Purchase_table_widget12299.setItem(row_position, 2, QTableWidgetItem(purchase_data.get('supplier_father_name', '')))
                self.Purchase_table_widget12299.setItem(row_position, 3, QTableWidgetItem(purchase_data.get('supplier_cnic', '')))
                self.Purchase_table_widget12299.setItem(row_position, 4, QTableWidgetItem(purchase_data.get('purchase_registration', '')))
                self.Purchase_table_widget12299.setItem(row_position, 5, QTableWidgetItem(purchase_data.get('purchase_model', '')))
                self.Purchase_table_widget12299.setItem(row_position, 6, QTableWidgetItem(purchase_data.get('purchase_color', '')))
                self.Purchase_table_widget12299.setItem(row_position, 7, QTableWidgetItem(purchase_data.get('purchase_chassis', '')))
                self.Purchase_table_widget12299.setItem(row_position, 8, QTableWidgetItem(purchase_data.get('purchase_date', '')))
                
                # Get purchase price and ensure it's an integer
                purchase_price = purchase_data.get('purchase_price', 0)
                try:
                    purchase_price = int(purchase_price)  # Ensure it's an integer
                except (ValueError, TypeError):
                    purchase_price = 0  # Default to 0 if conversion fails
                
                self.Purchase_table_widget12299.setItem(row_position, 9, QTableWidgetItem(str(purchase_price)))  # Convert to string
                
                # Accumulate total purchase price
                total_purchase_price += purchase_price
                
                self.Purchase_table_widget12299.setItem(row_position, 10, QTableWidgetItem(purchase_record.id))  # Firestore document ID

            # Resize columns to fit contents
            self.Purchase_table_widget12299.resizeColumnsToContents()
            
            # Display the total purchase price in the total_purchase_field
            self.My_total_purchase_field22.setText(f"{total_purchase_price}")  # Display total as an integer
            
        except Exception as e:
            self.show_custom_message(f"Error searching for records: {e}", "Error", is_success=False)
    ##########################################################################################################
    ###########################################################################3
    def calculate_total_purchases(self):
        """Count the total number of purchase records in Firestore."""
        total_purchases = 0  # Initialize counter for total purchases
        try:
            purchase_records = db.collection('Purchases').get()
            total_purchases = len(purchase_records)  # Count total records directly

        except Exception as e:
            self.show_custom_message(f"Error counting total purchases: {e}", "Error", is_success=False)
        return total_purchases
    ##################################################################
    def calculate_total_sale_records(self):
        """Count the total number of sale records in Firestore."""
        total_sale_records = 0  # Initialize counter for total sale records
        try:
            sale_records = db.collection('Car_Selling_Note').get()  # Fetch all sale records
            total_sale_records = len(sale_records)  # Count total records directly
        except Exception as e:
            self.show_custom_message(f"Error counting total sale records: {e}", "Error", is_success=False)
        return total_sale_records  # Return the total number of sale records
    ############################################################################
    def calculate_total_employee_records(self):
        """Count the total number of sale records in Firestore."""
        total_employee_records=0  # Initialize counter for total sale records
        try:
            employee_records=db.collection('Employees').get()  # Fetch all sale records
            total_employee_records=len(employee_records)  # Count total records directly
        except Exception as e:
            self.show_custom_message(f"Error counting total sale records: {e}", "Error", is_success=False)
        return total_employee_records #Return the total number of sale records
    ############################################################################
    def calculate_total_customer_records(self):
        """Count the total number of sale records in Firestore."""
        total_customer_records=0  # Initialize counter for total sale records
        try:
            customer_records=db.collection('Customers').get()  # Fetch all sale records
            total_customer_records=len(customer_records)  # Count total records directly
        except Exception as e:
            self.show_custom_message(f"Error counting total sale records: {e}", "Error", is_success=False)
        return total_customer_records #Return the total number of sale records
############################################(Report Management)###############################################################################################################################################################################################################
##################################################################################################################################################################################3
    def create_home_view(self):
        """Create the home view with the dashboard cards."""
        home_view=QWidget()
        home_layout=QVBoxLayout(home_view)
        # Set the base path
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(__file__))
        # Save the total purchase info before creating the home view
        total_purchase_price = self.calculate_total_purchases()
        total_sale_price = self.calculate_total_sale_records()
        total_employee = self.calculate_total_employee_records()
        total_customer = self.calculate_total_customer_records()
        # Dashboard cards setup
        card_layout = QHBoxLayout()
        card_layout.setContentsMargins(5,5,5,5)  # Set slight margins
        card_layout.setAlignment(Qt.AlignLeft)
        # Construct paths for card images
        cards = [
            ("No Of Sale Car", f"{total_sale_price}", os.path.join(base_path, "card_images/sold_2156144.png")),
            ("No Of Purchase Car", f"{total_purchase_price}", os.path.join(base_path, "card_images/car_554392.png")),
            ("No of Employees", f"{total_employee}", os.path.join(base_path, "card_images/borrow_3707947.png")),
            ("No of Customer", f"{total_customer}", os.path.join(base_path, "card_images/car_shop.png"))
        ]
        # Add dashboard cards to the layout
        for title, amount, image in cards:
            card = self.create_dashboard_card(title, amount, image)
            card_layout.addWidget(card)
        # Add card layout to the home layout
        home_layout.addLayout(card_layout)
        # Add a small spacing between the cards and the chart
        home_layout.addSpacing(0)
        # Create and add chart layout
        chart_layout = self.create_charts()
        home_layout.addLayout(chart_layout)
        return home_view
####################################################################################################################################################################################
########################################################################################################################################################################################
    def fetch_sales_and_purchases(self):
        """Fetch sales and purchase data based on date and aggregate totals."""
        sales_data = defaultdict(float)
        purchases_data = defaultdict(float)

        # Fetch sales data from Car_Selling_Note
        sales_ref = db.collection('Car_Selling_Note')
        sales_docs = sales_ref.stream()
        for doc in sales_docs:
            sale_price = float(doc.to_dict().get('sale_price', 0) or 0)  # Ensure sale_price is a float
            sale_date = doc.to_dict().get('date')  # Assuming date is in a format like 'YYYY-MM-DD'
            if sale_date:
                sales_data[sale_date] += sale_price

        # Fetch purchases data from Purchases
        purchases_ref = db.collection('Purchases')
        purchases_docs = purchases_ref.stream()
        for doc in purchases_docs:
            purchase_price = float(doc.to_dict().get('purchase_price', 0) or 0)  # Ensure purchase_price is a float
            purchase_date = doc.to_dict().get('purchase_date')  # Assuming date is in a format like 'YYYY-MM-DD'
            if purchase_date:
                purchases_data[purchase_date] += purchase_price
        return sales_data, purchases_data
    ##########################################################################################################3
    def create_dashboard_card(self, title, amount, image):
        """Create individual dashboard cards with fixed width and height."""
        card = QWidget()
        card.setStyleSheet("background-color: white; border-radius: 10px; padding: 20px;")
        
        # Set fixed size for the card
        card.setFixedSize(260,100)  # Adjust width and height as needed

        layout = QVBoxLayout(card)

        image_label = QLabel()
        pixmap = QPixmap(image).scaled(25,25, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        image_label.setPixmap(pixmap)
        layout.addWidget(image_label, alignment=Qt.AlignCenter)

        title_label = QLabel(title)
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setStyleSheet("color: #4B4D5A;")
        layout.addWidget(title_label, alignment=Qt.AlignCenter)

        amount_label = QLabel(amount)
        amount_label.setFont(QFont("Arial", 22, QFont.Bold))
        amount_label.setStyleSheet("color: #42a5f5;")
        layout.addWidget(amount_label, alignment=Qt.AlignCenter)

        layout.addStretch(1)
        return card
####################################################################################################3
    def create_charts(self):
        """Create the charts area with a bar chart for the dashboard."""
        chart_layout=QHBoxLayout()
        chart_layout.setContentsMargins(10,5,0,0)  # Reduced spacing on all sides
        ###############################################################################
        chart_layout.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        ##############################################################################
        # Fetch data from Firebase
        sales_data,purchases_data=self.fetch_sales_and_purchases()
        ##############################################
        # Prepare data for the chart
        categories = sorted(set(sales_data.keys()).union(purchases_data.keys()))
        sales_per_date = [sales_data[date] for date in categories]
        purchases_per_date = [purchases_data[date] for date in categories]
        #############################################################################
        # Bar chart data
        series = QBarSeries()
        set_sales = QBarSet("Sales")
        set_sales.append(sales_per_date)  # Add sales amounts for each date
        set_purchases = QBarSet("Purchases")
        set_purchases.append(purchases_per_date)  # Add purchase amounts for each date
        series.append(set_sales)
        series.append(set_purchases)
        ########################################################
        # Create chart and set axis
        chart = QChart()
        chart.addSeries(series)
        chart.setTitle("Sales vs Purchases by Date")
        axis_x = QBarCategoryAxis()
        axis_x.append(categories)  # Dates on X-axis
        chart.createDefaultAxes()
        chart.setAxisX(axis_x, series)
        ###############################################
        # Render chart
        chart_view = QChartView(chart)
        chart_view.setRenderHint(QPainter.Antialiasing)
        chart_view.setFixedSize(1000,600) 
        chart_layout.addWidget(chart_view)
        chart_layout.setAlignment(chart_view,Qt.AlignLeft|Qt.AlignTop)
        ####################################################################
        return chart_layout
#######################################################################################################################################################
########################################################################################################################################################################################
    def create_financial_management_view(self):
        # Create a main widget for the financial management view
        financial_management_widget = QWidget()
        layout = QVBoxLayout(financial_management_widget)
        layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)  # Align layout to the top-left
        layout.setSpacing(0)  # Remove vertical spacing between elements
        layout.setContentsMargins(0, 0, 0, 0)  # Remove margins
        financial_management_widget.setStyleSheet("background-color:white;")
        ####################################################################################
        #####################################################################
        # Title Label
        self.title_layout=QHBoxLayout()
        title_label = QLabel("Complete Financial Management System")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 26px;
            font-weight: bold;
            color: #2E86C1;
            margin-bottom: 20px;
        """)
        ######################
        ###############
        self.title_layout.addWidget(title_label)
        self.title_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        layout.addLayout(self.title_layout)
        ##########################################################################################################
        # Form Layout
        form_layout = QFormLayout()
        form_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        form_layout.setSpacing(10)
        form_layout.setContentsMargins(0,20,0,0)

        # Date fields
        date_label_start = QLabel("Start Date:")
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setFixedWidth(300)
        self.date_start.setFixedHeight(40)
        self.date_start.setStyleSheet("""
            QDateEdit {
                background-color: #ECDFCC;
                color: black;
                font-weight: bold;
                border-radius: 10px;
            }
        """)

        date_label_end = QLabel("End Date:")
        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setFixedHeight(40)
        self.date_end.setFixedWidth(300)
        self.date_end.setStyleSheet("""
            QDateEdit {
                background-color: #ECDFCC;
                color: black;
                font-weight: bold;
                border-radius: 10px;
            }
        """)

        form_layout.addRow(date_label_start, self.date_start)
        form_layout.addRow(date_label_end, self.date_end)

        # QLineEdit fields
        fields = [
            ("Total Sale Price:", "total_sale_price"),
            ("Purchase Price:", "purchase_price"),
            ("Total Employee Salary:", "total_employee_salary"),
            ("Final Price:", "final_price"),
            ("Business Status:", "business_status")
        ]

        for label_text, field_name in fields:
            label = QLabel(label_text)
            label.setStyleSheet("font-weight: bold;")
            label.setFixedHeight(30)

            line_edit = QLineEdit()
            line_edit.setStyleSheet("""
                QLineEdit {
                    background-color: #ECDFCC;
                    color: black;
                    font-weight: bold;
                    border-radius: 10px;
                }
            """)
            line_edit.setFixedHeight(40)
            line_edit.setFixedWidth(300)

            setattr(self, field_name, line_edit)

            form_layout.addRow(label, line_edit)
        layout.addLayout(form_layout)
        ##########################################################################################################
        # Button Layout
        button_layout = QHBoxLayout()
        button_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        button_layout.setSpacing(10)
        button_layout.setContentsMargins(0,0,0,0)

        clear_button = QPushButton("Get Data")
        save_button = QPushButton("Save")
        calculate_button = QPushButton("Calculate")

        clear_button.clicked.connect(self.get_data)
        calculate_button.clicked.connect(self.calculate_data)
        save_button.clicked.connect(self.save_financial_statement)

        # Style buttons
        button_style = "background-color: #4CAF50; color: white; font-size: 16px; border: none; border-radius: 5px; padding: 10px; height: 40px; width: 100px;"
        for button in [clear_button, save_button, calculate_button]:
            button.setStyleSheet(button_style)
            button.setFixedHeight(40)
            button.setFixedWidth(200)

        # Adding buttons to the button layout
        button_layout.addWidget(clear_button)
        button_layout.addWidget(save_button)
        button_layout.addWidget(calculate_button)
        layout.addLayout(button_layout)
        ##########################################################################################################
        # Update and Delete Buttons
        update_delete_layout = QHBoxLayout()
        update_delete_layout.setSpacing(10)  # Remove horizontal spacing
        update_delete_layout.setContentsMargins(0,0,0, 0)  # Remove margins

        delete_button = QPushButton("Delete")
        excel_report = QPushButton("Excel Report")
        pdf_report = QPushButton("PDF Report")

        for button in [delete_button, excel_report, pdf_report]:
            button.setStyleSheet("background-color: #F44336; color: white; font-size: 16px; border: none; border-radius: 5px; margin: 0px; padding: 10px;")
            button.setFixedHeight(40)
            button.setFixedWidth(200)

        delete_button.clicked.connect(self.delete_statement)
        excel_report.clicked.connect(self.excel_report)
        pdf_report.clicked.connect(self.pdf_report)

        update_delete_layout.addWidget(delete_button)
        update_delete_layout.addWidget(excel_report)
        update_delete_layout.addWidget(pdf_report)
        update_delete_layout.setAlignment(Qt.AlignLeft|Qt.AlignTop)
        layout.addLayout(update_delete_layout)
        layout.setSpacing(10)
        ##########################################################################################################
        # Table Layout
        table_layout = QVBoxLayout()
        table_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        table_layout.setSpacing(0)
        table_layout.setContentsMargins(0, 0, 0, 0)

        # Income Table
        self.income_table = QTableWidget()
        self.income_table.setColumnCount(8)
        self.income_table.setHorizontalHeaderLabels([
            "First Date", "End Date", "Sale Price", "Purchase Price",
            "Salary", "Final Price", "Status", "Document ID"
        ])

        # Modern table styling
        self.income_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px; 
                background-color: #FFFFFF; 
            }
            QHeaderView::section {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px; 
                text-align: center;
            }
            QTableWidget::item:selected {
                background-color: #B3E5FC; 
                color: #000000;
            }
        """)

        self.income_table.setFixedHeight(300)
        self.income_table.setFixedWidth(1050)
        self.income_table.setAlternatingRowColors(True)
        self.income_table.horizontalHeader().setStretchLastSection(True)
        self.income_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        table_layout.addWidget(self.income_table)

        ##########################################################################################################
        #Adding all layouts to the main layout in sequence
        layout.addLayout(table_layout)
        # Set the layout for the financial management widget
        financial_management_widget.setLayout(layout)

        self.load_financial_statements()
        return financial_management_widget
################################################################################################################################################################
############################################################################################################################################################
    def get_data(self):
        my_data=int(self.total_sale_field.text())
        my_data1=int(self.My_total_purchase_field22.text())
        my_data2=float(self.total_salary_field.text())
        ######################################################
        self.total_sale_price.setText(str(my_data))
        self.purchase_price.setText(str(my_data1))
        self.total_employee_salary .setText(str(my_data2))
    def calculate_data(self):
        my_record1 = int(self.total_sale_price.text())
        ########################################################
        my_record2 = int(self.purchase_price.text())
        my_record3 = float(self.total_employee_salary.text())
        ####################################
        sum_value = my_record2 + my_record3
        final_value = my_record1 - sum_value
        ##########################################
        self.final_price.setText(str(final_value))
        
        # (Set if my_record1 is less than sum, then show 'Loss' in business status; if not, show 'Profit')
        if my_record1 < sum_value:
            self.business_status.setText("Loss")
        else:
            self.business_status.setText("Profit")
    #############################################################
    def save_financial_statement(self):
        """Handle saving financial statement details with validation and Firebase storage."""
        # Collecting data from form fields
        start_date = self.date_start.date().toString("yyyy-MM-dd")
        end_date = self.date_end.date().toString("yyyy-MM-dd")
        total_sale_price = self.total_sale_price.text()
        purchase_price = self.purchase_price.text()
        total_employee_salary = self.total_employee_salary.text()
        final_price = self.final_price.text()
        business_status = self.business_status.text()

        # Check for empty fields
        if not start_date or not end_date or not total_sale_price or not purchase_price or not total_employee_salary or not final_price or not business_status:
            self.show_custom_message("All required fields must be filled.", "Error", is_success=False)
            return

        # Validate numeric fields
        try:
            total_sale_price = int(total_sale_price)  # Convert to integer
            purchase_price = int(purchase_price)      # Convert to integer
            total_employee_salary = float(total_employee_salary)  # Convert to float
            final_price = float(final_price)          # Convert to float
        except ValueError:
            self.show_custom_message("Please enter only numeric data.", "Error", is_success=False)
            return

        # Check for existing records based on start_date
        existing_start_date = db.collection('financial_statements').where('start_date', '==', start_date).get()
        if existing_start_date:
            self.show_custom_message("A financial statement already exists for the start date.", "Error", is_success=False)
            return

        # Check for existing records based on end_date
        existing_end_date = db.collection('financial_statements').where('end_date', '==', end_date).get()
        if existing_end_date:
            self.show_custom_message("A financial statement already exists for the end date.", "Error", is_success=False)
            return

        # Prepare the data as a dictionary to store in Firebase
        financial_data = {
            "start_date": start_date,
            "end_date": end_date,
            "total_sale_price": total_sale_price,
            "purchase_price": purchase_price,
            "total_employee_salary": total_employee_salary,
            "final_price": final_price,
            "business_status": business_status
        }

        # Try saving the data to Firebase Firestore
        try:
            # Add the financial data and retrieve a document reference
            doc_ref = db.collection('financial_statements').add(financial_data)
            document_id = doc_ref[1].id
            # Update the document with the document_id
            db.collection('financial_statements').document(document_id).update({'document_id': document_id})
            
            self.show_custom_message("Financial Statement saved successfully!", "Success", is_success=True)
            
            # Optionally, load the financial statements and clear the form fields after saving
            self.load_financial_statements()
            self.clear_form_fields()
            self.pdf_report()
            
        except Exception as e:
            self.show_custom_message("Please first connect to the internet before using the system.", "Error", is_success=False)

    ##################################################################################
    def load_financial_statements(self):
        """Load existing financial statements from Firebase Firestore and display them in the table."""
        try:
            # Retrieve financial data from Firestore
            financial_statements = db.collection('financial_statements').get()
            # Clear the table before loading new data
            self.income_table.setRowCount(0)
            # Check if there are any financial statements
            if not financial_statements:
                #self.show_custom_message("No financial statements found.", "Info", is_success=True)
                return
            # Loop through each record and insert it into the table
            for row_idx, statement in enumerate(financial_statements):
                statement_data = statement.to_dict()
                # Insert a new row
                self.income_table.insertRow(row_idx)
                # Populate each column with the respective data from Firestore
                self.income_table.setItem(row_idx, 0, QTableWidgetItem(statement_data.get('start_date', '')))
                self.income_table.setItem(row_idx, 1, QTableWidgetItem(statement_data.get('end_date', '')))
                self.income_table.setItem(row_idx, 2, QTableWidgetItem(str(statement_data.get('total_sale_price', 0))))
                self.income_table.setItem(row_idx, 3, QTableWidgetItem(str(statement_data.get('purchase_price', 0))))
                self.income_table.setItem(row_idx, 4, QTableWidgetItem(f"{statement_data.get('total_employee_salary', 0):.2f}"))  # Format as float
                self.income_table.setItem(row_idx, 5, QTableWidgetItem(f"{statement_data.get('final_price', 0):.2f}"))  # Format as float
                self.income_table.setItem(row_idx, 6, QTableWidgetItem(statement_data.get('business_status', '')))
                # Ensure the document ID is displayed correctly
                document_id = statement.id  # Get the document ID
                self.income_table.setItem(row_idx, 7, QTableWidgetItem(document_id))  # Set the document ID in the 8th column
            # Optionally, resize columns to fit contents
            self.income_table.resizeColumnsToContents()
        except Exception as e:
            self.show_custom_message("An error occurred while loading financial data: {}".format(e), "Error", is_success=False)
    ##########################################################################################################################################
    def excel_report(self):
        """Generate an Excel report from financial statements loaded from Firebase Firestore."""
        try:
            # Create Excel Workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Financial Statements Report"

            # Report title (centered across all columns)
            report_title = "Financial Statements Report"
            sheet.merge_cells('A1:H1')  # Adjust range as needed
            sheet["A1"] = report_title  # Set the title in the first cell
            sheet["A1"].font = openpyxl.styles.Font(size=14, bold=True)  # Apply styling to the title
            sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")  # Center alignment

            # Table header
            headers = ["Start Date", "End Date", "Total Sale Price", "Purchase Price", "Total Employee Salary", "Final Price", "Business Status", "Document ID"]
            sheet.append(headers)

            # Retrieve financial data from Firestore
            financial_statements = db.collection('financial_statements').get()

            # Prepare data for the Excel sheet
            for statement in financial_statements:
                statement_data = statement.to_dict()
                row = [
                    statement_data.get('start_date', ''),
                    statement_data.get('end_date', ''),
                    statement_data.get('total_sale_price', 0),
                    statement_data.get('purchase_price', 0),
                    f"{statement_data.get('total_employee_salary', 0):.2f}",
                    f"{statement_data.get('final_price', 0):.2f}",
                    statement_data.get('business_status', ''),
                    statement.id  # Get the document ID
                ]
                sheet.append(row)

            # Open a Save File dialog to select the path and file name
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_dialog = QFileDialog()
            file_dialog.setAcceptMode(QFileDialog.AcceptSave)
            file_dialog.setDefaultSuffix("xlsx")  # Default file extension
            save_path, _ = file_dialog.getSaveFileName(self, "Save Excel Report", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
            # If the user cancels, save_path will be empty
            if save_path:
                # Save the Excel file
                workbook.save(save_path)
                self.show_custom_message(f"Excel report generated: {save_path}", "Success", is_success=True)
            else:
                self.show_custom_message("Excel report generation was canceled.", "Canceled", is_success=False)
        except Exception as e:
            self.show_custom_message(f"Error generating Excel report: {e}", "Error", is_success=False)
    ###############################################################################################################################################
    def pdf_report(self):
        """Generate a PDF report of financial statements with charts in landscape, placing the table first and chart after."""
        try:
            # Open a file dialog for saving the PDF
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            # If no file path is selected, return
            if not file_path:
                self.show_custom_message("No file selected.", "Info", is_success=False)
                return

            # Ensure the file has a .pdf extension
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Retrieve financial data from Firestore
            financial_statements = db.collection('financial_statements').get()

            # Check if there are any financial statements
            if not financial_statements:
                self.show_custom_message("No financial statements found.", "Info", is_success=True)
                return

            # Create a PDF canvas in landscape orientation
            c = canvas.Canvas(file_path, pagesize=landscape(letter))
            width, height = landscape(letter)  # Get the width and height for landscape orientation

            # Title of the PDF
            c.setFont("Helvetica-Bold", 16)
            c.drawString(100, height - 50, "Financial Report - Landscape")

            # Prepare data for table and charts
            data = [["Start Date", "End Date", "Total Sale Price", "Purchase Price", "Total Employee Salary", "Final Price", "Business Status", "Document ID"]]
            total_sale_prices = []
            purchase_prices = []
            final_prices = []

            # Loop through each record
            for statement in financial_statements:
                statement_data = statement.to_dict()
                total_sale_price = float(statement_data.get('total_sale_price', 0))
                purchase_price = float(statement_data.get('purchase_price', 0))
                total_employee_salary = float(statement_data.get('total_employee_salary', 0))
                final_price = float(statement_data.get('final_price', 0))
                business_status = statement_data.get('business_status', '')
                document_id = statement.id

                # Append row data for table
                data.append([
                    statement_data.get('start_date', ''),
                    statement_data.get('end_date', ''),
                    f"{total_sale_price:.2f}",
                    f"{purchase_price:.2f}",
                    f"{total_employee_salary:.2f}",
                    f"{final_price:.2f}",
                    business_status,
                    document_id
                ])

                # Collect data for charts
                total_sale_prices.append(total_sale_price)
                purchase_prices.append(purchase_price)
                final_prices.append(final_price)

            # Draw the financial data table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            # Adjust table position to occupy the upper part of the PDF
            table.wrapOn(c, width, height)
            table_height = 300  # Adjust this value based on how tall the table should be
            table.drawOn(c, 30, height - table_height)

            # After the table, generate the bar chart
            fig, ax = plt.subplots()
            ax.bar(range(len(total_sale_prices)), total_sale_prices, color='green', label='Total Sale Price')
            ax.bar(range(len(purchase_prices)), purchase_prices, color='blue', label='Purchase Price', alpha=0.7)
            ax.set_title('Sales vs Purchases')
            ax.set_xlabel('Statements')
            ax.set_ylabel('Amount')
            ax.legend()

            # Save the chart to a byte buffer
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            chart_image = ImageReader(buf)

            # Draw chart image on the PDF, below the table
            c.drawImage(chart_image, 100, 50, width=500, height=250)  # Adjust size and position for landscape

            # Finalize PDF
            c.showPage()
            c.save()

            self.show_custom_message(f"PDF report saved to {file_path}", "Success", is_success=True)

        except Exception as e:
            self.show_custom_message(f"An error occurred while generating PDF: {e}", "Error", is_success=False)
        ##############################################################################################################################################################
    def clear_form_fields(self):
        self.total_sale_price.clear()
        self.purchase_price.clear()
        self.total_employee_salary.clear()
        self.final_price.clear()
        self.business_status.clear()
    #################################################################
    #############################################################################################################################33
    def delete_statement(self):
        """Delete the suppliers record based on the selected row in the table."""
        #Get the selected items in the table
        selected_items=self.income_table.selectedItems()
        if not selected_items:
            self.show_custom_message("Please select a row to delete.", "Error", is_success=False)
            return
        # Confirm the deletion action with the user
        reply = QMessageBox.question(self, "Confirm Deletion",
                                    "Are you sure you want to delete this Supplier?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # If the user confirms (Yes), proceed to delete
        if reply == QMessageBox.Yes:
            try:
                # Disable the delete button to prevent multiple clicks during operation
                self.delete_button.setEnabled(False)
                # Get the row index of the first selected item
                row = selected_items[0].row()
                # Retrieve the document_id from the 6th column (index 5)
                document_id=self.income_table.item(row,7).text()  # Changed index to 5
                # Call the delete_record_from_firestore function to delete the customer from Firestore
                self.delete_finacial_data(document_id)
                # Remove the row from the table
                self.income_table.removeRow(row)
                # Show a success message to the user
                self.show_custom_message("record deleted successfully!", "Success", is_success=True)
            except Exception as e:
                # In case of an error during deletion, show an error message
                self.show_custom_message(f"Error deleting customer: {e}", "Error", is_success=False)
            finally:
                # Re-enable the delete button after the operation is complete
                self.delete_button.setEnabled(True)
    ####################################################
    def delete_finacial_data(self, document_id):
        """Delete a record from Firestore based on the document_id."""
        try:
            # Reference to the document
            doc_ref=db.collection('financial_statements').document(document_id)
            # Delete the document
            doc_ref.delete()
            print(f'Document {document_id} deleted successfully.')
        except Exception as e:
            print(f'An error occurred while deleting the document: {e}')
            # Optionally, raise the exception to be handled by the caller
            raise

###############################################################################################################################################################################
#####################################################################################################################################################################
#####################################################################################################################################################################################
############################################(Report Management)###########################################################################################################################################
    def create_sale_purchase_stock_tab(self):
        # Create the main widget
        sales_management_widget1 = QWidget()
        sales_management_widget1.setStyleSheet("background-color:white;")
        # Create the main layout
        layout = QVBoxLayout(sales_management_widget1)
        layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)  # Align everything to the top center

        # Create the view widget
        view_widget = QWidget()
        view_layout = QVBoxLayout(view_widget)
        view_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)  # Align the view layout to the top center
        ################################################################################################
        # Add the title of the tab
        label_layout = QHBoxLayout()
        view_title_label = QLabel("Complete Sales and Purchases Summary")
        view_title_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #333;")
        label_layout.addWidget(view_title_label, alignment=Qt.AlignCenter)
        # Set margins for the title layout
        label_layout.setContentsMargins(0, 0, 10, 10)  # Set small margins for compactness
        view_layout.addLayout(label_layout)
        #####################################################################################################################
        # Add Search Records section
        search_grid_layout = QGridLayout()  # Create a grid layout for the search section
        search_grid_layout.setSpacing(5)  # Reduced spacing between elements
        search_grid_layout.setContentsMargins(0, 0, 10, 10)  # Set small margins for compactness

        # Search label
        search_label = QLabel("Search Records:")
        search_label.setStyleSheet("""
            QLabel {
                color: black;
                font-weight: bold;
                border-radius: 10px;
                font-size: 18px;
            }
        """)
        search_grid_layout.addWidget(search_label, 0, 0, alignment=Qt.AlignCenter)  # Place in grid

        # Search input field
        self.sale_purchase_search_field=QLineEdit()
        self.sale_purchase_search_field.setStyleSheet("""
            QLineEdit {
                background-color: #ECDFCC;
                color: black;
                font-weight: bold;
                border-radius: 10px;
                font-size: 16px;
            }
        """)
        self.sale_purchase_search_field.setFixedHeight(40)
        self.sale_purchase_search_field.setFixedWidth(500)
        search_grid_layout.addWidget(self.sale_purchase_search_field, 0, 1, alignment=Qt.AlignCenter)  # Place in grid
        ############################################################################################################################################
        #Search button
        sale_purchas_search_button = QPushButton("Search Records")
        sale_purchas_search_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        sale_purchas_search_button.setFixedWidth(150)
        sale_purchas_search_button.setFixedHeight(40)
        search_grid_layout.addWidget(sale_purchas_search_button, 0, 2, alignment=Qt.AlignCenter)  # Place in grid
        sale_purchas_search_button.clicked.connect(self.search_records_data)
        ###################################################
        #Search button
        sale_purchas_refresh_button=QPushButton("Refresh Table")
        sale_purchas_refresh_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        sale_purchas_refresh_button.setFixedWidth(150)
        sale_purchas_refresh_button.setFixedHeight(40)
        search_grid_layout.addWidget(sale_purchas_refresh_button,0,3, alignment=Qt.AlignCenter)  # Place in grid
        sale_purchas_refresh_button.clicked.connect(self.purchase_and_sales_refresher)
        #########################################################################################################################
        # report button
        Sale_and_Purchase_report=QPushButton("Create Report")
        Sale_and_Purchase_report.setStyleSheet("""
            QPushButton {
                background-color:#17a2b8;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        Sale_and_Purchase_report.setFixedWidth(150)
        Sale_and_Purchase_report.setFixedHeight(40)
        Sale_and_Purchase_report.clicked.connect(self.create_report_for_sale_and_purchase)
        search_grid_layout.addWidget(Sale_and_Purchase_report,1,2,alignment=Qt.AlignCenter)  # Place in grid
        ######################################################################################################################################
        view_layout.addLayout(search_grid_layout)  # Add grid layout to the view layout
        ###################################################################################################################################################################
        center_layout = QVBoxLayout()
        #######################################################################################################################################################
        #Purchase Table
        self.Purchase_table_widget1=QTableWidget()  # Make table_widget an instance attribute
        self.Purchase_table_widget1.setColumnCount(11)  # 11 columns including 'Document ID'
        self.Purchase_table_widget1.setHorizontalHeaderLabels([
            "Purchase_Invoice No", "Supplier Name", "Father Name", "CNIC NO", 
            "Registration No", "Model", "Color", "Chassis No", 
            "Purchase Date", "Purchase Price", "Document ID"
        ])
        self.Purchase_table_widget1.setFixedWidth(1050)
        self.Purchase_table_widget1.setFixedHeight(200)  # Set a fixed height for the table
        self.Purchase_table_widget1.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
            }
            QHeaderView::section {
                background-color: #007BFF;  /* Professional blue color */
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;  /* Center align text */
            }
        """)
        # Enable alternating row colors for better readability
        self.Purchase_table_widget1.setAlternatingRowColors(True)
        self.Purchase_table_widget1.horizontalHeader().setStretchLastSection(True)
        # Set fixed column widths
        column_widths =[170]*11  # Adjusted to 11 columns
        for i, width in enumerate(column_widths):
            self.Purchase_table_widget1.setColumnWidth(i, width)
        # Add the Purchase Table to the layout
        center_layout.addWidget(self.Purchase_table_widget1)
        #####################################################################################################################################################
        ######################################################################################################################################################
        #Using QVBoxLayout for vertical stacking
        # Sale Table
        self.table_widget1=QTableWidget()
        self.table_widget1.setColumnCount(11)
        self.table_widget1.setHorizontalHeaderLabels([
            "Sale_Invoice No", "Customer Name", "Father Name", "CNIC NO",
            "Registration No", "Model", "Color", "Sale Price", "Sale Date", "Make", "Document ID"
        ])
        self.table_widget1.setFixedWidth(1050)
        self.table_widget1.setFixedHeight(200)
        self.table_widget1.setStyleSheet("""
            QTableWidget {
                border: 1px solid #C0C0C0;
                border-radius: 5px;
                background-color: #FFFFFF;
            }
            QHeaderView::section {
                background-color: #007BFF;
                color: white;
                border: 1px solid #0056b3;
                padding: 5px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
            }
        """)
        self.table_widget1.setAlternatingRowColors(True)
        self.table_widget1.horizontalHeader().setStretchLastSection(True)
        #========================Set column widths===================================================================#
        # Set fixed column widths
        column_widths1=[170]*11  # Adjusted to 11 columns
        for j, width1 in enumerate(column_widths1):
            self.table_widget1.setColumnWidth(j,width1)
        #============================================================================================================#
        # Add the Sale Table to the layout
        center_layout.addWidget(self.table_widget1)
        #############################################################################################################################################################################################
        #####################################################################################################################################################
        # Set the alignment to center
        center_layout.setAlignment(Qt.AlignCenter)  # Center both tables horizontally
        center_layout.setAlignment(Qt.AlignTop)  # Center both tables vertically
        # Adjust the vertical stretch to push tables upwards
        center_layout.setStretch(0, 1)  # Allow the first table to take more space upwards
        center_layout.setStretch(1, 1)  # Allow the second table to also take space, but it will still stay under the first
        # Set contents margins to remove extra space
        center_layout.setContentsMargins(0, 0, 0, 0)  # Removed margins for a compact look
        # Add the center layout to your main layout
        view_layout.addLayout(center_layout)
        #############################################################################
        self.load_sale_Summary_records()
        self.Purchase_load_Summary_data()
        ###################################################################################################################################################################
        layout.addWidget(view_widget)
        sales_management_widget1.setLayout(layout)
        return sales_management_widget1
####################################################################################################################################################################################
############################################################################################################################################################
    def create_report_for_sale_and_purchase(self):
        """Generate a detailed purchase and sale report in PDF with tables, charts, and footer."""
        try:
            # Open a file dialog for saving the PDF
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf);;All Files (*)", options=options)

            if not file_path:
                self.show_custom_message("No file selected.", "Info", is_success=False)
                return
            if not file_path.endswith(".pdf"):
                file_path += ".pdf"

            # Retrieve data from Firestore
            purchase_records = db.collection('Purchases').get()
            sales_records = db.collection('Car_Selling_Note').get()

            if not purchase_records and not sales_records:
                self.show_custom_message("No purchase or sale data found.", "Info", is_success=True)
                return

            # Create a landscape A4 PDF canvas
            pdf_canvas = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf_canvas.setTitle("Sale and Purchase Report")

            # Get the page width for centering content
            page_width, page_height = landscape(A4)

            # Add company name
            pdf_canvas.setFont("Helvetica-Bold", 14)
            company_text = "CAR EXPERTS Sector-A, Bankers Town Ring Road, Near State Life Society Lahore"
            company_text_width = pdf_canvas.stringWidth(company_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - company_text_width) / 2, page_height - 80, company_text)

            # Add report title
            pdf_canvas.setFont("Helvetica-Bold", 16)
            report_title = "Complete Sale and Purchase Report Summary"
            report_title_width = pdf_canvas.stringWidth(report_title, "Helvetica-Bold", 16)
            pdf_canvas.drawString((page_width - report_title_width) / 2, page_height - 110, report_title)

            # Add current date
            report_date_text = f"Report Date - {datetime.datetime.now().strftime('%Y-%m-%d')}"
            report_date_width = pdf_canvas.stringWidth(report_date_text, "Helvetica-Bold", 14)
            pdf_canvas.drawString((page_width - report_date_width) / 2, 550, report_date_text)

            # Prepare purchase data for the table
            purchase_data = [["Purchase_Invoice No", "Supplier Name", "Father Name", "CNIC No", "Registration No", "Model", "Color", "Purchase Price", "Purchase Date"]]
            total_purchase_value = 0

            for record in purchase_records:
                pdata = record.to_dict()
                purchase_price = int(pdata.get('purchase_price', 0))
                total_purchase_value += purchase_price
                purchase_data.append([
                    pdata.get('invoice_number', ''),
                    pdata.get('supplier_name', ''),
                    pdata.get('supplier_father_name', ''),
                    pdata.get('supplier_cnic', ''),
                    pdata.get('purchase_registration', ''),
                    pdata.get('purchase_model', ''),
                    pdata.get('purchase_color', ''),
                    f"{purchase_price:.2f}",
                    pdata.get('purchase_date', '')
                ])

            # Add purchase table
            purchase_table = Table(purchase_data)
            purchase_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            # Adjust the table's position to left-align it and add vertical margin
            table_width, table_height = purchase_table.wrapOn(pdf_canvas, page_width, page_height - 150)
            purchase_table.drawOn(pdf_canvas, 30, page_height - table_height - 160)  # Changed x-coordinate to 30 for left alignment

            # Add total purchase value
            pdf_canvas.setFont("Helvetica", 12)
            pdf_canvas.drawString(30, page_height - table_height - 170, f"Total Purchase Records: {len(purchase_data) - 1}")
            pdf_canvas.drawString(30, page_height - table_height - 190, f"Total Purchase Value: {total_purchase_value:.2f}")

            # Prepare sale data for the table
            sales_data = [["Sale_Invoice No", "Customer Name", "Father Name", "CNIC No", "Registration No", "Model", "Color", "Sale Price", "Sale Date"]]
            total_sale_value = 0

            for record in sales_records:
                sdata = record.to_dict()
                sale_price = int(sdata.get('sale_price', 0))
                total_sale_value += sale_price
                sales_data.append([
                    sdata.get('invoice_number', ''),
                    sdata.get('customer_name', ''),
                    sdata.get('father_name', ''),
                    sdata.get('cnic', ''),
                    sdata.get('registration', ''),
                    sdata.get('model', ''),
                    sdata.get('color', ''),
                    f"{sale_price:.2f}",
                    sdata.get('date', '')
                ])

            # Add sale table
            sale_table = Table(sales_data)
            sale_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            # Adjust the sale table's position to left-align it and add vertical margin
            table_width, table_height = sale_table.wrapOn(pdf_canvas, page_width, page_height - table_height - 220)
            sale_table.drawOn(pdf_canvas, 30, page_height - table_height - 240)  # Changed x-coordinate to 30 for left alignment

            # Add total sale value
            pdf_canvas.setFont("Helvetica", 12)
            pdf_canvas.drawString(30, page_height - table_height - 250, f"Total Sale Records: {len(sales_data) - 1}")
            pdf_canvas.drawString(30, page_height - table_height - 270, f"Total Sale Value: {total_sale_value:.2f}")

            # Add page number
            pdf_canvas.setFont("Helvetica", 8)
            pdf_canvas.drawString(750, 30, f"Page {pdf_canvas.getPageNumber()}")

            # Save the PDF file
            pdf_canvas.save()
            # Notify the user that the PDF was successfully created
            self.show_custom_message(f"PDF report generated: {file_path}", "Success", is_success=True)
        except Exception as e:
            # Handle any errors that occur during the PDF generation
            self.show_custom_message(f"Error generating PDF report: {e}", "Error", is_success=False)
    ####################################################################################################################################################
    def purchase_and_sales_refresher(self):
        self.load_sale_Summary_records()
        self.Purchase_load_Summary_data()  
    #######################################################################
    def search_records_data(self):
        search_text=self.sale_purchase_search_field.text().strip().lower() 
        #Check if the search field is empty
        if not search_text:
            self.show_custom_message("Please enter a value to search Records.", "Warning", is_success=False)
            return
        try:
            #Define search parameters for sales and purchases
            sales_fields = ['invoice_number', 'registration', 'chassis']
            purchase_fields = ['invoice_number', 'purchase_registration', 'purchase_chassis']
            # Fetch broader records for partial matching
            sales_records = db.collection('Car_Selling_Note').get()
            purchase_records = db.collection('Purchases').get()
            #Filter sales records locally for partial matches
            matching_sales = []
            for record in sales_records:
                record_data = record.to_dict()
                if any(search_text in str(record_data.get(field, '')).lower() for field in sales_fields):
                    matching_sales.append(record)
            #Filter purchase records locally for partial matches
            matching_purchases = []
            for record in purchase_records:
                record_data = record.to_dict()
                if any(search_text in str(record_data.get(field, '')).lower() for field in purchase_fields):
                    matching_purchases.append(record)

            #Clear both tables before displaying search results
            self.table_widget1.setRowCount(0)  # Sales table
            self.Purchase_table_widget1.setRowCount(0)  # Purchases table

            # Populate sales table
            for record in matching_sales:
                record_data = record.to_dict()
                row_position = self.table_widget1.rowCount()
                self.table_widget1.insertRow(row_position)

                self.table_widget1.setItem(row_position, 0, QTableWidgetItem(record_data.get('invoice_number', '')))
                self.table_widget1.setItem(row_position, 1, QTableWidgetItem(record_data.get('customer_name', '')))
                self.table_widget1.setItem(row_position, 2, QTableWidgetItem(record_data.get('father_name', '')))
                self.table_widget1.setItem(row_position, 3, QTableWidgetItem(record_data.get('cnic', '')))
                self.table_widget1.setItem(row_position, 4, QTableWidgetItem(record_data.get('registration', '')))
                self.table_widget1.setItem(row_position, 5, QTableWidgetItem(record_data.get('model', '')))
                self.table_widget1.setItem(row_position, 6, QTableWidgetItem(record_data.get('color', '')))
                self.table_widget1.setItem(row_position, 7, QTableWidgetItem(str(record_data.get('sale_price', 0))))
                self.table_widget1.setItem(row_position, 8, QTableWidgetItem(record_data.get('date', '')))
                self.table_widget1.setItem(row_position, 9, QTableWidgetItem(record_data.get('make', '')))
                self.table_widget1.setItem(row_position, 10, QTableWidgetItem(record.id))
            self.table_widget1.resizeColumnsToContents()
            # Populate purchase table
            for record in matching_purchases:
                record_data = record.to_dict()
                row_position = self.Purchase_table_widget1.rowCount()
                self.Purchase_table_widget1.insertRow(row_position)
                self.Purchase_table_widget1.setItem(row_position, 0, QTableWidgetItem(record_data.get('invoice_number', '')))
                self.Purchase_table_widget1.setItem(row_position, 1, QTableWidgetItem(record_data.get('supplier_name', '')))
                self.Purchase_table_widget1.setItem(row_position, 2, QTableWidgetItem(record_data.get('supplier_father_name', '')))
                self.Purchase_table_widget1.setItem(row_position, 3, QTableWidgetItem(record_data.get('supplier_cnic', '')))
                self.Purchase_table_widget1.setItem(row_position, 4, QTableWidgetItem(record_data.get('purchase_registration', '')))
                self.Purchase_table_widget1.setItem(row_position, 5, QTableWidgetItem(record_data.get('purchase_model', '')))
                self.Purchase_table_widget1.setItem(row_position, 6, QTableWidgetItem(record_data.get('purchase_color', '')))
                self.Purchase_table_widget1.setItem(row_position, 7, QTableWidgetItem(record_data.get('purchase_chassis', '')))
                self.Purchase_table_widget1.setItem(row_position, 8, QTableWidgetItem(record_data.get('purchase_date', '')))
                self.Purchase_table_widget1.setItem(row_position, 9, QTableWidgetItem(str(record_data.get('purchase_price', 0))))
                self.Purchase_table_widget1.setItem(row_position, 10, QTableWidgetItem(record.id))
            self.Purchase_table_widget1.resizeColumnsToContents()
            # Show a message if no records were found in both tables
            if not matching_sales and not matching_purchases:
                self.show_custom_message("No records found for the given search criteria.", "Info", is_success=True)
                self.load_sale_Summary_records()
                self.Purchase_load_Summary_data()
        except Exception as e:
            self.show_custom_message(f"Error searching for records: {e}", "Error", is_success=False)
############################################################################################################################################################
    def load_sale_Summary_records(self):
        """Load existing selling car records from Firebase Firestore and display them in the table."""
        try:
            # Fetch records from Firestore
            sales_records = db.collection('Car_Selling_Note').get()
            # Clear existing rows in the table before loading new data
            self.table_widget1.setRowCount(0)

            total_sale_value = 0  # Initialize total sale value

            # Loop through each record and insert it into the table
            for record in sales_records:
                sale_data = record.to_dict()
                # Insert a new row in the table
                row_position = self.table_widget1.rowCount()
                self.table_widget1.insertRow(row_position)
                # Fill in the columns with data from Firestore
                self.table_widget1.setItem(row_position, 0, QTableWidgetItem(sale_data.get('invoice_number', '')))
                self.table_widget1.setItem(row_position, 1, QTableWidgetItem(sale_data.get('customer_name', '')))
                self.table_widget1.setItem(row_position, 2, QTableWidgetItem(sale_data.get('father_name', '')))
                self.table_widget1.setItem(row_position, 3, QTableWidgetItem(sale_data.get('cnic', '')))
                self.table_widget1.setItem(row_position, 4, QTableWidgetItem(sale_data.get('registration', '')))
                self.table_widget1.setItem(row_position, 5, QTableWidgetItem(sale_data.get('model', '')))
                self.table_widget1.setItem(row_position, 6, QTableWidgetItem(sale_data.get('color', '')))

                # Replace chassis with sale_price (integer)
                sale_price = sale_data.get('sale_price', 0)  # Default to 0 if sale_price is missing
                self.table_widget1.setItem(row_position, 7, QTableWidgetItem(str(sale_price)))  # Convert integer to string

                total_sale_value += sale_price  # Accumulate total sale value

                self.table_widget1.setItem(row_position, 8, QTableWidgetItem(sale_data.get('date', '')))
                self.table_widget1.setItem(row_position, 9, QTableWidgetItem(sale_data.get('make', '')))
                self.table_widget1.setItem(row_position, 10, QTableWidgetItem(record.id))  # Store document ID

            # Update the total sale field with the calculated total
            self.total_sale_field.setText(str(total_sale_value))  # Convert total to string for the text field

            # Optionally, resize columns to fit contents
            self.table_widget1.resizeColumnsToContents()
        except Exception as e:
            self.show_custom_message(f"Error loading sale records: {e}", "Error", is_success=False)
            #pass
############################################################################################################################
    def Purchase_load_Summary_data(self):
        """Load existing purchase records from Firebase Firestore and display them in the table."""
        try:
            # Fetch purchase records from Firebase Firestore
            purchase_records = db.collection('Purchases').get()
            
            # Clear existing rows in the table before loading new data
            self.Purchase_table_widget1.setRowCount(0)
            
            # Loop through each record and insert it into the table
            for record in purchase_records:
                purchase_data = record.to_dict()
                
                # Extract relevant fields for the table
                invoice_number = purchase_data.get('invoice_number', '')
                supplier_name = purchase_data.get('supplier_name', '')
                supplier_father_name = purchase_data.get('supplier_father_name', '')
                supplier_cnic = purchase_data.get('supplier_cnic', '')
                purchase_registration = purchase_data.get('purchase_registration', '')
                purchase_model = purchase_data.get('purchase_model', '')
                purchase_color = purchase_data.get('purchase_color', '')
                purchase_chassis = purchase_data.get('purchase_chassis', '')
                purchase_date = purchase_data.get('purchase_date', '')
                
                # Directly retrieve and format purchase_price as an integer string
                purchase_price = purchase_data.get('purchase_price', 0)
                formatted_purchase_price = str(int(purchase_price))  # Ensure it's an integer and then convert to string
                
                document_id = record.id  # Use Firestore's document ID for "Document ID"
                
                # Insert a new row in the table
                row_position = self.Purchase_table_widget1.rowCount()
                self.Purchase_table_widget1.insertRow(row_position)
                
                # Fill in the columns with data from the data source
                self.Purchase_table_widget1.setItem(row_position, 0, QTableWidgetItem(invoice_number))
                self.Purchase_table_widget1.setItem(row_position, 1, QTableWidgetItem(supplier_name))
                self.Purchase_table_widget1.setItem(row_position, 2, QTableWidgetItem(supplier_father_name))
                self.Purchase_table_widget1.setItem(row_position, 3, QTableWidgetItem(supplier_cnic))
                self.Purchase_table_widget1.setItem(row_position, 4, QTableWidgetItem(purchase_registration))
                self.Purchase_table_widget1.setItem(row_position, 5, QTableWidgetItem(purchase_model))
                self.Purchase_table_widget1.setItem(row_position, 6, QTableWidgetItem(purchase_color))
                self.Purchase_table_widget1.setItem(row_position, 7, QTableWidgetItem(purchase_chassis))
                self.Purchase_table_widget1.setItem(row_position, 8, QTableWidgetItem(purchase_date))
                self.Purchase_table_widget1.setItem(row_position, 9, QTableWidgetItem(formatted_purchase_price))  # Display as integer
                self.Purchase_table_widget1.setItem(row_position, 10, QTableWidgetItem(document_id))
            
            # Optionally, resize columns to fit contents
        except Exception as e:
            self.show_custom_message(f"Error loading purchase records: {e}", "Error", is_success=False)
            #pass
####################################################################################################################################
    def create_placeholder_view(self, title):
        """Create a placeholder view for various sections."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        label = QLabel(f"{title} View")
        label.setFont(QFont("Arial", 28, QFont.Bold))
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        return widget
##########################################################################
##########################################################################################
if __name__ == "__main__":
    app=QApplication(sys.argv)
    email="shams@gmail.com"
    window=DashboardWindow(email)
    window.delayed_logout()
    sys.exit(app.exec_())
